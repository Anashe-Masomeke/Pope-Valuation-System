import streamlit as st
import pandas as pd
import numpy as np
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import base64

def step(title: str, number: int):
    st.markdown(
        f"""
        <div class="fbc-step">
            <div class="fbc-step-badge">{number}</div>
            <div class="fbc-step-title">{title}</div>
        </div>
        """,
        unsafe_allow_html=True
    )
def add_watermark():
    logo_path = Path("assets") / "fbc_logo.png"
    if logo_path.exists():
        with open(logo_path, "rb") as f:
            logo_base64 = base64.b64encode(f.read()).decode()

        watermark_css = f"""
        <style>

        /* Make watermark very light */
        .stApp::before {{
            content: "";
            position: fixed;
            top: 40;
            left: 50;
            width: 100%;
            height: 100%;
            background-image: url("data:image/png;base64,{logo_base64}");
            background-repeat: no-repeat;
            background-position: center;
            background-size: 1500px;
            opacity: 0.07;   /* 🔥 control watermark visibility here */
            pointer-events: none;
            z-index: 0;
        }}

        .block-container {{
            position: relative;
            z-index: 1;
        }}
        </style>
        """
        st.markdown(watermark_css, unsafe_allow_html=True)

add_watermark()

# ---------------------------------------------------------
# PAGE CONFIG
# ---------------------------------------------------------
st.set_page_config(page_title="Dividend Discount Model (DDM)", layout="wide")

# ─── FBC DESIGN SYSTEM ─────────────────────────────────────────
st.markdown('''
<style>
/* ================================================================
   FBC INVESTMENT VALUATION SYSTEM — Design System v3.0
   ================================================================ */

/* ── 0. GOOGLE FONTS ──────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700;900&family=EB+Garamond:ital,wght@0,400;0,600;1,400&family=Material+Icons&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined');

/* ── 1. GLOBAL TYPOGRAPHY ─────────────────────────────────── */
html, body, .stApp, .block-container,
p, div, label, span,
h1, h2, h3, h4, h5, h6,
li, ul, ol, a, small,
.stDataFrame, .stTable {
  font-family: "EB Garamond", Georgia, "Times New Roman", serif !important;
  color: #1a1a2e;
}

/* Headings — Playfair Display */
h1, h2, h3, h4, .fbc-heading, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 {
  font-family: "Playfair Display", Georgia, serif !important;
  font-weight: 700 !important;
  letter-spacing: -0.01em !important;
}

/* ── 2. PAGE BACKGROUND ───────────────────────────────────── */
.stApp {
  background: #f5f7fb !important;
}
.main .block-container {
  background: #f5f7fb !important;
  padding-top: 1.5rem !important;
}

/* ── 3. SIDEBAR ───────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: linear-gradient(175deg, #001a5c 0%, #003399 45%, #0044cc 100%) !important;
    border-right: 2px solid rgba(245,180,0,0.25) !important;
    box-shadow: 4px 0 24px rgba(0,26,92,0.35) !important;
}
section[data-testid="stSidebar"]::before {
    content: "";
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #f5b400, #ffd040, #f5b400);
}
section[data-testid="stSidebar"] * {
    color: #e8f0ff !important;
    font-family: "EB Garamond", Georgia, serif !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
    font-family: "Playfair Display", serif !important;
    font-weight: 700 !important;
    text-shadow: 0 1px 4px rgba(0,0,0,0.3) !important;
}
section[data-testid="stSidebar"] .block-container {
    padding-top: 1.5rem !important;
}
section[data-testid="stSidebar"] a {
    color: #b8d0ff !important;
    transition: color 0.15s !important;
}
section[data-testid="stSidebar"] a:hover {
    color: #f5b400 !important;
}

/* sidebar hr divider */
section[data-testid="stSidebar"] hr {
    border: none !important;
    border-top: 1px solid rgba(245,180,0,0.25) !important;
    margin: 12px 0 !important;
}

/* ── 4. SIDEBAR COLLAPSE BUTTON ───────────────────────────── */
.material-icons,
span.material-icons,
i.material-icons,
.material-symbols-outlined,
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stSidebarCollapseButton"] i {
    font-family: "Material Icons", "Material Symbols Outlined" !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}
[data-testid="stSidebarCollapseButton"] button {
    background: linear-gradient(135deg, #f5b400, #ffd040) !important;
    border: none !important;
    border-radius: 50% !important;
    width: 44px !important; height: 44px !important;
    box-shadow: 0 4px 14px rgba(245,180,0,0.50) !important;
    transition: all 0.2s ease !important;
}
[data-testid="stSidebarCollapseButton"] button:hover {
    transform: translateY(-1px) scale(1.06) !important;
    box-shadow: 0 8px 20px rgba(245,180,0,0.60) !important;
}
[data-testid="stSidebarCollapseButton"] svg {
    width: 22px !important; height: 22px !important;
    fill: #001a5c !important;
}

/* ── 5. PAGE HEADER BANNER ────────────────────────────────── */
.fbc-page-header {
    background: linear-gradient(135deg, #001a5c 0%, #003399 50%, #0044cc 100%);
    border-radius: 18px;
    padding: 26px 32px;
    margin-bottom: 28px;
    border: 1px solid rgba(255,255,255,0.08);
    border-bottom: 3px solid #f5b400;
    box-shadow: 0 12px 40px rgba(0,26,92,0.28);
    position: relative;
    overflow: hidden;
}
.fbc-page-header::before {
    content: "";
    position: absolute;
    left: -40px; top: -40px;
    width: 200px; height: 200px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(245,180,0,0.12), transparent 70%);
    pointer-events: none;
}
.fbc-page-header::after {
    content: "";
    position: absolute;
    right: -30px; bottom: -30px;
    width: 160px; height: 160px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(255,255,255,0.06), transparent 65%);
    pointer-events: none;
}
.fbc-page-header-icon {
    font-size: 30px;
    margin-right: 12px;
    vertical-align: middle;
}
.fbc-page-header-title {
    font-family: "Playfair Display", serif !important;
    font-size: 28px !important;
    font-weight: 900 !important;
    color: #ffffff !important;
    display: inline !important;
    vertical-align: middle !important;
    letter-spacing: -0.01em !important;
}
.fbc-page-header-sub {
    font-size: 14px;
    color: rgba(255,255,255,0.78) !important;
    margin-top: 8px;
    font-style: italic;
}
.fbc-badge {
    display: inline-block;
    background: rgba(245,180,0,0.20);
    border: 1px solid rgba(245,180,0,0.50);
    color: #f5c842 !important;
    font-size: 10px;
    font-weight: 800;
    letter-spacing: 0.12em;
    padding: 3px 10px;
    border-radius: 999px;
    margin-left: 10px;
    vertical-align: middle;
    text-transform: uppercase;
    font-family: "EB Garamond", serif !important;
}

/* ── 6. SECTION HEADINGS ──────────────────────────────────── */
.fbc-section-heading {
    display: flex;
    align-items: center;
    gap: 12px;
    margin: 32px 0 16px 0;
    padding-bottom: 10px;
    border-bottom: 2px solid transparent;
    border-image: linear-gradient(90deg, #003399 0%, #f5b400 55%, transparent 90%) 1;
}
.fbc-section-heading-text {
    font-family: "Playfair Display", serif !important;
    font-size: 18px !important;
    font-weight: 700 !important;
    color: #001a5c !important;
    letter-spacing: 0.01em !important;
}
.fbc-section-heading-step {
    background: linear-gradient(135deg, #003399, #0044cc);
    color: white !important;
    font-size: 11px;
    font-weight: 900;
    min-width: 28px; height: 28px;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0;
    box-shadow: 0 3px 8px rgba(0,51,153,0.35);
}
.fbc-subsection-heading {
    font-family: "Playfair Display", serif !important;
    font-size: 15px !important;
    font-weight: 700 !important;
    color: #003399 !important;
    margin: 20px 0 10px 0 !important;
    padding-left: 12px !important;
    border-left: 3px solid #f5b400 !important;
}

/* ── 7. CARD / PANEL ──────────────────────────────────────── */
.fbc-card {
    background: #ffffff;
    border: 1px solid rgba(0,51,153,0.09);
    border-left: 5px solid #003399;
    border-radius: 16px;
    padding: 20px 24px;
    margin-bottom: 18px;
    box-shadow: 0 6px 18px rgba(0,26,92,0.07);
    transition: box-shadow 0.2s, transform 0.2s;
}
.fbc-card:hover {
    box-shadow: 0 12px 32px rgba(0,51,153,0.14);
    transform: translateY(-2px);
}
.fbc-card h3, .fbc-card h4 {
    font-family: "Playfair Display", serif !important;
    color: #001a5c !important;
    margin: 0 0 10px 0 !important;
}
.fbc-card-gold {
    border-left-color: #f5b400 !important;
}
.fbc-card-green {
    border-left-color: #10b981 !important;
}

/* sub-card (nested) */
.fbc-subcard {
    background: rgba(0,51,153,0.03);
    border: 1px solid rgba(0,51,153,0.10);
    border-radius: 12px;
    padding: 16px 18px;
    margin-top: 12px;
}

/* ── 8. KPI METRIC CARDS ──────────────────────────────────── */
.fbc-kpi {
    background: linear-gradient(135deg, #f0f5ff, #fff8e6);
    border: 1px solid rgba(0,51,153,0.12);
    border-radius: 16px;
    padding: 16px 18px;
    text-align: center;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    transition: transform 0.2s, box-shadow 0.2s;
}
.fbc-kpi:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 20px rgba(0,51,153,0.12);
}
.fbc-kpi-label {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 0.10em;
    text-transform: uppercase;
    color: #5a7099 !important;
    margin-bottom: 6px;
    font-family: "EB Garamond", serif !important;
}
.fbc-kpi-value {
    font-family: "Playfair Display", serif !important;
    font-size: 24px !important;
    font-weight: 800 !important;
    color: #001a5c !important;
    line-height: 1.2;
}
.fbc-kpi-unit {
    font-size: 11px;
    color: #7a90b8 !important;
    margin-top: 3px;
}

/* DCF card / kpi aliases */
.dcf-card {
    background: #ffffff;
    border: 1px solid rgba(0,0,0,0.08);
    border-left: 5px solid #003399;
    border-radius: 16px;
    padding: 20px 22px;
    box-shadow: 0 6px 18px rgba(0,0,0,0.06);
    margin-top: 12px;
    margin-bottom: 16px;
}
.dcf-card h3 { margin: 0 0 10px 0; font-family: "Playfair Display", serif !important; color: #001a5c !important; }
.dcf-subcard {
    background: rgba(0,51,153,0.03);
    border: 1px solid rgba(0,51,153,0.10);
    border-radius: 14px;
    padding: 14px;
    margin-top: 10px;
}
.dcf-kpi {
    background: linear-gradient(135deg, rgba(0,51,153,0.08), rgba(245,180,0,0.08));
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: 16px;
    padding: 14px 16px;
    margin: 6px 0;
    transition: transform 0.2s;
}
.dcf-kpi:hover { transform: translateY(-1px); }
.dcf-kpi-title { font-size: 11px; opacity: 0.7; margin-bottom: 3px; text-transform: uppercase; letter-spacing: 0.07em; }
.dcf-kpi-value { font-size: 20px; font-weight: 800; font-family: "Playfair Display", serif !important; color: #001a5c !important; }

/* ── 9. RESET / UTILITY CARDS ─────────────────────────────── */
.fbc-reset-card {
    background: linear-gradient(135deg, #001a5c 0%, #003399 55%, #0055cc 100%);
    padding: 22px 28px;
    border-radius: 16px;
    color: white !important;
    box-shadow: 0 8px 24px rgba(0,26,92,0.30);
    margin-bottom: 22px;
    border-bottom: 3px solid #f5b400;
    position: relative;
    overflow: hidden;
}
.fbc-reset-card::after {
    content: "";
    position: absolute;
    right: -20px; top: -20px;
    width: 120px; height: 120px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(245,180,0,0.15), transparent 70%);
    pointer-events: none;
}
.fbc-reset-title {
    font-family: "Playfair Display", serif !important;
    font-size: 20px !important;
    font-weight: 700 !important;
    margin-bottom: 6px;
    color: white !important;
}
.fbc-reset-sub { font-size: 14px; opacity: 0.88; margin-bottom: 14px; color: rgba(255,255,255,0.85) !important; }

/* ── 10. BUTTONS ──────────────────────────────────────────── */
.stButton > button[kind="primary"],
.stButton > button {
    background: linear-gradient(135deg, #003399, #0044cc) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-family: "EB Garamond", Georgia, serif !important;
    padding: 10px 20px !important;
    transition: all 0.2s ease !important;
    box-shadow: 0 4px 12px rgba(0,51,153,0.28) !important;
    letter-spacing: 0.01em !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #0044cc, #0055ee) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 22px rgba(0,51,153,0.38) !important;
}
.fbc-reset-btn .stButton > button {
    background: linear-gradient(135deg, #f5b400, #ffd040) !important;
    color: #001a5c !important;
    box-shadow: 0 4px 12px rgba(245,180,0,0.32) !important;
}
.fbc-reset-btn .stButton > button:hover {
    background: linear-gradient(135deg, #ffd040, #ffe070) !important;
    box-shadow: 0 8px 20px rgba(245,180,0,0.45) !important;
}
.fbc-nav-btn button,
.fbc-nav-btn .stButton > button {
    background: linear-gradient(135deg, #003399, #001a4d) !important;
    color: white !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    padding: 8px 18px !important;
    border: none !important;
    transition: all 0.25s ease-in-out !important;
}
.fbc-nav-btn button:hover,
.fbc-nav-btn .stButton > button:hover {
    background: linear-gradient(135deg, #0055cc, #003399) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 14px rgba(0,0,0,0.25) !important;
}
.stDownloadButton > button {
    background: linear-gradient(135deg, #003399, #0044cc) !important;
    color: white !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    border: none !important;
    transition: all 0.2s ease !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #0044cc, #0055ee) !important;
    transform: translateY(-1px) !important;
}

/* ── 11. INPUTS ───────────────────────────────────────────── */
.stNumberInput input, .stTextInput input {
    border: 1.5px solid rgba(0,51,153,0.18) !important;
    border-radius: 10px !important;
    font-family: "EB Garamond", Georgia, serif !important;
    background: #ffffff !important;
    transition: border-color 0.15s, box-shadow 0.15s !important;
    padding: 8px 12px !important;
}
.stNumberInput input:focus, .stTextInput input:focus {
    border-color: #003399 !important;
    box-shadow: 0 0 0 3px rgba(0,51,153,0.10) !important;
}
.stSelectbox > div > div {
    border: 1.5px solid rgba(0,51,153,0.18) !important;
    border-radius: 10px !important;
    font-family: "EB Garamond", Georgia, serif !important;
}

/* ── 12. TABS ─────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    border-bottom: 2px solid rgba(0,51,153,0.12);
    padding-bottom: 0;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px 10px 0 0 !important;
    font-weight: 700 !important;
    font-family: "EB Garamond", Georgia, serif !important;
    color: #5a7099 !important;
    padding: 10px 20px !important;
    transition: all 0.15s !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #003399, #0044cc) !important;
    color: white !important;
    box-shadow: 0 -2px 0 0 #f5b400 inset !important;
}

/* ── 13. EXPANDERS ────────────────────────────────────────── */
.streamlit-expanderHeader {
    font-weight: 700 !important;
    color: #001a5c !important;
    font-family: "Playfair Display", Georgia, serif !important;
    border-radius: 10px !important;
    background: rgba(0,51,153,0.04) !important;
    padding: 12px 16px !important;
    border-left: 3px solid #f5b400 !important;
}
/* Fix: hide raw "keyboard_arrow_right" text Streamlit injects into expander headers */
[data-testid="stExpander"] details > summary > div > p > span[style],
[data-testid="stExpander"] details > summary p span[data-testid="stMarkdownContainer"] > p > span,
details > summary p > span[style*="font-weight"] {
    display: none !important;
}
[data-testid="stExpanderToggleIcon"] svg,
details > summary svg {
    font-family: "Material Icons", "Material Symbols Outlined" !important;
}

/* ── 14. METRICS (st.metric) ──────────────────────────────── */
[data-testid="metric-container"] {
    background: linear-gradient(135deg, #f0f5ff, #fff8e6) !important;
    border: 1px solid rgba(0,51,153,0.12) !important;
    border-radius: 16px !important;
    padding: 16px 18px !important;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05) !important;
    transition: transform 0.2s, box-shadow 0.2s !important;
}
[data-testid="metric-container"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 20px rgba(0,51,153,0.10) !important;
}
[data-testid="metric-container"] label {
    font-size: 10px !important;
    font-weight: 700 !important;
    letter-spacing: 0.10em !important;
    text-transform: uppercase !important;
    color: #5a7099 !important;
    font-family: "EB Garamond", serif !important;
}
[data-testid="stMetricValue"] {
    font-family: "Playfair Display", serif !important;
    font-size: 24px !important;
    font-weight: 800 !important;
    color: #001a5c !important;
}
[data-testid="stMetricDelta"] {
    font-family: "EB Garamond", serif !important;
}

/* ── 15. ALERTS & INFO BOXES ──────────────────────────────── */
.stAlert {
    border-radius: 14px !important;
    font-family: "EB Garamond", Georgia, serif !important;
    border-left-width: 4px !important;
}

/* ── 16. PROGRESS BAR ─────────────────────────────────────── */
.stProgress > div > div > div {
    background: linear-gradient(90deg, #003399, #0044cc, #f5b400) !important;
    border-radius: 999px !important;
}

/* ── 17. SCROLLBAR ────────────────────────────────────────── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f0f4ff; border-radius: 999px; }
::-webkit-scrollbar-thumb { background: linear-gradient(#003399, #0044cc); border-radius: 999px; }
::-webkit-scrollbar-thumb:hover { background: #0055ee; }

/* ── 18. FOOTER ───────────────────────────────────────────── */
.fbc-footer {
    text-align: center;
    padding: 28px;
    margin-top: 48px;
    color: #5a7099 !important;
    font-size: 13px;
    border-top: 1px solid rgba(0,51,153,0.10);
    font-style: italic;
}
.fbc-footer b { color: #003399 !important; font-style: normal; }

/* ── 19. FEATURE CARDS (dashboard) ───────────────────────── */
.feature-box {
    background: #ffffff;
    padding: 22px 26px;
    border-radius: 16px;
    border-left: 6px solid #003399;
    box-shadow: 0 4px 14px rgba(0,0,0,0.07);
    transition: all 0.25s ease;
    margin-bottom: 16px;
    font-family: "EB Garamond", serif !important;
}
.feature-box:hover {
    background: linear-gradient(135deg, #f4f8ff, #fffdf0);
    box-shadow: 0 10px 28px rgba(0,51,153,0.16);
    transform: translateY(-3px);
    border-left-color: #f5b400;
}
.feature-icon { font-size: 24px; margin-right: 10px; }

/* ── 20. SMALL UTILITIES ──────────────────────────────────── */
.small-note {
    font-size: 12px;
    color: #7a90b8 !important;
    font-style: italic;
}
.gold-accent { color: #b87c00 !important; font-weight: 700; }
.blue-accent  { color: #003399 !important; font-weight: 700; }

/* ── 21. DATAFRAMES ───────────────────────────────────────── */
.stDataFrame thead th {
    background: linear-gradient(135deg, #001a5c, #003399) !important;
    color: white !important;
    font-weight: 700 !important;
    font-family: "Playfair Display", serif !important;
    letter-spacing: 0.02em !important;
}
.stDataFrame tbody tr:hover td {
    background: rgba(0,51,153,0.04) !important;
}
.stDataFrame tbody tr:nth-child(even) td {
    background: rgba(0,51,153,0.02) !important;
}

/* ── 22. TOP NAV (dashboard) ──────────────────────────────── */
.top-nav {
    position: fixed; top: 0; left: 0; width: 100%; height: 68px;
    background: linear-gradient(90deg, #001a5c, #003399, #0044cc);
    color: white; display: flex; align-items: center;
    padding: 0 28px; z-index: 99999;
    box-shadow: 0 4px 16px rgba(0,0,0,0.30);
    border-bottom: 2px solid #f5b400;
}
.top-title {
    font-family: "Playfair Display", serif !important;
    font-size: 26px; font-weight: 900; margin-left: 16px; letter-spacing: -0.01em;
    color: white;
}

/* ── 23. DIVIDER ──────────────────────────────────────────── */
.fbc-divider {
    height: 2px;
    background: linear-gradient(90deg, transparent, #f5b400, #003399, transparent);
    border: none;
    margin: 28px 0;
    border-radius: 999px;
}

/* ── 24. RADIO / CHECKBOX ─────────────────────────────────── */
.stRadio > label, .stCheckbox > label {
    font-family: "EB Garamond", Georgia, serif !important;
    color: #1a1a2e !important;
}

/* ── 25. SELECTBOX OPTIONS ────────────────────────────────── */
div[data-baseweb="select"] span {
    font-family: "EB Garamond", Georgia, serif !important;
}

/* ── 26. PEER PICKER (Comparables) ───────────────────────── */
.peer-picker-wrap {
    border: 1px solid rgba(0,51,153,0.14);
    border-radius: 18px;
    padding: 18px 18px 12px 18px;
    background: #f8fbff;
    box-shadow: 0 6px 20px rgba(0,26,92,0.07);
    margin-top: 12px;
    margin-bottom: 14px;
}
.peer-picker-head {
    font-family: "Playfair Display", serif !important;
    font-size: 17px;
    font-weight: 700;
    color: #001a5c;
    margin-bottom: 8px;
}
.peer-picker-sub {
    font-size: 13px;
    color: #475569;
    margin-bottom: 14px;
    font-style: italic;
}

/* ── 27. SENS TABLE OVERRIDES ─────────────────────────────── */
.sens-outer { border: 2px solid rgba(0,51,153,0.15) !important; border-radius: 16px !important; }
.sens-table { font-family: "EB Garamond", Georgia, serif !important; }
.sens-table thead th { background: #001a5c !important; }

/* ── 28. FOOTER BANNER ────────────────────────────────────── */
.footer {
    text-align: center;
    padding: 24px;
    margin-top: 40px;
    color: #5a7099 !important;
    font-size: 13px;
    border-top: 1px solid rgba(0,51,153,0.10);
    font-style: italic;
}
.footer b { color: #003399 !important; font-style: normal; }

</style>
''', unsafe_allow_html=True)

st.markdown('''
<div class="fbc-page-header">
    <span class="fbc-page-header-icon">💰</span>
    <span class="fbc-page-header-title">Dividend Discount Model</span>
    <span class="fbc-badge">FBC Securities</span>
    <div class="fbc-page-header-sub">Gordon Growth & Multi-Stage DDM — equity value from dividend streams.</div>
</div>
''', unsafe_allow_html=True)
# ────────────────────────────────────────────────────────────────

# ---------------------------------------------------------
# ✅ FIX SIDEBAR COLLAPSE ARROW (Material Icons)
# ---------------------------------------------------------
st.markdown("""
<style>

/* ✅ Load Material Icons so Streamlit's sidebar collapse icon renders correctly */
@import url('https://fonts.googleapis.com/icon?family=Material+Icons');

/* ✅ Make sure ONLY icons use the Material Icons font (prevents 'keyboard_double_arrow_right' text) */
.material-icons, 
span.material-icons,
i.material-icons,
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stSidebarCollapseButton"] i {
    font-family: 'Material Icons' !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    display: inline-block !important;
    white-space: nowrap !important;
    word-wrap: normal !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}

/* ✅ Style the collapse/expand button nicely */
[data-testid="stSidebarCollapseButton"] button {
    background: #003399 !important;
    border: 1px solid rgba(255,255,255,0.25) !important;
    border-radius: 999px !important;
    width: 44px !important;
    height: 44px !important;
    box-shadow: 0 6px 18px rgba(0, 51, 153, 0.35) !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease !important;
}

[data-testid="stSidebarCollapseButton"] button:hover {
    transform: translateY(-1px) !important;
    background: #0047d6 !important;
    box-shadow: 0 10px 22px rgba(0, 71, 214, 0.35) !important;
}

[data-testid="stSidebarCollapseButton"] svg {
    width: 22px !important;
    height: 22px !important;
    fill: white !important;
}
/* ===== SIDEBAR GLASS STYLE ===== */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #003399 0%, #001a4d 100%) !important;
    color: white !important;
    border-right: 1px solid rgba(255,255,255,0.15);
    backdrop-filter: blur(8px);
}

/* Sidebar text */
section[data-testid="stSidebar"] * {
    color: white !important;
}

/* Sidebar headings */
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
}

/* Remove default sidebar padding spacing */
section[data-testid="stSidebar"] .block-container {
    padding-top: 1rem !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
html, body, .stApp, .block-container,
p, div, label,
h1, h2, h3, h4, h5, h6,
li, ul, ol, a, small {
  font-family: Georgia, "Times New Roman", serif !important;
}

</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>

/* =========================================================
   FBC STEP COMPONENT — FINAL & RELIABLE
   ========================================================= */

.fbc-step {
    display: flex;
    align-items: center;
    gap: 14px;

    padding: 16px 20px;
    margin: 28px 0 18px 0;

    background: linear-gradient(
        135deg,
        rgba(0, 51, 153, 0.06),
        rgba(245, 180, 0, 0.05)
    );

    border-left: 6px solid #003399;
    border-radius: 14px;

    box-shadow: 0 4px 14px rgba(0, 26, 92, 0.08);
}

.fbc-step-badge {
    min-width: 34px;
    height: 34px;
    border-radius: 50%;

    background: linear-gradient(135deg, #003399, #0044cc);
    color: white;

    font-weight: 900;
    font-size: 14px;

    display: flex;
    align-items: center;
    justify-content: center;

    box-shadow: 0 3px 8px rgba(0, 51, 153, 0.35);
}

.fbc-step-title {
    font-family: "Playfair Display", serif !important;
    font-size: 20px;
    font-weight: 800;
    color: #001a5c !important;
    letter-spacing: -0.01em;
}

</style>
""", unsafe_allow_html=True)
st.markdown(
    """
This module values equity using the Gordon Growth DDM:

### **P₀ = D₁ / (Re − g)**  

Where:  
- **D₁** = Dividend next year  
- **Re** = Cost of Equity  
- **g** = Long-term dividend growth rate  
"""
)

# ---------------------------------------------------------
# SMALL HELPERS
# ---------------------------------------------------------
def init(key, value):
    """Initialize a session_state key once."""
    if key not in st.session_state:
        st.session_state[key] = value

# ---------------------------------------------------------
# STEP 1 — DIVIDEND HISTORY
# ---------------------------------------------------------
step("Dividend History", 1)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
# Initialise once if missing
init("ddm_start_year", 2021)
init("ddm_end_year", 2025)

col1, col2 = st.columns(2)

with col1:
    start_year_input = st.number_input(
        "Start Year",
        value=int(st.session_state["ddm_start_year"]),
        step=1,
        key="ddm_start_year_input",
    )
    st.session_state["ddm_start_year"] = int(start_year_input)

with col2:
    end_year_input = st.number_input(
        "End Year",
        value=int(st.session_state["ddm_end_year"]),
        step=1,
        key="ddm_end_year_input",
    )
    st.session_state["ddm_end_year"] = int(end_year_input)

start_year = st.session_state["ddm_start_year"]
end_year = st.session_state["ddm_end_year"]

if start_year > end_year:
    st.error("❌ Start year cannot be greater than end year.")
    st.stop()

years = list(range(start_year, end_year + 1))

# Persistent dividend storage per year
for y in years:
    if f"ddm_div_{y}" not in st.session_state:
        st.session_state[f"ddm_div_{y}"] = 0.01  # default once

step("Enter Dividends",2)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
dividends = []
for y in years:
    div = st.number_input(
        f"Dividend for {y}",
        value=float(st.session_state[f"ddm_div_{y}"]),
        step=0.00001,
        format="%.5f",
        key=f"ddm_div_input_{y}",
    )
    st.session_state[f"ddm_div_{y}"] = div
    dividends.append(div)

# Store full dividend history for AI / summary pages
st.session_state["ddm_dividends"] = {
    str(y): float(d) for y, d in zip(years, dividends)
}

# Display table
df_history = pd.DataFrame({"Year": years, "Dividend": dividends})
st.dataframe(df_history, width='stretch')

# ---------------------------------------------------------
# STEP 2 — GROWTH CALCULATION RANGE
# ---------------------------------------------------------
step("Growth Calculation Range", 3)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
init("ddm_g_start", years[0])
init("ddm_g_end", years[-1])

c1, c2 = st.columns(2)
with c1:
    g_start = st.selectbox("Growth start year:", years, key="ddm_g_start")
with c2:
    g_end = st.selectbox("Growth end year:", years, key="ddm_g_end")

if g_start > g_end:
    st.error("❌ Growth start year must be earlier or equal to end year.")
    st.stop()

D_start = dividends[years.index(g_start)]
D_end = dividends[years.index(g_end)]

# ---------------------------------------------------------
# STEP 3 — DIVIDEND GROWTH RATE (g)
# ---------------------------------------------------------
step("Dividend Growth", 4)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
if g_start == g_end:
    g = 0.0
elif D_start > 0:
    # CAGR between selected years
    g = (D_end / D_start) ** (1 / (g_end - g_start)) - 1
else:
    # Fallback if starting dividend is zero
    g = 0.02

st.success(f"Growth rate (g): **{g:.2%}**")

D1 = D_end * (1 + g)
st.metric("Next year's dividend (D₁)", f"{D1:,.5f}")

# ---------------------------------------------------------
# STEP 4 — COST OF EQUITY (Re)
# ---------------------------------------------------------
step("Cost of Equity Inputs", 5)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
# Pull live values from DCF page where possible
rf = st.session_state.get("dcf_rf_pct", st.session_state.get("rf", 0.0)) / 100
mrp = st.session_state.get("dcf_mrp_pct", st.session_state.get("erp", 0.0)) / 100
tax_rate = (
    st.session_state.get("dcf_tax_pct", st.session_state.get("tax_rate", 0.0)) / 100
)
unlevered_beta = st.session_state.get(
    "dcf_unlevered_beta", st.session_state.get("unlevered_beta", 0.0)
)
de_ratio = st.session_state.get("de_ratio", 0.0)

# Store back normalised keys
st.session_state["rf"] = rf
st.session_state["erp"] = mrp
st.session_state["tax_rate"] = tax_rate
st.session_state["unlevered_beta"] = unlevered_beta

st.write("Using parameters loaded from the DCF page (you can override them below).")

use_custom = st.checkbox(
    "Manually override parameters",
    value=st.session_state.get("ddm_use_custom_params", False),
    key="ddm_use_custom_params",
)

if use_custom:
    cA, cB = st.columns(2)

    with cA:
        unlevered_beta = st.number_input(
            "Unlevered Beta",
            value=float(unlevered_beta),
            step=0.001,
            format="%.4f",
            key="ddm_unlevered_beta",
        )

        de_ratio = st.number_input(
            "Debt/Equity Ratio (D/E)",
            value=float(de_ratio),
            step=0.001,
            format="%.4f",
            key="ddm_de_ratio",
        )

    with cB:
        tax_rate = (
            st.number_input(
                "Tax Rate (%)",
                value=float(tax_rate * 100),
                step=0.01,
                format="%.2f",
                key="ddm_tax_rate",
            )
            / 100
        )

        rf = (
            st.number_input(
                "Risk-Free Rate (%)",
                value=float(rf * 100),
                step=0.01,
                format="%.2f",
                key="ddm_rf",
            )
            / 100
        )

        mrp = (
            st.number_input(
                "Equity Risk Premium (%)",
                value=float(mrp * 100),
                step=0.01,
                format="%.2f",
                key="ddm_erp",
            )
            / 100
        )

# Save final values
st.session_state["rf"] = rf
st.session_state["erp"] = mrp
st.session_state["tax_rate"] = tax_rate
st.session_state["unlevered_beta"] = unlevered_beta
st.session_state["de_ratio"] = de_ratio

# CAPM Re
levered_beta = unlevered_beta * (1 + (1 - tax_rate) * de_ratio)
Re = rf + levered_beta * mrp

st.metric("Levered Beta", f"{levered_beta:.4f}")
st.metric("Cost of Equity (Re)", f"{Re * 100:.2f}%")

# ---------------------------------------------------------
# STEP 5 — VALUE PER SHARE
# ---------------------------------------------------------
step("Equity Value per Share", 6)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
if Re <= g:
    st.error("❌ Re must be greater than g for the Gordon Growth DDM to work.")
    P0 = np.nan
else:
    P0 = D1 / (Re - g)
    st.success(f"Equity Value / Share = **{P0:,.4f} USD**")

# Store for AI / summary pages
st.session_state["ddm_g"] = float(g)
st.session_state["ddm_Re"] = float(Re)
st.session_state["ddm_P0"] = None if np.isnan(P0) else float(P0)

# ---------------------------------------------------------
# STEP 6 — TOTAL EQUITY VALUE
# ---------------------------------------------------------
step("Total Equity Value", 7)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
init("num_shares", 0.0)
st.markdown("""
<div class="fbc-forecast-label">
    Number Of Shares
</div>

<style>
.fbc-forecast-label {
    font-family: "Playfair Display", serif;
    font-size: 18px;
    font-weight: 800;
    color: #001a5c;

    display: inline-block;
    padding-left: 10px;
    border-left: 4px solid #003399;

    margin-bottom: 8px;
}

.fbc-forecast-label::after {
    content: "";
    display: block;
    width: 60px;
    height: 2px;
    margin-top: 6px;
    background: linear-gradient(90deg, #003399, transparent);
    border-radius: 2px;
}
</style>
""", unsafe_allow_html=True)
num_shares = st.number_input(
    "Number of Shares",
    value=float(st.session_state["num_shares"]),
    step=1000.0,
    format="%.0f",
    key="ddm_num_shares",
    label_visibility="collapsed"
)

if num_shares > 0 and not np.isnan(P0):
    equity_value = P0 * num_shares
    st.success(f"Total Equity Value = **{equity_value:,.2f} USD**")

    st.session_state["num_shares"] = float(num_shares)
    st.session_state["equity_value_ddm"] = float(equity_value)
else:
    st.warning("Enter a valid number of shares to compute total equity value.")

# =========================================================
# ✅ FULL DDM EXCEL EXPORT (ALWAYS VISIBLE)
# =========================================================

def _excel_col(n: int) -> str:
    return get_column_letter(n)

def build_full_ddm_excel_model(
    years, dividends,
    g_start, g_end,
    rf, mrp, tax_rate,
    unlevered_beta, de_ratio,
    num_shares,
):
    wb = Workbook()

    BLUE = "003399"
    DARK = "071426"
    LIGHT_BG = "F7FAFF"
    GRID = "D9E2EF"

    thin = Side(style="thin", color=GRID)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_title(ws, title, end_col=6):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        c = ws.cell(1, 1, title)
        c.font = Font(bold=True, color="FFFFFF", size=14)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

    def style_header(ws, r, c1, c2):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
        ws.row_dimensions[r].height = 20

    # =========================================================
    # SHEET 1: Dividend History
    # =========================================================
    wsH = wb.active
    wsH.title = "DividendHistory"
    style_title(wsH, "DDM - Step 1: Dividend History", end_col=4)

    wsH["A3"], wsH["B3"] = "Year", "Dividend"
    style_header(wsH, 3, 1, 2)

    r0 = 4
    for i, (y, d) in enumerate(zip(years, dividends)):
        r = r0 + i
        wsH.cell(r, 1, int(y)).border = border_all
        wsH.cell(r, 2, float(d)).border = border_all
        wsH.cell(r, 2).number_format = "0.00000"

    wsH.column_dimensions["A"].width = 10
    wsH.column_dimensions["B"].width = 16
    wsH.freeze_panes = "A4"

    last_row = r0 + len(years) - 1

    # =========================================================
    # SHEET 2: Growth
    # =========================================================
    wsG = wb.create_sheet("Growth")
    style_title(wsG, "DDM - Steps 2 & 3: Growth Range & g", end_col=6)

    wsG["A3"], wsG["B3"] = "Input", "Value"
    style_header(wsG, 3, 1, 2)

    wsG["A4"], wsG["B4"] = "Growth start year", int(g_start)
    wsG["A5"], wsG["B5"] = "Growth end year", int(g_end)

    wsG["A7"], wsG["B7"] = "D_start", (
        '=INDEX(DividendHistory!$B$4:$B$%d, MATCH($B$4, DividendHistory!$A$4:$A$%d, 0))'
        % (last_row, last_row)
    )
    wsG["A8"], wsG["B8"] = "D_end", (
        '=INDEX(DividendHistory!$B$4:$B$%d, MATCH($B$5, DividendHistory!$A$4:$A$%d, 0))'
        % (last_row, last_row)
    )

    wsG["A10"], wsG["B10"] = "Growth rate (g)", (
        '=IF($B$4=$B$5,0,IF($B$7>0,POWER($B$8/$B$7,1/($B$5-$B$4))-1,0.02))'
    )
    wsG["B10"].number_format = "0.00%"

    wsG["A11"], wsG["B11"] = "Next dividend (D1)", "=$B$8*(1+$B$10)"
    wsG["B11"].number_format = "0.00000"

    wsG.column_dimensions["A"].width = 22
    wsG.column_dimensions["B"].width = 28
    wsG.freeze_panes = "A4"

    # =========================================================
    # SHEET 3: Parameters (CAPM)
    # =========================================================
    wsP = wb.create_sheet("Parameters")
    style_title(wsP, "DDM - Step 4: Cost of Equity (CAPM)", end_col=6)

    wsP["A3"], wsP["B3"] = "Parameter", "Value"
    style_header(wsP, 3, 1, 2)

    wsP["A4"], wsP["B4"] = "Risk-free rate (RF)", float(rf)
    wsP["A5"], wsP["B5"] = "Equity risk premium (MRP)", float(mrp)
    wsP["A6"], wsP["B6"] = "Tax rate", float(tax_rate)
    wsP["A7"], wsP["B7"] = "Unlevered beta (βu)", float(unlevered_beta)
    wsP["A8"], wsP["B8"] = "Debt/Equity (D/E)", float(de_ratio)

    wsP["B4"].number_format = "0.00%"
    wsP["B5"].number_format = "0.00%"
    wsP["B6"].number_format = "0.00%"
    wsP["B7"].number_format = "0.0000"
    wsP["B8"].number_format = "0.0000"

    wsP["A10"], wsP["B10"] = "Levered beta (βL)", "=$B$7*(1+(1-$B$6)*$B$8)"
    wsP["B10"].number_format = "0.0000"

    wsP["A11"], wsP["B11"] = "Cost of Equity (Re)", "=$B$4 + $B$10*$B$5"
    wsP["B11"].number_format = "0.00%"

    wsP.column_dimensions["A"].width = 26
    wsP.column_dimensions["B"].width = 18
    wsP.freeze_panes = "A4"

    # =========================================================
    # SHEET 4: Valuation
    # =========================================================
    wsV = wb.create_sheet("Valuation")
    style_title(wsV, "DDM - Steps 5 & 6: Valuation", end_col=6)

    wsV["A3"], wsV["B3"] = "Metric", "Value"
    style_header(wsV, 3, 1, 2)

    wsV["A4"], wsV["B4"] = "g (from Growth sheet)", "=Growth!$B$10"
    wsV["A5"], wsV["B5"] = "D1 (from Growth sheet)", "=Growth!$B$11"
    wsV["A6"], wsV["B6"] = "Re (from Parameters)", "=Parameters!$B$11"
    wsV["B4"].number_format = "0.00%"
    wsV["B5"].number_format = "0.00000"
    wsV["B6"].number_format = "0.00%"

    wsV["A8"], wsV["B8"] = "Equity Value / Share (P0)", "=IF($B$6<=$B$4,NA(),$B$5/($B$6-$B$4))"
    wsV["B8"].number_format = "#,##0.0000"

    wsV["A10"], wsV["B10"] = "Number of shares", float(num_shares)
    wsV["B10"].number_format = "#,##0"

    wsV["A11"], wsV["B11"] = "Total Equity Value", "=IF(ISNUMBER($B$8),$B$8*$B$10,NA())"
    wsV["B11"].number_format = "#,##0.00"

    wsV.column_dimensions["A"].width = 28
    wsV.column_dimensions["B"].width = 22
    wsV.freeze_panes = "A4"

    # =========================================================
    # SHEET 5: Summary
    # =========================================================
    wsS = wb.create_sheet("Summary")
    style_title(wsS, "DDM Summary", end_col=6)

    wsS["A3"], wsS["B3"], wsS["C3"] = "Metric", "Value", "Unit"
    style_header(wsS, 3, 1, 3)

    rows = [
        ("Growth rate (g)", "=Growth!$B$10", "%"),
        ("Next dividend (D1)", "=Growth!$B$11", "USD"),
        ("Cost of Equity (Re)", "=Parameters!$B$11", "%"),
        ("Value per share (P0)", "=Valuation!$B$8", "USD"),
        ("Number of shares", "=Valuation!$B$10", "shares"),
        ("Total equity value", "=Valuation!$B$11", "USD"),
    ]

    r0 = 4
    for i, (m, v, u) in enumerate(rows):
        r = r0 + i
        wsS.cell(r, 1, m).border = border_all
        wsS.cell(r, 2, v).border = border_all
        wsS.cell(r, 3, u).border = border_all
        if u == "USD":
            wsS.cell(r, 2).number_format = "#,##0.00"
        elif u == "%":
            wsS.cell(r, 2).number_format = "0.00%"
        elif u == "shares":
            wsS.cell(r, 2).number_format = "#,##0"

    wsS.column_dimensions["A"].width = 26
    wsS.column_dimensions["B"].width = 18
    wsS.column_dimensions["C"].width = 10
    wsS.freeze_panes = "A4"

    return wb

def workbook_to_bytes(wb: Workbook) -> bytes:
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

st.markdown("---")
step("⬇️ Download FULL DDM Excel Model (All Steps + Formulas)",8)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
if "ddm_excel_bytes" not in st.session_state:
    st.session_state["ddm_excel_bytes"] = None

if st.button("📥 Generate / Update FULL DDM Excel Model", key="ddm_generate_excel"):
    wb = build_full_ddm_excel_model(
        years=years,
        dividends=dividends,
        g_start=int(g_start),
        g_end=int(g_end),
        rf=float(rf),
        mrp=float(mrp),
        tax_rate=float(tax_rate),
        unlevered_beta=float(unlevered_beta),
        de_ratio=float(de_ratio),
        num_shares=float(num_shares),
    )
    st.session_state["ddm_excel_bytes"] = workbook_to_bytes(wb)

st.download_button(
    "⬇️ Download FULL_DDM_Model.xlsx",
    data=st.session_state["ddm_excel_bytes"] or b"",
    file_name="FULL_DDM_Model.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    disabled=st.session_state["ddm_excel_bytes"] is None,
    key="ddm_download_excel",
)
