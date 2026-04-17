import streamlit as st
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import date
import io
from pandas.io.formats.style import Styler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import base64
# ---------------------------------------------------------
# STREAMLIT APP
# ---------------------------------------------------------
st.set_page_config(
    page_title="Forecast + DCF (IS + BS + CF)",
    layout="wide"
)

# ─── FBC DESIGN SYSTEM ─────────────────────────────────────────
st.markdown("""
<style>
/* ── FIX: Hide raw Material Icon text in expander headers ── */
[data-testid="stExpander"] summary p,
[data-testid="stExpander"] details summary p {
    display: none !important;
}

/* Re-show only the actual text content inside the summary */
[data-testid="stExpander"] summary [data-testid="stMarkdownContainer"] p {
    display: block !important;
}

/* Target the specific span that renders the raw icon text */
details > summary > div > p > span,
[data-testid="stExpander"] details > summary span[style*="font-family"] {
    display: none !important;
}

/* Ensure the SVG arrow icon still shows */
[data-testid="stExpanderToggleIcon"] {
    display: flex !important;
}
</style>
""", unsafe_allow_html=True)
st.markdown('''
<style>
/* ================================================================
   FBC INVESTMENT VALUATION SYSTEM — Design System v3.0
   ================================================================ */

/* ── 0. GOOGLE FONTS ──────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700;900&family=EB+Garamond:ital,wght@0,400;0,600;1,400&family=Material+Icons&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined');

/* ── 1. GLOBAL TYPOGRAPHY ─────────────────────────────────── */
html, body, .stApp, .block-container,
p, div, label,
h1, h2, h3, h4, h5, h6,
li, ul, ol, a, small,
.stDataFrame, .stTable {
  font-family: "EB Garamond", Georgia, "Times New Roman", serif !important;
  color: #1a1a2e;
}

/* Headings — Playfair Display */
h1, h2, h3, h4, .fbc-heading, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 {
  font-family: "Playfair Display", Georgia, serif !important;
  font-weight: 700 !important;
  letter-spacing: -0.01em !important;
}

/* ── 2. PAGE BACKGROUND ───────────────────────────────────── */
.stApp {
  background: #f5f7fb !important;
}
.main .block-container {
  background: #f5f7fb !important;
  padding-top: 1.5rem !important;
}

/* ── 3. SIDEBAR ───────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: linear-gradient(175deg, #001a5c 0%, #003399 45%, #0044cc 100%) !important;
    border-right: 2px solid rgba(245,180,0,0.25) !important;
    box-shadow: 4px 0 24px rgba(0,26,92,0.35) !important;
}
section[data-testid="stSidebar"]::before {
    content: "";
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #f5b400, #ffd040, #f5b400);
}
section[data-testid="stSidebar"] * {
    color: #e8f0ff !important;
    font-family: "EB Garamond", Georgia, serif !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
    font-family: "Playfair Display", serif !important;
    font-weight: 700 !important;
    text-shadow: 0 1px 4px rgba(0,0,0,0.3) !important;
}
section[data-testid="stSidebar"] .block-container {
    padding-top: 1.5rem !important;
}
section[data-testid="stSidebar"] a {
    color: #b8d0ff !important;
    transition: color 0.15s !important;
}
section[data-testid="stSidebar"] a:hover {
    color: #f5b400 !important;
}

/* sidebar hr divider */
section[data-testid="stSidebar"] hr {
    border: none !important;
    border-top: 1px solid rgba(245,180,0,0.25) !important;
    margin: 12px 0 !important;
}

/* ✅ Ensure Material Icons render correctly */
.material-icons,
.material-icons-outlined,
.material-symbols-outlined,
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stSidebarCollapseButton"] i {
    font-family: 'Material Icons', 'Material Symbols Outlined' !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    display: inline-block !important;
    white-space: nowrap !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}

[data-testid="stSidebarCollapseButton"] button {
    background: linear-gradient(135deg, #f5b400, #ffd040) !important;
    border: none !important;
    border-radius: 50% !important;
    width: 44px !important; height: 44px !important;
    box-shadow: 0 4px 14px rgba(245,180,0,0.50) !important;
    transition: all 0.2s ease !important;
}
[data-testid="stSidebarCollapseButton"] button:hover {
    transform: translateY(-1px) scale(1.06) !important;
    box-shadow: 0 8px 20px rgba(245,180,0,0.60) !important;
}
[data-testid="stSidebarCollapseButton"] svg {
    width: 22px !important; height: 22px !important;
    fill: #001a5c !important;
}
/* =========================================================
   🔥 PREMIUM SIDEBAR COLLAPSE BUTTON (FBC STYLE)
   ========================================================= */

/* Button container */
[data-testid="stSidebarCollapseButton"] {
    position: relative;
    margin-top: 10px;
}

/* Main button */
[data-testid="stSidebarCollapseButton"] button {
    background: linear-gradient(135deg, #003399, #0055ee) !important;
    border: none !important;
    border-radius: 50% !important;
    width: 46px !important;
    height: 46px !important;

    box-shadow: 0 6px 18px rgba(0, 51, 153, 0.35) !important;

    display: flex !important;
    align-items: center !important;
    justify-content: center !important;

    transition: all 0.25s ease !important;
}


/* ── 5. PAGE HEADER BANNER ────────────────────────────────── */
.fbc-page-header {
    background: linear-gradient(135deg, #001a5c 0%, #003399 50%, #0044cc 100%);
    border-radius: 18px;
    padding: 26px 32px;
    margin-bottom: 28px;
    border: 1px solid rgba(255,255,255,0.08);
    border-bottom: 3px solid #f5b400;
    box-shadow: 0 12px 40px rgba(0,26,92,0.28);
    position: relative;
    overflow: hidden;
}
.fbc-page-header::before {
    content: "";
    position: absolute;
    left: -40px; top: -40px;
    width: 200px; height: 200px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(245,180,0,0.12), transparent 70%);
    pointer-events: none;
}
.fbc-page-header::after {
    content: "";
    position: absolute;
    right: -30px; bottom: -30px;
    width: 160px; height: 160px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(255,255,255,0.06), transparent 65%);
    pointer-events: none;
}
.fbc-page-header-icon {
    font-size: 30px;
    margin-right: 12px;
    vertical-align: middle;
}
.fbc-page-header-title {
    font-family: "Playfair Display", serif !important;
    font-size: 28px !important;
    font-weight: 900 !important;
    color: #ffffff !important;
    display: inline !important;
    vertical-align: middle !important;
    letter-spacing: -0.01em !important;
}
.fbc-page-header-sub {
    font-size: 14px;
    color: rgba(255,255,255,0.78) !important;
    margin-top: 8px;
    font-style: italic;
}
.fbc-badge {
    display: inline-block;
    background: rgba(245,180,0,0.20);
    border: 1px solid rgba(245,180,0,0.50);
    color: #f5c842 !important;
    font-size: 10px;
    font-weight: 800;
    letter-spacing: 0.12em;
    padding: 3px 10px;
    border-radius: 999px;
    margin-left: 10px;
    vertical-align: middle;
    text-transform: uppercase;
    font-family: "EB Garamond", serif !important;
}

/* ── 6. SECTION HEADINGS ──────────────────────────────────── */
.fbc-section-heading {
    display: flex;
    align-items: center;
    gap: 12px;
    margin: 32px 0 16px 0;
    padding-bottom: 10px;
    border-bottom: 2px solid transparent;
    border-image: linear-gradient(90deg, #003399 0%, #f5b400 55%, transparent 90%) 1;
}
.fbc-section-heading-text {
    font-family: "Playfair Display", serif !important;
    font-size: 18px !important;
    font-weight: 700 !important;
    color: #001a5c !important;
    letter-spacing: 0.01em !important;
}
.fbc-section-heading-step {
    background: linear-gradient(135deg, #003399, #0044cc);
    color: white !important;
    font-size: 11px;
    font-weight: 900;
    min-width: 28px; height: 28px;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0;
    box-shadow: 0 3px 8px rgba(0,51,153,0.35);
}
.fbc-subsection-heading {
    font-family: "Playfair Display", serif !important;
    font-size: 15px !important;
    font-weight: 700 !important;
    color: #003399 !important;
    margin: 20px 0 10px 0 !important;
    padding-left: 12px !important;
    border-left: 3px solid #f5b400 !important;
}

/* ── 7. CARD / PANEL ──────────────────────────────────────── */
.fbc-card {
    background: #ffffff;
    border: 1px solid rgba(0,51,153,0.09);
    border-left: 5px solid #003399;
    border-radius: 16px;
    padding: 20px 24px;
    margin-bottom: 18px;
    box-shadow: 0 6px 18px rgba(0,26,92,0.07);
    transition: box-shadow 0.2s, transform 0.2s;
}
.fbc-card:hover {
    box-shadow: 0 12px 32px rgba(0,51,153,0.14);
    transform: translateY(-2px);
}
.fbc-card h3, .fbc-card h4 {
    font-family: "Playfair Display", serif !important;
    color: #001a5c !important;
    margin: 0 0 10px 0 !important;
}
.fbc-card-gold {
    border-left-color: #f5b400 !important;
}
.fbc-card-green {
    border-left-color: #10b981 !important;
}

/* sub-card (nested) */
.fbc-subcard {
    background: rgba(0,51,153,0.03);
    border: 1px solid rgba(0,51,153,0.10);
    border-radius: 12px;
    padding: 16px 18px;
    margin-top: 12px;
}

/* ── 8. KPI METRIC CARDS ──────────────────────────────────── */
.fbc-kpi {
    background: linear-gradient(135deg, #f0f5ff, #fff8e6);
    border: 1px solid rgba(0,51,153,0.12);
    border-radius: 16px;
    padding: 16px 18px;
    text-align: center;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    transition: transform 0.2s, box-shadow 0.2s;
}
.fbc-kpi:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 20px rgba(0,51,153,0.12);
}
.fbc-kpi-label {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 0.10em;
    text-transform: uppercase;
    color: #5a7099 !important;
    margin-bottom: 6px;
    font-family: "EB Garamond", serif !important;
}
.fbc-kpi-value {
    font-family: "Playfair Display", serif !important;
    font-size: 24px !important;
    font-weight: 800 !important;
    color: #001a5c !important;
    line-height: 1.2;
}
.fbc-kpi-unit {
    font-size: 11px;
    color: #7a90b8 !important;
    margin-top: 3px;
}

/* DCF card / kpi aliases */
.dcf-card {
    background: #ffffff;
    border: 1px solid rgba(0,0,0,0.08);
    border-left: 5px solid #003399;
    border-radius: 16px;
    padding: 20px 22px;
    box-shadow: 0 6px 18px rgba(0,0,0,0.06);
    margin-top: 12px;
    margin-bottom: 16px;
}
.dcf-card h3 { margin: 0 0 10px 0; font-family: "Playfair Display", serif !important; color: #001a5c !important; }
.dcf-subcard {
    background: rgba(0,51,153,0.03);
    border: 1px solid rgba(0,51,153,0.10);
    border-radius: 14px;
    padding: 14px;
    margin-top: 10px;
}
.dcf-kpi {
    background: linear-gradient(135deg, rgba(0,51,153,0.08), rgba(245,180,0,0.08));
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: 16px;
    padding: 14px 16px;
    margin: 6px 0;
    transition: transform 0.2s;
}
.dcf-kpi:hover { transform: translateY(-1px); }
.dcf-kpi-title { font-size: 11px; opacity: 0.7; margin-bottom: 3px; text-transform: uppercase; letter-spacing: 0.07em; }
.dcf-kpi-value { font-size: 20px; font-weight: 800; font-family: "Playfair Display", serif !important; color: #001a5c !important; }

/* ── 9. RESET / UTILITY CARDS ─────────────────────────────── */
.fbc-reset-card {
    background: linear-gradient(135deg, #001a5c 0%, #003399 55%, #0055cc 100%);
    padding: 22px 28px;
    border-radius: 16px;
    color: white !important;
    box-shadow: 0 8px 24px rgba(0,26,92,0.30);
    margin-bottom: 22px;
    border-bottom: 3px solid #f5b400;
    position: relative;
    overflow: hidden;
}
.fbc-reset-card::after {
    content: "";
    position: absolute;
    right: -20px; top: -20px;
    width: 120px; height: 120px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(245,180,0,0.15), transparent 70%);
    pointer-events: none;
}
.fbc-reset-title {
    font-family: "Playfair Display", serif !important;
    font-size: 20px !important;
    font-weight: 700 !important;
    margin-bottom: 6px;
    color: white !important;
}
.fbc-reset-sub { font-size: 14px; opacity: 0.88; margin-bottom: 14px; color: rgba(255,255,255,0.85) !important; }

/* ── 10. BUTTONS ──────────────────────────────────────────── */
.stButton > button[kind="primary"],
.stButton > button {
    background: linear-gradient(135deg, #003399, #0044cc) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-family: "EB Garamond", Georgia, serif !important;
    padding: 10px 20px !important;
    transition: all 0.2s ease !important;
    box-shadow: 0 4px 12px rgba(0,51,153,0.28) !important;
    letter-spacing: 0.01em !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #0044cc, #0055ee) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 22px rgba(0,51,153,0.38) !important;
}
.fbc-reset-btn .stButton > button {
    background: linear-gradient(135deg, #f5b400, #ffd040) !important;
    color: #001a5c !important;
    box-shadow: 0 4px 12px rgba(245,180,0,0.32) !important;
}
.fbc-reset-btn .stButton > button:hover {
    background: linear-gradient(135deg, #ffd040, #ffe070) !important;
    box-shadow: 0 8px 20px rgba(245,180,0,0.45) !important;
}
.fbc-nav-btn button,
.fbc-nav-btn .stButton > button {
    background: linear-gradient(135deg, #003399, #001a4d) !important;
    color: white !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    padding: 8px 18px !important;
    border: none !important;
    transition: all 0.25s ease-in-out !important;
}
.fbc-nav-btn button:hover,
.fbc-nav-btn .stButton > button:hover {
    background: linear-gradient(135deg, #0055cc, #003399) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 14px rgba(0,0,0,0.25) !important;
}
.stDownloadButton > button {
    background: linear-gradient(135deg, #003399, #0044cc) !important;
    color: white !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    border: none !important;
    transition: all 0.2s ease !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #0044cc, #0055ee) !important;
    transform: translateY(-1px) !important;
}

/* ── 11. INPUTS ───────────────────────────────────────────── */
.stNumberInput input, .stTextInput input {
    border: 1.5px solid rgba(0,51,153,0.18) !important;
    border-radius: 10px !important;
    font-family: "EB Garamond", Georgia, serif !important;
    background: #ffffff !important;
    transition: border-color 0.15s, box-shadow 0.15s !important;
    padding: 8px 12px !important;
}
.stNumberInput input:focus, .stTextInput input:focus {
    border-color: #003399 !important;
    box-shadow: 0 0 0 3px rgba(0,51,153,0.10) !important;
}
.stSelectbox > div > div {
    border: 1.5px solid rgba(0,51,153,0.18) !important;
    border-radius: 10px !important;
    font-family: "EB Garamond", Georgia, serif !important;
}

/* ── 12. TABS ─────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    border-bottom: 2px solid rgba(0,51,153,0.12);
    padding-bottom: 0;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px 10px 0 0 !important;
    font-weight: 700 !important;
    font-family: "EB Garamond", Georgia, serif !important;
    color: #5a7099 !important;
    padding: 10px 20px !important;
    transition: all 0.15s !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #003399, #0044cc) !important;
    color: white !important;
    box-shadow: 0 -2px 0 0 #f5b400 inset !important;
}

/* ── 13. EXPANDERS ────────────────────────────────────────── */
.streamlit-expanderHeader {
    font-weight: 700 !important;
    color: #001a5c !important;
    font-family: "Playfair Display", Georgia, serif !important;
    border-radius: 10px !important;
    background: rgba(0,51,153,0.04) !important;
    padding: 12px 16px !important;
    border-left: 3px solid #f5b400 !important;
}
/* Fix: hide raw "keyboard_arrow_right" text Streamlit injects into expander headers */
[data-testid="stExpander"] details > summary > div > p > span[style],
[data-testid="stExpander"] details > summary p span[data-testid="stMarkdownContainer"] > p > span,
details > summary p > span[style*="font-weight"] {
    display: none !important;
}
[data-testid="stExpanderToggleIcon"] svg,
details > summary svg {
    font-family: "Material Icons", "Material Symbols Outlined" !important;
}

/* ── 14. METRICS (st.metric) ──────────────────────────────── */
[data-testid="metric-container"] {
    background: linear-gradient(135deg, #f0f5ff, #fff8e6) !important;
    border: 1px solid rgba(0,51,153,0.12) !important;
    border-radius: 16px !important;
    padding: 16px 18px !important;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05) !important;
    transition: transform 0.2s, box-shadow 0.2s !important;
}
[data-testid="metric-container"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 20px rgba(0,51,153,0.10) !important;
}
[data-testid="metric-container"] label {
    font-size: 10px !important;
    font-weight: 700 !important;
    letter-spacing: 0.10em !important;
    text-transform: uppercase !important;
    color: #5a7099 !important;
    font-family: "EB Garamond", serif !important;
}
[data-testid="stMetricValue"] {
    font-family: "Playfair Display", serif !important;
    font-size: 24px !important;
    font-weight: 800 !important;
    color: #001a5c !important;
}
[data-testid="stMetricDelta"] {
    font-family: "EB Garamond", serif !important;
}

/* ── 15. ALERTS & INFO BOXES ──────────────────────────────── */
.stAlert {
    border-radius: 14px !important;
    font-family: "EB Garamond", Georgia, serif !important;
    border-left-width: 4px !important;
}

/* ── 16. PROGRESS BAR ─────────────────────────────────────── */
.stProgress > div > div > div {
    background: linear-gradient(90deg, #003399, #0044cc, #f5b400) !important;
    border-radius: 999px !important;
}

/* ── 17. SCROLLBAR ────────────────────────────────────────── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f0f4ff; border-radius: 999px; }
::-webkit-scrollbar-thumb { background: linear-gradient(#003399, #0044cc); border-radius: 999px; }
::-webkit-scrollbar-thumb:hover { background: #0055ee; }

/* ── 18. FOOTER ───────────────────────────────────────────── */
.fbc-footer {
    text-align: center;
    padding: 28px;
    margin-top: 48px;
    color: #5a7099 !important;
    font-size: 13px;
    border-top: 1px solid rgba(0,51,153,0.10);
    font-style: italic;
}
.fbc-footer b { color: #003399 !important; font-style: normal; }

/* ── 19. FEATURE CARDS (dashboard) ───────────────────────── */
.feature-box {
    background: #ffffff;
    padding: 22px 26px;
    border-radius: 16px;
    border-left: 6px solid #003399;
    box-shadow: 0 4px 14px rgba(0,0,0,0.07);
    transition: all 0.25s ease;
    margin-bottom: 16px;
    font-family: "EB Garamond", serif !important;
}
.feature-box:hover {
    background: linear-gradient(135deg, #f4f8ff, #fffdf0);
    box-shadow: 0 10px 28px rgba(0,51,153,0.16);
    transform: translateY(-3px);
    border-left-color: #f5b400;
}
.feature-icon { font-size: 24px; margin-right: 10px; }

/* ── 20. SMALL UTILITIES ──────────────────────────────────── */
.small-note {
    font-size: 12px;
    color: #7a90b8 !important;
    font-style: italic;
}
.gold-accent { color: #b87c00 !important; font-weight: 700; }
.blue-accent  { color: #003399 !important; font-weight: 700; }

/* ── 21. DATAFRAMES ───────────────────────────────────────── */
.stDataFrame thead th {
    background: linear-gradient(135deg, #001a5c, #003399) !important;
    color: white !important;
    font-weight: 700 !important;
    font-family: "Playfair Display", serif !important;
    letter-spacing: 0.02em !important;
}
.stDataFrame tbody tr:hover td {
    background: rgba(0,51,153,0.04) !important;
}
.stDataFrame tbody tr:nth-child(even) td {
    background: rgba(0,51,153,0.02) !important;
}

/* ── 22. TOP NAV (dashboard) ──────────────────────────────── */
.top-nav {
    position: fixed; top: 0; left: 0; width: 100%; height: 68px;
    background: linear-gradient(90deg, #001a5c, #003399, #0044cc);
    color: white; display: flex; align-items: center;
    padding: 0 28px; z-index: 99999;
    box-shadow: 0 4px 16px rgba(0,0,0,0.30);
    border-bottom: 2px solid #f5b400;
}
.top-title {
    font-family: "Playfair Display", serif !important;
    font-size: 26px; font-weight: 900; margin-left: 16px; letter-spacing: -0.01em;
    color: white;
}

/* ── 23. DIVIDER ──────────────────────────────────────────── */
.fbc-divider {
    height: 2px;
    background: linear-gradient(90deg, transparent, #f5b400, #003399, transparent);
    border: none;
    margin: 28px 0;
    border-radius: 999px;
}

/* ── 24. RADIO / CHECKBOX ─────────────────────────────────── */
.stRadio > label, .stCheckbox > label {
    font-family: "EB Garamond", Georgia, serif !important;
    color: #1a1a2e !important;
}

/* ── 25. SELECTBOX OPTIONS ────────────────────────────────── */
div[data-baseweb="select"] span {
    font-family: "EB Garamond", Georgia, serif !important;
}

/* ── 26. PEER PICKER (Comparables) ───────────────────────── */
.peer-picker-wrap {
    border: 1px solid rgba(0,51,153,0.14);
    border-radius: 18px;
    padding: 18px 18px 12px 18px;
    background: #f8fbff;
    box-shadow: 0 6px 20px rgba(0,26,92,0.07);
    margin-top: 12px;
    margin-bottom: 14px;
}
.peer-picker-head {
    font-family: "Playfair Display", serif !important;
    font-size: 17px;
    font-weight: 700;
    color: #001a5c;
    margin-bottom: 8px;
}
.peer-picker-sub {
    font-size: 13px;
    color: #475569;
    margin-bottom: 14px;
    font-style: italic;
}

/* ── 27. SENS TABLE OVERRIDES ─────────────────────────────── */
.sens-outer { border: 2px solid rgba(0,51,153,0.15) !important; border-radius: 16px !important; }
.sens-table { font-family: "EB Garamond", Georgia, serif !important; }
.sens-table thead th { background: #001a5c !important; }

/* ── 28. FOOTER BANNER ────────────────────────────────────── */
.footer {
    text-align: center;
    padding: 24px;
    margin-top: 40px;
    color: #5a7099 !important;
    font-size: 13px;
    border-top: 1px solid rgba(0,51,153,0.10);
    font-style: italic;
}
.footer b { color: #003399 !important; font-style: normal; }

</style>
''', unsafe_allow_html=True)

st.markdown('''
<div class="fbc-page-header">
    <span class="fbc-page-header-icon">📊</span>
    <span class="fbc-page-header-title">Discounted Cash Flow Model</span>
    <span class="fbc-badge">FBC Securities</span>
    <div class="fbc-page-header-sub">Forecast financial statements, compute WACC, UFCF, and intrinsic equity value.</div>
</div>
''', unsafe_allow_html=True)
# ────────────────────────────────────────────────────────────────

# ---------------------------------------------------------
# ✅ FIX SIDEBAR COLLAPSE ARROW (Material Icons)
# ---------------------------------------------------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/icon?family=Material+Icons');

/* Force ONLY the collapse icon area to render with the Material Icons font */
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stSidebarCollapseButton"] i,
[data-testid="stSidebarCollapseButton"] svg,
.material-icons,
span.material-icons,
i.material-icons {
    font-family: "Material Icons" !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}
</style>
""", unsafe_allow_html=True)
st.markdown("""
<style>

/* =========================================================
   FBC CLEAN SECTION HEADER (NO SUBTITLE, NO STEPS)
   ========================================================= */

.fbc-section {
    display: flex;
    align-items: center;
    gap: 14px;

    padding: 16px 22px;
    margin: 28px 0 18px 0;

    background: linear-gradient(
        135deg,
        rgba(0, 51, 153, 0.08),
        rgba(245, 180, 0, 0.05)
    );

    border-left: 6px solid #003399;
    border-radius: 14px;

    box-shadow: 0 4px 14px rgba(0, 26, 92, 0.08);
}

/* Left indicator (circle like your UI) */
.fbc-section {
    display: block;
    padding: 16px 0;
    margin: 28px 0 18px 0;
    border-bottom: 2px solid rgba(0,51,153,0.15);
    transition: all 0.25s ease;
}

.fbc-section-title {
    font-size: 21px;
    font-weight: 800;
    color: #001a5c;
    position: relative;
}

/* animated underline */
.fbc-section-title::after {
    content: "";
    position: absolute;
    left: 0;
    bottom: -6px;
    width: 40px;
    height: 3px;
    background: #003399;
    transition: width 0.3s ease;
}

.fbc-section:hover .fbc-section-title::after {
    width: 100%;
}

/* Title only */
.fbc-section-title {
    font-family: "Playfair Display", serif !important;
    font-size: 21px;
    font-weight: 800;
    color: #001a5c !important;
    letter-spacing: -0.01em;
}

</style>
""", unsafe_allow_html=True)
def section(title: str):
    st.markdown(
        f"""
        <div class="fbc-section">
            <div class="fbc-section-dot"></div>
            <div class="fbc-section-title">{title}</div>
        </div>
        """,
        unsafe_allow_html=True
    )
def add_watermark():
    logo_path = Path("assets") / "fbc_logo.png"
    if logo_path.exists():
        with open(logo_path, "rb") as f:
            logo_base64 = base64.b64encode(f.read()).decode()

        watermark_css = f"""
        <style>

        /* Make watermark very light */
        .stApp::before {{
            content: "";
            position: fixed;
            top: 40;
            left: 50;
            width: 100%;
            height: 100%;
            background-image: url("data:image/png;base64,{logo_base64}");
            background-repeat: no-repeat;
            background-position: center;
            background-size: 1500px;
            opacity: 0.07;   /* 🔥 control watermark visibility here */
            pointer-events: none;
            z-index: 0;
        }}

        .block-container {{
            position: relative;
            z-index: 1;
        }}
        </style>
        """
        st.markdown(watermark_css, unsafe_allow_html=True)

add_watermark()
PROJECT_ROOT = Path(__file__).resolve().parents[1]   # <-- go up from /pages to project root
DATA_DIR = PROJECT_ROOT / "data"

DCF_PARAMS_PATH = DATA_DIR / "dcf_parameters.xlsx"
UNLEVERED_BETAS_PATH = DATA_DIR / "unlevered_betas.xlsx"

# ---------------------------------------------------------
# HELPERS
# ---------------------------------------------------------
st.markdown("""
<style>

.fbc-reset-card {
    background: linear-gradient(135deg, #003399 0%, #0055cc 100%);
    padding: 20px 24px;
    border-radius: 14px;
    color: white;
    box-shadow: 0 6px 16px rgba(0,0,0,0.15);
    margin-bottom: 25px;
}

.fbc-reset-title {
    font-size: 20px;
    font-weight: 700;
    margin-bottom: 6px;
}

.fbc-reset-sub {
    font-size: 14px;
    opacity: 0.9;
    margin-bottom: 14px;
}

/* ✅ Load Material Icons so Streamlit's sidebar collapse icon renders correctly */
@import url('https://fonts.googleapis.com/icon?family=Material+Icons');

/* ✅ Make sure ONLY icons use the Material Icons font (prevents 'keyboard_double_arrow_right' text) */
.material-icons, 
span.material-icons,
i.material-icons,
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stSidebarCollapseButton"] i {
    font-family: 'Material Icons' !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    display: inline-block !important;
    white-space: nowrap !important;
    word-wrap: normal !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}

/* ✅ Style the collapse/expand button nicely */
[data-testid="stSidebarCollapseButton"] button {
    background: #003399 !important;
    border: 1px solid rgba(255,255,255,0.25) !important;
    border-radius: 999px !important;
    width: 44px !important;
    height: 44px !important;
    box-shadow: 0 6px 18px rgba(0, 51, 153, 0.35) !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease !important;
}

[data-testid="stSidebarCollapseButton"] button:hover {
    transform: translateY(-1px) !important;
    background: #0047d6 !important;
    box-shadow: 0 10px 22px rgba(0, 71, 214, 0.35) !important;
}

[data-testid="stSidebarCollapseButton"] svg {
    width: 22px !important;
    height: 22px !important;
    fill: white !important;
}
/* ===== SIDEBAR GLASS STYLE ===== */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #003399 0%, #001a4d 100%) !important;
    color: white !important;
    border-right: 1px solid rgba(255,255,255,0.15);
    backdrop-filter: blur(8px);
}

/* Sidebar text */
section[data-testid="stSidebar"] * {
    color: white !important;
}

/* Sidebar headings */
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
}

/* Remove default sidebar padding spacing */
section[data-testid="stSidebar"] .block-container {
    padding-top: 1rem !important;
}

.fbc-reset-btn button {
    background-color: #f5b400 !important;   /* FBC gold */
    color: #002266 !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    padding: 10px 20px !important;
    border: none !important;
    transition: all 0.25s ease-in-out;
}

.fbc-reset-btn button:hover {
    background-color: #ffd24d !important;
    transform: translateY(-2px);
    box-shadow: 0 6px 14px rgba(0,0,0,0.25);
}
</style>
""", unsafe_allow_html=True)
st.markdown("""
<style>
.dcf-card{
  background: #ffffff;
  border: 1px solid rgba(0,0,0,0.08);
  border-radius: 16px;
  padding: 18px 18px;
  box-shadow: 0 6px 18px rgba(0,0,0,0.06);
  margin-top: 10px;
  margin-bottom: 14px;
}
.dcf-card h3{
  margin: 0 0 8px 0;
}
.dcf-subcard{
  background: rgba(0,51,153,0.03);
  border: 1px solid rgba(0,51,153,0.10);
  border-radius: 14px;
  padding: 14px;
  margin-top: 10px;
}
.dcf-kpi{
  background: linear-gradient(135deg, rgba(0,51,153,0.10), rgba(245,180,0,0.10));
  border: 1px solid rgba(0,0,0,0.08);
  border-radius: 16px;
  padding: 12px 14px;
  margin: 6px 0;
}
.dcf-kpi-title{
  font-size: 12px;
  opacity: 0.75;
  margin-bottom: 2px;
}
.dcf-kpi-value{
  font-size: 18px;
  font-weight: 800;
}
.small-note{
  font-size: 12px;
  opacity: 0.75;
}
</style>
""", unsafe_allow_html=True)
st.markdown("""
<style>

/* =========================================
   FORCE WHITE TEXT INSIDE ALL BLUE BOXES
   ========================================= */

/* 2️⃣ Navigation buttons (Back / Next) */
.fbc-nav-btn button,
.fbc-nav-btn button * {
    color: #ffffff !important;
}

/* 3️⃣ Primary Streamlit buttons */
.stButton > button,
.stButton > button * {
    color: #ffffff !important;
}

/* 4️⃣ Blue section headers / wizard headers */
.fbc-section-heading,
.fbc-section-heading *,
.streamlit-expanderHeader,
.streamlit-expanderHeader * {
    color: #ffffff !important;
}

/* 5️⃣ Radio & checkbox labels when inside blue areas */
.fbc-nav-btn label,
.fbc-nav-btn span,
.fbc-nav-btn p {
    color: #ffffff !important;
}

/* 6️⃣ Progress bar text (if any overlays) */
.stProgress * {
    color: #ffffff !important;
}

</style>
""", unsafe_allow_html=True)
st.markdown("""
<style>

/* ✅ DCF KPI title — muted blue-gray */
.dcf-kpi-title {
    color: #475569 !important;
    font-weight: 600;
}

/* ✅ DCF KPI values — strong finance blue */
.dcf-kpi-value {
    color: #1e3a8a !important;
    font-weight: 800;
}

</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>

/* =========================================
   FILE UPLOADER — DCF STYLE
   ========================================= */
/* =========================================
   FILE UPLOADER — MAKE TEXT WHITE
   ========================================= */

/* Main uploader container text */
[data-testid="stFileUploader"] {
    color: white !important;
}

/* "Drag and drop file here" text */
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] div,
[data-testid="stFileUploader"] span,
[data-testid="stFileUploader"] p {
    color: white !important;
}

/* Uploaded filename */
[data-testid="stFileUploaderFileName"] {
    color: white !important;
}

/* Browse button text */
[data-testid="stFileUploader"] button {
    color: white !important;
}

/* Optional: make border visible on dark background */
[data-testid="stFileUploader"] section {
    border: 1px solid rgba(255,255,255,0.3) !important;
}
/* Uploader container */
[data-testid="stFileUploader"] {
    background: rgba(0, 51, 153, 0.04);
    border: 1px dashed rgba(0, 51, 153, 0.25);
    border-radius: 14px;
    padding: 18px 20px;
}

/* Upload button */
[data-testid="stFileUploader"] button {
    background: linear-gradient(135deg, #003399, #1e3a8a) !important;
    color: #ffffff !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    padding: 8px 18px !important;
    border: none !important;
    box-shadow: 0 4px 14px rgba(0, 51, 153, 0.35) !important;
    transition: all 0.2s ease-in-out !important;
}

/* Hover effect */
[data-testid="stFileUploader"] button:hover {
    background: linear-gradient(135deg, #1d4ed8, #003399) !important;
    transform: translateY(-1px);
    box-shadow: 0 8px 20px rgba(0, 51, 153, 0.45) !important;
}

/* Helper text (file size, type) */
[data-testid="stFileUploader"] small {
    color: #475569 !important;
    font-style: italic;
}

/* Uploaded filename */
[data-testid="stFileUploader"] span {
    color: #1e3a8a !important;
    font-weight: 600;
}

</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------
# RESET DCF SESSION STATE (SAFE & CONTROLLED)
# ---------------------------------------------------------
def reset_dcf_state():
    keys_to_clear = [
        # file & parsed data
        "dcf_uploaded_file", "dcf_is_df", "dcf_bs_df", "dcf_cf_df",

        # ✅ ADD THESE (so reset truly clears the file)
        "dcf_file_bytes", "dcf_file_name",

        # FX
        "dcf_fx_file", "dcf_fx_raw", "dcf_yearly_fx",
        "dcf_fx_applied", "dcf_apply_fx_bs", "dcf_fx_column",
        "dcf_closing_fx_rate",
        "dcf_conversion_method", "dcf_currency",
        "dcf_fx_bytes", "dcf_fx_name",
        "dcf_factor_enabled", "dcf_zig_factor", "dcf_factor_year_ranges",
        "dcf_bs_fx_dirty",

        # mappings
        "dcf_mapping", "is_core_mapping",
        "bs_map_step", "bs_jump_radio", "bs_widget_reset",
        "cf_map_step", "cf_jump_radio", "cf_widget_reset",

        # forecasts
        "dcf_rev_forecast", "dcf_ebitda_all", "dcf_ebitda_forecast",
        "dcf_profit_all",
        # parameters + widgets
        "dcf_rf_pct", "dcf_mrp_pct", "dcf_tax_pct", "dcf_unlevered_beta", "dcf_terminal_g_pct",
        "dcf_rf_pct_input", "dcf_mrp_pct_input", "dcf_tax_pct_input",
        "dcf_unlevered_beta_input", "dcf_terminal_g_pct_input",
        "dcf_use_auto_params", "dcf_use_auto_params_ui",
        "dcf_country_select", "dcf_zim_avg_cost_debt_pct", "dcf_zim_avg_cost_debt_pct_input",
        "dcf_beta_manual_mode",
        "dcf_beta_manual_value",
        "dcf_beta_auto_last",
        "dcf_beta_mode_radio",

        # working capital
        "dcf_fcff_array", "dcf_pv_fcff_array", "dcf_discount_periods_n",
        "dcf_is_base", "dcf_bs_base", "dcf_cf_base",
        "dcf_fx_signature", "dcf_bs_fx_rates", "dcf_bs_closing_dates",

        # valuation outputs
        "enterprise_value_dcf", "equity_value_dcf", "equity_value",

        # parameters
        "dcf_init", "dcf_timing_init",
    ]

    for k in keys_to_clear:
        if k in st.session_state:
            del st.session_state[k]
def sort_year_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # identify year columns (keep Item separate)
    item_col = "Item"
    year_cols = [c for c in df.columns if c != item_col]

    # convert to int safely for sorting
    def try_int(x):
        try:
            return int(str(x))
        except:
            return x

    sorted_years = sorted(year_cols, key=try_int)

    # reorder dataframe
    return df[[item_col] + sorted_years]
def clean_numeric_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c) for c in df.columns]
    first_col = df.columns[0]
    df.rename(columns={first_col: "Item"}, inplace=True)

    for col in df.columns[1:]:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.replace("(", "-", regex=False)
            .str.replace(")", "", regex=False)
            .str.strip()
        )
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def get_year_cols(df: pd.DataFrame):
    return [c for c in df.columns if c != "Item"]


def avg_revenue_growth(revenue_row: pd.DataFrame, year_cols) -> float:
    vals = revenue_row[year_cols].values.flatten().astype(float)
    growth = []
    for i in range(1, len(vals)):
        prev_, curr_ = vals[i - 1], vals[i]
        if curr_ != 0:
            g = (curr_ - prev_) / curr_
            if -0.5 < g < 0.5:
                growth.append(g)
    return float(np.mean(growth)) if growth else 0.05


def ratio_to_revenue(row_vals: np.ndarray, rev_vals: np.ndarray) -> float:
    mask = (~np.isnan(row_vals)) & (~np.isnan(rev_vals)) & (rev_vals != 0)
    if not mask.any():
        return 0.0
    ratios = row_vals[mask] / rev_vals[mask]
    ratios = ratios[(ratios > -5) & (ratios < 5)]
    return float(np.mean(ratios)) if len(ratios) else 0.0


def find_row_indices(df: pd.DataFrame, keywords):
    if df.empty:
        return []
    s = df["Item"].astype(str).str.lower()
    mask = False
    for kw in keywords:
        mask = mask | s.str.contains(kw, na=False)
    return list(df[mask].index)


def find_single_row(df: pd.DataFrame, keywords):
    idx_list = find_row_indices(df, keywords)
    return (idx_list[0], df.iloc[idx_list[0]]) if idx_list else (None, None)


def convert_df_yearwise(df: pd.DataFrame, year_rates: dict) -> pd.DataFrame:
    df2 = df.copy()
    for col in df2.columns:
        if col == "Item":
            continue
        if str(col) in year_rates and year_rates[str(col)] != 0:
            df2[col] = df2[col] / year_rates[str(col)]
    return df2
def get_fx_asof_date(fx_df, fx_col, closing_date):
    """
    Returns the last available FX rate on or before the closing_date
    """
    fx_before = fx_df[fx_df["Date"] <= pd.Timestamp(closing_date)]
    if fx_before.empty:
        return None
    return float(fx_before.sort_values("Date").iloc[-1][fx_col])


def load_fx_yearly_from_excel(fx_file) -> dict:
    """
    Excel must contain:
    Date | Interbank | Alternative | Premium  (or similar)
    """
    fx = pd.read_excel(fx_file)
    fx.columns = [str(c) for c in fx.columns]

    date_col = fx.columns[0]
    fx[date_col] = pd.to_datetime(fx[date_col], errors="coerce")

    rate_col = st.selectbox(
        "Which FX rate column should be used?",
        fx.columns[1:]
    )

    fx[rate_col] = pd.to_numeric(fx[rate_col], errors="coerce")
    fx = fx.dropna(subset=[date_col, rate_col])

    fx["Year"] = fx[date_col].dt.year
    yearly = fx.groupby("Year")[rate_col].mean()

    return {str(int(y)): float(v) for y, v in yearly.items()}


def option_labels_from_items(items):
    return [f"{i+1}: {name}" for i, name in enumerate(items)]


def indices_from_labels(labels):
    idx = []
    for s in labels:
        try:
            idx.append(int(s.split(":", 1)[0]) - 1)
        except:
            pass
    return idx

@st.cache_data(show_spinner=False)
def _load_unlevered_betas_any(file_or_path, file_mtime: float = 0.0) -> pd.DataFrame:
    """
    Excel required columns (flexible match):
      Industry Name | Unlevered beta
    Also supports: Column1 (industry) + Column6 (beta)

    file_mtime is used ONLY to invalidate Streamlit cache when the Excel file changes.
    """
    df = pd.read_excel(file_or_path)
    df.columns = [str(c).strip() for c in df.columns]

    possible_industry_cols = [c for c in df.columns if c.lower() in ["industry name", "industry", "column1"]]
    possible_beta_cols = [c for c in df.columns if c.lower() in ["unlevered beta", "unlevered_beta", "beta", "column6"]]

    if not possible_industry_cols or not possible_beta_cols:
        raise ValueError("Excel must have 'Industry Name' and 'Unlevered beta' (or Column1 + Column6).")

    ind_col = possible_industry_cols[0]
    beta_col = possible_beta_cols[0]

    out = df[[ind_col, beta_col]].copy()
    out.columns = ["Industry", "UnleveredBeta"]
    out["Industry"] = out["Industry"].astype(str).str.strip()
    out["UnleveredBeta"] = pd.to_numeric(out["UnleveredBeta"], errors="coerce")
    out = out.dropna(subset=["Industry", "UnleveredBeta"]).sort_values("Industry").reset_index(drop=True)
    return out

st.markdown("""
<style>

/* ===============================
   FBC NAVIGATION ARROW BUTTONS
   =============================== */

.fbc-nav-btn button {
    background: linear-gradient(135deg, #003399, #001a4d) !important;
    color: white !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    padding: 8px 18px !important;
    border: none !important;
    transition: all 0.25s ease-in-out;
}

.fbc-nav-btn button:hover {
    background: linear-gradient(135deg, #0055cc, #003399) !important;
    transform: translateY(-2px);
    box-shadow: 0 6px 14px rgba(0,0,0,0.25);
}

.fbc-nav-btn button:disabled {
    background: #6b7280 !important;
    color: #e5e7eb !important;
    cursor: not-allowed;
}

</style>
""", unsafe_allow_html=True)
st.markdown("""
<style>
html, body, .stApp, .block-container,
p, div, label,
h1, h2, h3, h4, h5, h6,
li, ul, ol, a, small {
  font-family: Georgia, "Times New Roman", serif !important;
}

/* ✅ Keep the sidebar collapse arrow as an icon */
[data-testid="stSidebarCollapseButton"] * {
  font-family: "Material Icons" !important;
}


/* ---------------------------
   PROTECT ALL ICON FONTS
   (Fixes keyboard_double_arrow_right issue)
---------------------------- */

/* Material Icons */
.material-icons,
.material-symbols-outlined,
.material-symbols-rounded,
.material-symbols-sharp,
[class*="material-icons"],
[class*="material-symbols"] {
  font-family: "Material Icons" !important;
}

/* Bootstrap Icons */
.bi,
[class^="bi-"], [class*=" bi-"] {
  font-family: "bootstrap-icons" !important;
}

/* Font Awesome */
.fa, .fas, .far, .fal, .fab,
[class^="fa-"], [class*=" fa-"] {
  font-family: "Font Awesome 6 Free" !important;
}
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------
# COMPANY NAME (persistent across pages/models)
# ---------------------------------------------------------
if "company_name" not in st.session_state:
    st.session_state["company_name"] = "Selected Company"

# ---------------------------------------------------------
# 🔄 START NEW VALUATION — FBC STYLE
# ---------------------------------------------------------
st.markdown("""
<div class="fbc-reset-card">
    <div class="fbc-reset-title">🔄 Start New Valuation</div>
    <div class="fbc-reset-sub">
        Reset the workspace and upload a new set of financial statements.
    </div>
</div>
""", unsafe_allow_html=True)

col_reset_left, col_reset_right = st.columns([1, 3])

with col_reset_left:
    st.markdown('<div class="fbc-reset-btn">', unsafe_allow_html=True)
    if st.button("🗂️ Clear & Upload New File", width='stretch'):
        reset_dcf_state()

        # ✅ Increment uploader key to force Streamlit to forget previous upload
        st.session_state["dcf_uploader_key"] = st.session_state.get("dcf_uploader_key", 0) + 1

        st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------------------------------------
# FILE UPLOAD
# -------------------------------------------------------


section("Upload Financial Statements")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
# --- Persist uploaded file bytes so the app doesn't "forget" after inactivity/rerun ---
if "dcf_file_bytes" not in st.session_state:
    st.session_state["dcf_file_bytes"] = None
if "dcf_file_name" not in st.session_state:
    st.session_state["dcf_file_name"] = None

# ✅ Initialize uploader key BEFORE using it
if "dcf_uploader_key" not in st.session_state:
    st.session_state["dcf_uploader_key"] = 0

st.markdown(
    '<p style="color:#003399; font-weight:700; font-size:16px;">Upload Excel with IS, BS, CF</p>',
    unsafe_allow_html=True
)

uploaded_file = st.file_uploader("", type=["xlsx"])


# Save bytes once
if uploaded_file is not None:
    st.session_state["dcf_file_bytes"] = uploaded_file.getvalue()
    st.session_state["dcf_file_name"] = uploaded_file.name

# Rebuild a file-like object from bytes
if st.session_state["dcf_file_bytes"] is None:
    st.info("⬆️ Please upload an Excel file to begin.")
    st.stop()

file_like = io.BytesIO(st.session_state["dcf_file_bytes"])

# Now use this in ExcelFile
xls = pd.ExcelFile(file_like)


# ---------------------------------------------------------
# LOAD & CACHE PARSED STATEMENTS (ONCE)
# ---------------------------------------------------------
if "dcf_is_df" not in st.session_state:
    xls = pd.ExcelFile(io.BytesIO(st.session_state["dcf_file_bytes"]))  # ✅ always valid

    st.session_state["dcf_is_df"] = sort_year_columns(
        clean_numeric_cols(xls.parse(xls.sheet_names[0]))
    )

    st.session_state["dcf_bs_df"] = sort_year_columns(
        clean_numeric_cols(xls.parse(xls.sheet_names[1]))
    )

    st.session_state["dcf_cf_df"] = sort_year_columns(
        clean_numeric_cols(xls.parse(xls.sheet_names[2]))
    )

is_df = st.session_state["dcf_is_df"]
bs_df = st.session_state["dcf_bs_df"]
cf_df = st.session_state["dcf_cf_df"]

# ---------------------------------------------------------
# STORE BASE (ORIGINAL) STATEMENTS ONCE (for re-conversion)
# ---------------------------------------------------------
if "dcf_is_base" not in st.session_state:
    st.session_state["dcf_is_base"] = st.session_state["dcf_is_df"].copy()
if "dcf_bs_base" not in st.session_state:
    st.session_state["dcf_bs_base"] = st.session_state["dcf_bs_df"].copy()
if "dcf_cf_base" not in st.session_state:
    st.session_state["dcf_cf_base"] = st.session_state["dcf_cf_df"].copy()

year_cols_is = get_year_cols(is_df)
year_cols_bs = get_year_cols(bs_df)
year_cols_cf = get_year_cols(cf_df)

# FX SECTION — EXCEL-BASED (ZWG → USD) — FINAL & CORRECT
# ---------------------------------------------------------
section("💱 Currency & Exchange Rates")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
# ✅ Persist conversion method across tabs/pages
# conversion_method can be: "NO_FX" or "FX_EXCEL"
if "dcf_conversion_method" not in st.session_state:
    st.session_state["dcf_conversion_method"] = "NO_FX"

if "dcf_currency" not in st.session_state:
    st.session_state["dcf_currency"] = "USD (already converted)"

# -------------------------------------------------
# 1️⃣ Currency selector (persistent)
# -------------------------------------------------
currency = st.selectbox(
    "Currency of uploaded statements:",
    ["USD (already converted)", "ZWG (convert using FX Excel)"],
    index=0 if st.session_state.get("dcf_conversion_method") == "NO_FX" else 1,
    key="dcf_currency_select"
)

# ✅ Store BOTH: the label + a stable method flag
st.session_state["dcf_currency"] = currency
st.session_state["dcf_conversion_method"] = "NO_FX" if currency.startswith("USD") else "FX_EXCEL"


# -------------------------------------------------
# 2️⃣ FX Excel upload — SHOW ONLY IF ZWG
# -------------------------------------------------
# ✅ Persist FX file across reruns/tabs using bytes
if "dcf_fx_bytes" not in st.session_state:
    st.session_state["dcf_fx_bytes"] = None
if "dcf_fx_name" not in st.session_state:
    st.session_state["dcf_fx_name"] = None
if st.session_state["dcf_conversion_method"] == "FX_EXCEL":

    st.markdown("""
    <div style="
        border: 1px dashed #f5b400;
        padding: 18px;
        border-radius: 12px;
        background-color: #fffaf0;
        margin-bottom: 15px;
    ">
        <strong>📥 FX Data Required</strong><br>
        Upload exchange rates to convert ZWG → USD
    </div>
    """, unsafe_allow_html=True)

    fx_file = st.file_uploader(
        "Upload FX Excel (Date + FX columns)",
        type=["xlsx"],
        key="dcf_fx_uploader"
    )

    # ✅ Save FX bytes once
    if fx_file is not None:
        st.session_state["dcf_fx_bytes"] = fx_file.getvalue()
        st.session_state["dcf_fx_name"] = fx_file.name

else:
    # USD selected → clear FX bytes
    st.session_state["dcf_fx_bytes"] = None
    st.session_state["dcf_fx_name"] = None

# -------------------------------------------------
# 3️⃣ If USD → skip FX
# -------------------------------------------------
if currency.startswith("USD"):
    st.success("✅ Data assumed to be in USD. No FX conversion applied.")

else:
    st.warning("ZWG detected. Upload FX Excel with Dates and Interbank Rates to convert to USD.")

    if st.session_state["dcf_fx_bytes"] is None:
        st.stop()

    # -------------------------------------------------
    # 4️⃣ Load FX Excel ONCE
    # -------------------------------------------------
    if "dcf_fx_raw" not in st.session_state:
        fx_raw = pd.read_excel(io.BytesIO(st.session_state["dcf_fx_bytes"]))
        fx_raw.columns = [str(c).strip() for c in fx_raw.columns]
        st.session_state["dcf_fx_raw"] = fx_raw
    else:
        fx_raw = st.session_state["dcf_fx_raw"]

    st.subheader("Raw FX data (preview)")
    st.dataframe(fx_raw.head(), width='stretch')
    st.subheader("📊 Balance Sheet Closing FX Rates Used")
    bs_fx_rates = st.session_state.get("dcf_bs_fx_rates", {})

    # Build BS FX confirmation table safely
    bs_fx_table = pd.DataFrame([
        {
            "Year": y,
            "Closing Date": st.session_state["dcf_bs_closing_dates"][y],
            "FX Rate Used": bs_fx_rates[y],
        }
        for y in bs_fx_rates.keys()
    ])

    st.dataframe(bs_fx_table, width='stretch')

    # -------------------------------------------------
    # 5️⃣ Validate required columns
    # -------------------------------------------------
    if "Date" not in fx_raw.columns:
        st.error("❌ FX Excel must contain a column named 'Date'.")
        st.stop()

    fx_df = fx_raw.copy()

    fx_df["Date"] = pd.to_datetime(
        fx_df["Date"],
        errors="coerce",
        dayfirst=True
    )

    fx_df = fx_df.dropna(subset=["Date"])

    # -------------------------------------------------
    # 6️⃣ FX column selector (restricted + persistent)
    # -------------------------------------------------
    allowed_fx_cols = ["Interbank", "Alternative", "Premium"]
    available_fx_cols = [c for c in allowed_fx_cols if c in fx_df.columns]

    if not available_fx_cols:
        st.error("❌ FX Excel must contain Interbank / Alternative / Premium columns.")
        st.stop()

    if "dcf_fx_column" not in st.session_state:
        st.session_state["dcf_fx_column"] = available_fx_cols[0]

    fx_col = st.selectbox(
        "Which FX rate column should be used?",
        available_fx_cols,
        index=available_fx_cols.index(st.session_state["dcf_fx_column"]),
        key="dcf_fx_column_select"
    )

    st.session_state["dcf_fx_column"] = fx_col

    fx_df[fx_col] = pd.to_numeric(fx_df[fx_col], errors="coerce")
    fx_df = fx_df.dropna(subset=[fx_col])
    # -------------------------------------------------
    # 🪙 Apply conversion factor by selected Year(s) + Date Ranges
    # -------------------------------------------------
    st.markdown("### 🪙 Apply ZWG→ZiG factor by Year + Range")

    # years available from statements
    available_years = sorted({str(int(y)) for y in year_cols_is})

    if "dcf_factor_enabled" not in st.session_state:
        st.session_state["dcf_factor_enabled"] = False

    if "dcf_zig_factor" not in st.session_state:
        st.session_state["dcf_zig_factor"] = 2498.7242

    if "dcf_factor_year_ranges" not in st.session_state:
        # {"2024": [{"start": date(...), "end": date(...)}], ...}
        st.session_state["dcf_factor_year_ranges"] = {}
    # ✅ Persist selected years across page/tab switches
    if "dcf_factor_years_selected_vals" not in st.session_state:
        st.session_state["dcf_factor_years_selected_vals"] = []

    enable_factor = st.checkbox(
        "Enable manual factor (for mixed ZWG/ZiG periods)",
        value=st.session_state["dcf_factor_enabled"],
        key="dcf_factor_enabled_ui"
    )
    st.session_state["dcf_factor_enabled"] = enable_factor

    zig_factor = st.number_input(
        "ZWG → ZiG conversion factor (divide FX by this inside selected ranges)",
        value=float(st.session_state["dcf_zig_factor"]),
        step=0.0001,
        format="%.6f",
        key="dcf_zig_factor_ui2"
    )
    st.session_state["dcf_zig_factor"] = zig_factor

    if enable_factor:
        years_selected = st.multiselect(
            "Select the year(s) where you want to apply the factor",
            available_years,
            default=[y for y in st.session_state["dcf_factor_years_selected_vals"] if y in available_years],
            key="dcf_factor_years_selected_ui"
        )

        # ✅ store it explicitly (this survives page/tab switching better)
        st.session_state["dcf_factor_years_selected_vals"] = years_selected

        for y in years_selected:
            st.session_state["dcf_factor_year_ranges"].setdefault(y, [])

            st.markdown(f"#### Ranges for {y}")

            if st.button(f"➕ Add range for {y}", key=f"add_range_{y}"):
                st.session_state["dcf_factor_year_ranges"][y].append({
                    "start": date(int(y), 1, 1),
                    "end": date(int(y), 12, 31),
                })

            ranges = st.session_state["dcf_factor_year_ranges"][y]
            for i, r in enumerate(ranges):
                c1, c2, c3 = st.columns([2, 2, 1])

                with c1:
                    new_start = st.date_input(
                        f"{y} range {i + 1} start",
                        value=r["start"],
                        key=f"{y}_r{i}_start"
                    )
                with c2:
                    new_end = st.date_input(
                        f"{y} range {i + 1} end",
                        value=r["end"],
                        key=f"{y}_r{i}_end"
                    )

                if new_end < new_start:
                    st.error("❌ End date cannot be before start date.")
                else:
                    st.session_state["dcf_factor_year_ranges"][y][i]["start"] = new_start
                    st.session_state["dcf_factor_year_ranges"][y][i]["end"] = new_end

                with c3:
                    if st.button("🗑️ Delete", key=f"{y}_r{i}_del"):
                        st.session_state["dcf_factor_year_ranges"][y].pop(i)
                        st.rerun()

        # Apply factor to FX rows in selected ranges
        if zig_factor <= 0:
            st.error("❌ Factor must be > 0.")
            st.stop()

        fx_df["_factor_applied"] = False

        for y in years_selected:
            for r in st.session_state["dcf_factor_year_ranges"].get(y, []):
                s = pd.Timestamp(r["start"])
                e = pd.Timestamp(r["end"])
                mask = (fx_df["Date"] >= s) & (fx_df["Date"] <= e)
                if mask.any():
                    fx_df.loc[mask, fx_col] = fx_df.loc[mask, fx_col] / float(zig_factor)
                    fx_df.loc[mask, "_factor_applied"] = True

        st.success(f"✅ Factor applied to {int(fx_df['_factor_applied'].sum()):,} FX rows.")
        st.dataframe(fx_df.loc[fx_df["_factor_applied"], ["Date", fx_col]].head(20), width='stretch')

    # -------------------------------------------------
    # 7️⃣ Compute YEARLY AVERAGE FX (Income Statement)
    # -------------------------------------------------
    fx_df["Year"] = fx_df["Date"].dt.year.astype(int)

    yearly_fx = (
        fx_df
        .groupby("Year")[fx_col]
        .mean()
        .round(6)
        .to_dict()
    )

    yearly_fx = {str(y): float(v) for y, v in yearly_fx.items()}
    st.session_state["dcf_yearly_fx"] = yearly_fx
    bs_fx_rates = st.session_state.get("dcf_bs_fx_rates", {})

    st.subheader("📊 Yearly FX averages (Income Statement and Cash Flow Statement)")
    st.dataframe(
        pd.DataFrame({
            "Year": yearly_fx.keys(),
            "FX Rate": yearly_fx.values()
        }),
        width='stretch'    )

    # -------------------------------------------------
    # 8️⃣ Balance Sheet FX OPTION (closing rate)
    # -------------------------------------------------
    if "dcf_apply_fx_bs" not in st.session_state:
        st.session_state["dcf_apply_fx_bs"] = False

    apply_fx_bs = st.checkbox(
        "Apply FX to Balance Sheet using closing rate?",
        value=st.session_state["dcf_apply_fx_bs"],
        help="Uses ONE FX rate (latest available date)",
        key="dcf_fx_bs_checkbox"
    )

    st.session_state["dcf_apply_fx_bs"] = apply_fx_bs
    # -------------------------------------------------
    # 8️⃣ Balance Sheet FX — PER-YEAR CLOSING DATES (NEW)
    # -------------------------------------------------
    st.markdown("### 📌 Balance Sheet FX — Closing Dates (per year)")

    # ✅ INIT FIRST (CRITICAL)
    if "dcf_bs_closing_dates" not in st.session_state:
        st.session_state["dcf_bs_closing_dates"] = {}

    # ✅ Dirty flag init
    if "dcf_bs_fx_dirty" not in st.session_state:
        st.session_state["dcf_bs_fx_dirty"] = False


    bs_years = [str(y) for y in year_cols_bs]

    for y in bs_years:
        default_date = st.session_state["dcf_bs_closing_dates"].get(
            y, date(int(y), 12, 31)
        )

        chosen_date = st.date_input(
            f"Closing date for Balance Sheet {y}",
            value=default_date,
            key=f"bs_close_date_{y}"
        )

        # ✅ Detect change immediately (fixes double click)
        if st.session_state["dcf_bs_closing_dates"].get(y) != chosen_date:
            st.session_state["dcf_bs_closing_dates"][y] = chosen_date
            st.session_state["dcf_bs_fx_dirty"] = True

    # -------------------------------------------------
    # 9️⃣ COMPUTE BALANCE SHEET FX RATES (PER YEAR)
    # -------------------------------------------------
    bs_fx_rates = {}

    for y in bs_years:
        closing_date = st.session_state["dcf_bs_closing_dates"][y]

        fx_rate = get_fx_asof_date(
            fx_df=fx_df,
            fx_col=fx_col,
            closing_date=closing_date
        )

        if fx_rate is None:
            st.error(f"❌ No FX rate found on or before {closing_date} for year {y}")
            st.stop()

        bs_fx_rates[y] = fx_rate

    # ✅ STORE IN SESSION STATE
    st.session_state["dcf_bs_fx_rates"] = bs_fx_rates

    # -------------------------------------------------
    # 🔟 Validate FX coverage for IS years
    # -------------------------------------------------
    statement_years = set(year_cols_is)
    fx_years = set(yearly_fx.keys())

    missing_years = sorted(statement_years - fx_years)

    if missing_years:
        st.error(
            f"❌ Missing FX data for statement years: {', '.join(missing_years)}"
        )
        st.stop()

    # -------------------------------------------------
    # 1️⃣1️⃣ APPLY FX CONVERSION (RE-RUN IF SETTINGS CHANGE)
    # -------------------------------------------------
    factor_signature = (
        st.session_state.get("dcf_factor_enabled", False),
        st.session_state.get("dcf_zig_factor", None),
        str(st.session_state.get("dcf_factor_year_ranges", {}))
    )

    fx_signature = (
        currency,
        fx_col,
        factor_signature,
        tuple((y, str(st.session_state["dcf_bs_closing_dates"][y])) for y in bs_years)
    )

    # Recompute if signature changed (or first run)
    if (
            st.session_state.get("dcf_fx_signature") != fx_signature
            or st.session_state.get("dcf_bs_fx_dirty")
    ):

        # Always start from BASE statements (pre-conversion)
        is_base = st.session_state["dcf_is_base"].copy()
        bs_base = st.session_state["dcf_bs_base"].copy()
        cf_base = st.session_state["dcf_cf_base"].copy()

        # Income Statement → YEARLY AVERAGE FX
        is_converted = convert_df_yearwise(is_base, yearly_fx)

        # Balance Sheet → PER-YEAR CLOSING FX
        bs_converted = convert_df_yearwise(bs_base, bs_fx_rates)

        # Cash Flow → SAME YEARLY AVERAGE FX AS IS
        cf_converted = convert_df_yearwise(cf_base, yearly_fx)

        # Save converted versions
        st.session_state["dcf_is_df"] = is_converted
        st.session_state["dcf_bs_df"] = bs_converted
        st.session_state["dcf_cf_df"] = cf_converted

        # Save the signature so we don't reconvert unnecessarily
        st.session_state["dcf_fx_signature"] = fx_signature
        st.session_state["dcf_bs_fx_dirty"] = False

        # Optional: for debugging / clarity
        st.info("🔁 FX conversion refreshed (settings changed).")
    else:
         st.success("✅ FX conversion applied correctly (IS = yearly average, BS = per-year closing rates)")

# ---------------------------------------------------------
# SHOW CLEANED STATEMENTS
# ---------------------------------------------------------
section("Income Statement (cleaned, in USD)")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
st.dataframe(is_df, width='stretch')

section("Balance Sheet (cleaned, in USD)")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
st.dataframe(bs_df, width='stretch')

section("Cash Flow Statement (cleaned, in USD)")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
st.dataframe(cf_df, width='stretch')

# Re-detect year columns (as strings)
year_cols_is = get_year_cols(is_df)
year_cols_bs = get_year_cols(bs_df)
year_cols_cf = get_year_cols(cf_df)

if len(year_cols_is) < 2:
    st.error("❌ Need at least 2 historical year columns in Income Statement.")
    st.stop()

# Prepare year ints/labels
last_hist_label = year_cols_is[-1]           # string label e.g. "2025"
last_hist_year = int(str(last_hist_label))   # int 2025


# --- Persistent dictionary for DCF row mappings ---
if "dcf_mapping" not in st.session_state:
    st.session_state["dcf_mapping"] = {
        "debt": [],
        "cash": [],
        "ca": [],
        "cl": [],
        "dep": [],
        "capex": [],
        "interest": []
    }
def clean_defaults(default_list, options):
    """
    Keep only those default values that still exist in options.
    Prevents Streamlit error: 'default value ... is not part of the options'.
    """
    if not isinstance(default_list, (list, tuple)):
        return []
    return [x for x in default_list if x in options]

# ---------------------------------------------------------
# BALANCE SHEET — OPTION C WIZARD (multi-select + preview)
# ---------------------------------------------------------
BS_LINES = [
    ("debt",   "Total Debt / Borrowings (multi-select)"),
    ("cash",   "Cash & Cash Equivalents (multi-select)"),
    ("ca",     "Current Assets (for Working Capital) (multi-select)"),
    ("cl",     "Current Liabilities (for Working Capital) (multi-select)"),
    ("equity", "Equity (multi-select)"),
]

def map_bs_wizard(bs_df, year_cols_bs):
    section("Balance Sheet — Mapping")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    bs_items = list(bs_df["Item"].astype(str))
    bs_labels = option_labels_from_items(bs_items)

    if "dcf_mapping" not in st.session_state:
        st.session_state["dcf_mapping"] = {}
    for k, _ in BS_LINES:
        st.session_state["dcf_mapping"].setdefault(k, [])

    if "bs_map_step" not in st.session_state:
        st.session_state["bs_map_step"] = 0

    # --- progress
    mapped = sum(1 for k, _ in BS_LINES if len(st.session_state["dcf_mapping"].get(k, [])) > 0)
    st.progress(mapped / len(BS_LINES))
    st.caption(f"Mapped: {mapped}/{len(BS_LINES)}")

    step_names = [name for _, name in BS_LINES]

    # ✅ make radio fully controlled
    if "bs_jump_radio" not in st.session_state:
        st.session_state["bs_jump_radio"] = step_names[st.session_state["bs_map_step"]]

    def _set_step(i: int):
        i = max(0, min(i, len(BS_LINES) - 1))
        st.session_state["bs_map_step"] = i
        st.session_state["bs_jump_radio"] = step_names[i]   # ✅ move red dot
        st.rerun()

    c1, c2, _ = st.columns([1, 1, 2])

    with c1:
        st.markdown('<div class="fbc-nav-btn">', unsafe_allow_html=True)
        if st.button("⬅️ Back (BS)", disabled=st.session_state["bs_map_step"] == 0):
            _set_step(st.session_state["bs_map_step"] - 1)
        st.markdown('</div>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="fbc-nav-btn">', unsafe_allow_html=True)
        if st.button("➡️ Next unmapped (BS)"):
            target = None
            for i, (k, _) in enumerate(BS_LINES):
                if len(st.session_state["dcf_mapping"].get(k, [])) == 0:
                    target = i
                    break
            if target is None:
                target = len(BS_LINES) - 1
            _set_step(target)
        st.markdown('</div>', unsafe_allow_html=True)

    # ✅ radio drives step too
    chosen_step = st.radio(
        "Jump to BS line:",
        step_names,
        key="bs_jump_radio",
        horizontal=True
    )
    st.session_state["bs_map_step"] = step_names.index(chosen_step)

    k, title = BS_LINES[st.session_state["bs_map_step"]]
    stored = clean_defaults(st.session_state["dcf_mapping"].get(k, []), bs_labels)

    # ✅ widget reset counters (versioned key)
    if "bs_widget_reset" not in st.session_state:
        st.session_state["bs_widget_reset"] = {}
    st.session_state["bs_widget_reset"].setdefault(k, 0)

    widget_key = f"bs_pick_{k}_{st.session_state['bs_widget_reset'][k]}"

    with st.container(border=True):
        st.markdown(f"#### {title}")

        sel = st.multiselect(
            "Select row(s):",
            bs_labels,
            default=stored,
            key=widget_key
        )

        st.session_state["dcf_mapping"][k] = sel

        if st.button("🧹 Clear selection", key=f"bs_clear_{k}"):
            st.session_state["dcf_mapping"][k] = []
            st.session_state["bs_widget_reset"][k] += 1  # ✅ forces a fresh widget
            st.rerun()

        # preview
        if sel:
            idx_list = indices_from_labels(sel)
            preview_vals = bs_df.loc[idx_list, year_cols_bs].sum(axis=0)
            st.caption("Preview (sum of selected rows):")
            st.dataframe(
                pd.DataFrame({"Year": year_cols_bs, "Total": preview_vals.values}),
                hide_index=True,
                width='stretch'
            )

    out = {}
    for kk, _ in BS_LINES:
        out[kk] = indices_from_labels(st.session_state["dcf_mapping"].get(kk, []))
    return out

bs_idx = map_bs_wizard(bs_df, year_cols_bs)

debt_idx_list   = bs_idx["debt"]
cash_idx_list   = bs_idx["cash"]
ca_idx_list     = bs_idx["ca"]
cl_idx_list     = bs_idx["cl"]
equity_idx_list = bs_idx["equity"]
# ---------------------------------------------------------
# CASH FLOW — OPTION C WIZARD (multi-select + preview)
# ---------------------------------------------------------
CF_LINES = [
    ("dep",      "Depreciation & Amortisation (multi-select)"),
    ("capex",    "Capex  (multi-select)"),
    ("interest", "Interest paid (if using CF for interest) (multi-select)"),
]

def map_cf_wizard(cf_df, year_cols_cf):
    section("Cash Flow — Mapping")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    cf_items = list(cf_df["Item"].astype(str))
    cf_labels = option_labels_from_items(cf_items)

    if "dcf_mapping" not in st.session_state:
        st.session_state["dcf_mapping"] = {}
    for k, _ in CF_LINES:
        st.session_state["dcf_mapping"].setdefault(k, [])

    if "cf_map_step" not in st.session_state:
        st.session_state["cf_map_step"] = 0

    mapped = sum(1 for k, _ in CF_LINES if len(st.session_state["dcf_mapping"].get(k, [])) > 0)
    st.progress(mapped / len(CF_LINES))
    st.caption(f"Mapped: {mapped}/{len(CF_LINES)}")

    step_names = [name for _, name in CF_LINES]

    if "cf_jump_radio" not in st.session_state:
        st.session_state["cf_jump_radio"] = step_names[st.session_state["cf_map_step"]]

    def _set_step(i: int):
        i = max(0, min(i, len(CF_LINES) - 1))
        st.session_state["cf_map_step"] = i
        st.session_state["cf_jump_radio"] = step_names[i]   # ✅ move red dot
        st.rerun()

    c1, c2, _ = st.columns([1, 1, 2])

    with c1:
        st.markdown('<div class="fbc-nav-btn">', unsafe_allow_html=True)
        if st.button("⬅️ Back (CF)", disabled=st.session_state["cf_map_step"] == 0):
            _set_step(st.session_state["cf_map_step"] - 1)
        st.markdown('</div>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="fbc-nav-btn">', unsafe_allow_html=True)
        if st.button("➡️ Next unmapped (CF)"):
            target = None
            for i, (k, _) in enumerate(CF_LINES):
                if len(st.session_state["dcf_mapping"].get(k, [])) == 0:
                    target = i
                    break
            if target is None:
                target = len(CF_LINES) - 1
            _set_step(target)
        st.markdown('</div>', unsafe_allow_html=True)

    chosen_step = st.radio(
        "Jump to CF line:",
        step_names,
        key="cf_jump_radio",
        horizontal=True
    )
    st.session_state["cf_map_step"] = step_names.index(chosen_step)

    k, title = CF_LINES[st.session_state["cf_map_step"]]
    stored = clean_defaults(st.session_state["dcf_mapping"].get(k, []), cf_labels)

    if "cf_widget_reset" not in st.session_state:
        st.session_state["cf_widget_reset"] = {}
    st.session_state["cf_widget_reset"].setdefault(k, 0)

    widget_key = f"cf_pick_{k}_{st.session_state['cf_widget_reset'][k]}"

    with st.container(border=True):
        st.markdown(f"#### {title}")

        sel = st.multiselect(
            "Select row(s):",
            cf_labels,
            default=stored,
            key=widget_key
        )

        st.session_state["dcf_mapping"][k] = sel

        if st.button("🧹 Clear selection", key=f"cf_clear_{k}"):
            st.session_state["dcf_mapping"][k] = []
            st.session_state["cf_widget_reset"][k] += 1
            st.rerun()

        if sel:
            idx_list = indices_from_labels(sel)
            preview_vals = cf_df.loc[idx_list, year_cols_cf].sum(axis=0)
            st.caption("Preview (sum of selected rows):")
            st.dataframe(
                pd.DataFrame({"Year": year_cols_cf, "Total": preview_vals.values}),
                hide_index=True,
                width='stretch'
            )

    out = {}
    for kk, _ in CF_LINES:
        out[kk] = indices_from_labels(st.session_state["dcf_mapping"].get(kk, []))
    return out

cf_idx = map_cf_wizard(cf_df, year_cols_cf)

dep_cf_idx_list    = cf_idx["dep"]
capex_cf_idx_list  = cf_idx["capex"]
int_cf_idx_list    = cf_idx["interest"]

# ---------------------------------------------------------
# INCOME STATEMENT FORECASTING
# ---------------------------------------------------------
# ---------------------------------------------------------
# INCOME STATEMENT — OPTION C (WIZARD: steps + progress + next unmapped)
# ---------------------------------------------------------
CORE_LINES = [
    ("rev", "Revenue"),
    ("cos", "Cost of Sales / Raw Materials (optional)"),
    ("gp", "Gross Profit"),
    ("ebitda", "EBITDA"),
    # ✅ ADD THIS LINE
    ("dep", "Depreciation & Amortisation (IS line)"),
    ("op", "Operating Profit / EBIT"),
    ("pbt", "Profit Before Tax"),
    ("tax", "Income Tax (Tax expense)"),
    ("np", "Profit for the Year"),
]


def _labels_from_items(items):
    return ["N/A (not in statement)"] + [f"{i+1}: {str(name)}" for i, name in enumerate(items)]

def map_core_is_totals_wizard(is_df, year_cols_is):
    section("Income Statement — Core Totals Mapping")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    items = list(is_df["Item"].astype(str))
    options = _labels_from_items(items)

    # init state once
    if "is_core_mapping" not in st.session_state:
        st.session_state["is_core_mapping"] = {k: None for k, _ in CORE_LINES}
    if "is_map_step" not in st.session_state:
        st.session_state["is_map_step"] = 0

    # progress
    mapped = sum(1 for k, _ in CORE_LINES if st.session_state["is_core_mapping"].get(k))
    st.progress(mapped / len(CORE_LINES))
    st.caption(f"Mapped: {mapped}/{len(CORE_LINES)}")

    c1, c2, c3 = st.columns([1, 1, 2])

    with c1:
        st.markdown('<div class="fbc-nav-btn">', unsafe_allow_html=True)
        if st.button("⬅️ Back", disabled=st.session_state["is_map_step"] == 0):
            st.session_state["is_map_step"] -= 1
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="fbc-nav-btn">', unsafe_allow_html=True)
        if st.button("➡️ Next unmapped"):
            for i, (k, _) in enumerate(CORE_LINES):
                if not st.session_state["is_core_mapping"].get(k):
                    st.session_state["is_map_step"] = i
                    st.rerun()
            st.session_state["is_map_step"] = len(CORE_LINES) - 1
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # step selector (feels interactive + reduces page length)
    step_names = [name for _, name in CORE_LINES]
    step = st.radio(
        "Jump to line:",
        step_names,
        index=int(st.session_state["is_map_step"]),
        horizontal=True
    )
    st.session_state["is_map_step"] = step_names.index(step)

    # current step UI
    k, title = CORE_LINES[st.session_state["is_map_step"]]
    stored = st.session_state["is_core_mapping"].get(k)
    default = stored if stored in options else "N/A (not in statement)"
    default_index = options.index(default)

    chosen = st.selectbox(
        f"Select statement line for: **{title}**",
        options,
        index=default_index,
        key=f"is_pick_{k}"
    )

    st.session_state["is_core_mapping"][k] = None if chosen.startswith("N/A") else chosen

    # small preview for selected row (makes it feel alive)
    if not chosen.startswith("N/A"):
        idx = int(chosen.split(":", 1)[0]) - 1
        row_vals = is_df.iloc[idx][year_cols_is]
        st.dataframe(
            pd.DataFrame({"Year": year_cols_is, "Value": row_vals.values}),
            hide_index=True,
            width='stretch'
        )

    # convert to indices
    idx_map = {}
    for kk, _ in CORE_LINES:
        v = st.session_state["is_core_mapping"].get(kk)
        idx_map[kk] = (int(v.split(":", 1)[0]) - 1) if v else None

    # required check
    if idx_map["rev"] is None:
        st.error("❌ Revenue must be selected.")
        st.stop()

    return idx_map

# ✅ use this instead of your old mapping call
core_idx = map_core_is_totals_wizard(is_df, year_cols_is)
rev_idx    = core_idx["rev"]
cos_idx    = core_idx["cos"]
gp_idx     = core_idx["gp"]
ebitda_idx = core_idx["ebitda"]
dep_idx    = core_idx["dep"]
op_idx     = core_idx["op"]
pbt_idx    = core_idx["pbt"]
tax_idx    = core_idx["tax"]
np_idx     = core_idx["np"]

# ---------------------------------------------------------
# VALIDATE MAPPING ORDER (top-to-bottom in statement)
# ---------------------------------------------------------
order_pairs = [
    ("Revenue", rev_idx),
    ("Gross Profit", gp_idx),
    ("EBITDA", ebitda_idx),
    ("Operating Profit", op_idx),
    ("Profit Before Tax", pbt_idx),
    ("Tax", tax_idx),
    ("Profit for the Year", np_idx),
]
mapped = [(n, i) for (n, i) in order_pairs if isinstance(i, int)]

# Check monotonic increasing indices
bad = []
for a, b in zip(mapped, mapped[1:]):
    if b[1] <= a[1]:
        bad.append((a, b))

if bad:
    st.error("❌ Mapping order problem: totals must appear top-to-bottom in the statement.")
    st.write("Current mapped order:", mapped)
    st.stop()
# ✅ Only Revenue is mandatory
if rev_idx is None:
    st.error("❌ Revenue must be selected.")
    st.stop()

# ✅ Cost of Sales is OPTIONAL
if cos_idx is None:
    st.warning("⚠️ Cost of Sales / Raw Materials not selected. Forecast will run using other lines as % of revenue.")


revenue_row = is_df.iloc[[rev_idx]]

# Calculate historical growth
calculated_g = avg_revenue_growth(revenue_row, year_cols_is)
# ---------------------------------------------------------
# FORECAST HORIZON (USER-DEFINED)
# ---------------------------------------------------------
if "dcf_forecast_years" not in st.session_state:
    st.session_state["dcf_forecast_years"] = 5
st.markdown("""
<div class="fbc-forecast-label">
    Number of years to forecast
</div>

<style>
.fbc-forecast-label {
    font-family: "Playfair Display", serif;
    font-size: 18px;
    font-weight: 800;
    color: #001a5c;

    display: inline-block;
    padding-left: 10px;
    border-left: 4px solid #003399;

    margin-bottom: 8px;
}

.fbc-forecast-label::after {
    content: "";
    display: block;
    width: 60px;
    height: 2px;
    margin-top: 6px;
    background: linear-gradient(90deg, #003399, transparent);
    border-radius: 2px;
}
</style>
""", unsafe_allow_html=True)
forecast_horizon = st.number_input(
    "Number of years to forecast",
    min_value=1,
    max_value=15,
    value=int(st.session_state["dcf_forecast_years"]),
    step=1,
    key="dcf_forecast_years_input",
    label_visibility="collapsed"
)

st.session_state["dcf_forecast_years"] = forecast_horizon

forecast_years_int = [
    last_hist_year + i
    for i in range(1, forecast_horizon + 1)
]

forecast_cols = [str(y) for y in forecast_years_int]
st.markdown(f"📌 **Calculated Avg Revenue Growth:** {calculated_g:.2%}")

# --- Persistent revenue growth override ---
if "dcf_rev_growth_override" not in st.session_state:
    st.session_state["dcf_rev_growth_override"] = None   # means "not overridden yet"


# Determine what value to display in the input
default_display_value = (
    st.session_state["dcf_rev_growth_override"] * 100
    if st.session_state["dcf_rev_growth_override"] is not None
    else calculated_g * 100
)

# User override input
override_input = st.number_input(
    "Override revenue growth (%) if needed:",
    value=float(default_display_value),
    step=0.1,
    format="%.2f",
)

# Save to session_state as DECIMAL
st.session_state["dcf_rev_growth_override"] = override_input / 100

# Use final revenue growth for forecasting
avg_g = (
    st.session_state["dcf_rev_growth_override"]
    if st.session_state["dcf_rev_growth_override"] is not None
    else calculated_g
)


# ---------------------------------------------------------
# BUILD FORECAST INCOME STATEMENT
# ---------------------------------------------------------
forecast_is = is_df.copy()

# 🔥 ENSURE forecast columns exist
for col in forecast_cols:
    if col not in forecast_is.columns:
        forecast_is[col] = np.nan

# revenue forecast (ALLOW YEAR-BY-YEAR GROWTH)
rev_hist_vals = revenue_row[year_cols_is].values.flatten().astype(float)


# INCOME TAX RATE (Income Tax / Profit Before Tax) → AVERAGE
# ---------------------------------------------------------
avg_tax_ratio = 0.0

if isinstance(tax_idx, int) and isinstance(pbt_idx, int):

    tax_hist_vals = forecast_is.iloc[tax_idx][year_cols_is].values.astype(float)
    pbt_hist_vals = forecast_is.iloc[pbt_idx][year_cols_is].values.astype(float)

    mask = (~np.isnan(tax_hist_vals)) & (~np.isnan(pbt_hist_vals)) & (pbt_hist_vals > 0)

    if mask.any():
        pbt_valid = pbt_hist_vals[mask]
        tax_valid = tax_hist_vals[mask]

        # only profitable years
        profit_mask = pbt_valid > 0

        # KEEP SIGN: tax is negative, so ratio should be negative
        ratios = tax_valid[profit_mask] / pbt_valid[profit_mask]

        # sane cap for negative tax ratios (-60% to 0%)
        ratios = ratios[(ratios <= 0) & (ratios >= -1.50)]

        if len(ratios):
            avg_tax_ratio = float(np.mean(ratios))

# --- UI choice: same growth vs year-by-year growth ---
section("📈 Revenue Growth Method")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
if "dcf_rev_growth_mode" not in st.session_state:
    st.session_state["dcf_rev_growth_mode"] = "Uniform (same % each year)"

growth_mode = st.radio(
    "Choose how you want to apply revenue growth:",
    ["Uniform (same % each year)", "Different growth per year"],
    index=0 if st.session_state["dcf_rev_growth_mode"].startswith("Uniform") else 1,
    key="dcf_rev_growth_mode_radio"
)
st.session_state["dcf_rev_growth_mode"] = growth_mode

# ---------------------------------------------------------
# AUTO YEAR-BY-YEAR GROWTH ENGINE (fade to long-run)
# ---------------------------------------------------------
def auto_growth_curve(start_g: float, terminal_g: float, n: int, speed: float = 0.55):
    """
    Returns a list of n growth rates that fade from start_g to terminal_g.
    speed in (0,1): higher = faster fade.
    """
    out = []
    g = start_g
    for _ in range(n):
        # move part-way toward terminal each year
        g = terminal_g + (g - terminal_g) * (1 - speed)
        out.append(g)
    return out

# --- If year-by-year, store a % for each forecast year ---
if "dcf_yearly_growth_pct" not in st.session_state:
    st.session_state["dcf_yearly_growth_pct"] = {}

yearly_g = {}

if growth_mode == "Different growth per year":
    st.markdown("#### Enter growth for each forecast year (%)")
    for y in forecast_years_int:
        default_pct = st.session_state["dcf_yearly_growth_pct"].get(str(y), avg_g * 100)
        pct = st.number_input(
            f"Growth for {y} (%)",
            value=float(default_pct),
            step=0.1,
            format="%.2f",
            key=f"growth_{y}"
        )
        st.session_state["dcf_yearly_growth_pct"][str(y)] = pct
        yearly_g[y] = pct / 100.0
else:
    # Uniform growth uses avg_g from your existing logic
    for y in forecast_years_int:
        yearly_g[y] = avg_g

# --- Now forecast revenue using the selected growth rates ---
rev_forecast = {}
current_rev = float(rev_hist_vals[-1])

for y in forecast_years_int:
    g_y = yearly_g[y]
    current_rev = current_rev * (1 + g_y)
    rev_forecast[y] = current_rev
    forecast_is.iat[rev_idx, forecast_is.columns.get_loc(str(y))] = current_rev


# ---------------------------------------------------------
# COST / GROSS PROFIT HANDLING (COS OPTIONAL)
# ---------------------------------------------------------
has_cos = isinstance(cos_idx, int)
has_gp  = isinstance(gp_idx, int)

# 1) If GP exists, compute historical GP margin
avg_gp_margin = None
if has_gp:
    gp_hist_vals = forecast_is.iloc[gp_idx][year_cols_is].values.astype(float)
    mask = (rev_hist_vals != 0) & (~np.isnan(gp_hist_vals)) & (~np.isnan(rev_hist_vals))
    gp_margins = gp_hist_vals[mask] / rev_hist_vals[mask]
    gp_margins = gp_margins[(gp_margins > -5) & (gp_margins < 5)]
    avg_gp_margin = float(np.mean(gp_margins)) if len(gp_margins) else 0.30

# ✅ CASE A: GP + COS exist → forecast COS using GP margin (your original approach)
if has_gp and has_cos and avg_gp_margin is not None:

    last_cos_hist = float(forecast_is.iloc[cos_idx][last_hist_label])
    cos_sign = -1 if last_cos_hist < 0 else 1

    for y in forecast_years_int:
        forecast_is.iat[cos_idx, forecast_is.columns.get_loc(str(y))] = (
            cos_sign * rev_forecast[y] * (1 - avg_gp_margin)
        )

    st.success(f"✅ COS forecasted using average GP margin = {avg_gp_margin:.2%}")

# ✅ CASE B: GP exists but COS missing → forecast GP directly
elif has_gp and (not has_cos) and avg_gp_margin is not None:

    for y in forecast_years_int:
        forecast_is.iat[gp_idx, forecast_is.columns.get_loc(str(y))] = (
            rev_forecast[y] * avg_gp_margin
        )

    st.info(f"ℹ️ COS not selected. GP forecasted using average GP margin = {avg_gp_margin:.2%}")

# ✅ CASE C: COS exists but GP missing → forecast COS as % of revenue
elif has_cos and (not has_gp):

    cos_hist_vals = forecast_is.iloc[cos_idx][year_cols_is].values.astype(float)
    cos_ratio = ratio_to_revenue(cos_hist_vals, rev_hist_vals)

    for y in forecast_years_int:
        forecast_is.iat[cos_idx, forecast_is.columns.get_loc(str(y))] = (
            rev_forecast[y] * cos_ratio
        )

    st.info(f"ℹ️ GP not selected. COS forecasted as % of revenue (avg ratio = {cos_ratio:.2%})")

# ✅ CASE D: Neither GP nor COS exists → do nothing special (rest of rows will still forecast)
else:
    st.info("ℹ️ GP and COS not selected. Forecast will rely on other rows as % of revenue.")

# ---------------------------------------------------------
# FORECAST ALL NON-TOTAL ROWS AS % OF REVENUE (NO KEYWORDS)
# ---------------------------------------------------------

# 1) Define "total rows" from mapping (these must be CALCULATED, not forecasted)
total_rows = [gp_idx, ebitda_idx, op_idx, pbt_idx, np_idx]
total_rows = [i for i in total_rows if isinstance(i, int)]
total_set = set(total_rows)

# 2) Revenue is never forecasted as % of itself
protected_set = set([rev_idx]) | total_set

# 3) If COS was forecasted using GP-margin (Case A), don't overwrite COS
gp_cos_mode = (has_gp and has_cos and (avg_gp_margin is not None))

for idx in range(len(forecast_is)):

    # Skip revenue + totals
    if idx in protected_set:
        continue

    # Skip COS if already handled via GP-margin method
    if gp_cos_mode and isinstance(cos_idx, int) and idx == cos_idx:
        continue

    # Forecast row as % of revenue
    row_hist = forecast_is.iloc[idx][year_cols_is].values.astype(float)
    ratio = ratio_to_revenue(row_hist, rev_hist_vals)

    for y in forecast_years_int:
        forecast_is.iat[idx, forecast_is.columns.get_loc(str(y))] = rev_forecast[y] * ratio
# ---------------------------------------------------------
# TOTALS CHAIN ENGINE (previous mapped total -> next mapped total)
# ---------------------------------------------------------

def safe_sum(df, start_i, end_i, col):
    """Sum from start_i to end_i-1 inclusive."""
    if not (isinstance(start_i, int) and isinstance(end_i, int)):
        return np.nan
    if end_i <= start_i:
        return np.nan
    return df.loc[start_i:end_i - 1, col].sum(skipna=True)

# Build totals chain based on the lines YOU mapped
chain = [
    ("REV", rev_idx),
    ("GP", gp_idx),
    ("EBITDA", ebitda_idx),
    ("OP", op_idx),
    ("PBT", pbt_idx),
    ("NP", np_idx),
]

# Keep only mapped items and ensure correct order top-to-bottom in the statement
chain = [(name, idx) for (name, idx) in chain if isinstance(idx, int)]
chain = sorted(chain, key=lambda x: x[1])

for col in forecast_cols:

    # 1) GP special case: if both Revenue and COS exist, derive GP = Revenue + COS
    if isinstance(gp_idx, int) and isinstance(cos_idx, int):
        rev_val = pd.to_numeric(
            forecast_is.iat[rev_idx, forecast_is.columns.get_loc(col)],
            errors="coerce"
        )
        cos_val = pd.to_numeric(
            forecast_is.iat[cos_idx, forecast_is.columns.get_loc(col)],
            errors="coerce"
        )
        rev_val = 0.0 if pd.isna(rev_val) else float(rev_val)
        cos_val = 0.0 if pd.isna(cos_val) else float(cos_val)

        forecast_is.iat[gp_idx, forecast_is.columns.get_loc(col)] = rev_val + cos_val

    # 2) For each mapped total: sum from previous mapped total down to just above current total
    for j in range(1, len(chain)):
        prev_name, prev_idx = chain[j - 1]
        curr_name, curr_idx = chain[j]

        # Skip overwriting GP if we already computed it using Revenue + COS above
        if curr_name == "GP" and isinstance(cos_idx, int):
            continue

        val = safe_sum(forecast_is, prev_idx, curr_idx, col)
        if np.isfinite(val):
            forecast_is.iat[curr_idx, forecast_is.columns.get_loc(col)] = val

    # 3) Tax derived from PBT (if both mapped)
    if isinstance(tax_idx, int) and isinstance(pbt_idx, int):
        pbt_val = pd.to_numeric(
            forecast_is.iat[pbt_idx, forecast_is.columns.get_loc(col)],
            errors="coerce"
        )
        pbt_val = 0.0 if pd.isna(pbt_val) else float(pbt_val)

        forecast_is.iat[tax_idx, forecast_is.columns.get_loc(col)] = pbt_val * avg_tax_ratio

    # 4) Profit for the Year (PAT/NP) MUST be after tax:
    #    NP = PBT + Tax + any after-tax lines between Tax and NP
    if isinstance(np_idx, int) and isinstance(pbt_idx, int):

        pbt_val = pd.to_numeric(
            forecast_is.iat[pbt_idx, forecast_is.columns.get_loc(col)],
            errors="coerce"
        )
        pbt_val = 0.0 if pd.isna(pbt_val) else float(pbt_val)

        tax_val = 0.0
        if isinstance(tax_idx, int):
            tax_val = pd.to_numeric(
                forecast_is.iat[tax_idx, forecast_is.columns.get_loc(col)],
                errors="coerce"
            )
            tax_val = 0.0 if pd.isna(tax_val) else float(tax_val)

        # add any after-tax adjustments between tax line and NP line
        extra_after_tax = 0.0
        if isinstance(tax_idx, int) and (tax_idx + 1) <= (np_idx - 1):
            extra_after_tax = forecast_is.loc[tax_idx + 1: np_idx - 1, col].sum(skipna=True)

        forecast_is.iat[np_idx, forecast_is.columns.get_loc(col)] = pbt_val + tax_val + extra_after_tax


# 5) Cleanup: remove None and force numeric
all_year_cols = [c for c in forecast_is.columns if c != "Item"]
forecast_is[all_year_cols] = forecast_is[all_year_cols].replace(
    {None: np.nan, "None": np.nan, "none": np.nan}
)
forecast_is[all_year_cols] = forecast_is[all_year_cols].apply(pd.to_numeric, errors="coerce")

# ---------------------------------------------------------
# STORE FORECASTED NET PROFIT (Profit for the Year) FOR COMPARABLES
# ---------------------------------------------------------
dcf_np_forecast = {}

if np_idx is not None:
    for y in forecast_years_int:
        col = str(y)
        if isinstance(np_idx, int):
            val = forecast_is.iat[np_idx, forecast_is.columns.get_loc(col)]
            val = float(val) if pd.notna(val) else 0.0
        else:
            val = 0.0
        dcf_np_forecast[col] = val
else:
    dcf_np_forecast = {}

# =========================================================
# STORE ALL NET PROFIT VALUES (HISTORICAL + FORECAST)
# =========================================================
dcf_profit_all = {}

# 1. Include historical
if np_idx is not None:
    for col in year_cols_is:  # historical labels (strings)
        val = float(is_df.iloc[np_idx][col])
        dcf_profit_all[col] = val

# 2. Include forecast
for y in forecast_years_int:
    col = str(y)
    if np_idx is not None:
        val = float(forecast_is.iloc[np_idx][col])
        dcf_profit_all[col] = val

# Save to session_state
st.session_state["dcf_profit_all"] = dcf_profit_all
FORECAST_BLUE = "#1d4ed8"  # strong blue (tailwind blue-700)

def style_forecast_columns(styler, forecast_cols):
    # color ONLY the forecast year columns
    def _blue_forecast(col_name):
        if str(col_name) in set(map(str, forecast_cols)):
            return f"color: {FORECAST_BLUE}; font-weight: 700;"
        return ""

    return styler.apply(
        lambda df: pd.DataFrame(
            {c: [_blue_forecast(c)] * len(df) for c in df.columns},
            index=df.index
        ),
        axis=None
    )
# ---------------------------------------------------------
# STYLING: Historical vs Forecast columns
# ---------------------------------------------------------
HIST_COLOR = "#111827"     # near-black / dark blue-gray
FORECAST_COLOR = "#1d4ed8" # light blue (forecast)

def style_hist_vs_forecast(styler, hist_cols, forecast_cols):
    hist_cols = set(map(str, hist_cols))
    forecast_cols = set(map(str, forecast_cols))

    def apply_colors(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)

        for c in df.columns:
            if c in forecast_cols:
                styles[c] = f"color: {FORECAST_COLOR}; font-weight: 700;"
            elif c in hist_cols:
                styles[c] = f"color: {HIST_COLOR};"

        return styles

    return styler.apply(apply_colors, axis=None)
section(
    f"📘 Forecasted Income Statement ({forecast_horizon} years, USD)"
)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
def style_forecast_headers(styler, forecast_cols, color="#1d4ed8"):
    """
    Color ONLY the column headers for forecast years.
    Works with pandas Styler in Streamlit.
    """
    cols = list(styler.data.columns)
    forecast_set = set(map(str, forecast_cols))

    styles = []
    for i, c in enumerate(cols):
        if str(c) in forecast_set:
            # header selector for this column
            styles.append({
                "selector": f"th.col_heading.level0.col{i}",
                "props": [("color", color), ("font-weight", "700")]
            })
    return styler.set_table_styles(styles, overwrite=False)

# ---------------------------------------------------------
# DISPLAY: Historical vs Forecast styled table
# ---------------------------------------------------------

fmt_map = {
    c: "{:,.0f}".format
    for c in forecast_is.select_dtypes(include=[np.number]).columns
}
styled_is = (
    forecast_is.style
        .format(fmt_map, na_rep="")
)

styled_is = style_hist_vs_forecast(
    styled_is,
    hist_cols=year_cols_is,
    forecast_cols=forecast_cols
)
# ✅ NEW: make forecast year headers blue too
styled_is = style_forecast_headers(
    styled_is,
    forecast_cols=forecast_cols,
    color=FORECAST_COLOR
)

# ----------------------------
# DOWNLOADS (because HTML table has no Streamlit toolbar)
# ----------------------------
c_dl1, c_dl2, _ = st.columns([1, 1, 3])

# Excel download
buf_xlsx = io.BytesIO()
forecast_is.to_excel(buf_xlsx, index=False, sheet_name="Forecast_IS")
buf_xlsx.seek(0)

with c_dl1:
    st.download_button(
        "⬇️ Download IS (Excel)",
        data=buf_xlsx,
        file_name="Forecasted_Income_Statement.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_forecast_is_xlsx"
    )
# ----------------------------
# DISPLAY (Streamlit-native like other tables)
# ----------------------------
st.dataframe(
    styled_is,
    width='stretch',
    hide_index=True
)

# Extract EBITDA row for forecast years
if isinstance(ebitda_idx, int):
    ebitda_forecast_vals = np.array(
        [float(forecast_is.iat[ebitda_idx, forecast_is.columns.get_loc(str(y))])
         for y in forecast_years_int],
        dtype=float
    )
else:
    ebitda_forecast_vals = np.zeros(len(forecast_years_int))
# ---------------------------------------------------------
# SAVE ALL EBITDA VALUES (HISTORICAL + FORECAST)
# ---------------------------------------------------------

dcf_all_ebitda = {}

# 1️⃣ Save historical EBITDA
if isinstance(ebitda_idx, int):
    for y in year_cols_is:
        col_idx = forecast_is.columns.get_loc(str(y))
        val = forecast_is.iat[ebitda_idx, col_idx]
        dcf_all_ebitda[str(y)] = float(val) if pd.notna(val) else 0.0

# 2️⃣ Save forecast EBITDA
if isinstance(ebitda_idx, int):
    for y in forecast_years_int:
        col_idx = forecast_is.columns.get_loc(str(y))
        val = forecast_is.iat[ebitda_idx, col_idx]
        dcf_all_ebitda[str(y)] = float(val) if pd.notna(val) else 0.0

# 3️⃣ Store into session_state (BOTH KEYS)
st.session_state["dcf_ebitda_all"] = dcf_all_ebitda
st.session_state["dcf_ebitda_forecast"] = dcf_all_ebitda   # <-- backward compatibility

# Save EVERYTHING to session_state
st.session_state["dcf_ebitda_all"] = dcf_all_ebitda
# ---------------------------------------------------------
# ✅ DEPRECIATION (prefer mapped IS dep line, else fallback)
# ---------------------------------------------------------
dep_forecast_vals = None

# 1) Best: mapped Depreciation row from IS wizard
if isinstance(dep_idx, int):
    dep_forecast_vals = np.array(
        [pd.to_numeric(forecast_is.iat[dep_idx, forecast_is.columns.get_loc(str(y))], errors="coerce")
         for y in forecast_years_int],
        dtype=float
    )
    dep_forecast_vals = np.nan_to_num(dep_forecast_vals, nan=0.0)

else:
    # 2) Fallback: if user mapped Dep in Cash Flow wizard
    if dep_cf_idx_list:
        dep_forecast_vals = np.array(
            [float(cf_df.loc[dep_cf_idx_list, str(y)].sum(skipna=True)) if str(y) in cf_df.columns else 0.0
             for y in forecast_years_int],
            dtype=float
        )
        dep_forecast_vals = np.nan_to_num(dep_forecast_vals, nan=0.0)

    else:
        # 3) Last fallback: ratio to revenue from historical CF dep (or 0)
        common = [c for c in year_cols_cf if c in year_cols_is]
        if common and dep_cf_idx_list:
            dep_ratio = ratio_to_revenue(
                cf_df.loc[dep_cf_idx_list, common].sum(axis=0).values.astype(float),
                revenue_row[common].values.flatten().astype(float)
            )
        else:
            dep_ratio = 0.0

        dep_forecast_vals = np.array(
            [rev_forecast[y] * dep_ratio for y in forecast_years_int],
            dtype=float
        )

# Optional: store for other pages / exports
st.session_state["dcf_dep_forecast"] = {str(y): float(dep_forecast_vals[i]) for i, y in enumerate(forecast_years_int)}

# After building rev_forecast dict
st.session_state["dcf_rev_forecast"] = {str(y): float(rev_forecast[y]) for y in forecast_years_int}
st.session_state["forecast_is_df"] = forecast_is.copy()

# ---------------------------------------------------------
# CAPITAL STRUCTURE FROM BS: Total Debt, Cash, CA, CL
# ---------------------------------------------------------
common_hist_bs = [c for c in year_cols_bs if c in year_cols_is]
bs_year_used_label = common_hist_bs[-1] if common_hist_bs else year_cols_bs[-1]

total_debt = 0.0
if debt_idx_list:
    total_debt = float(bs_df.loc[debt_idx_list, bs_year_used_label].sum(skipna=True))

cash_bal = 0.0
if cash_idx_list:
    cash_bal = float(bs_df.loc[cash_idx_list, bs_year_used_label].sum(skipna=True))

# equity: try some standard keywords
total_equity = 0.0
if equity_idx_list:
    total_equity = float(bs_df.loc[equity_idx_list, bs_year_used_label].sum(skipna=True))


net_debt = total_debt - cash_bal
de_ratio = (total_debt / total_equity) if total_equity != 0 else 0.0
c_cap5 = st.columns(1)[0]


# Save BS capital structure into session_state for other pages
st.session_state["total_debt"] = float(total_debt)
st.session_state["cash_balance"] = float(cash_bal)
st.session_state["net_debt"] = float(net_debt)
st.session_state["book_equity"] = float(total_equity)
st.session_state["de_ratio"] = float(de_ratio)

# ---------------------------------------------------------
# 🟦 WORKING CAPITAL MODULE (HISTORICAL → WC% → FORECAST → ΔWC)
# ---------------------------------------------------------
section("📘 Working Capital Calculation (Historical & Forecast)")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
delta_wc_forecast_vals = np.zeros(len(forecast_years_int))

if ca_idx_list and cl_idx_list:

    # -------- 1️⃣ HISTORICAL WC (CA - CL)
    section("Historical Working Capital (CA - CL)")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    ca_hist = bs_df.loc[ca_idx_list, year_cols_bs].sum(axis=0)
    cl_hist = bs_df.loc[cl_idx_list, year_cols_bs].sum(axis=0)
    wc_hist = ca_hist - cl_hist

    df_wc_hist = pd.DataFrame({
        "Year": year_cols_bs,
        "Current Assets": ca_hist.values,
        "Current Liabilities": cl_hist.values,
        "Working Capital (CA-CL)": wc_hist.values,
    })

    st.dataframe(
        df_wc_hist.style.format({
            "Current Assets": "{:,.0f}",
            "Current Liabilities": "{:,.0f}",
            "Working Capital (CA-CL)": "{:,.0f}",
        }),
        width='stretch'    )
    # 2️⃣ WC% OF SALES
    section("Historical Working Capital as % of Sales")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    common_hist = [c for c in year_cols_is if c in wc_hist.index]

    wc_vals_hist = wc_hist[common_hist].astype(float).values
    rev_vals_hist = revenue_row[common_hist].values.flatten().astype(float)

    # avoid divide-by-zero warnings
    wc_percent_hist = np.where(rev_vals_hist != 0, wc_vals_hist / rev_vals_hist, 0.0)

    df_wc_pct = pd.DataFrame({
        "Year": common_hist,
        "Working Capital": wc_vals_hist,
        "Revenue": rev_vals_hist,
        "WC % of Sales": wc_percent_hist,
    })

    # ---------------------------------------------------------
    # ✅ ANALYST FILTER (Include/Exclude each WC% year)
    # ---------------------------------------------------------
    # session store: which years are included
    if "dcf_wc_include_years" not in st.session_state:
        st.session_state["dcf_wc_include_years"] = {str(y): True for y in common_hist}

    # ensure new years added automatically
    for y in common_hist:
        st.session_state["dcf_wc_include_years"].setdefault(str(y), True)

    df_wc_pct_editor = df_wc_pct.copy()
    df_wc_pct_editor["Include"] = [bool(st.session_state["dcf_wc_include_years"].get(str(y), True)) for y in
                                   common_hist]

    edited_wc = st.data_editor(
        df_wc_pct_editor,
        width="stretch",
        hide_index=True,
        disabled=["Year", "Working Capital", "Revenue", "WC % of Sales"],
        column_config={
            "Include": st.column_config.CheckboxColumn("Include",
                                                       help="Tick to include this year's WC% in the average"),
            "Working Capital": st.column_config.NumberColumn("Working Capital", format="%,d"),
            "Revenue": st.column_config.NumberColumn("Revenue", format="%,d"),
            "WC % of Sales": st.column_config.NumberColumn("WC % of Sales", format="%.2f%%"),
        },
        key="dcf_wc_pct_editor"
    )

    # write back checkbox state to session
    st.session_state["dcf_wc_include_years"] = {
        str(row["Year"]): bool(row["Include"])
        for _, row in edited_wc.iterrows()
    }

    # ---------------------------------------------------------
    # 3️⃣ WC% OF SALES — USER CHOICE (Average vs Most Recent) [PERSISTENT]
    # ---------------------------------------------------------
    include_mask = edited_wc["Include"].astype(bool).values
    wc_percent_array = edited_wc["WC % of Sales"].astype(float).values

    # sanity filter + include filter
    mask_valid = (wc_percent_array > -5) & (wc_percent_array < 5) & include_mask
    wc_percent_clean = wc_percent_array[mask_valid]

    wc_percent_mean = float(np.mean(wc_percent_clean)) if len(wc_percent_clean) else 0.0

    # most recent INCLUDED year
    included_years = edited_wc.loc[edited_wc["Include"] == True, "Year"].astype(str).tolist()

    if included_years:
        last_year = included_years[-1]
        last_wc = float(wc_hist[last_year])
        last_rev = float(revenue_row[last_year].values[0])
        wc_percent_last = (last_wc / last_rev) if last_rev != 0 else 0.0
    else:
        # if user unticks all, fallback safely
        last_year = common_hist[-1]
        last_wc = float(wc_hist[last_year])
        last_rev = float(revenue_row[last_year].values[0])
        wc_percent_last = (last_wc / last_rev) if last_rev != 0 else 0.0
        st.warning("⚠️ You excluded all years. Using the last available year as fallback for 'Most recent'.")

    section("✅ Working Capital Assumption (WC % of Sales)")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    # ✅ INITIALIZE SESSION STATE ONCE
    if "dcf_wc_pct_method" not in st.session_state:
        st.session_state["dcf_wc_pct_method"] = "last"  # default = most recent

    wc_choice = st.radio(
        "Which WC% of Sales should be used for forecasting?",
        [
            f"Use average of historical WC% ({wc_percent_mean:.2%})",
            f"Use most recent WC% ({last_year}) = {wc_percent_last:.2%}"
        ],
        index=0 if st.session_state["dcf_wc_pct_method"] == "average" else 1,
        key="dcf_wc_pct_method_radio"
    )

    # ✅ UPDATE SESSION STATE EXPLICITLY
    if "average" in wc_choice.lower():
        st.session_state["dcf_wc_pct_method"] = "average"
        wc_percent_avg = wc_percent_mean
        st.success(f"✅ Using historical average WC% of Sales = {wc_percent_avg:.2%}")
    else:
        st.session_state["dcf_wc_pct_method"] = "last"
        wc_percent_avg = wc_percent_last
        st.info(f"📌 Using most recent WC% of Sales ({last_year}) = {wc_percent_avg:.2%}")

    # 4️⃣ FORECAST WC
    section("Forecast Working Capital")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    wc_forecast_vals = np.array(
        [rev_forecast[y] * wc_percent_avg for y in forecast_years_int],
        dtype=float
    )

    df_wc_forecast = pd.DataFrame({
        "Year": forecast_years_int,
        "Forecast Revenue": [rev_forecast[y] for y in forecast_years_int],
        "Forecast WC": wc_forecast_vals,
    })

    st.dataframe(
        df_wc_forecast.style.format({
            "Forecast Revenue": "{:,.0f}",
            "Forecast WC": "{:,.0f}",
        }),
        width='stretch'    )

    # 5️⃣ ΔWC = OLD – NEW
    section("Change in Working Capital (ΔWC = Old – New)")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    last_wc_hist_value = float(wc_hist[common_hist[-1]])

    prev_wc = last_wc_hist_value
    delta_list = []

    for wc_new in wc_forecast_vals:
        delta_list.append(prev_wc - wc_new)  # Old – New
        prev_wc = wc_new

    delta_wc_forecast_vals = np.array(delta_list, dtype=float)

    df_delta_wc = pd.DataFrame({
        "Year": forecast_years_int,
        "Forecast WC": wc_forecast_vals,
        "ΔWC (Old – New)": delta_wc_forecast_vals,
    })

    st.dataframe(
        df_delta_wc.style.format({
            "Forecast WC": "{:,.0f}",
            "ΔWC (Old – New)": "{:,.0f}",
        }),
        width='stretch'    )

else:
    st.warning("⚠️ Please select Current Assets and Current Liabilities rows first.")

# Capital structure summary
section("Capital Structure & Working Capital (from Balance Sheet)")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
c_cap1, c_cap2, c_cap3, c_cap4 = st.columns(4)
with c_cap1:
    st.metric(f"Total Debt ({bs_year_used_label})", f"{total_debt:,.0f}")
with c_cap2:
    st.metric(f"Cash & Equivalents ({bs_year_used_label})", f"{cash_bal:,.0f}")
with c_cap3:
    st.metric("Net Debt", f"{net_debt:,.0f}")
with c_cap4:
    st.metric("D/E Ratio", f"{de_ratio:.2f}x")
with c_cap5:
    st.metric(f"Equity ({bs_year_used_label})", f"{total_equity:,.0f}")


# ---------------------------------------------------------
# COST OF DEBT (Interest / Debt)
# ---------------------------------------------------------
int_is_idx_list = find_row_indices(is_df, ["net finance costs","net finance cost", "finance costs", "interest expense", "interest paid"])
if int_is_idx_list:
    interest_last = float(is_df.loc[int_is_idx_list, last_hist_label].sum(skipna=True))
else:
    if int_cf_idx_list:
        interest_last = float(cf_df.loc[int_cf_idx_list, bs_year_used_label].sum(skipna=True))
    else:
        interest_last = 0.0

if total_debt != 0:
    cost_of_debt = abs(interest_last) / abs(total_debt)
else:
    cost_of_debt = 0.0

rd_auto = cost_of_debt

# ---------------------------------------------------------
# DCF PARAMETERS — AUTO + OVERRIDE (WITH 2 OPTIONAL UPLOADS)
# ---------------------------------------------------------
st.markdown("---")
section("💰 DCF Parameters (Auto + Override)")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
# =============== helpers ===============
def _to_decimal(x):
    """Accepts 0.15 or 15; returns decimal 0.15"""
    try:
        x = float(x)
    except Exception:
        return None
    return x / 100.0 if x > 1.5 else x

def _load_country_params_df(file_or_path) -> pd.DataFrame:
    """
    Excel required columns (flexible match):
      Country | ERP | Default Spread
    Returns normalized df with: Country, ERP, DefaultSpread
    """
    df = pd.read_excel(file_or_path)
    df.columns = [str(c).strip() for c in df.columns]

    col_country = [c for c in df.columns if c.lower() == "country"]
    col_erp = [c for c in df.columns if c.lower() in ["erp", "equity risk premium", "equity_risk_premium"]]
    col_spread = [c for c in df.columns if c.lower() in ["default spread", "default_spread", "spread"]]

    if not (col_country and col_erp and col_spread):
        raise ValueError("Excel must contain columns: Country, ERP, Default Spread")

    out = df[[col_country[0], col_erp[0], col_spread[0]]].copy()
    out.columns = ["Country", "ERP", "DefaultSpread"]
    out["Country"] = out["Country"].astype(str).str.strip()
    return out


def init_widget_key(widget_key: str, master_key: str, default_val: float):
    """
    IMPORTANT: only set widget key if it doesn't exist yet
    (prevents StreamlitAPIException).
    """
    if master_key not in st.session_state:
        st.session_state[master_key] = float(default_val)
    if widget_key not in st.session_state:
        st.session_state[widget_key] = float(st.session_state[master_key])

# =============== layout ===============
left, right = st.columns([1.15, 1.0], vertical_alignment="top")

# =========================================================
# LEFT: Country ERP + Default Spread upload toggle
# =========================================================
with left:
    section("🌍 Country ERP & Default Spread (Auto RF + MRP)")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    # init upload states
    st.session_state.setdefault("dcf_country_upload_enabled", False)
    st.session_state.setdefault("dcf_country_params_bytes", None)
    st.session_state.setdefault("dcf_country_params_name", None)

    # toggle
    country_upload = st.checkbox(
        "📤 Upload Country ERP + Default Spread Excel (optional)",
        value=st.session_state["dcf_country_upload_enabled"],
        key="dcf_country_upload_enabled_ui"
    )
    st.session_state["dcf_country_upload_enabled"] = country_upload

    if country_upload:
        st.caption("Required Excel columns: **Country, ERP, Default Spread**")
        up_country = st.file_uploader(
            "Upload Country params Excel",
            type=["xlsx"],
            key="dcf_country_params_uploader"
        )
        if up_country is not None:
            st.session_state["dcf_country_params_bytes"] = up_country.getvalue()
            st.session_state["dcf_country_params_name"] = up_country.name
    else:
        st.session_state["dcf_country_params_bytes"] = None
        st.session_state["dcf_country_params_name"] = None

    # load params df (uploaded takes precedence)
    df_params = None
    params_source = None

    try:
        if country_upload and st.session_state["dcf_country_params_bytes"] is not None:
            df_params = _load_country_params_df(io.BytesIO(st.session_state["dcf_country_params_bytes"]))
            params_source = f"Uploaded: {st.session_state.get('dcf_country_params_name','(file)')}"
        else:
            if DCF_PARAMS_PATH.exists():
                df_params = _load_country_params_df(DCF_PARAMS_PATH)
                params_source = f"Default file: {DCF_PARAMS_PATH.name}"
            else:
                st.warning(f"⚠️ Missing default file: {DCF_PARAMS_PATH}. Upload a file above.")
    except Exception as e:
        st.error(f"❌ Country params file error: {e}")
        df_params = None

    if params_source:
        st.caption(f"Source: **{params_source}**")

    # choose country and get ERP + spread
    auto_erp_dec = None
    auto_spread_dec = None

    if df_params is not None and not df_params.empty:
        country_list = sorted(df_params["Country"].dropna().astype(str).unique().tolist())
        default_country = "Zimbabwe" if "Zimbabwe" in country_list else (country_list[0] if country_list else None)

        if default_country is not None:
            chosen_country = st.selectbox(
                "Select country (auto ERP + Default Spread):",
                country_list,
                index=country_list.index(default_country),
                key="dcf_country_select"
            )

            row = df_params[df_params["Country"].astype(str) == str(chosen_country)]
            if not row.empty:
                auto_erp_dec = _to_decimal(row.iloc[0]["ERP"])               # ERP -> MRP
                auto_spread_dec = _to_decimal(row.iloc[0]["DefaultSpread"])  # used in RF formula

    # Zimbabwe Avg Cost of Debt (USD)
    st.session_state.setdefault("dcf_zim_avg_cost_debt_pct", 18.0)
    zim_avg_cod_pct = st.number_input(
        "Average cost of debt Zimbabwe (US$) (%)",
        value=float(st.session_state["dcf_zim_avg_cost_debt_pct"]),
        step=0.1,
        key="dcf_zim_avg_cost_debt_pct_input"
    )
    st.session_state["dcf_zim_avg_cost_debt_pct"] = zim_avg_cod_pct
    zim_avg_cod = zim_avg_cod_pct / 100.0

    # Derive Auto RF & Auto MRP (MRP=ERP)
    auto_mrp_pct = (auto_erp_dec * 100) if auto_erp_dec is not None else None
    auto_rf_pct = ((zim_avg_cod - auto_spread_dec) * 100) if (auto_spread_dec is not None) else None

    if auto_mrp_pct is not None and auto_rf_pct is not None:
        st.success(
            f"✅ Auto from Excel: MRP={auto_mrp_pct:.2f}% | "
            f"RF=(Avg CoD ZW USD − Spread)={auto_rf_pct:.2f}%"
        )
    else:
        st.info("ℹ️ Auto values not available yet (check Excel columns/values).")


# =========================================================
# RIGHT: Industry Betas upload toggle
# =========================================================
with right:
    section(" 🧩 Industry Unlevered Betas (βu)")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    st.session_state.setdefault("dcf_beta_upload_enabled", False)
    st.session_state.setdefault("dcf_beta_file_bytes", None)
    st.session_state.setdefault("dcf_beta_file_name", None)

    beta_upload = st.checkbox(
        "📤 Upload Industry Betas Excel (optional)",
        value=st.session_state["dcf_beta_upload_enabled"],
        key="dcf_beta_upload_enabled_ui"
    )
    st.session_state["dcf_beta_upload_enabled"] = beta_upload

    if beta_upload:
        st.caption("Required Excel columns: **Industry Name, Unlevered beta**")
        up_beta = st.file_uploader(
            "Upload Industry betas Excel",
            type=["xlsx"],
            key="dcf_beta_uploader"
        )
        if up_beta is not None:
            st.session_state["dcf_beta_file_bytes"] = up_beta.getvalue()
            st.session_state["dcf_beta_file_name"] = up_beta.name
    else:
        st.session_state["dcf_beta_file_bytes"] = None
        st.session_state["dcf_beta_file_name"] = None


# =========================================================
# INIT defaults once (master keys + states)
# =========================================================
if "dcf_init" not in st.session_state:
    st.session_state["dcf_rf_pct"] = float(auto_rf_pct) if auto_rf_pct is not None else 11.61
    st.session_state["dcf_mrp_pct"] = float(auto_mrp_pct) if auto_mrp_pct is not None else 13.82
    st.session_state["dcf_tax_pct"] = 25.0
    st.session_state["dcf_unlevered_beta"] = 1.00
    st.session_state["dcf_terminal_g_pct"] = 5.0

    st.session_state["dcf_use_auto_params"] = True

    # beta states
    st.session_state["dcf_industries_selected"] = []
    st.session_state["dcf_beta_blend_method"] = "Simple average"
    st.session_state["dcf_industry_weights"] = {}
    st.session_state["dcf_beta_manual_mode"] = False
    st.session_state["dcf_beta_manual_value"] = None
    st.session_state["dcf_beta_auto_last"] = None

    st.session_state["dcf_init"] = True

# =========================================================
# Auto vs Override toggle (RF & MRP)  ✅ HARD SYNC FIX
#    When auto is ON and we have auto values:
#      - force RF textbox to auto_rf_pct
#      - force MRP textbox to auto_mrp_pct
# =========================================================
use_auto = st.checkbox(
    "Use Auto (from Excel) for RF & MRP",
    value=bool(st.session_state.get("dcf_use_auto_params", True)),
    key="dcf_use_auto_params_ui"
)
st.session_state["dcf_use_auto_params"] = use_auto

# Build signature of the auto sources
auto_signature = (
    float(auto_rf_pct) if auto_rf_pct is not None else None,
    float(auto_mrp_pct) if auto_mrp_pct is not None else None,
    st.session_state.get("dcf_country_select", None),
    float(st.session_state.get("dcf_zim_avg_cost_debt_pct", 0.0)),
)

# Track previous signature so we only "snap" inputs when auto data changes
st.session_state.setdefault("dcf_auto_signature", None)

should_snap_to_auto = (
    use_auto
    and (auto_rf_pct is not None)
    and (auto_mrp_pct is not None)
    and (auto_signature != st.session_state["dcf_auto_signature"])
)

# If auto is ON and signature changed -> update BOTH master keys AND widget keys
# so the textboxes visually change immediately.
if should_snap_to_auto:
    # master
    st.session_state["dcf_rf_pct"] = float(auto_rf_pct)
    st.session_state["dcf_mrp_pct"] = float(auto_mrp_pct)

    # widget keys (these control the textbox displayed values)
    st.session_state["dcf_rf_pct_input"] = float(auto_rf_pct)
    st.session_state["dcf_mrp_pct_input"] = float(auto_mrp_pct)

    st.session_state["dcf_auto_signature"] = auto_signature
elif not use_auto:
    # If auto OFF, don't overwrite manual values; just reset signature so next ON snaps again
    st.session_state["dcf_auto_signature"] = None

# =========================================================
# Backfill widget keys (safe)
# =========================================================
init_widget_key("dcf_rf_pct_input", "dcf_rf_pct", 11.61)
init_widget_key("dcf_mrp_pct_input", "dcf_mrp_pct", 13.82)
init_widget_key("dcf_tax_pct_input", "dcf_tax_pct", 25.0)
init_widget_key("dcf_unlevered_beta_input", "dcf_unlevered_beta", 1.0)
init_widget_key("dcf_terminal_g_pct_input", "dcf_terminal_g_pct", 5.0)

# =========================================================
# Main input widgets
# =========================================================
col1, col2 = st.columns(2)

with col1:
    rf_input = st.number_input("Risk-free rate (%)", step=0.1, key="dcf_rf_pct_input")
    mrp_input = st.number_input("Market risk premium (%)", step=0.1, key="dcf_mrp_pct_input")
    tax_input = st.number_input("Tax rate (%)", step=0.5, key="dcf_tax_pct_input")

with col2:
    # Load betas df: uploaded takes precedence
    betas_df = None
    beta_source = None
    try:
        if st.session_state.get("dcf_beta_upload_enabled") and st.session_state.get("dcf_beta_file_bytes") is not None:
            betas_df = _load_unlevered_betas_any(io.BytesIO(st.session_state["dcf_beta_file_bytes"]))
            beta_source = f"Uploaded: {st.session_state.get('dcf_beta_file_name','(file)')}"
        else:
            if UNLEVERED_BETAS_PATH.exists():
                mtime = UNLEVERED_BETAS_PATH.stat().st_mtime  # ✅ changes when you save Excel
                betas_df = _load_unlevered_betas_any(UNLEVERED_BETAS_PATH, file_mtime=mtime)
                beta_source = f"Default file: {UNLEVERED_BETAS_PATH.name}"

            else:
                st.warning(f"⚠️ Missing default file: {UNLEVERED_BETAS_PATH}. Upload a file above.")
    except Exception as e:
        st.warning(f"⚠️ Could not load industry betas: {e}")
        betas_df = None

    if beta_source:
        st.caption(f"Source: **{beta_source}**")

    # Multi-industry selector for blended beta
    if betas_df is not None and not betas_df.empty:
        industry_list = betas_df["Industry"].tolist()

        selected = st.multiselect(
            "Select Industry / Industries (for blended βu):",
            industry_list,
            default=[i for i in st.session_state.get("dcf_industries_selected", []) if i in industry_list],
            key="dcf_industries_multiselect"
        )
        st.session_state["dcf_industries_selected"] = selected

        blend_method = st.radio(
            "How should industries be combined?",
            ["Simple average", "Weighted average"],
            index=0 if st.session_state.get("dcf_beta_blend_method", "Simple average") == "Simple average" else 1,
            key="dcf_beta_blend_method_radio",
            horizontal=True
        )
        st.session_state["dcf_beta_blend_method"] = blend_method

        beta_u_auto = None
        if selected:
            sub = betas_df[betas_df["Industry"].isin(selected)].copy()

            if blend_method == "Simple average":
                beta_u_auto = float(sub["UnleveredBeta"].mean())
            else:
                st.markdown("#### Enter weights (they will be normalized to 100%)")
                weights = []
                for ind in selected:
                    default_w = float(st.session_state.get("dcf_industry_weights", {}).get(ind, 1.0))
                    w = st.number_input(
                        f"Weight for {ind}",
                        min_value=0.0,
                        value=default_w,
                        step=1.0,
                        key=f"w_{ind}"
                    )
                    st.session_state.setdefault("dcf_industry_weights", {})
                    st.session_state["dcf_industry_weights"][ind] = w
                    weights.append(w)

                total_w = float(sum(weights))
                if total_w <= 0:
                    st.error("❌ Total weight must be > 0.")
                else:
                    sub = sub.sort_values("Industry").reset_index(drop=True)
                    w_norm = np.array([st.session_state["dcf_industry_weights"][ind] for ind in sub["Industry"]]) / total_w
                    beta_u_auto = float(np.sum(sub["UnleveredBeta"].values * w_norm))

                    dfw = pd.DataFrame({
                        "Industry": sub["Industry"].values,
                        "UnleveredBeta": sub["UnleveredBeta"].values,
                        "Weight (raw)": [st.session_state["dcf_industry_weights"][i] for i in sub["Industry"]],
                        "Weight (norm %)": (w_norm * 100).round(2)
                    })
                    st.dataframe(dfw, width="stretch", hide_index=True)

            if beta_u_auto is not None and np.isfinite(beta_u_auto):
                st.session_state["dcf_beta_auto_last"] = float(beta_u_auto)
                st.caption(f"Blended industry βu (auto): **{beta_u_auto:.2f}**")
        else:
            st.info("Select at least 1 industry to auto-fill βu.")

    # Auto vs manual beta mode
    st.session_state.setdefault("dcf_beta_manual_mode", False)
    beta_mode = st.radio(
        "Unlevered beta mode:",
        ["Use Auto (from industries)", "Manual override (type my own βu)"],
        index=1 if st.session_state["dcf_beta_manual_mode"] else 0,
        key="dcf_beta_mode_radio",
        horizontal=True
    )
    st.session_state["dcf_beta_manual_mode"] = beta_mode.startswith("Manual")

    # Apply auto beta button
    if not st.session_state["dcf_beta_manual_mode"]:
        auto_beta = st.session_state.get("dcf_beta_auto_last")
        if auto_beta is not None and np.isfinite(auto_beta):
            st.caption(f"Auto βu available: {auto_beta:.2f}")
            if st.button("✅ Apply Auto βu to input", key="apply_auto_beta_btn"):
                st.session_state["dcf_unlevered_beta_input"] = float(auto_beta)

    beta_u_input = st.number_input("Unlevered beta (asset beta)", step=0.05, key="dcf_unlevered_beta_input")
    if st.session_state.get("dcf_beta_manual_mode", False):
        st.session_state["dcf_beta_manual_value"] = float(beta_u_input)

    g_input = st.number_input("Terminal growth rate (%)", step=0.1, key="dcf_terminal_g_pct_input")

# =========================================================
# Save user inputs to master keys (USED BY FORMULAS)
# =========================================================
st.session_state["dcf_rf_pct"] = float(rf_input)
st.session_state["dcf_mrp_pct"] = float(mrp_input)
st.session_state["dcf_tax_pct"] = float(tax_input)
st.session_state["dcf_unlevered_beta"] = float(beta_u_input)
st.session_state["dcf_terminal_g_pct"] = float(g_input)

# Decimals
rf = st.session_state["dcf_rf_pct"] / 100
mrp = st.session_state["dcf_mrp_pct"] / 100
tax = st.session_state["dcf_tax_pct"] / 100
g = st.session_state["dcf_terminal_g_pct"] / 100
# =========================================================
# Cost of Debt (Rd): Auto vs Manual override (with signature)
#   - Auto % comes from rd_auto (computed earlier)
#   - Manual persists across reruns and model/tab switches
# =========================================================

# 0) Ensure we have an auto rd as decimal
auto_rd_dec = float(rd_auto if 'rd_auto' in locals() else (cost_of_debt if 'cost_of_debt' in locals() else 0.0))
auto_rd_pct = float(auto_rd_dec * 100.0)

# 1) Init state (do not overwrite if already set)
st.session_state.setdefault("dcf_rd_manual_mode", False)   # False = Auto by default
if "dcf_rd_pct" not in st.session_state:
    st.session_state["dcf_rd_pct"] = float(auto_rd_pct)
st.session_state.setdefault("dcf_rd_auto_signature", None) # tracks when to snap to auto
init_widget_key("dcf_rd_pct_input", "dcf_rd_pct", auto_rd_pct)

# ---------------------------------------------------------
# Cost of Debt (Rd): Auto vs Manual override (FINAL STABLE)
# ---------------------------------------------------------

auto_rd_dec = float(cost_of_debt)
auto_rd_pct = auto_rd_dec * 100.0

# Initialize once only
if "dcf_rd_manual_mode" not in st.session_state:
    st.session_state["dcf_rd_manual_mode"] = False

if "dcf_rd_manual_value" not in st.session_state:
    st.session_state["dcf_rd_manual_value"] = auto_rd_pct

section(" 🧮 Cost of Debt (Rd) — Auto + Override")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
rd_mode = st.radio(
    "Cost of Debt (Rd) mode:",
    ["Use Auto (Interest / Debt)", "Manual override (%)"],
    index=1 if st.session_state["dcf_rd_manual_mode"] else 0,
    key="dcf_rd_mode_radio",
    horizontal=True
)

st.session_state["dcf_rd_manual_mode"] = rd_mode.startswith("Manual")

# ----- MANUAL MODE -----
if st.session_state["dcf_rd_manual_mode"]:

    rd_manual_pct = st.number_input(
        "Cost of Debt (Rd) — manual (%)",
        min_value=0.0,
        max_value=100.0,
        step=0.1,
        value=float(st.session_state["dcf_rd_manual_value"]),
        key="dcf_rd_manual_input"
    )

    st.session_state["dcf_rd_manual_value"] = float(rd_manual_pct)
    rd = float(rd_manual_pct) / 100.0

    st.caption(f"Auto reference (Interest ÷ Debt): {auto_rd_pct:.2f}%")

# ----- AUTO MODE -----
else:
    rd = auto_rd_dec
    st.caption(f"Auto Rd (from statements): {auto_rd_pct:.2f}%")

# CAPM & WACC
beta_levered = st.session_state["dcf_unlevered_beta"] * (1 + (1 - tax) * de_ratio)

if de_ratio <= 0:
    w_e, w_d = 1, 0
else:
    w_d = de_ratio / (1 + de_ratio)
    w_e = 1 / (1 + de_ratio)

re = rf + beta_levered * mrp
wacc = w_e * re + w_d * rd * (1 - tax)

# Save computed
st.session_state["levered_beta"] = float(beta_levered)
st.session_state["wacc"] = float(wacc)

# =========================================================
# OUTPUT HEADER (STOP HERE)
# =========================================================
st.markdown('<div class="dcf-card">', unsafe_allow_html=True)
section(" 📌 DCF Output")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
k1, k2, k3, k4 = st.columns(4)

with k1:
    st.markdown(f"""
    <div class="dcf-kpi">
      <div class="dcf-kpi-title">Cost of Debt (Rd)</div>
      <div class="dcf-kpi-value">{rd*100:.2f}%</div>
    </div>
    """, unsafe_allow_html=True)

with k2:
    st.markdown(f"""
    <div class="dcf-kpi">
      <div class="dcf-kpi-title">Levered Beta</div>
      <div class="dcf-kpi-value">{beta_levered:.2f}</div>
    </div>
    """, unsafe_allow_html=True)

with k3:
    st.markdown(f"""
    <div class="dcf-kpi">
      <div class="dcf-kpi-title">Cost of Equity (Re)</div>
      <div class="dcf-kpi-value">{re*100:.2f}%</div>
    </div>
    """, unsafe_allow_html=True)

with k4:
    st.markdown(f"""
    <div class="dcf-kpi">
      <div class="dcf-kpi-title">WACC</div>
      <div class="dcf-kpi-value">{wacc*100:.2f}%</div>
    </div>
    """, unsafe_allow_html=True)

st.markdown(f'<div class="small-note">Terminal growth (g): {g*100:.2f}% • D/E: {de_ratio:.2f}x</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------------------------------------
# DATE-BASED DISCOUNTING (FULLY PERSISTENT — NO RESETTING)
# ---------------------------------------------------------
section("📅 Valuation Timing & Mid-point")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
# 1️⃣ INITIALIZE DEFAULTS (only ONCE)
if "dcf_timing_init" not in st.session_state:

    st.session_state["dcf_valuation_date"] = date.today()
    st.session_state["dcf_first_fs_date"] = date(last_hist_year + 1, 12, 31)
    st.session_state["dcf_use_midyear"] = False

    st.session_state["dcf_timing_init"] = True
# 2️⃣ WIDGETS (using separate keys so they do NOT overwrite session_state)
valuation_date_input = st.date_input(
    "Valuation date (today / deal date)",
    value=st.session_state["dcf_valuation_date"],
    key="dcf_valuation_date_input"
)

first_fs_date_input = st.date_input(
    "Financial statement year-end date for forecasts (first forecast year)",
    value=st.session_state["dcf_first_fs_date"],
    key="dcf_first_fs_date_input"
)

use_midyear_input = st.checkbox(
    "Use mid-year (0.5 year earlier) convention?",
    value=st.session_state["dcf_use_midyear"],
    key="dcf_use_midyear_input"
)

# 3️⃣ UPDATE session_state values explicitly
st.session_state["dcf_valuation_date"] = valuation_date_input
st.session_state["dcf_first_fs_date"] = first_fs_date_input
st.session_state["dcf_use_midyear"] = use_midyear_input

# 4️⃣ CALCULATE DISCOUNT PERIODS USING STORED VALUES
valuation_date = st.session_state["dcf_valuation_date"]
first_forecast_fs_date = st.session_state["dcf_first_fs_date"]
use_midyear = st.session_state["dcf_use_midyear"]

gap_days = (first_forecast_fs_date - valuation_date).days
gap_years = gap_days / 365.25

n0 = max(gap_years, 0.0)
if use_midyear:
    n0 = max(n0 - 0.5, 0.0)

# discount periods for each forecast year
discount_periods_n = np.array([n0 + i for i in range(len(forecast_years_int))], dtype=float)

# DF0
midpoint_df0 = (1 / (1 + wacc) ** n0) if wacc > 0 else 1.0

# 5️⃣ DISPLAY SUMMARY TABLE
midpoint_table = pd.DataFrame(
    {
        "Valuation date": [valuation_date],
        "FS date (first forecast year)": [first_forecast_fs_date],
        "Gap (days)": [gap_days],
        "Discount period n₀ (years)": [n0],
        "Mid-point DF₀ = 1/(1+WACC)ⁿ⁰": [midpoint_df0],
    }
)

st.dataframe(midpoint_table, width='stretch')
# ---------------------------------------------------------
# CAPEX: use selected CF rows directly, allow excluding outlier years (PERSISTENT)
# ---------------------------------------------------------
avg_capex = 0.0

if capex_cf_idx_list:
    # 1) Build historical CAPEX series by year (sum selected rows)
    capex_by_year = cf_df.loc[capex_cf_idx_list, year_cols_cf].sum(axis=0)

    capex_hist_years = [str(y) for y in capex_by_year.index.tolist()]
    capex_hist_vals = capex_by_year.values.astype(float)

    # -------------------------------------------------
    # ✅ Persistent store (SURVIVES page/model switches)
    # -------------------------------------------------
    if "dcf_capex_excluded_years_store" not in st.session_state:
        st.session_state["dcf_capex_excluded_years_store"] = []

    # keep store clean if years changed
    st.session_state["dcf_capex_excluded_years_store"] = [
        y for y in st.session_state["dcf_capex_excluded_years_store"] if y in capex_hist_years
    ]

    # widget key (separate) so it never overwrites store unexpectedly
    if "dcf_capex_excluded_years_widget" not in st.session_state:
        st.session_state["dcf_capex_excluded_years_widget"] = list(st.session_state["dcf_capex_excluded_years_store"])

    def _sync_capex_exclusions():
        st.session_state["dcf_capex_excluded_years_store"] = list(st.session_state["dcf_capex_excluded_years_widget"])

    # 2) UI (BEFORE averaging) — reads from store, writes back to store
    if "show_capex_expander" not in st.session_state:
        st.session_state["show_capex_expander"] = False

    if st.button(
            "🧹 Show/Hide CAPEX History: Exclude outlier years before averaging",
            key="toggle_capex_expander_btn"
    ):
        st.session_state["show_capex_expander"] = not st.session_state["show_capex_expander"]

    if st.session_state["show_capex_expander"]:
        st.multiselect(
            "Select historical years to EXCLUDE from CAPEX average:",
            options=capex_hist_years,
            default=list(st.session_state["dcf_capex_excluded_years_store"]),
            key="dcf_capex_excluded_years_widget",
            on_change=_sync_capex_exclusions
        )

        if st.button("↩️ Reset CAPEX exclusions", key="reset_capex_exclusions_btn"):
            st.session_state["dcf_capex_excluded_years_store"] = []
            if "dcf_capex_excluded_years_widget" in st.session_state:
                del st.session_state["dcf_capex_excluded_years_widget"]
            st.rerun()

    excluded_years = set(st.session_state["dcf_capex_excluded_years_store"])

    # 3) Filter out excluded years + remove NaNs
    mask_keep = np.array([y not in excluded_years for y in capex_hist_years], dtype=bool)

    capex_hist_vals_used = capex_hist_vals[mask_keep]
    capex_hist_years_used = np.array(capex_hist_years)[mask_keep]

    # drop NaNs
    capex_hist_vals_used = capex_hist_vals_used[~np.isnan(capex_hist_vals_used)]

    # 4) Safety fallback if all excluded
    if capex_hist_vals_used.size == 0:
        st.warning("⚠️ You excluded all CAPEX years (or all were NaN). Using ALL historical years instead.")
        capex_hist_vals_used = capex_hist_vals[~np.isnan(capex_hist_vals)]
        capex_hist_years_used = np.array(capex_hist_years)[~np.isnan(capex_hist_vals)]

    # 5) Preview
    st.dataframe(
        pd.DataFrame({
            "Year": capex_hist_years,
            "CAPEX": capex_hist_vals,
            "Included?": ["✅" if y in capex_hist_years_used.tolist() else "❌" for y in capex_hist_years],
        }),
        width="stretch",
        hide_index=True
    )

    # 6) Compute average CAPEX from remaining years
    if capex_hist_vals_used.size > 0:
        avg_capex = float(np.mean(capex_hist_vals_used))

# Forecast capex = average of historical (negative number preserved)
capex_forecast_vals = np.full(len(forecast_years_int), avg_capex, dtype=float)

# ---------------------------------------------------------
# FCFF / UFCF
# ---------------------------------------------------------
ebitda_after_tax = ebitda_forecast_vals * (1 - tax)
dep_tax_vals = -dep_forecast_vals * tax

# UFCF = EBITDA(1-T) + Dep×T + ΔWC + Capex
fcff_vals = ebitda_after_tax + dep_tax_vals + delta_wc_forecast_vals + capex_forecast_vals

# Discount factors using date-based n
discount_factors = np.array([(1 / (1 + wacc) ** n) for n in discount_periods_n])
pv_fcff = fcff_vals * discount_factors

st.session_state["dcf_fcff_array"] = fcff_vals.tolist()
st.session_state["dcf_pv_fcff_array"] = pv_fcff.tolist()
st.session_state["dcf_discount_periods_n"] = discount_periods_n.tolist()

# ---------------------------------------------------------
# TERMINAL VALUE
# ---------------------------------------------------------
if wacc <= g:
    terminal_value = np.nan
    pv_terminal = np.nan
else:
    terminal_value = fcff_vals[-1] * (1 + g) / (wacc - g)
    discount_factor_terminal = float(discount_factors[-1])
    pv_terminal = terminal_value * discount_factor_terminal

enterprise_value = np.nansum(pv_fcff) + (0 if np.isnan(pv_terminal) else pv_terminal)
equity_value = enterprise_value - net_debt
st.session_state["dcf_terminal_value"] = float(terminal_value) if not np.isnan(terminal_value) else None
st.session_state["dcf_pv_terminal"] = float(pv_terminal) if not np.isnan(pv_terminal) else None
st.session_state["dcf_pv_fcff_sum"] = float(np.nansum(pv_fcff))

# Save DCF valuation outputs into session_state
st.session_state["enterprise_value_dcf"] = float(enterprise_value)
st.session_state["equity_value"] = float(equity_value)          # generic key used by COMPARABLES
st.session_state["equity_value_dcf"] = float(equity_value)      # explicit DCF key

# ---------------------------------------------------------
# DCF TABLE (UFCF style)
# ---------------------------------------------------------
section("📉 DCF Cashflows (UFCF) — Date-based Discounting")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
df_dcf = pd.DataFrame(
    {
        "Year": [str(y) for y in forecast_years_int],
        "Discount period n (years)": discount_periods_n,
        "EBITDA × (1−T)": ebitda_after_tax,
        "Depreciation × Tax": dep_tax_vals,
        "Δ Working capital": delta_wc_forecast_vals,
        "Capex": capex_forecast_vals,
        "UFCF": fcff_vals,
        "Discount factor": discount_factors,
        "PV of UFCF": pv_fcff,
    }
)

num_cols_dcf = df_dcf.select_dtypes(include=[np.number]).columns
fmt_dict = {c: "{:,.0f}".format for c in num_cols_dcf if c not in ["Discount period n (years)", "Discount factor"]}
fmt_dict["Discount period n (years)"] = "{:.3f}".format
fmt_dict["Discount factor"] = "{:.3f}".format

styled_dcf = df_dcf.style.format(fmt_dict, na_rep="")
st.dataframe(styled_dcf, width='stretch')

# Terminal summary
section("Terminal Value and Present Value:")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
df_term = pd.DataFrame(
    {
        "Terminal Value": [terminal_value],
        "Discount factor (last year)": [discount_factors[-1]],
        "PV of Terminal Value": [pv_terminal],
    }
)

fmt_term = {}
for c in df_term.columns:
    if c == "Discount factor (last year)":
        fmt_term[c] = "{:.3f}".format
    else:
        fmt_term[c] = "{:,.0f}".format

st.dataframe(
    df_term.style.format(fmt_term, na_rep=""),
    width='stretch',
)
st.session_state["df_dcf_export"] = df_dcf.copy()
st.session_state["rd"] = float(rd)  # so the download page can use Rd

# ---------------------------------------------------------
# SUMMARY (STYLED TABLE)
# ---------------------------------------------------------
section("📌 Valuation Summary")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
summary_rows = [
    ("Enterprise Value (EV)", enterprise_value, "USD"),
    ("Net Debt", net_debt, "USD"),
    ("Equity Value", equity_value, "USD"),
    ("WACC", wacc * 100, "%"),
    ("Terminal Growth Rate (g)", g * 100, "%"),
]
df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value", "Unit"])
def _fmt_row(v, unit):
    if pd.isna(v):
        return ""
    if unit == "USD":
        return f"{v:,.0f}"
    if unit == "%":
        return f"{v:.2f}%"
    return str(v)

df_summary["Value"] = [
    _fmt_row(v, u) for v, u in zip(df_summary["Value"], df_summary["Unit"])
]
styled_summary = (
    df_summary[["Metric", "Value"]]
    .style
    .set_properties(**{
        "font-size": "15px",
        "padding": "10px",
    })
    .set_table_styles([
        {"selector": "thead th", "props": [
            ("background-color", "#071426"),
            ("color", "white"),
            ("font-weight", "700"),
            ("text-align", "left"),
            ("padding", "12px"),
        ]},
        {"selector": "tbody td", "props": [
            ("border-bottom", "1px solid rgba(0,0,0,0.08)"),
        ]},
        {"selector": "tbody tr:hover", "props": [
            ("background-color", "rgba(0,51,153,0.06)"),
        ]},
    ])
)
st.dataframe(styled_summary, width="stretch", hide_index=True)
# =========================================================
# 📊 SENSITIVITY (PERSIST ACROSS PAGES — inputs DON'T reset)
# =========================================================
import streamlit.components.v1 as components

section("📊 Sensitivity of Equity Value to changes in WACC and Terminal Growth Rate")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
# -------------------------
# Persistent storage keys (NOT widget keys)
# -------------------------
SENS_STORE_DEFAULTS = {
    "sens_store_wacc_points": 5,
    "sens_store_g_points": 7,
    "sens_store_wacc_step_pct": 5.00,
    "sens_store_g_step_pct": 0.50,
}

for k, v in SENS_STORE_DEFAULTS.items():
    st.session_state.setdefault(k, v)

# -------------------------
# Helper: bind widget -> store (survives page switches)
# -------------------------
def bind_number_input(label, store_key, widget_key, *, cast=int, **kwargs):
    """
    cast=int  -> for integer widgets (points)
    cast=float -> for decimal widgets (step %)
    """
    def _sync():
        st.session_state[store_key] = cast(st.session_state[widget_key])

    # value must match min/max/step types
    val = st.session_state.get(store_key)

    # Ensure stored value has the right type
    if val is None:
        val = cast(0)
    else:
        val = cast(val)

    # Also ensure kwargs numeric args match the same type
    for k in ["min_value", "max_value", "step"]:
        if k in kwargs and kwargs[k] is not None:
            kwargs[k] = cast(kwargs[k])

    return st.number_input(
        label,
        value=val,
        key=widget_key,
        on_change=_sync,
        **kwargs
    )

# --- base values from your DCF (decimals)
base_wacc = float(wacc)
base_g = float(g)

# --------------------------
# User controls (WIDGET KEYS are separate)
# --------------------------
cA, cB, cC = st.columns([1, 1, 2])
with cA:
    wacc_points = bind_number_input(
        "WACC points (rows)",
        store_key="sens_store_wacc_points",
        widget_key="sens_wacc_points_w",
        cast=int,
        min_value=3, max_value=9, step=2
    )
with cB:
    g_points = bind_number_input(
        "g points (columns)",
        store_key="sens_store_g_points",
        widget_key="sens_g_points_w",
        cast=int,
        min_value=3, max_value=11, step=2
    )

cD, cE, _ = st.columns([1, 1, 2])
with cD:
    wacc_step_pct = bind_number_input(
        "WACC step (%)",
        store_key="sens_store_wacc_step_pct",
        widget_key="sens_wacc_step_pct_w",
        cast=float,
        min_value=0.5, max_value=10.0, step=0.5
    )
with cE:
    g_step_pct = bind_number_input(
        "g step (%)",
        store_key="sens_store_g_step_pct",
        widget_key="sens_g_step_pct_w",
        cast=float,
        min_value=0.1, max_value=5.0, step=0.1
    )
# Use persistent stored values (NOT widget keys)
wacc_step = float(st.session_state["sens_store_wacc_step_pct"]) / 100.0
g_step = float(st.session_state["sens_store_g_step_pct"]) / 100.0

def build_centered_range(center: float, step: float, points: int):
    points = int(points)
    if points % 2 == 0:
        points += 1
    half = points // 2
    arr = np.array([center + (i - half) * step for i in range(points)], dtype=float)
    return np.sort(arr)

wacc_range = build_centered_range(base_wacc, wacc_step, int(st.session_state["sens_store_wacc_points"]))
g_range = build_centered_range(base_g, g_step, int(st.session_state["sens_store_g_points"]))
g_range = np.clip(g_range, -0.50, 0.50)

def equity_value_sensitivity(fcff_vals, discount_periods, net_debt, wacc_, g_):
    wacc_ = float(wacc_)
    g_ = float(g_)
    if (1 + wacc_) <= 0:
        return np.nan
    if wacc_ <= g_:
        return np.nan

    fcff_arr = np.array(fcff_vals, dtype=float)
    n_arr = np.array(discount_periods, dtype=float)

    dfs = 1.0 / (1.0 + wacc_) ** n_arr
    pv_fcff = float(np.nansum(fcff_arr * dfs))

    tv = float(fcff_arr[-1]) * (1.0 + g_) / (wacc_ - g_)
    pv_tv = float(tv * dfs[-1])

    ev = pv_fcff + pv_tv
    return float(ev - float(net_debt))

def pct_label(x, decimals=2):
    return f"{x*100:.{decimals}f}%"

label_decimals = 2
row_labels = [pct_label(w, label_decimals) for w in wacc_range]
col_labels = [pct_label(gg, label_decimals) for gg in g_range]

if len(set(row_labels)) != len(row_labels):
    st.error("❌ WACC labels collided (duplicate % labels). Increase decimals or step.")
    st.stop()

if len(set(col_labels)) != len(col_labels):
    st.error("❌ Terminal growth labels collided (duplicate % labels). Increase decimals or step.")
    st.stop()

sens_table = pd.DataFrame(index=row_labels, columns=col_labels, dtype=float)
for w_ in wacc_range:
    rlab = pct_label(w_, label_decimals)
    for gg_ in g_range:
        clab = pct_label(gg_, label_decimals)
        sens_table.loc[rlab, clab] = equity_value_sensitivity(
            fcff_vals=fcff_vals,
            discount_periods=discount_periods_n,
            net_debt=net_debt,
            wacc_=w_,
            g_=gg_,
        )

base_row = pct_label(base_wacc, label_decimals)
base_col = pct_label(base_g, label_decimals)

vals = sens_table.values.astype(float)
finite_mask = np.isfinite(vals)

min_rc = max_rc = None
min_val = max_val = np.nan
if finite_mask.any():
    min_val = float(np.min(vals[finite_mask]))
    max_val = float(np.max(vals[finite_mask]))
    min_pos = np.argwhere(vals == min_val)
    max_pos = np.argwhere(vals == max_val)
    if min_pos.shape[0] > 0:
        min_rc = (int(min_pos[0, 0]), int(min_pos[0, 1]))
    if max_pos.shape[0] > 0:
        max_rc = (int(max_pos[0, 0]), int(max_pos[0, 1]))

nan_count = int(np.isnan(vals).sum())
if nan_count > 0:
    st.warning(
        f"⚠️ Some cells are blank because Terminal Value is invalid when WACC ≤ g. "
        f"Blanks found: {nan_count}. Reduce g range or increase WACC range."
    )

def fmt_num(x):
    if x is None:
        return ""
    try:
        if np.isnan(x) or not np.isfinite(x):
            return ""
    except Exception:
        pass
    return f"{float(x):,.0f}"

display_df = sens_table.map(fmt_num)
cols = list(display_df.columns)
rows = list(display_df.index)

html_parts = []
html_parts.append(f"""
<style>
.sens-outer{{border:2px solid #000;border-radius:10px;padding:10px 12px 12px;background:#fff;overflow-x:auto;}}
.sens-table{{border-collapse:collapse;width:100%;min-width:760px;font-size:14px;font-family:Georgia,Georgia;}}
.sens-table th,.sens-table td{{border-bottom:1px solid rgba(0,0,0,0.08);border-right:1px solid rgba(0,0,0,0.08);padding:10px 12px;white-space:nowrap;}}
.sens-table thead th{{background:#071426;color:#fff;font-weight:900;text-align:center;}}
.sens-table .rowhdr{{background:#f1f5f9;font-weight:900;text-align:left;}}
.sens-table td{{text-align:right;}}
.sens-title{{background:#fff !important;color:#000 !important;font-weight:900 !important;font-size:15px !important;border-bottom:0 !important;}}
.sens-vwacc{{writing-mode:vertical-rl;transform:rotate(180deg);font-weight:900;text-align:center !important;background:#fff !important;color:#000 !important;border-bottom:0 !important;padding:0 6px !important;}}
</style>

<div class="sens-outer">
<table class="sens-table">
<thead>
<tr>
  <th class="sens-vwacc" rowspan="2">WACC(%)</th>
  <th class="sens-title" colspan="{len(cols)}">Terminal Growth Rate</th>
</tr>
<tr>
""")

for c in cols:
    html_parts.append(f"<th>{c}</th>")
html_parts.append("</tr></thead><tbody>")

for i, r in enumerate(rows):
    html_parts.append("<tr>")
    html_parts.append(f'<th class="rowhdr">{r}</th>')
    for j, c in enumerate(cols):
        val = display_df.iloc[i, j]
        is_base = (r == base_row and c == base_col)
        is_min = (min_rc is not None and (i, j) == min_rc and not is_base)
        is_max = (max_rc is not None and (i, j) == max_rc and not is_base)

        bg = "#ffffff"
        fw = "400"
        if is_base:
            bg, fw = "#7dd3fc", "900"
        elif is_min:
            bg, fw = "#fde047", "900"
        elif is_max:
            bg, fw = "#fdba74", "900"

        html_parts.append(f'<td style="background:{bg};font-weight:{fw};">{val}</td>')
    html_parts.append("</tr>")

html_parts.append("</tbody></table></div>")
components.html("".join(html_parts), height=460, scrolling=True)

base_eq = np.nan
if base_row in sens_table.index and base_col in sens_table.columns:
    base_eq = float(sens_table.loc[base_row, base_col])

cK1, cK2, cK3 = st.columns(3)
with cK1:
    show_base = base_eq if np.isfinite(base_eq) else float(equity_value)
    st.metric("Base Equity Value (current WACC & g)", f"{show_base:,.0f}")
with cK2:
    st.metric("Lowest (grid)", f"{min_val:,.0f}" if np.isfinite(min_val) else "—")
with cK3:
    st.metric("Highest (grid)", f"{max_val:,.0f}" if np.isfinite(max_val) else "—")

st.caption(
    f"Base case highlighted in blue (WACC={base_wacc*100:.2f}% , g={base_g*100:.2f}%). "
    f"Min highlighted in yellow, Max highlighted in orange."
)

# =========================================================
# ✅ FULL DCF EXCEL EXPORT (FULL INCOME STATEMENT + FORMULAS + SENSITIVITY)
# =========================================================
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def _excel_col(n: int) -> str:
    return get_column_letter(n)

def build_full_dcf_excel_model(
    is_df,
    forecast_is_df,
    year_cols_is,
    forecast_years_int,
    forecast_cols,

    rev_idx, cos_idx, gp_idx, ebitda_idx, op_idx, pbt_idx, tax_idx, np_idx,

    growth_mode,
    yearly_g_dict,
    avg_g,
    avg_tax_ratio,
    avg_gp_margin,
    cos_ratio,

    wc_percent_used,
    last_wc_hist_value,

    discount_periods_n,
    dep_forecast_vals,
    capex_forecast_vals,
    wacc, tax, g, net_debt,

    rf, mrp,
    dcf_unlevered_beta=1.0,
    rd=0.0,
    total_debt=0.0,
    cash_balance=0.0,
    book_equity=0.0,
    de_ratio=0.0,

    # ✅ IMPORTANT: these MUST match your *persistent* sensitivity store keys
    sens_wacc_points=5,
    sens_g_points=7,
    sens_wacc_step_pct=5.0,
    sens_g_step_pct=0.5,
):
    wb = Workbook()

    # -----------------------------
    # Theme (FBC-ish)
    # -----------------------------
    BLUE = "003399"
    DARK = "071426"
    LIGHT_BG = "F7FAFF"
    GRID = "D9E2EF"
    ORANGE = "F5B400"
    SKY = "7DD3FC"
    YELLOW = "FDE047"

    thin = Side(style="thin", color=GRID)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_title(ws, title, end_col=8):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        c = ws.cell(1, 1, title)
        c.font = Font(bold=True, color="FFFFFF", size=14)
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

    def style_header_row(ws, r, c1, c2):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
        ws.row_dimensions[r].height = 20

    def style_body(ws, r1, r2, c1, c2, money_cols=None, pct_cols=None, dec_cols=None):
        money_cols = set(money_cols or [])
        pct_cols = set(pct_cols or [])
        dec_cols = set(dec_cols or [])
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(r, c)
                cell.border = border_all
                cell.alignment = Alignment(vertical="center")
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=LIGHT_BG)
                if c in money_cols:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if c in pct_cols:
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if c in dec_cols:
                    cell.number_format = '0.000'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

    # =========================================================
    # 1) PARAMETERS SHEET
    # =========================================================
    wsP = wb.active
    wsP.title = "Parameters"

    all_years = list(map(str, year_cols_is)) + [str(y) for y in forecast_years_int]
    n_fore = len(forecast_years_int)

    style_title(wsP, "DCF Parameters & Assumptions", end_col=8)

    wsP["A3"] = "Key Inputs"
    wsP["A3"].font = Font(bold=True)

    # NOTE: Inputs sheet is created later; formulas still work fine
    key_rows = [
        ("WACC",
         "=((1/(1+Inputs!$F$13))*(Inputs!$F$4 + (Inputs!$F$6*(1+(1-Inputs!$F$7)*Inputs!$F$13))*Inputs!$F$5)) + ((Inputs!$F$13/(1+Inputs!$F$13))*Inputs!$F$9*(1-Inputs!$F$7))",
         "decimal"),
        ("Tax rate (DCF)", "=Inputs!$F$7", "decimal"),
        ("Terminal growth (g)", "=Inputs!$F$8", "decimal"),
        ("Net debt", "=(Inputs!$F$10-Inputs!$F$11)", "money"),
        ("Income tax ratio (IS): Tax/PBT", float(avg_tax_ratio), "decimal"),
        ("Revenue growth mode", growth_mode, "text"),
        ("Uniform growth (if used)", float(avg_g), "decimal"),
        ("GP margin (if used)", float(avg_gp_margin) if avg_gp_margin is not None else 0.0, "decimal"),
        ("COS ratio (if used)", float(cos_ratio), "decimal"),
        ("WC % of Sales used", float(wc_percent_used), "decimal"),
        ("Last historical WC used", float(last_wc_hist_value), "money"),

        # ✅ Sensitivity settings recorded here too
        ("Sensitivity: WACC points (rows)", int(sens_wacc_points), "int"),
        ("Sensitivity: g points (cols)", int(sens_g_points), "int"),
        ("Sensitivity: WACC step", float(sens_wacc_step_pct) / 100.0, "decimal"),
        ("Sensitivity: g step", float(sens_g_step_pct) / 100.0, "decimal"),
    ]

    start_r = 5
    wsP["A4"], wsP["B4"], wsP["C4"] = "Item", "Value", "Format"
    style_header_row(wsP, 4, 1, 3)

    for i, (k, v, kind) in enumerate(key_rows):
        r = start_r + i
        wsP.cell(r, 1, k)
        if kind == "text":
            wsP.cell(r, 2, str(v))
        elif kind == "int":
            wsP.cell(r, 2, int(v))
        else:
            if isinstance(v, str) and v.startswith("="):
                wsP.cell(r, 2, v)
            else:
                wsP.cell(r, 2, float(v))
        wsP.cell(r, 3, kind)

        for c in range(1, 4):
            wsP.cell(r, c).border = border_all

        if kind == "money":
            wsP.cell(r, 2).number_format = '#,##0'
        elif kind == "decimal":
            wsP.cell(r, 2).number_format = '0.00%'

    wsP.column_dimensions["A"].width = 34
    wsP.column_dimensions["B"].width = 22
    wsP.column_dimensions["C"].width = 14

    # ---- Growth table (per forecast year) ----
    growth_table_r = start_r + len(key_rows) + 2
    wsP.cell(growth_table_r, 1, "Forecast Growth Rates (editable)")
    wsP.cell(growth_table_r, 1).font = Font(bold=True)

    hdr = growth_table_r + 1
    wsP.cell(hdr, 1, "Year")
    wsP.cell(hdr, 2, "Growth (decimal)")
    style_header_row(wsP, hdr, 1, 2)

    for j, y in enumerate(forecast_years_int):
        r = hdr + 1 + j
        wsP.cell(r, 1, int(y))
        wsP.cell(r, 2, float(yearly_g_dict.get(y, avg_g)))
        wsP.cell(r, 2).number_format = "0.00%"
        wsP.cell(r, 1).border = border_all
        wsP.cell(r, 2).border = border_all

    # ---- Ratio table (each IS row as % of revenue) ----
    ratio_table_r = hdr + 1 + n_fore + 2
    wsP.cell(ratio_table_r, 1, "Income Statement Row Ratios (Row / Revenue) (editable)")
    wsP.cell(ratio_table_r, 1).font = Font(bold=True)

    rrh = ratio_table_r + 1
    wsP.cell(rrh, 1, "IS Row #")
    wsP.cell(rrh, 2, "Item")
    wsP.cell(rrh, 3, "Ratio to Revenue")
    style_header_row(wsP, rrh, 1, 3)

    rev_hist_vals = forecast_is_df.iloc[rev_idx][year_cols_is].values.astype(float)

    def ratio_to_rev_hist(row_vals, rev_vals):
        import numpy as np
        mask = (~np.isnan(row_vals)) & (~np.isnan(rev_vals)) & (rev_vals != 0)
        if not mask.any():
            return 0.0
        ratios = row_vals[mask] / rev_vals[mask]
        ratios = ratios[(ratios > -5) & (ratios < 5)]
        return float(np.mean(ratios)) if len(ratios) else 0.0

    protected = set([rev_idx])
    for x in [gp_idx, ebitda_idx, op_idx, pbt_idx, np_idx]:
        if isinstance(x, int):
            protected.add(x)
    if isinstance(tax_idx, int):
        protected.add(tax_idx)

    rr = rrh + 1
    for i in range(len(forecast_is_df)):
        item = str(forecast_is_df.iloc[i]["Item"])
        wsP.cell(rr, 1, int(i))
        wsP.cell(rr, 2, item)

        if i in protected:
            ratio = 0.0
        else:
            row_hist = forecast_is_df.iloc[i][year_cols_is].values.astype(float)
            ratio = ratio_to_rev_hist(row_hist, rev_hist_vals)

        wsP.cell(rr, 3, float(ratio))
        wsP.cell(rr, 3).number_format = "0.0000"
        for c in range(1, 4):
            wsP.cell(rr, c).border = border_all
        rr += 1

    wsP.freeze_panes = "A5"

    # =========================================================
    # 2) INCOME STATEMENT SHEET
    # =========================================================
    wsIS = wb.create_sheet("IncomeStatement")

    end_col = 2 + len(all_years)
    style_title(wsIS, "Forecast Income Statement (Full, with formulas)", end_col=end_col)

    hdr_row = 3
    wsIS.cell(hdr_row, 1, "Item")
    wsIS.cell(hdr_row, 2, "Section")
    for j, y in enumerate(all_years):
        wsIS.cell(hdr_row, 3 + j, str(y))
    style_header_row(wsIS, hdr_row, 1, end_col)

    base_row = hdr_row + 1

    for i in range(len(forecast_is_df)):
        r = base_row + i
        wsIS.cell(r, 1, str(forecast_is_df.iloc[i]["Item"]))
        wsIS.cell(r, 2, "IS")
        for j, y in enumerate(year_cols_is):
            v = forecast_is_df.iloc[i][y]
            wsIS.cell(r, 3 + j, float(v) if v == v else None)

    wsIS.column_dimensions["A"].width = 42
    wsIS.column_dimensions["B"].width = 10
    for j in range(len(all_years)):
        wsIS.column_dimensions[_excel_col(3 + j)].width = 14

    rev_row_excel = base_row + rev_idx

    growth_data_start = (start_r + len(key_rows) + 2) + 2  # hdr+1
    ratio_hdr_row = (start_r + len(key_rows) + 2) + 2 + n_fore + 2 + 1
    ratio_data_start = ratio_hdr_row + 1

    avg_tax_ratio_cell = f"Parameters!$B${start_r + 4}"
    gp_margin_cell = f"Parameters!$B${start_r + 7}"
    cos_ratio_cell = f"Parameters!$B${start_r + 8}"

    has_cos = isinstance(cos_idx, int)
    has_gp = isinstance(gp_idx, int)

    for f_i, y_int in enumerate(forecast_years_int):
        col = 3 + len(year_cols_is) + f_i
        colL = _excel_col(col)

        if f_i == 0:
            prev_col = 3 + len(year_cols_is) - 1
            prevL = _excel_col(prev_col)
            growth_cell = f"Parameters!$B${growth_data_start + f_i}"
            wsIS[f"{colL}{rev_row_excel}"] = f"={prevL}{rev_row_excel}*(1+{growth_cell})"
        else:
            prevL = _excel_col(col - 1)
            growth_cell = f"Parameters!$B${growth_data_start + f_i}"
            wsIS[f"{colL}{rev_row_excel}"] = f"={prevL}{rev_row_excel}*(1+{growth_cell})"

        # COS / GP special handling
        if has_gp and has_cos and avg_gp_margin is not None:
            cos_row_excel = base_row + cos_idx
            gp_row_excel = base_row + gp_idx
            last_cos_hist = float(forecast_is_df.iloc[cos_idx][year_cols_is[-1]]) if forecast_is_df.iloc[cos_idx][year_cols_is[-1]] == forecast_is_df.iloc[cos_idx][year_cols_is[-1]] else 0.0
            cos_sign = -1 if last_cos_hist < 0 else 1
            wsIS[f"{colL}{cos_row_excel}"] = f"={cos_sign}*{colL}{rev_row_excel}*(1-{gp_margin_cell})"
            wsIS[f"{colL}{gp_row_excel}"] = f"={colL}{rev_row_excel}+{colL}{cos_row_excel}"

        elif has_gp and (not has_cos) and avg_gp_margin is not None:
            gp_row_excel = base_row + gp_idx
            wsIS[f"{colL}{gp_row_excel}"] = f"={colL}{rev_row_excel}*{gp_margin_cell}"

        elif has_cos and (not has_gp):
            cos_row_excel = base_row + cos_idx
            wsIS[f"{colL}{cos_row_excel}"] = f"={colL}{rev_row_excel}*{cos_ratio_cell}"

        # Other rows from ratios
        for i in range(len(forecast_is_df)):
            r = base_row + i

            if i == rev_idx:
                continue
            if i in [gp_idx, ebitda_idx, op_idx, pbt_idx, np_idx] and isinstance(i, int):
                continue
            if isinstance(tax_idx, int) and i == tax_idx:
                continue

            if has_gp and has_cos and avg_gp_margin is not None and isinstance(cos_idx, int) and i == cos_idx:
                continue
            if has_gp and has_cos and avg_gp_margin is not None and isinstance(gp_idx, int) and i == gp_idx:
                continue
            if has_gp and (not has_cos) and avg_gp_margin is not None and isinstance(gp_idx, int) and i == gp_idx:
                continue
            if has_cos and (not has_gp) and isinstance(cos_idx, int) and i == cos_idx:
                continue

            ratio_cell = f"Parameters!$C${ratio_data_start + i}"
            wsIS[f"{colL}{r}"] = f"={colL}{rev_row_excel}*{ratio_cell}"

        # totals chain
        chain = [("REV", rev_idx), ("GP", gp_idx), ("EBITDA", ebitda_idx), ("OP", op_idx), ("PBT", pbt_idx), ("NP", np_idx)]
        chain = [(nm, idx) for nm, idx in chain if isinstance(idx, int)]
        chain = sorted(chain, key=lambda x: x[1])

        for j in range(1, len(chain)):
            prev_nm, prev_idx0 = chain[j - 1]
            cur_nm, cur_idx0 = chain[j]
            prev_r = base_row + prev_idx0
            cur_r = base_row + cur_idx0

            if cur_nm == "GP" and has_cos and isinstance(gp_idx, int) and isinstance(cos_idx, int):
                continue

            wsIS[f"{colL}{cur_r}"] = f"=SUM({colL}{prev_r}:{colL}{cur_r-1})"

        if isinstance(tax_idx, int) and isinstance(pbt_idx, int):
            tax_r = base_row + tax_idx
            pbt_r = base_row + pbt_idx
            wsIS[f"{colL}{tax_r}"] = f"={colL}{pbt_r}*{avg_tax_ratio_cell}"

        if isinstance(np_idx, int) and isinstance(pbt_idx, int):
            np_r = base_row + np_idx
            pbt_r = base_row + pbt_idx
            if isinstance(tax_idx, int):
                tax_r = base_row + tax_idx
                if (tax_r + 1) <= (np_r - 1):
                    wsIS[f"{colL}{np_r}"] = f"={colL}{pbt_r}+{colL}{tax_r}+SUM({colL}{tax_r+1}:{colL}{np_r-1})"
                else:
                    wsIS[f"{colL}{np_r}"] = f"={colL}{pbt_r}+{colL}{tax_r}"
            else:
                wsIS[f"{colL}{np_r}"] = f"={colL}{pbt_r}"

    money_cols = list(range(3, end_col + 1))
    style_body(wsIS, base_row, base_row + len(forecast_is_df) - 1, 1, end_col, money_cols=money_cols)
    wsIS.freeze_panes = f"C{base_row}"

    # =========================================================
    # 3) WORKING CAPITAL SHEET
    # =========================================================
    wsWC = wb.create_sheet("WorkingCapital")
    style_title(wsWC, "Working Capital (Forecast & ΔWC)", end_col=8)

    wsWC["A3"], wsWC["B3"], wsWC["C3"], wsWC["D3"] = "Year", "Revenue", "WC (Rev*WC%)", "ΔWC (Old-New)"
    style_header_row(wsWC, 3, 1, 4)

    wc_pct_cell = f"Parameters!$B${start_r + 9}"
    last_wc_cell = f"Parameters!$B${start_r + 10}"

    r0 = 4
    for i, y in enumerate(forecast_years_int):
        r = r0 + i
        wsWC.cell(r, 1, int(y))

        col_index_is = 3 + len(year_cols_is) + i
        colL_is = _excel_col(col_index_is)
        wsWC.cell(r, 2, f"=IncomeStatement!{colL_is}{rev_row_excel}")

        wsWC.cell(r, 3, f"=B{r}*{wc_pct_cell}")

        if i == 0:
            wsWC.cell(r, 4, f"={last_wc_cell}-C{r}")
        else:
            wsWC.cell(r, 4, f"=C{r-1}-C{r}")

    wsWC.column_dimensions["A"].width = 10
    wsWC.column_dimensions["B"].width = 16
    wsWC.column_dimensions["C"].width = 18
    wsWC.column_dimensions["D"].width = 18

    style_body(wsWC, 4, 4 + n_fore - 1, 1, 4, money_cols=[2, 3, 4])
    wsWC.freeze_panes = "A4"

    # =========================================================
    # 4) INPUTS SHEET
    # =========================================================
    wsI = wb.create_sheet("Inputs")
    style_title(wsI, "Model Inputs (Used by DCF)", end_col=6)

    wsI["A3"], wsI["B3"], wsI["C3"] = "Year", "Depreciation (forecast)", "Capex (forecast)"
    style_header_row(wsI, 3, 1, 3)

    wsI["E3"], wsI["F3"] = "DCF Input", "Value"
    style_header_row(wsI, 3, 5, 6)

    wsI["E4"], wsI["F4"] = "Risk-free rate (RF)", float(rf)
    wsI["E5"], wsI["F5"] = "Market risk premium (MRP)", float(mrp)
    wsI["E6"], wsI["F6"] = "Unlevered beta (βu)", float(dcf_unlevered_beta)
    wsI["E7"], wsI["F7"] = "Tax rate (T)", float(tax)
    wsI["E8"], wsI["F8"] = "Terminal growth (g)", float(g)
    wsI["E9"], wsI["F9"] = "Cost of debt (Rd)", float(rd)
    wsI["E10"], wsI["F10"] = "Total Debt", float(total_debt)
    wsI["E11"], wsI["F11"] = "Cash", float(cash_balance)
    wsI["E12"], wsI["F12"] = "Book Equity", float(book_equity)
    wsI["E13"], wsI["F13"] = "D/E ratio (book)", float(de_ratio)

    for rr in range(4, 14):
        wsI.cell(rr, 5).border = border_all
        wsI.cell(rr, 6).border = border_all

    wsI["F4"].number_format = "0.00%"
    wsI["F5"].number_format = "0.00%"
    wsI["F6"].number_format = "0.00"
    wsI["F7"].number_format = "0.00%"
    wsI["F8"].number_format = "0.00%"
    wsI["F9"].number_format = "0.00%"
    wsI["F10"].number_format = "#,##0"
    wsI["F11"].number_format = "#,##0"
    wsI["F12"].number_format = "#,##0"
    wsI["F13"].number_format = "0.00"

    r0 = 4
    for i, y in enumerate(forecast_years_int):
        r = r0 + i
        wsI.cell(r, 1, int(y))
        wsI.cell(r, 2, float(dep_forecast_vals[i]))
        wsI.cell(r, 3, float(capex_forecast_vals[i]))
        wsI.cell(r, 2).number_format = "#,##0"
        wsI.cell(r, 3).number_format = "#,##0"
        for c in range(1, 4):
            wsI.cell(r, c).border = border_all

    wsI.column_dimensions["A"].width = 10
    wsI.column_dimensions["B"].width = 22
    wsI.column_dimensions["C"].width = 18
    wsI.freeze_panes = "A4"

    # =========================================================
    # 5) DCF SHEET
    # =========================================================
    wsD = wb.create_sheet("DCF")
    end_col = 2 + n_fore
    style_title(wsD, "DCF Valuation (Formulas)", end_col=end_col)

    wsD.cell(3, 1, "Line")
    wsD.cell(3, 2, "Unit")
    for j, y in enumerate(forecast_years_int):
        wsD.cell(3, 3 + j, str(y))
    style_header_row(wsD, 3, 1, end_col)

    lines = [
        ("EBITDA × (1−T)", "USD"),
        ("Depreciation × Tax", "USD"),
        ("Δ Working capital", "USD"),
        ("Capex", "USD"),
        ("UFCF", "USD"),
        ("Discount factor", "x"),
        ("PV of UFCF", "USD"),
    ]
    r_start = 4
    for i, (nm, unit) in enumerate(lines):
        wsD.cell(r_start + i, 1, nm)
        wsD.cell(r_start + i, 2, unit)

    TAX_DCF = f"Parameters!$B${start_r + 1}"
    WACC = f"Parameters!$B${start_r + 0}"
    G = f"Parameters!$B${start_r + 2}"
    NETDEBT = f"Parameters!$B${start_r + 3}"

    ebitda_row_excel = (base_row + ebitda_idx) if isinstance(ebitda_idx, int) else None

    for j in range(n_fore):
        col = 3 + j
        colL = _excel_col(col)
        is_col = _excel_col(3 + len(year_cols_is) + j)

        if ebitda_row_excel is not None:
            wsD[f"{colL}{r_start+0}"] = f"=IncomeStatement!{is_col}{ebitda_row_excel}*(1-{TAX_DCF})"
        else:
            wsD[f"{colL}{r_start+0}"] = f"=0*(1-{TAX_DCF})"

        wsD[f"{colL}{r_start+1}"] = f"=-Inputs!B{4+j}*{TAX_DCF}"
        wsD[f"{colL}{r_start+2}"] = f"=WorkingCapital!D{4+j}"
        wsD[f"{colL}{r_start+3}"] = f"=Inputs!C{4+j}"

        wsD[f"{colL}{r_start+4}"] = f"={colL}{r_start+0}+{colL}{r_start+1}+{colL}{r_start+2}+{colL}{r_start+3}"

        n_row = r_start + 8
        wsD[f"{colL}{n_row}"] = float(discount_periods_n[j])
        wsD[f"{colL}{r_start+5}"] = f"=1/(1+{WACC})^{colL}{n_row}"
        wsD[f"{colL}{r_start+6}"] = f"={colL}{r_start+4}*{colL}{r_start+5}"

    wsD.cell(r_start + 8, 1, "Discount period n (hidden)")
    wsD.row_dimensions[r_start + 8].hidden = True

    term_r = r_start + 10
    wsD.cell(term_r, 1, "Terminal Value")
    wsD.cell(term_r, 2, "USD")
    last_colL = _excel_col(3 + n_fore - 1)

    wsD[f"{last_colL}{term_r}"] = f"={last_colL}{r_start+4}*(1+{G})/({WACC}-{G})"

    pv_term_r = term_r + 1
    wsD.cell(pv_term_r, 1, "PV of Terminal Value")
    wsD.cell(pv_term_r, 2, "USD")
    wsD[f"{last_colL}{pv_term_r}"] = f"={last_colL}{term_r}*{last_colL}{r_start+5}"

    ev_r = pv_term_r + 2
    wsD.cell(ev_r, 1, "Enterprise Value (EV)")
    wsD.cell(ev_r, 2, "USD")
    for c in [1, 2]:
        wsD.cell(ev_r, c).font = Font(bold=True, color="FFFFFF")
        wsD.cell(ev_r, c).fill = PatternFill("solid", fgColor=DARK)

    first_pv = f"C{r_start+6}"
    last_pv = f"{last_colL}{r_start+6}"
    wsD[f"C{ev_r}"] = f"=SUM({first_pv}:{last_pv})+{last_colL}{pv_term_r}"

    BLACK = "000000"

    eq_r = ev_r + 1
    wsD.cell(eq_r, 1, "Equity Value")
    wsD.cell(eq_r, 2, "USD")
    for c in [1, 2]:
        wsD.cell(eq_r, c).font = Font(bold=True, color="FFFFFF")
        wsD.cell(eq_r, c).fill = PatternFill("solid", fgColor=DARK)
    wsD[f"C{eq_r}"] = f"=C{ev_r}-{NETDEBT}"

    money_cols = list(range(3, end_col + 1))
    style_body(wsD, r_start, ev_r - 1, 1, end_col, money_cols=money_cols)
    for j in range(n_fore):
        wsD.cell(r_start + 5, 3 + j).number_format = "0.000"

    wsD.column_dimensions["A"].width = 28
    wsD.column_dimensions["B"].width = 10
    for j in range(n_fore):
        wsD.column_dimensions[_excel_col(3 + j)].width = 14

    wsD.freeze_panes = f"C{r_start}"

    # =========================================================
    # 6) SUMMARY SHEET
    # =========================================================
    wsS = wb.create_sheet("Summary")
    style_title(wsS, "DCF Summary", end_col=6)

    wsS["A3"], wsS["B3"], wsS["C3"] = "Metric", "Value", "Unit"
    style_header_row(wsS, 3, 1, 3)

    rows = [
        ("Enterprise Value (EV)", f"=DCF!C{ev_r}", "USD"),
        ("Net Debt", f"={NETDEBT}", "USD"),
        ("Equity Value", f"=DCF!C{eq_r}", "USD"),
        ("WACC", f"={WACC}", "%"),
        ("Terminal growth (g)", f"={G}", "%"),
        ("Tax rate (DCF)", f"={TAX_DCF}", "%"),
        ("Tax/PBT ratio (IS)", f"=Parameters!$B${start_r + 4}", "%"),
        ("WC % of Sales used", f"=Parameters!$B${start_r + 9}", "%"),
    ]

    r0 = 4
    for i, (m, v, u) in enumerate(rows):
        r = r0 + i
        wsS.cell(r, 1, m)
        wsS.cell(r, 2, v)
        wsS.cell(r, 3, u)
        for c in range(1, 4):
            wsS.cell(r, c).border = border_all
        wsS.cell(r, 2).number_format = "#,##0" if u == "USD" else "0.00%"

    wsS.column_dimensions["A"].width = 28
    wsS.column_dimensions["B"].width = 18
    wsS.column_dimensions["C"].width = 8
    style_body(wsS, 4, 4 + len(rows) - 1, 1, 3, money_cols=[2])
    wsS.freeze_panes = "A4"

    # =========================================================
    # 7) ✅ SENSITIVITY SHEET (kept + formula-driven)
    # =========================================================
    wsSens = wb.create_sheet("Sensitivity")
    style_title(wsSens, "Sensitivity: Equity Value vs WACC & Terminal Growth (g)", end_col=9)

    wsSens["A2"] = "Tip: Edit WACC (rows) or g (columns) to instantly see equity value impact."
    wsSens["A2"].font = Font(italic=True, color="4B5563")
    wsSens.merge_cells("A2:I2")

    wsSens["A3"] = "WACC \\ g"
    wsSens["A3"].font = Font(bold=True, color="FFFFFF")
    wsSens["A3"].fill = PatternFill("solid", fgColor=BLUE)
    wsSens["A3"].alignment = Alignment(horizontal="center", vertical="center")
    wsSens["A3"].border = border_all

    # Steps in decimals
    wacc_step_dec = float(sens_wacc_step_pct) / 100.0
    g_step_dec = float(sens_g_step_pct) / 100.0

    R = int(sens_wacc_points)
    C = int(sens_g_points)

    w_mid = (R - 1) / 2.0
    g_mid = (C - 1) / 2.0

    # Header g values (row 3)
    for j in range(C):
        g_val = float(g) + (j - g_mid) * g_step_dec
        cell = wsSens.cell(3, 2 + j, g_val)
        cell.number_format = "0.00%"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all

    # WACC values (col A)
    for i in range(R):
        w_val = float(wacc) + (i - w_mid) * wacc_step_dec
        cell = wsSens.cell(4 + i, 1, w_val)
        cell.number_format = "0.00%"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all

    # Grid formulas based on UFCF + discount periods in DCF sheet
    ufcf_row = r_start + 4
    n_row = r_start + 8
    last_year_col = 3 + n_fore - 1
    last_year_colL = _excel_col(last_year_col)

    ufcf_rng = f"DCF!$C${ufcf_row}:${last_year_colL}${ufcf_row}"
    n_rng = f"DCF!$C${n_row}:${last_year_colL}${n_row}"
    last_ufcf_cell = f"DCF!${last_year_colL}${ufcf_row}"
    last_n_cell = f"DCF!${last_year_colL}${n_row}"

    for i in range(R):
        wacc_hdr = f"$A{4+i}"
        for j in range(C):
            g_hdr = f"{_excel_col(2+j)}$3"

            pv_explicit = f"SUMPRODUCT({ufcf_rng}/(1+{wacc_hdr})^{n_rng})"
            tv = f"({last_ufcf_cell}*(1+{g_hdr}))/({wacc_hdr}-{g_hdr})"
            pv_tv = f"({tv})/(1+{wacc_hdr})^{last_n_cell}"
            formula = f"=({pv_explicit})+({pv_tv})-{NETDEBT}"

            cell = wsSens.cell(4 + i, 2 + j, formula)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = border_all

    # Base highlight (center)
    base_i = int(round(w_mid))
    base_j = int(round(g_mid))
    base_cell = wsSens.cell(4 + base_i, 2 + base_j)
    base_cell.fill = PatternFill("solid", fgColor=SKY)
    base_cell.font = Font(bold=True)

    # Min/Max summary
    stats_r = 6 + R
    wsSens[f"A{stats_r}"] = "Base Equity Value"
    wsSens[f"C{stats_r}"] = "Lowest (grid)"
    wsSens[f"E{stats_r}"] = "Highest (grid)"
    for c in ["A", "C", "E"]:
        wsSens[f"{c}{stats_r}"].font = Font(bold=True)

    grid_top_left = "B4"
    grid_bot_right = f"{_excel_col(1+C)}{3+R}"
    grid_range = f"{grid_top_left}:{grid_bot_right}"

    wsSens[f"A{stats_r+1}"] = f"={_excel_col(2+base_j)}{4+base_i}"
    wsSens[f"C{stats_r+1}"] = f"=MIN({grid_range})"
    wsSens[f"E{stats_r+1}"] = f"=MAX({grid_range})"
    for c in ["A", "C", "E"]:
        wsSens[f"{c}{stats_r+1}"].number_format = "#,##0"
        wsSens[f"{c}{stats_r+1}"].font = Font(bold=True, size=16)

    wsSens.column_dimensions["A"].width = 12
    for j in range(C):
        wsSens.column_dimensions[_excel_col(2 + j)].width = 16
    wsSens.freeze_panes = "B4"

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

st.markdown("---")
section("⬇️ Download FULL DCF Excel Model (Formulas + Sensitivity)")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
# ✅ pull sensitivity settings from your NEW persistent store keys
sens_wacc_points = int(st.session_state.get("sens_store_wacc_points", 5))
sens_g_points = int(st.session_state.get("sens_store_g_points", 7))
sens_wacc_step_pct = float(st.session_state.get("sens_store_wacc_step_pct", 5.0))
sens_g_step_pct = float(st.session_state.get("sens_store_g_step_pct", 0.5))

# safe locals
_wc_percent_used = float(wc_percent_avg) if "wc_percent_avg" in globals() else 0.0
_last_wc_hist_value = float(last_wc_hist_value) if "last_wc_hist_value" in globals() else 0.0
_cos_ratio = float(cos_ratio) if "cos_ratio" in globals() else 0.0

if st.button("📥 Generate FULL Excel Model", key="gen_full_dcf_excel_btn"):
    wb = build_full_dcf_excel_model(
        is_df=is_df,
        forecast_is_df=forecast_is,
        year_cols_is=year_cols_is,
        forecast_years_int=forecast_years_int,
        forecast_cols=forecast_cols,

        rev_idx=rev_idx, cos_idx=cos_idx, gp_idx=gp_idx, ebitda_idx=ebitda_idx,
        op_idx=op_idx, pbt_idx=pbt_idx, tax_idx=tax_idx, np_idx=np_idx,

        growth_mode=growth_mode,
        yearly_g_dict=yearly_g,
        avg_g=avg_g,
        avg_tax_ratio=avg_tax_ratio,
        avg_gp_margin=avg_gp_margin if ("avg_gp_margin" in globals() and avg_gp_margin is not None) else None,
        cos_ratio=_cos_ratio,

        wc_percent_used=_wc_percent_used,
        last_wc_hist_value=_last_wc_hist_value,

        discount_periods_n=discount_periods_n,
        dep_forecast_vals=dep_forecast_vals,
        capex_forecast_vals=capex_forecast_vals,
        wacc=wacc, tax=tax, g=g, net_debt=net_debt,

        rf=rf,
        mrp=mrp,
        dcf_unlevered_beta=float(st.session_state.get("dcf_unlevered_beta", 1.0)),
        rd=float(st.session_state.get("rd", 0.0)),
        total_debt=float(st.session_state.get("total_debt", 0.0)),
        cash_balance=float(st.session_state.get("cash_balance", 0.0)),
        book_equity=float(st.session_state.get("book_equity", 0.0)),
        de_ratio=float(st.session_state.get("de_ratio", 0.0)),

        # ✅ EXACT sensitivity inputs (won’t reset)
        sens_wacc_points=sens_wacc_points,
        sens_g_points=sens_g_points,
        sens_wacc_step_pct=sens_wacc_step_pct,
        sens_g_step_pct=sens_g_step_pct,
    )

    xbytes = workbook_to_bytes(wb)

    st.download_button(
        "✅ Download FULL_DCF_Model.xlsx",
        data=xbytes,
        file_name="FULL_DCF_Model.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_full_dcf_model_btn",
    )
