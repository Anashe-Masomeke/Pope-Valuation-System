import streamlit as st
import pandas as pd
import numpy as np
import re
from pathlib import Path
from bs4 import BeautifulSoup
from io import StringIO
import requests
import time
import random
import base64
FMP_API_KEY = st.secrets["FMP_API_KEY"]
st.set_page_config(page_title="Comparables Valuation", layout="wide")

# ---------------------------------------------------------
# GLOBAL STYLING
# ---------------------------------------------------------
def add_watermark():
    logo_path = Path("assets") / "fbc_logo.png"
    if logo_path.exists():
        with open(logo_path, "rb") as f:
            logo_base64 = base64.b64encode(f.read()).decode()
        st.markdown(f"""
        <style>
        .stApp::before {{
            content: "";
            position: fixed;
            top: 40px;
            left: 50px;
            width: 100%;
            height: 100%;
            background-image: url("data:image/png;base64,{logo_base64}");
            background-repeat: no-repeat;
            background-position: center;
            background-size: 1500px;
            opacity: 0.07;
            pointer-events: none;
            z-index: 0;
        }}
        .block-container {{
            position: relative;
            z-index: 1;
        }}
        </style>
        """, unsafe_allow_html=True)

add_watermark()

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/icon?family=Material+Icons');
.material-icons,
span.material-icons,
i.material-icons,
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stSidebarCollapseButton"] i {
    font-family: 'Material Icons' !important;
}
[data-testid="stSidebarCollapseButton"] button {
    background: #003399 !important;
    border: 1px solid rgba(255,255,255,0.25) !important;
    border-radius: 999px !important;
    width: 44px !important;
    height: 44px !important;
    box-shadow: 0 6px 18px rgba(0, 51, 153, 0.35) !important;
}
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #003399 0%, #001a4d 100%) !important;
    color: white !important;
}
section[data-testid="stSidebar"] * {
    color: white !important;
}
html, body, .stApp, .block-container,
p, div, label,
h1, h2, h3, h4, h5, h6 {
  font-family: Georgia, "Times New Roman", serif !important;
}
</style>
""", unsafe_allow_html=True)

st.title("📊 Comparables Valuation – EV/EBITDA, P/B, P/E")
st.markdown("All values + inputs are **saved in session_state**, so switching pages keeps your work.")

if st.button("🔄 Refresh ratios (fix missing data)"):
    st.cache_data.clear()
    st.rerun()

S = st.session_state


# =========================================================
# HELPERS
# =========================================================
@st.cache_data(show_spinner=False, ttl=60*60*24)
def fmp_ratios(symbol: str) -> dict:
    sym = normalize_peer_ticker(symbol)

    out = {
        "Ticker": sym,
        "P/E": np.nan,
        "P/B": np.nan,
        "EV/EBITDA": np.nan,
        "ratio_source": "",
        "ratio_note": ""
    }

    if not sym:
        return out

    try:
        url = f"https://financialmodelingprep.com/api/v3/ratios/{sym}?apikey={FMP_API_KEY}"
        r = requests.get(url, timeout=10)
        data = r.json()

        if isinstance(data, list) and len(data) > 0:
            d = data[0]

            out["P/E"] = _clean_num(d.get("priceEarningsRatio"))
            out["P/B"] = _clean_num(d.get("priceToBookRatio"))
            out["EV/EBITDA"] = _clean_num(d.get("enterpriseValueMultiple"))

            if not _all_nan_ratio_dict(out):
                out["ratio_source"] = "FMP API"
                out["ratio_note"] = "Fetched from Financial Modeling Prep API"

    except Exception as e:
        out["ratio_note"] = f"FMP failed: {repr(e)}"

    return out
def format_numeric_columns(df):
    fmt = {}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            fmt[col] = "{:,.2f}"
    return df.style.format(fmt)


def _clean_text(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return str(x).strip()


def _norm_text(x: str) -> str:
    return re.sub(r"[^a-z0-9]", "", _clean_text(x).lower())


def _clean_num(x):
    try:
        if x is None or x == "":
            return np.nan
        return float(x)
    except Exception:
        return np.nan


def _num_input_default(x, fallback=0.0):
    try:
        if x is None or pd.isna(x):
            return float(fallback)
        return float(x)
    except Exception:
        return float(fallback)


def normalize_peer_ticker(symbol: str) -> str:
    sym = _clean_text(symbol).upper()
    fixes = {
        "MTNN": "MTNN.NG",
        "SCOM": "SCOM.KE",
        "SAFARICOM": "SCOM.KE",
        "EQTY": "EQTY.KE",
        "KCB": "KCB.KE",
        "VODACOM": "VOD.JO",
        "MTN": "MTN.JO",
    }
    return fixes.get(sym, sym)


def make_yahoo_profile_url(symbol: str) -> str:
    sym = normalize_peer_ticker(symbol)
    return f"https://finance.yahoo.com/quote/{sym}/profile/" if sym else ""


def make_yahoo_statistics_url(symbol: str) -> str:
    sym = normalize_peer_ticker(symbol)
    return f"https://finance.yahoo.com/quote/{sym}/key-statistics?p={sym}" if sym else ""


# FIX: single, correct filtered_average — no duplicate, no missing return
def filtered_average(series):
    series = pd.to_numeric(series, errors="coerce")
    series = series.dropna()
    if len(series) == 0:
        return np.nan
    return float(series.mean())


def _std_colname(x: str) -> str:
    x = _clean_text(x).lower()
    x = re.sub(r"[^a-z0-9]+", "_", x)
    return x.strip("_")


def split_keywords(text: str):
    out = []
    for x in _clean_text(text).split(","):
        x = _clean_text(x).lower()
        if x:
            out.append(x)
    return out


def _all_nan_ratio_dict(d: dict) -> bool:
    return (
        pd.isna(d.get("P/E", np.nan)) and
        pd.isna(d.get("P/B", np.nan)) and
        pd.isna(d.get("EV/EBITDA", np.nan))
    )


# =========================================================
# UNIVERSE FILE
# =========================================================
def find_universe_file() -> str:
    candidates = [
        "data/africa_yahoo_peer_universe_strict_final.xlsx",
        "data/africa_yahoo_peer_universe_starter.xlsx",
        "africa_yahoo_peer_universe_strict_final.xlsx",
        "africa_yahoo_peer_universe_starter.xlsx",
        "/mnt/data/africa_yahoo_peer_universe_strict_final.xlsx",
        "/mnt/data/africa_yahoo_peer_universe_starter.xlsx",
    ]
    for p in candidates:
        if Path(p).exists():
            return p
    return ""


def _find_sheet_name(xls: pd.ExcelFile, wanted_names: list, fallback_contains: list = None):
    sheets = list(xls.sheet_names)
    norm_map = {_std_colname(s): s for s in sheets}
    for w in wanted_names:
        wn = _std_colname(w)
        if wn in norm_map:
            return norm_map[wn]
    if fallback_contains:
        for s in sheets:
            sn = _std_colname(s)
            if all(tok in sn for tok in fallback_contains):
                return s
    return None


@st.cache_data(show_spinner=False)
def load_peer_universe_bundle(path: str):
    if not path or not Path(path).exists():
        return None, None, None, {"error": "Universe file not found."}
    try:
        xls = pd.ExcelFile(path)
    except Exception as e:
        return None, None, None, {"error": f"Could not open Excel file: {e}"}

    universe_sheet = _find_sheet_name(
        xls,
        wanted_names=["CODE_READY_UNIVERSE", "code_ready_universe", "UNIVERSE", "peer_universe"],
        fallback_contains=["universe"]
    )
    zim_map_sheet = _find_sheet_name(
        xls,
        wanted_names=["ZIM_TARGET_MAP", "zim_target_map", "TARGET_MAP"],
        fallback_contains=["target", "map"]
    )
    alias_sheet = _find_sheet_name(
        xls,
        wanted_names=["SECTOR_ALIAS_MAP", "sector_alias_map", "ALIAS_MAP"],
        fallback_contains=["alias"]
    )

    debug = {
        "file": path,
        "sheet_names": list(xls.sheet_names),
        "universe_sheet": universe_sheet,
        "zim_map_sheet": zim_map_sheet,
        "alias_sheet": alias_sheet,
    }

    if universe_sheet is None:
        return None, None, None, {**debug, "error": "Could not find the universe sheet in the Excel file."}

    universe = pd.read_excel(xls, sheet_name=universe_sheet)
    zim_map = pd.read_excel(xls, sheet_name=zim_map_sheet) if zim_map_sheet else pd.DataFrame()
    alias_map = pd.read_excel(xls, sheet_name=alias_sheet) if alias_sheet else pd.DataFrame()

    universe.columns = [_std_colname(c) for c in universe.columns]
    zim_map.columns = [_std_colname(c) for c in zim_map.columns]
    alias_map.columns = [_std_colname(c) for c in alias_map.columns]

    rename_universe = {
        "symbol": "ticker", "peer_ticker": "ticker", "peer_symbol": "ticker",
        "company_name": "company", "peer_company": "company", "peer_name": "company",
        "country_name": "country", "sector_name": "sector", "industry_name": "industry",
        "keywords": "sector_keywords", "peer_keywords": "sector_keywords",
        "priority": "match_priority", "yahoo_confirmed": "yahoo_status",
        "status": "yahoo_status", "exchange_name": "exchange",
    }
    universe = universe.rename(columns={k: v for k, v in rename_universe.items() if k in universe.columns})

    rename_zim = {
        "symbol": "target_symbol", "company": "target_company",
        "sector": "preferred_sector", "industry": "preferred_industry",
        "keywords": "preferred_peer_keywords",
    }
    zim_map = zim_map.rename(columns={k: v for k, v in rename_zim.items() if k in zim_map.columns})

    rename_alias = {
        "alias": "input_alias", "sector": "preferred_sector", "industry": "preferred_industry",
    }
    alias_map = alias_map.rename(columns={k: v for k, v in rename_alias.items() if k in alias_map.columns})

    for col in ["ticker", "company", "country", "exchange", "sector", "industry",
                "sector_keywords", "match_priority", "yahoo_status"]:
        if col not in universe.columns:
            universe[col] = ""
    for col in ["target_symbol", "target_company", "preferred_sector", "preferred_industry",
                "preferred_peer_keywords", "search_aliases"]:
        if col not in zim_map.columns:
            zim_map[col] = ""
    for col in ["input_alias", "preferred_sector", "preferred_industry"]:
        if col not in alias_map.columns:
            alias_map[col] = ""

    for df in [universe, zim_map, alias_map]:
        for col in df.columns:
            df[col] = df[col].map(_clean_text)

    universe["ticker"] = universe["ticker"].map(normalize_peer_ticker)
    zim_map["target_symbol"] = zim_map["target_symbol"].map(lambda x: _clean_text(x).upper())
    alias_map["input_alias"] = alias_map["input_alias"].map(lambda x: _clean_text(x).lower())
    universe = universe[universe["ticker"].map(lambda x: _clean_text(x) != "")].copy()

    debug["universe_columns"] = list(universe.columns)
    debug["zim_map_columns"] = list(zim_map.columns)
    debug["alias_map_columns"] = list(alias_map.columns)
    debug["universe_rows"] = len(universe)
    debug["zim_map_rows"] = len(zim_map)
    debug["alias_map_rows"] = len(alias_map)

    return universe, zim_map, alias_map, debug


UNIVERSE_FILE = find_universe_file()
UNIVERSE_DF, ZIM_TARGET_MAP_DF, SECTOR_ALIAS_DF, UNIVERSE_DEBUG = load_peer_universe_bundle(UNIVERSE_FILE)

# =========================================================
# FALLBACK ZIM TARGET MAP
# =========================================================
FALLBACK_ZIM_TARGETS = [
    {"target_symbol": "ECOZIM", "target_company": "Econet Wireless Zimbabwe",
     "preferred_sector": "Telecommunications", "preferred_industry": "Mobile Telecoms",
     "preferred_peer_keywords": "telecommunications,telecom,mobile,wireless,communications,network,broadband,data"},
    {"target_symbol": "CBZ", "target_company": "CBZ", "preferred_sector": "Banking",
     "preferred_industry": "Commercial Banks",
     "preferred_peer_keywords": "banking,bank,commercial bank,lending,deposits"},
    {"target_symbol": "FBC", "target_company": "FBC", "preferred_sector": "Banking",
     "preferred_industry": "Commercial Banks",
     "preferred_peer_keywords": "banking,bank,commercial bank,lending,deposits"},
    {"target_symbol": "NMBZ", "target_company": "NMBZ", "preferred_sector": "Banking",
     "preferred_industry": "Commercial Banks",
     "preferred_peer_keywords": "banking,bank,commercial bank,lending,deposits"},
    {"target_symbol": "ZB", "target_company": "ZB", "preferred_sector": "Banking",
     "preferred_industry": "Commercial Banks",
     "preferred_peer_keywords": "banking,bank,commercial bank,lending,deposits"},
    {"target_symbol": "PADENGA", "target_company": "Padenga", "preferred_sector": "Mining",
     "preferred_industry": "Gold Mining",
     "preferred_peer_keywords": "gold mining,gold,gold producer,gold miner,mining,minerals,precious metals"},
    {"target_symbol": "CMCL", "target_company": "Caledonia", "preferred_sector": "Mining",
     "preferred_industry": "Gold Mining",
     "preferred_peer_keywords": "gold mining,gold,gold producer,gold miner,mining,minerals,precious metals"},
    {"target_symbol": "DELTA", "target_company": "Delta", "preferred_sector": "Consumer Staples",
     "preferred_industry": "Beverages",
     "preferred_peer_keywords": "beverages,brewery,beer,spirits,distillery,drinks"},
    {"target_symbol": "AFDIS", "target_company": "Afdis", "preferred_sector": "Consumer Staples",
     "preferred_industry": "Beverages",
     "preferred_peer_keywords": "beverages,brewery,beer,spirits,distillery,drinks"},
    {"target_symbol": "INNSCOR", "target_company": "Innscor Africa", "preferred_sector": "Consumer Staples",
     "preferred_industry": "Food Producers",
     "preferred_peer_keywords": "food,consumer,food processing,packaged foods,brands"},
    {"target_symbol": "SIMBISA", "target_company": "Simbisa", "preferred_sector": "Consumer Discretionary",
     "preferred_industry": "Restaurants",
     "preferred_peer_keywords": "restaurants,quick service restaurants,foodservice,fast food"},
    {"target_symbol": "WESTPROP", "target_company": "WestProp", "preferred_sector": "Real Estate",
     "preferred_industry": "Property",
     "preferred_peer_keywords": "real estate,property,reit,property development"},
    {"target_symbol": "RTG", "target_company": "Rainbow Tourism Group",
     "preferred_sector": "Consumer Discretionary", "preferred_industry": "Hotels",
     "preferred_peer_keywords": "hotels,lodging,leisure,tourism"},
    {"target_symbol": "ASUN", "target_company": "African Sun", "preferred_sector": "Consumer Discretionary",
     "preferred_industry": "Hotels",
     "preferred_peer_keywords": "hotels,lodging,leisure,tourism"},
]

if ZIM_TARGET_MAP_DF is None or ZIM_TARGET_MAP_DF.empty:
    ZIM_TARGET_MAP_DF = pd.DataFrame(FALLBACK_ZIM_TARGETS)

if SECTOR_ALIAS_DF is None or SECTOR_ALIAS_DF.empty:
    SECTOR_ALIAS_DF = pd.DataFrame([
        {"input_alias": "telecom", "preferred_sector": "Telecommunications",
         "preferred_industry": "Mobile Telecoms"},
        {"input_alias": "telecommunications", "preferred_sector": "Telecommunications",
         "preferred_industry": "Mobile Telecoms"},
        {"input_alias": "communication services", "preferred_sector": "Telecommunications",
         "preferred_industry": "Mobile Telecoms"},
        {"input_alias": "mobile telecoms", "preferred_sector": "Telecommunications",
         "preferred_industry": "Mobile Telecoms"},
        {"input_alias": "banking", "preferred_sector": "Banking", "preferred_industry": "Commercial Banks"},
        {"input_alias": "banks", "preferred_sector": "Banking", "preferred_industry": "Commercial Banks"},
        {"input_alias": "gold mining", "preferred_sector": "Mining", "preferred_industry": "Gold Mining"},
        {"input_alias": "mining", "preferred_sector": "Mining", "preferred_industry": ""},
        {"input_alias": "beverages", "preferred_sector": "Consumer Staples", "preferred_industry": "Beverages"},
        {"input_alias": "real estate", "preferred_sector": "Real Estate", "preferred_industry": "Property"},
    ])


# =========================================================
# TARGET RESOLUTION
# =========================================================
def find_target_row(query: str):
    q = _clean_text(query)
    qn = _norm_text(q)
    if not qn or ZIM_TARGET_MAP_DF is None or ZIM_TARGET_MAP_DF.empty:
        return None
    for _, r in ZIM_TARGET_MAP_DF.iterrows():
        if _norm_text(r.get("target_symbol", "")) == qn:
            return r.to_dict()
    for _, r in ZIM_TARGET_MAP_DF.iterrows():
        if _norm_text(r.get("target_company", "")) == qn:
            return r.to_dict()
    for _, r in ZIM_TARGET_MAP_DF.iterrows():
        if qn and qn in _norm_text(r.get("target_company", "")):
            return r.to_dict()
    return None


def normalize_sector_override(sector_text: str):
    s = _clean_text(sector_text).lower()
    if not s:
        return "", ""
    for _, r in SECTOR_ALIAS_DF.iterrows():
        if s == _clean_text(r.get("input_alias")).lower():
            return _clean_text(r.get("preferred_sector")), _clean_text(r.get("preferred_industry"))
    return sector_text.strip(), ""


def get_target_profile(target_query: str, manual_sector_override: str = ""):
    row = find_target_row(target_query)
    if row:
        target_symbol = _clean_text(row.get("target_symbol")).upper()
        target_company = _clean_text(row.get("target_company"))
        preferred_sector = _clean_text(row.get("preferred_sector"))
        preferred_industry = _clean_text(row.get("preferred_industry"))
        preferred_keywords = split_keywords(row.get("preferred_peer_keywords"))
        search_aliases = split_keywords(row.get("search_aliases", ""))
        source = "ZIM_TARGET_MAP"
    else:
        target_symbol = _clean_text(target_query).upper()
        target_company = _clean_text(target_query)
        preferred_sector = ""
        preferred_industry = ""
        preferred_keywords = []
        search_aliases = []
        source = "manual"

    if manual_sector_override.strip():
        sec, ind = normalize_sector_override(manual_sector_override)
        preferred_sector = sec or preferred_sector
        preferred_industry = ind
        source = f"{source} + manual_sector_override"

    if not preferred_keywords:
        preferred_keywords = split_keywords(f"{preferred_sector},{preferred_industry}")

    return {
        "target_symbol": target_symbol,
        "target_company": target_company,
        "preferred_sector": preferred_sector,
        "preferred_industry": preferred_industry,
        "preferred_peer_keywords": preferred_keywords,
        "search_aliases": search_aliases,
        "source": source,
    }


# =========================================================
# UNIVERSE FILTERING
# =========================================================
def strict_peer_score(peer_row: dict, target_profile: dict) -> int:
    score = 0
    peer_sector = _clean_text(peer_row.get("sector")).lower()
    peer_industry = _clean_text(peer_row.get("industry")).lower()
    peer_keywords = _clean_text(peer_row.get("sector_keywords")).lower()
    peer_company = _clean_text(peer_row.get("company")).lower()
    peer_country = _clean_text(peer_row.get("country")).lower()
    peer_priority = _clean_num(peer_row.get("match_priority"))

    tgt_sector = _clean_text(target_profile.get("preferred_sector")).lower()
    tgt_industry = _clean_text(target_profile.get("preferred_industry")).lower()
    tgt_keywords = [k.lower() for k in target_profile.get("preferred_peer_keywords", [])]

    combo = f"{peer_sector} | {peer_industry} | {peer_keywords} | {peer_company}"

    if tgt_sector and peer_sector == tgt_sector:
        score += 50
    elif tgt_sector and (tgt_sector in peer_sector or peer_sector in tgt_sector):
        score += 35
    elif tgt_sector and tgt_sector in combo:
        score += 20

    if tgt_industry and peer_industry == tgt_industry:
        score += 60
    elif tgt_industry and (tgt_industry in peer_industry or peer_industry in tgt_industry):
        score += 40
    elif tgt_industry and tgt_industry in combo:
        score += 20

    for kw in tgt_keywords:
        if kw and kw in peer_industry:
            score += 15
        if kw and kw in peer_keywords:
            score += 12
        if kw and kw in peer_sector:
            score += 10
        if kw and kw in peer_company:
            score += 4

    if not pd.isna(peer_priority):
        score += int(peer_priority) * 5

    if peer_country in ["south africa", "kenya", "nigeria", "zimbabwe", "botswana",
                        "egypt", "ghana", "mauritius"]:
        score += 3

    return score


def strict_universe_filter(target_profile: dict, max_peers: int = 8):
    if UNIVERSE_DF is None or UNIVERSE_DF.empty:
        return pd.DataFrame()

    df = UNIVERSE_DF.copy()

    tgt_symbol = _clean_text(target_profile.get("target_symbol")).upper()
    tgt_company = _clean_text(target_profile.get("target_company"))
    tgt_sector = _clean_text(target_profile.get("preferred_sector")).lower()
    tgt_industry = _clean_text(target_profile.get("preferred_industry")).lower()
    tgt_keywords = [k.lower() for k in target_profile.get("preferred_peer_keywords", [])]

    df = df[df["ticker"].map(lambda x: _clean_text(x).upper()) != tgt_symbol].copy()
    df = df[df["company"].map(lambda x: _norm_text(x)) != _norm_text(tgt_company)].copy()

    df["sector_l"] = df["sector"].map(lambda x: _clean_text(x).lower())
    df["industry_l"] = df["industry"].map(lambda x: _clean_text(x).lower())
    df["keywords_l"] = df["sector_keywords"].map(lambda x: _clean_text(x).lower())
    df["company_l"] = df["company"].map(lambda x: _clean_text(x).lower())

    same_sector = df[df["sector_l"] == tgt_sector].copy() if tgt_sector else df.copy()

    same_industry = pd.DataFrame()
    if not same_sector.empty and tgt_industry:
        same_industry = same_sector[
            same_sector["industry_l"].str.contains(re.escape(tgt_industry), na=False) |
            same_sector["industry_l"].eq(tgt_industry)
        ].copy()

    keyword_match = pd.DataFrame()
    if not same_sector.empty and tgt_keywords:
        keyword_match = same_sector[
            same_sector.apply(
                lambda r: any(
                    kw in f"{r['sector_l']} | {r['industry_l']} | {r['keywords_l']} | {r['company_l']}"
                    for kw in tgt_keywords if kw
                ),
                axis=1
            )
        ].copy()

    combined = pd.concat([same_industry, keyword_match, same_sector], axis=0, ignore_index=True)
    if combined.empty:
        combined = df.copy()

    combined = combined.drop_duplicates(subset=["ticker"]).copy()
    combined["SimilarityScore"] = combined.apply(
        lambda r: strict_peer_score(r.to_dict(), target_profile), axis=1
    )

    if "match_priority" not in combined.columns:
        combined["match_priority"] = ""

    combined["match_priority_num"] = pd.to_numeric(
        combined["match_priority"], errors="coerce"
    ).fillna(0)

    combined = combined.sort_values(
        by=["SimilarityScore", "match_priority_num", "company"],
        ascending=[False, False, True]
    ).reset_index(drop=True)

    S["debug_filter_stage_counts"] = {
        "universe_rows": len(df),
        "same_sector_rows": len(same_sector),
        "same_industry_rows": len(same_industry),
        "keyword_match_rows": len(keyword_match),
        "combined_rows": len(combined),
    }
    S["debug_strict_candidates_preview"] = combined[[
        "ticker", "company", "sector", "industry", "sector_keywords",
        "match_priority", "SimilarityScore"
    ]].head(30)

    return combined.head(max_peers).reset_index(drop=True)


# =========================================================
# NETWORK HELPERS
# =========================================================
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://finance.yahoo.com/",
    "Origin": "https://finance.yahoo.com",
    "X-Requested-With": "XMLHttpRequest",
    "Connection": "keep-alive",
}

YAHOO_QUOTESUMMARY_URL = "https://query2.finance.yahoo.com/v10/finance/quoteSummary/{symbol}"


def get_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def _safe_get(url, params=None, timeout=15, tries=3, headers=None):
    last_err = None
    use_headers = headers or HEADERS
    for attempt in range(int(tries)):
        try:
            session = get_session()
            r = session.get(url, params=params, timeout=timeout,
                            headers=use_headers, allow_redirects=True)
            if r.status_code == 429:
                # FIX: longer back-off on rate limit
                time.sleep(3.0 + attempt * 2.0 + random.random())
                continue
            if r.status_code == 403:
                raise requests.HTTPError(
                    f"403 Forbidden: {r.url}", response=r)
            r.raise_for_status()
            return r
        except Exception as e:
            last_err = e
            time.sleep(4.0 + attempt * 2.5 + random.random())
    raise last_err


# FIX: yahoo_warmup only fires ONCE per session, not on every rerun
def yahoo_warmup():
    if st.session_state.get("yahoo_warmed_up", False):
        return
    try:
        _safe_get(
            "https://finance.yahoo.com/",
            timeout=15,
            tries=3,
            headers={
                "User-Agent": HEADERS["User-Agent"],
                "Accept-Language": "en-US,en;q=0.9",
                "Referer": "https://finance.yahoo.com/",
            },
        )
        _safe_get(
            "https://query2.finance.yahoo.com/v1/finance/search",
            params={"q": "test", "quotesCount": 1, "newsCount": 0},
            timeout=15,
            tries=3,
        )
    except Exception:
        pass
    st.session_state["yahoo_warmed_up"] = True


yahoo_warmup()


# =========================================================
# INVESTING.COM FALLBACK
# =========================================================
INVESTING_SYMBOL_MAP = {
    "MTNN.NG": {"slug": "mtn-nigeria-com"},
    "MTN.RW": {"slug": "mtn-rwandacell"},
    "MTN.JO": {"slug": "mtn-group-ltd"},
    "VOD.JO": {"slug": "vodacom-group-ltd"},
    "SCOM.KE": {"slug": "safaricom"},
    "CTL.RW": {"slug": "crystal-telecom-ltd"},
    "ANH.JO": {"slug": "anheuser-busch-inbev-sa-nv"},
    "BLR.RW": {"slug": "bralirwa"},
}


def make_investing_url(symbol: str) -> str:
    sym = normalize_peer_ticker(symbol)
    meta = INVESTING_SYMBOL_MAP.get(sym, {})
    slug = meta.get("slug", "")
    if not slug:
        return ""
    return f"https://www.investing.com/equities/{slug}"


def _first_non_null(d: dict, keys):
    for k in keys:
        v = d.get(k)
        if v is not None and v != "":
            return v
    return None


# =========================================================
# RATIO FETCHERS — all cached, NO st.* calls inside
# =========================================================

def _parse_ratio_value(x) -> float:
    s = str(x).strip()
    if s in ["", "N/A", "NaN", "None", "-", "--"]:
        return np.nan
    s = s.replace(",", "").replace("x", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return np.nan
    try:
        return float(m.group(0))
    except Exception:
        return np.nan


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def investing_ratios(symbol: str) -> dict:
    sym = normalize_peer_ticker(symbol)
    out = {
        "Ticker": sym, "Company": "", "Exchange": "", "Country": "",
        "Sector": "", "Industry": "", "P/E": np.nan, "P/B": np.nan,
        "EV/EBITDA": np.nan, "ratio_source": "", "ratio_note": "",
        "profile_url": "", "stats_url": "",
    }
    if not sym:
        out["ratio_note"] = "Blank symbol after normalization."
        return out

    investing_url = make_investing_url(sym)
    if not investing_url:
        out["ratio_note"] = f"No Investing mapping configured for {sym}"
        return out

    out["profile_url"] = investing_url
    out["stats_url"] = investing_url

    try:
        HEADERS = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": "https://finance.yahoo.com/",
            "Connection": "keep-alive",
        }
        r = _safe_get(investing_url, timeout=15, tries=3, headers=HEADERS)
        html = r.text or ""
        soup = BeautifulSoup(html, "html.parser")

        h1 = soup.find("h1")
        if h1:
            out["Company"] = h1.get_text(" ", strip=True)

        lines = [x.strip() for x in soup.get_text("\n", strip=True).split("\n") if x.strip()]

        def extract_metric(lines_in, labels):
            for i, line in enumerate(lines_in):
                line_l = line.lower().strip()
                for label in labels:
                    if label in line_l:
                        v = _parse_ratio_value(line)
                        if not pd.isna(v):
                            return v
                        for j in range(i + 1, min(i + 6, len(lines_in))):
                            v = _parse_ratio_value(lines_in[j])
                            if not pd.isna(v):
                                return v
            return np.nan

        out["P/E"] = extract_metric(lines, ["p/e ratio", "p/e"])
        out["P/B"] = extract_metric(lines, ["price/book", "price / book", "price to book"])
        out["EV/EBITDA"] = extract_metric(lines, ["ev/ebitda", "ev / ebitda", "enterprise value/ebitda"])

        if pd.isna(out["P/E"]):
            m = re.search(r"P/E Ratio.*?([0-9]+\.[0-9]+)", html, re.I | re.S)
            if m:
                out["P/E"] = _parse_ratio_value(m.group(1))
        if pd.isna(out["P/B"]):
            m = re.search(r"Price/Book.*?([0-9]+\.[0-9]+)", html, re.I | re.S)
            if m:
                out["P/B"] = _parse_ratio_value(m.group(1))
        if pd.isna(out["EV/EBITDA"]):
            m = re.search(r"EV/EBITDA.*?([0-9]+\.[0-9]+)", html, re.I | re.S)
            if m:
                out["EV/EBITDA"] = _parse_ratio_value(m.group(1))

        if not _all_nan_ratio_dict(out):
            out["ratio_source"] = "Investing"
            out["ratio_note"] = "Fetched from Investing.com."
        else:
            out["ratio_note"] = "Investing page loaded but ratios were not extracted."

    except requests.HTTPError as e:
        msg = str(e)
        out["ratio_note"] = f"Investing blocked (403): {investing_url}" if "403" in msg else f"Investing HTTP error: {repr(e)}"
    except Exception as e:
        out["ratio_note"] = f"Investing fetch failed: {repr(e)}"

    return out


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def yahoo_profile_and_metrics(symbol: str) -> dict:
    sym = normalize_peer_ticker(symbol)
    out = {
        "Ticker": sym, "Company": "", "Exchange": "", "Country": "",
        "Sector": "", "Industry": "", "Description": "",
        "P/E": np.nan, "P/B": np.nan, "EV/EBITDA": np.nan,
        "ratio_source": "", "ratio_note": "", "quote_exists": False,
    }
    if not sym:
        return out

    def _raw(v):
        if isinstance(v, dict):
            return v.get("raw", v.get("fmt"))
        return v

    try:
        url = YAHOO_QUOTESUMMARY_URL.format(symbol=sym)
        r = _safe_get(
            url,
            params={"modules": "price,summaryDetail,defaultKeyStatistics,financialData,assetProfile"},
            timeout=15,
            tries=3,
        )
        data = r.json()
        res = (((data.get("quoteSummary") or {}).get("result")) or [])
        if not res:
            out["ratio_note"] = "Yahoo quoteSummary returned no result."
            return out

        root = res[0]
        price = root.get("price") or {}
        summary = root.get("summaryDetail") or {}
        dks = root.get("defaultKeyStatistics") or {}
        fin = root.get("financialData") or {}
        ap = root.get("assetProfile") or {}

        out["quote_exists"] = True
        out["Company"] = _raw(price.get("longName")) or _raw(price.get("shortName")) or sym
        out["Exchange"] = _raw(price.get("exchangeName")) or _raw(price.get("fullExchangeName")) or ""
        out["Country"] = ap.get("country") or ""
        out["Sector"] = ap.get("sector") or ""
        out["Industry"] = ap.get("industry") or ""
        out["Description"] = ap.get("longBusinessSummary") or ""

        forward_pe = _raw(summary.get("forwardPE")) or _raw(dks.get("forwardPE")) or _raw(fin.get("forwardPE"))
        trailing_pe = _raw(summary.get("trailingPE")) or _raw(dks.get("trailingPE")) or _raw(fin.get("trailingPE"))
        pb = _raw(dks.get("priceToBook")) or _raw(summary.get("priceToBook")) or _raw(fin.get("priceToBook"))
        evebitda = _raw(fin.get("enterpriseToEbitda")) or _raw(dks.get("enterpriseToEbitda"))

        out["P/E"] = _clean_num(forward_pe) if not pd.isna(_clean_num(forward_pe)) else _clean_num(trailing_pe)
        out["P/B"] = _clean_num(pb)
        out["EV/EBITDA"] = _clean_num(evebitda)

        if not _all_nan_ratio_dict(out):
            out["ratio_source"] = "Yahoo quoteSummary"
            out["ratio_note"] = "Fetched from Yahoo quoteSummary."
        else:
            out["ratio_note"] = "Yahoo quoteSummary found company, but ratios were blank."

    except Exception as e:
        out["ratio_note"] = f"Yahoo quoteSummary failed: {repr(e)}"

    return out


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def yahoo_stats_table_fallback(symbol: str) -> dict:
    sym = normalize_peer_ticker(symbol)
    out = {
        "Ticker": sym, "P/E": np.nan, "P/B": np.nan, "EV/EBITDA": np.nan,
        "ratio_source": "", "ratio_note": "", "page_exists": False,
    }
    if not sym:
        out["ratio_note"] = "Blank symbol after normalization."
        return out

    url = f"https://finance.yahoo.com/quote/{sym}/key-statistics?p={sym}"
    html_headers = {
        "User-Agent": HEADERS["User-Agent"],
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": "https://finance.yahoo.com/",
        "Connection": "keep-alive",
    }

    try:
        r = _safe_get(url, timeout=15, tries=3, headers=html_headers)
        html = r.text or ""
        if html:
            out["page_exists"] = True

        tables = pd.read_html(StringIO(html))
        best_hits = {"forward_pe": np.nan, "trailing_pe": np.nan, "pb": np.nan, "evebitda": np.nan}

        for t in tables:
            if t is None or t.empty or len(t.columns) < 2:
                continue
            t = t.copy()
            t.columns = [str(c).strip() for c in t.columns]
            label_col = t.columns[0]
            current_col = None
            for c in t.columns:
                if "current" in str(c).strip().lower():
                    current_col = c
                    break
            if current_col is None:
                continue
            for _, row in t.iterrows():
                label = str(row.get(label_col, "")).strip().lower()
                value = row.get(current_col, "")
                if label == "forward p/e" and pd.isna(best_hits["forward_pe"]):
                    best_hits["forward_pe"] = _parse_ratio_value(value)
                elif label == "trailing p/e" and pd.isna(best_hits["trailing_pe"]):
                    best_hits["trailing_pe"] = _parse_ratio_value(value)
                elif label == "price/book" and pd.isna(best_hits["pb"]):
                    best_hits["pb"] = _parse_ratio_value(value)
                elif label == "enterprise value/ebitda" and pd.isna(best_hits["evebitda"]):
                    best_hits["evebitda"] = _parse_ratio_value(value)

        out["P/E"] = best_hits["forward_pe"] if not pd.isna(best_hits["forward_pe"]) else best_hits["trailing_pe"]
        out["P/B"] = best_hits["pb"]
        out["EV/EBITDA"] = best_hits["evebitda"]

        if not _all_nan_ratio_dict(out):
            pe_note = "Used Forward P/E." if not pd.isna(best_hits["forward_pe"]) else "Used Trailing P/E fallback."
            out["ratio_source"] = "Yahoo Statistics"
            out["ratio_note"] = f"Fetched from Yahoo Statistics. {pe_note}"
        else:
            out["ratio_note"] = "Yahoo statistics page loaded but valuation rows not found."

    except Exception as e:
        out["ratio_note"] = f"Yahoo stats fetch failed: {repr(e)}"

    return out


@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def yahoo_html_ratio_fallback(symbol: str) -> dict:
    sym = normalize_peer_ticker(symbol)
    out = {
        "Ticker": sym, "P/E": np.nan, "P/B": np.nan, "EV/EBITDA": np.nan,
        "ratio_source": "", "ratio_note": "", "page_exists": False,
    }
    if not sym:
        out["ratio_note"] = "Blank symbol after normalization."
        return out

    try:
        url = f"https://finance.yahoo.com/quote/{sym}"
        r = _safe_get(url, timeout=20, tries=2)
        html = r.text or ""
        if html:
            out["page_exists"] = True

        patterns = {
            "P/E": [
                r'"trailingPE"\s*:\s*\{"raw"\s*:\s*([\-0-9.]+)',
                r'"trailingPE"\s*:\s*([\-0-9.]+)',
            ],
            "P/B": [
                r'"priceToBook"\s*:\s*\{"raw"\s*:\s*([\-0-9.]+)',
                r'"priceToBook"\s*:\s*([\-0-9.]+)',
            ],
            "EV/EBITDA": [
                r'"enterpriseToEbitda"\s*:\s*\{"raw"\s*:\s*([\-0-9.]+)',
                r'"enterpriseToEbitda"\s*:\s*([\-0-9.]+)',
            ],
        }
        for field, pats in patterns.items():
            for pat in pats:
                m = re.search(pat, html)
                if m:
                    out[field] = _clean_num(m.group(1))
                    break

        if not _all_nan_ratio_dict(out):
            out["ratio_source"] = "Yahoo HTML fallback"
            out["ratio_note"] = "Fetched from Yahoo HTML fallback."
        else:
            out["ratio_note"] = "Yahoo quote page loaded but embedded ratios not found."

    except Exception as e:
        out["ratio_note"] = f"Yahoo HTML fallback failed: {repr(e)}"

    return out


# FIX: retry_fetch — clean, no st.* calls, correct empty-dict guard
def retry_fetch(func, symbol: str, retries: int = 2) -> dict:
    last_result = {}
    for i in range(retries):
        try:
            result = func(symbol)
            if result and not _all_nan_ratio_dict(result):
                return result
            last_result = result or {}
        except Exception:
            pass
        time.sleep(1.5 * (i + 1) + random.random())
    return last_result


# FIX: get_live_peer_row — NO st.* calls inside, NO ThreadPoolExecutor
# Sequential fetches with polite delays — cache still works correctly
@st.cache_data(show_spinner=False, ttl=60 * 60 * 6)
def get_live_peer_row(
    symbol: str,
    fallback_company: str = "",
    fallback_country: str = "",
    fallback_exchange: str = "",
    fallback_sector: str = "",
    fallback_industry: str = ""
) -> dict:
    sym = normalize_peer_ticker(symbol)

    out = {
        "Company": fallback_company or sym,
        "Ticker": sym,
        "Exchange": fallback_exchange,
        "Country": fallback_country,
        "Sector": fallback_sector,
        "Industry": fallback_industry,
        "EV/EBITDA": np.nan,
        "P/B": np.nan,
        "P/E": np.nan,
        "Source": "",
        "RatioNote": "",
        "YahooProfile": make_yahoo_profile_url(sym),
        "YahooStats": make_yahoo_statistics_url(sym),
        "NeedsManualInvesting": False,
    }

    if not sym:
        return out

    # 🔥 Always fetch both
    fmp = retry_fetch(fmp_ratios, sym)
    yh = retry_fetch(yahoo_profile_and_metrics, sym)

    # Only fetch deeper Yahoo layers if needed
    ystats = {}
    yhtml = {}

    if _all_nan_ratio_dict(yh):
        ystats = retry_fetch(yahoo_stats_table_fallback, sym)

    if _all_nan_ratio_dict(yh) and _all_nan_ratio_dict(ystats):
        yhtml = retry_fetch(yahoo_html_ratio_fallback, sym)

    # Layer 2: Yahoo Stats table (only if layer 1 failed)
    ystats = {}
    if _all_nan_ratio_dict(yh):
        ystats = retry_fetch(yahoo_stats_table_fallback, sym)

    # Layer 3: Yahoo HTML regex (only if layers 1+2 failed)
    yhtml = {}
    if _all_nan_ratio_dict(yh) and _all_nan_ratio_dict(ystats):
        yhtml = retry_fetch(yahoo_html_ratio_fallback, sym)

    # Merge company info from Yahoo
    company = _clean_text(yh.get("Company")) or fallback_company or sym
    exchange = _clean_text(yh.get("Exchange")) or fallback_exchange
    country = _clean_text(yh.get("Country")) or fallback_country
    sector = _clean_text(yh.get("Sector")) or fallback_sector
    industry = _clean_text(yh.get("Industry")) or fallback_industry

    # Best available value wins

    pe = fmp.get("P/E", np.nan)
    if pd.isna(pe):
        pe = yh.get("P/E", np.nan)
    if pd.isna(pe):
        pe = ystats.get("P/E", np.nan)
    if pd.isna(pe):
        pe = yhtml.get("P/E", np.nan)

    pb = fmp.get("P/B", np.nan)
    if pd.isna(pb):
        pb = yh.get("P/B", np.nan)
    if pd.isna(pb):
        pb = ystats.get("P/B", np.nan)
    if pd.isna(pb):
        pb = yhtml.get("P/B", np.nan)

    ev_ebitda = fmp.get("EV/EBITDA", np.nan)
    if pd.isna(ev_ebitda):
        ev_ebitda = yh.get("EV/EBITDA", np.nan)
    if pd.isna(ev_ebitda):
        ev_ebitda = ystats.get("EV/EBITDA", np.nan)
    if pd.isna(ev_ebitda):
        ev_ebitda = yhtml.get("EV/EBITDA", np.nan)

    has_yahoo_ratio = not (pd.isna(pe) and pd.isna(pb) and pd.isna(ev_ebitda))

    # Layer 4: Investing.com (only if all Yahoo layers failed)
    inv = {}
    if not has_yahoo_ratio:
        inv = retry_fetch(investing_ratios, sym)
        pe = inv.get("P/E", pe)
        pb = inv.get("P/B", pb)
        ev_ebitda = inv.get("EV/EBITDA", ev_ebitda)

    # Source label
    yahoo_exists = bool(yh.get("quote_exists")) or bool(ystats.get("page_exists")) or bool(yhtml.get("page_exists"))
    has_fmp_ratio = not _all_nan_ratio_dict(fmp)
    has_yahoo_ratio = not (pd.isna(pe) and pd.isna(pb) and pd.isna(ev_ebitda))

    if has_fmp_ratio:
        source = "FMP API"
        ratio_note = fmp.get("ratio_note")

    elif has_yahoo_ratio:
        source = (
                ystats.get("ratio_source") or
                yh.get("ratio_source") or
                yhtml.get("ratio_source") or
                "Yahoo Finance"
        )
        ratio_note = (
                ystats.get("ratio_note") or
                yh.get("ratio_note") or
                yhtml.get("ratio_note") or
                "Yahoo ratios fetched."
        )
    elif inv.get("ratio_source"):
        source = inv["ratio_source"]
        ratio_note = inv.get("ratio_note", "Fetched from Investing.com.")
    else:
        source = ""
        ratio_note = (
            ystats.get("ratio_note") or yh.get("ratio_note") or yhtml.get("ratio_note") or
            "No usable ratios found from any source."
        )

    out.update({
        "Company": company,
        "Exchange": exchange,
        "Country": country,
        "Sector": sector,
        "Industry": industry,
        "EV/EBITDA": _clean_num(ev_ebitda),
        "P/B": _clean_num(pb),
        "P/E": _clean_num(pe),
        "Source": source,
        "RatioNote": ratio_note,
        "YahooProfile": make_yahoo_profile_url(sym),
        "YahooStats": make_yahoo_statistics_url(sym),
        "NeedsManualInvesting": False,
    })

    return out


def _is_yahoo_usable_status(x):
    if pd.isna(x):
        return False
    return str(x).lower() not in ["error", "failed", "no data", "not found", "invalid", ""]


@st.cache_data(show_spinner=False)
def get_precomputed_target_peers(
    target_symbol: str,
    target_company: str,
    preferred_sector: str,
    preferred_industry: str,
    preferred_keywords_str: str,
    max_peers: int = 9
) -> pd.DataFrame:
    # FIX: accept flat hashable args instead of dict — ensures cache keys are stable
    global TARGET_PEER_MATCHES_DF

    if "TARGET_PEER_MATCHES_DF" not in globals() or TARGET_PEER_MATCHES_DF is None or TARGET_PEER_MATCHES_DF.empty:
        return pd.DataFrame()

    target_key = _norm_text(target_symbol or target_company)
    df = TARGET_PEER_MATCHES_DF.copy()

    target_col = None
    for c in ["target_symbol", "target_company", "target", "company"]:
        if c in df.columns:
            target_col = c
            break
    if target_col is None:
        return pd.DataFrame()

    df["_target_norm"] = df[target_col].map(_norm_text)
    df = df[df["_target_norm"] == target_key].copy()

    if df.empty and "target_company" in TARGET_PEER_MATCHES_DF.columns:
        df = TARGET_PEER_MATCHES_DF.copy()
        df["_target_norm"] = df["target_company"].map(_norm_text)
        df = df[df["_target_norm"] == _norm_text(target_company)].copy()

    if df.empty:
        return pd.DataFrame()

    for col in ["ticker", "company", "country", "exchange", "sector", "industry",
                "sector_keywords", "match_priority", "yahoo_status"]:
        if col not in df.columns:
            df[col] = ""

    df["country_l"] = df["country"].map(lambda x: _clean_text(x).lower())
    df = df[df["country_l"] != "zimbabwe"].copy()
    df["ticker"] = df["ticker"].map(normalize_peer_ticker)
    df["ticker_l"] = df["ticker"].map(lambda x: _clean_text(x).lower())
    df = df[
        ~df["ticker_l"].str.endswith(".zw", na=False) &
        ~df["ticker_l"].str.endswith(".vx", na=False)
    ].copy()

    if "yahoo_status" in df.columns:
        df = df[df["yahoo_status"].map(_is_yahoo_usable_status)].copy()

    # Rebuild profile dict inside for scoring
    target_profile = {
        "target_symbol": target_symbol,
        "target_company": target_company,
        "preferred_sector": preferred_sector,
        "preferred_industry": preferred_industry,
        "preferred_peer_keywords": [k.strip() for k in preferred_keywords_str.split(",") if k.strip()],
    }

    df["SimilarityScore"] = df.apply(
        lambda r: strict_peer_score(r.to_dict(), target_profile), axis=1
    )
    df["match_priority_num"] = pd.to_numeric(df.get("match_priority", 0), errors="coerce").fillna(0)
    df = df.sort_values(
        by=["SimilarityScore", "match_priority_num", "company"],
        ascending=[False, False, True]
    ).drop_duplicates(subset=["ticker"]).reset_index(drop=True)

    return df.head(max_peers).copy()


# FIX: build_live_comps_from_target — sequential fetches, polite delays, progress bar
def build_live_comps_from_target(
    target_query: str,
    max_peers: int = 5,
    manual_sector_override: str = ""
):
    target_profile = get_target_profile(target_query, manual_sector_override)

    # FIX: pass flat args to cached get_precomputed_target_peers
    preferred_keywords_str = ",".join(target_profile.get("preferred_peer_keywords", []))
    strict_df = get_precomputed_target_peers(
        target_symbol=target_profile.get("target_symbol", ""),
        target_company=target_profile.get("target_company", ""),
        preferred_sector=target_profile.get("preferred_sector", ""),
        preferred_industry=target_profile.get("preferred_industry", ""),
        preferred_keywords_str=preferred_keywords_str,
        max_peers=max_peers + 2,
    )
    peer_source = "TARGET_PEER_MATCHES"

    if strict_df is None or strict_df.empty:
        strict_df = strict_universe_filter(target_profile, max_peers=max_peers + 2)
        peer_source = "Africa universe Excel"

    S["debug_strict_df_shape"] = strict_df.shape if strict_df is not None else (0, 0)
    S["debug_strict_df_preview"] = strict_df.head(20) if strict_df is not None and not strict_df.empty else pd.DataFrame()

    if strict_df is None or strict_df.empty:
        return pd.DataFrame(), {
            "error": f"No peers found for {target_profile.get('target_symbol') or target_query}.",
            "target": target_profile,
            "peer_source": peer_source,
        }

    rows = []
    peer_rows = list(strict_df.iterrows())
    total = len(peer_rows)

    # FIX: sequential fetch with progress bar and polite delay — no ThreadPoolExecutor
    progress_bar = st.progress(0, text="Fetching peer ratios...")

    for idx, (_, r) in enumerate(peer_rows):
        try:
            live = get_live_peer_row(
                symbol=r.get("ticker", ""),
                fallback_company=r.get("company", ""),
                fallback_country=r.get("country", ""),
                fallback_exchange=r.get("exchange", ""),
                fallback_sector=r.get("sector", ""),
                fallback_industry=r.get("industry", ""),
            )

            live["SimilarityScore"] = _clean_num(r.get("SimilarityScore"))
            if pd.isna(live["SimilarityScore"]):
                live["SimilarityScore"] = strict_peer_score(r.to_dict(), target_profile)

            live["UniverseSector"] = _clean_text(r.get("sector"))
            live["UniverseIndustry"] = _clean_text(r.get("industry"))
            live["UniverseKeywords"] = _clean_text(r.get("sector_keywords"))
            live["YahooStatus"] = _clean_text(r.get("yahoo_status"))
            live["MatchPriority"] = _clean_text(r.get("match_priority"))

            rows.append(live)

        except Exception:
            pass

        # Polite delay between requests — prevents IP bans on Streamlit Cloud
        if idx < total - 1:
            time.sleep(0.8 + random.random() * 0.5)

        progress_bar.progress(
            int((idx + 1) / total * 100),
            text=f"Fetching peer {idx + 1} of {total}..."
        )

    progress_bar.empty()

    df = pd.DataFrame(rows).drop_duplicates(subset=["Ticker"]).reset_index(drop=True)

    if df.empty:
        return pd.DataFrame(), {
            "error": f"Peers matched but live fetch returned no rows for {target_profile.get('target_symbol') or target_query}.",
            "target": target_profile,
            "peer_source": peer_source,
        }

    # Hard filter: never allow Zimbabwe peers through
    df["Country_l"] = df["Country"].map(lambda x: _clean_text(x).lower())
    df = df[df["Country_l"] != "zimbabwe"].copy()
    df["Ticker_l"] = df["Ticker"].map(lambda x: _clean_text(x).lower())
    df = df[
        ~df["Ticker_l"].str.endswith(".zw", na=False) &
        ~df["Ticker_l"].str.endswith(".vx", na=False)
    ].copy()
    df = df.drop(columns=["Country_l", "Ticker_l"], errors="ignore")

    df["RatioCount"] = (
        df["EV/EBITDA"].notna().astype(int) +
        df["P/B"].notna().astype(int) +
        df["P/E"].notna().astype(int)
    )

    df = df.sort_values(
        by=["SimilarityScore", "RatioCount", "Company"],
        ascending=[False, False, True]
    ).head(max_peers).reset_index(drop=True)

    meta = {
        "target": target_profile,
        "peer_source": peer_source,
        "peer_count": len(df),
        "target_sector": target_profile.get("preferred_sector", ""),
        "target_industry": target_profile.get("preferred_industry", ""),
    }

    return df, meta


def apply_live_comps_to_session(df_live: pd.DataFrame):
    if df_live is None or df_live.empty:
        return
    S.setdefault("comps", {})
    n = len(df_live)
    S["num_comps"] = n

    for i, (_, r) in enumerate(df_live.iterrows()):
        S[f"comp_name_{i}"] = _clean_text(r.get("Company")) or _clean_text(r.get("Ticker"))
        S[f"comp_ticker_{i}"] = _clean_text(r.get("Ticker"))
        S[f"comp_source_{i}"] = _clean_text(r.get("Source"))
        S[f"comp_profile_{i}"] = _clean_text(r.get("YahooProfile"))
        S[f"comp_ev_{i}"] = np.nan if pd.isna(r.get("EV/EBITDA")) else float(r["EV/EBITDA"])
        S[f"comp_pb_{i}"] = np.nan if pd.isna(r.get("P/B")) else float(r["P/B"])
        S[f"comp_pe_{i}"] = np.nan if pd.isna(r.get("P/E")) else float(r["P/E"])
        S[f"inc_ev_{i}"] = not pd.isna(r.get("EV/EBITDA"))
        S[f"inc_pb_{i}"] = not pd.isna(r.get("P/B"))
        S[f"inc_pe_{i}"] = not pd.isna(r.get("P/E"))
        S["comps"].setdefault(i, {})
        S["comps"][i].update({
            "name": S[f"comp_name_{i}"],
            "ticker": S[f"comp_ticker_{i}"],
            "source": S[f"comp_source_{i}"],
            "profile": S[f"comp_profile_{i}"],
            "ev": S[f"comp_ev_{i}"],
            "pb": S[f"comp_pb_{i}"],
            "pe": S[f"comp_pe_{i}"],
            "inc_ev": S[f"inc_ev_{i}"],
            "inc_pb": S[f"inc_pb_{i}"],
            "inc_pe": S[f"inc_pe_{i}"],
        })


def render_peer_picker_table(df_live: pd.DataFrame):
    if df_live is None or df_live.empty:
        return pd.DataFrame()

    S.setdefault("peer_picker_selected_map", {})
    S.setdefault("live_comps_df_selected", pd.DataFrame())

    st.markdown("""
    <style>
    .peer-picker-wrap {
        border: 1px solid #dbe4ee; border-radius: 16px;
        padding: 16px 16px 10px 16px; background: #f8fbff;
        box-shadow: 0 6px 18px rgba(15,23,42,0.05);
        margin-top: 10px; margin-bottom: 12px;
    }
    .peer-picker-head { font-size: 16px; font-weight: 700; color: #0f172a; margin-bottom: 10px; }
    .peer-picker-sub  { font-size: 12px; color: #475569; margin-bottom: 14px; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="peer-picker-wrap">
        <div class="peer-picker-head">Choose peers to send into Step 1</div>
        <div class="peer-picker-sub">Tick the peers you want to use for your comparables set.</div>
    </div>
    """, unsafe_allow_html=True)

    df_reset = df_live.reset_index(drop=True)
    selected_idx = []

    for i, r in df_reset.iterrows():
        ticker = _clean_text(r.get("Ticker"))
        company = _clean_text(r.get("Company"))
        country = _clean_text(r.get("Country"))
        sector = _clean_text(r.get("Sector"))

        ev_txt = "—" if pd.isna(r.get("EV/EBITDA")) else f"{float(r.get('EV/EBITDA')):.2f}"
        pb_txt = "—" if pd.isna(r.get("P/B")) else f"{float(r.get('P/B')):.2f}"
        pe_txt = "—" if pd.isna(r.get("P/E")) else f"{float(r.get('P/E')):.2f}"

        if ticker not in S["peer_picker_selected_map"]:
            S["peer_picker_selected_map"][ticker] = (i < min(6, len(df_reset)))

        widget_key = f"peer_pick_tbl_{ticker}"
        if widget_key not in S:
            S[widget_key] = bool(S["peer_picker_selected_map"][ticker])

        c1, c2, c3 = st.columns([0.7, 4.2, 2.1])
        with c1:
            st.checkbox("Use", key=widget_key, label_visibility="collapsed")
            S["peer_picker_selected_map"][ticker] = bool(S[widget_key])
        with c2:
            st.markdown(f"**{company}**  \n{ticker} | {country} | {sector}")
        with c3:
            st.markdown(f"""
            <div style="border-radius:12px;padding:12px 14px;background:#0a3554;
                        color:white;line-height:1.9;font-size:13px;font-weight:500;
                        border:2px solid #16d6d6;box-shadow:0 4px 12px rgba(0,0,0,0.18);">
                <div><b>EV/EBITDA:</b> {ev_txt}</div>
                <div><b>P/B:</b> {pb_txt}</div>
                <div><b>P/E:</b> {pe_txt}</div>
            </div>""", unsafe_allow_html=True)

        if S["peer_picker_selected_map"][ticker]:
            selected_idx.append(i)

    selected_df = df_reset.loc[selected_idx].copy() if selected_idx else pd.DataFrame(columns=df_live.columns)
    S["live_comps_df_selected"] = selected_df.copy()
    return selected_df


# =========================================================
# STEP 1 — INPUT COMPARABLE COMPANIES & MULTIPLES
# =========================================================
st.header("Step 1 — Input Comparable Companies & Multiples")
st.subheader("Auto Peer Suggestions from Strict Africa Universe Excel")

if UNIVERSE_FILE:
    st.caption(f"Universe file loaded: {UNIVERSE_FILE}")
else:
    st.error("❌ Africa peer universe Excel was not found. Put it inside a data folder, "
             "e.g. data/africa_yahoo_peer_universe_strict_final.xlsx")
    st.stop()

S.setdefault("target_company", "")
S.setdefault("auto_peer_count", 8)
S.setdefault("manual_sector_override", "")
S.setdefault("live_comps_df", pd.DataFrame())
S.setdefault("live_comps_meta", {})
S.setdefault("live_comps_df_selected", pd.DataFrame())
S.setdefault("peer_picker_selected_map", {})

cA, cB, cC = st.columns([2.2, 1, 1.2])
with cA:
    target_company = st.text_input(
        "Company you are valuing (Zimbabwe / VFEX / JSE / any ticker in your universe)",
        value=S["target_company"],
        key="target_company_input",
        placeholder="e.g. econet, padenga, cbz, delta, innscor ...",
    )
with cB:
    peer_count = st.number_input(
        "Peers to suggest",
        min_value=3, max_value=15,
        value=int(S["auto_peer_count"]),
        step=1, key="auto_peer_count_input",
    )
with cC:
    st.caption(" ")
    auto_apply = st.checkbox("Auto-fill Step 1 names", value=True, key="auto_apply_peers")

S["target_company"] = target_company
S["auto_peer_count"] = int(peer_count)

manual_sector = st.text_input(
    "Optional manual sector override",
    value=S["manual_sector_override"],
    key="manual_sector_override_input",
    placeholder="e.g. telecommunications, banking, mining, beverages",
)
S["manual_sector_override"] = manual_sector

st.subheader("Live peer search and ratio fill")
st.caption("Peers come from your Africa universe Excel first, then ratios are fetched "
           "from Yahoo Finance Statistics tab sequentially.")

live_peer_limit = st.slider(
    "Live peers to import",
    min_value=3, max_value=12,
    value=min(int(S["auto_peer_count"]), 12),
    step=1, key="live_peer_limit"
)

run_live_comps = st.button("⚡ Auto-search live peers and ratios")

if st.button("Clear ratio cache"):
    st.cache_data.clear()
    st.session_state["yahoo_warmed_up"] = False   # also reset warmup so it re-runs
    st.success("Cache cleared. Run live peer search again.")

if run_live_comps:
    if not target_company.strip():
        st.warning("Enter the company name or ticker first.")
    else:
        S["peer_picker_selected_map"] = {}
        S["live_comps_df_selected"] = pd.DataFrame()
        live_df, meta = build_live_comps_from_target(
            target_query=target_company,
            max_peers=int(live_peer_limit),
            manual_sector_override=manual_sector,
        )
        S["live_comps_df"] = live_df
        S["live_comps_meta"] = meta
        if live_df is not None and not live_df.empty:
            st.success(f"{len(live_df)} peers found. Select which ones to use below.")

live_df = S.get("live_comps_df", pd.DataFrame())
live_meta = S.get("live_comps_meta", {})

if live_meta:
    tgt = live_meta.get("target", {})
    st.caption(
        f"Resolved target: {tgt.get('target_company', '')} ({tgt.get('target_symbol', '')}) "
        f"via {tgt.get('source', '')} | Peer source: {live_meta.get('peer_source', '')} "
        f"| Manual sector override: {S.get('manual_sector_override', '') or 'None'}"
    )

if live_df is not None and not live_df.empty:
    df_show = live_df.copy()
    ratio_cols = ["EV/EBITDA", "P/B", "P/E"]
    for c in ratio_cols:
        df_show[c] = pd.to_numeric(df_show[c], errors="coerce")

    display_cols = [
        "YahooStats", "YahooProfile", "Company", "Ticker", "Exchange",
        "Country", "Sector", "Industry", "EV/EBITDA", "P/B", "P/E",
        "Source", "RatioNote",
    ]
    st.dataframe(
        df_show[display_cols],
        use_container_width=True,
        column_config={
            "YahooStats": st.column_config.LinkColumn("Stats Page", display_text="Open Stats"),
            "YahooProfile": st.column_config.LinkColumn("Profile Page", display_text="Open"),
            "EV/EBITDA": st.column_config.NumberColumn("EV/EBITDA", format="%.2f"),
            "P/B": st.column_config.NumberColumn("P/B", format="%.2f"),
            "P/E": st.column_config.NumberColumn("P/E", format="%.2f"),
        }
    )
    st.markdown("---")

    selected_live_df = render_peer_picker_table(df_show)

    pick_c1, pick_c2 = st.columns([1.3, 2])
    with pick_c1:
        use_selected_btn = st.button("✅ Use Selected Peers", key="use_selected_peers_btn")
    with pick_c2:
        saved_selected = S.get("live_comps_df_selected", pd.DataFrame())
        st.caption(f"Selected peers: {0 if saved_selected is None else len(saved_selected)}")

    if use_selected_btn:
        saved_selected = S.get("live_comps_df_selected", pd.DataFrame())
        if saved_selected is None or saved_selected.empty:
            st.warning("Please select at least one peer first.")
        else:
            apply_live_comps_to_session(saved_selected)
            st.success(f"{len(saved_selected)} selected peers were loaded into Step 1.")

    missing_all = df_show[ratio_cols].isna().all(axis=1)
    if missing_all.any():
        st.warning("Some peers were found, but ratio fields are still unavailable from live sources.")

    manual_rows = (
        df_show[df_show["NeedsManualInvesting"] == True].copy()
        if "NeedsManualInvesting" in df_show.columns else pd.DataFrame()
    )
    if not manual_rows.empty:
        st.subheader("Manual Investing.com ratio entry")
        for _, rr in manual_rows.iterrows():
            tkr = rr["Ticker"]
            company_name = rr["Company"]
            src_link = rr["YahooStats"]
            st.markdown(f"**{company_name} ({tkr})**")
            if src_link:
                st.markdown(f"[Open Investing page]({src_link})")
            c1, c2, c3 = st.columns(3)
            with c1:
                man_ev = st.number_input(f"{tkr} Manual EV/EBITDA", min_value=0.0,
                                         value=float(S.get(f"manual_ev_{tkr}", 0.0)),
                                         step=0.01, key=f"manual_ev_{tkr}")
            with c2:
                man_pb = st.number_input(f"{tkr} Manual P/B", min_value=0.0,
                                         value=float(S.get(f"manual_pb_{tkr}", 0.0)),
                                         step=0.01, key=f"manual_pb_{tkr}")
            with c3:
                man_pe = st.number_input(f"{tkr} Manual P/E", min_value=0.0,
                                         value=float(S.get(f"manual_pe_{tkr}", 0.0)),
                                         step=0.01, key=f"manual_pe_{tkr}")
            if man_ev > 0:
                live_df.loc[live_df["Ticker"] == tkr, "EV/EBITDA"] = man_ev
            if man_pb > 0:
                live_df.loc[live_df["Ticker"] == tkr, "P/B"] = man_pb
            if man_pe > 0:
                live_df.loc[live_df["Ticker"] == tkr, "P/E"] = man_pe
            if man_ev > 0 or man_pb > 0 or man_pe > 0:
                live_df.loc[live_df["Ticker"] == tkr, "Source"] = "Manual from Investing page"
                live_df.loc[live_df["Ticker"] == tkr, "RatioNote"] = "Manually entered from Investing page."
                live_df.loc[live_df["Ticker"] == tkr, "NeedsManualInvesting"] = False
        S["live_comps_df"] = live_df

with st.expander("Debug peer search"):
    st.write("Universe debug:", UNIVERSE_DEBUG)
    if live_meta:
        st.write("Target profile:", live_meta.get("target", {}))
    st.write("Strict df shape:", S.get("debug_strict_df_shape"))
    st.write("Filter stage counts:", S.get("debug_filter_stage_counts", {}))
    st.write("Strict df preview:", S.get("debug_strict_df_preview"))
    st.write("Candidate preview:", S.get("debug_strict_candidates_preview", pd.DataFrame()))
    if live_df is not None and not live_df.empty:
        st.write("Returned peer tickers:", live_df["Ticker"].tolist())
        debug_cols = ["Ticker", "Source", "RatioNote", "EV/EBITDA", "P/B", "P/E"]
        available_debug_cols = [c for c in debug_cols if c in live_df.columns]
        st.dataframe(live_df[available_debug_cols], use_container_width=True)
        st.write("Ratio notes by ticker:")
        for _, rr in live_df.iterrows():
            st.write(f"{rr.get('Ticker', '')}: {rr.get('RatioNote', '')}")


# =========================================================
# STEP 1 MANUAL COMPS INPUT
# =========================================================
S.setdefault("num_comps", 3)
S.setdefault("comps", {})

num_comps = st.number_input(
    "How many comparables?",
    min_value=1, max_value=20,
    value=int(S.get("num_comps", 3)),
    key="num_comps_input",
)
S["num_comps"] = int(num_comps)

for i in range(int(num_comps)):
    S["comps"].setdefault(i, {
        "name": f"Comp {i + 1}", "ev": np.nan, "pb": np.nan, "pe": np.nan,
        "inc_ev": True, "inc_pb": True, "inc_pe": True,
    })

rows = []
for i in range(int(num_comps)):
    st.subheader(f"Comparable {i + 1}")
    c1, c2, c3, c4, c5 = st.columns([2, 1, 1, 1, 1.2])

    with c1:
        default_name = S.get(f"comp_name_{i}", S["comps"][i]["name"])
        name = st.text_input(f"Company {i + 1} name", value=str(default_name), key=f"comp_name_{i}")
        S["comps"][i]["name"] = name
    with c2:
        default_ev = _num_input_default(S.get(f"comp_ev_{i}", S["comps"][i]["ev"]), 0.0)
        ev = st.number_input(f"{name} EV/EBITDA", value=default_ev, step=0.01,
                             format="%.2f", key=f"comp_ev_{i}")
        S["comps"][i]["ev"] = ev
    with c3:
        default_pb = _num_input_default(S.get(f"comp_pb_{i}", S["comps"][i]["pb"]), 0.0)
        pb = st.number_input(f"{name} P/B", value=default_pb, step=0.01,
                             format="%.2f", key=f"comp_pb_{i}")
        S["comps"][i]["pb"] = pb
    with c4:
        default_pe = _num_input_default(S.get(f"comp_pe_{i}", S["comps"][i]["pe"]), 0.0)
        pe = st.number_input(f"{name} P/E", value=default_pe, step=0.01,
                             format="%.2f", key=f"comp_pe_{i}")
        S["comps"][i]["pe"] = pe

    ev_key = f"inc_ev_{i}"
    pb_key = f"inc_pb_{i}"
    pe_key = f"inc_pe_{i}"
    if ev_key not in S:
        S[ev_key] = bool(S["comps"][i].get("inc_ev", True))
    if pb_key not in S:
        S[pb_key] = bool(S["comps"][i].get("inc_pb", True))
    if pe_key not in S:
        S[pe_key] = bool(S["comps"][i].get("inc_pe", True))

    with c5:
        st.caption("Analyst filter")
        st.checkbox("Include EV/EBITDA", key=ev_key)
        st.checkbox("Include P/B", key=pb_key)
        st.checkbox("Include P/E", key=pe_key)

    ticker_val = S.get(f"comp_ticker_{i}", "")
    source_val = S.get(f"comp_source_{i}", "")
    profile_val = S.get(f"comp_profile_{i}", "")
    if ticker_val or source_val:
        st.caption(f"Ticker: {ticker_val} | Ratio source: {source_val}")
        if profile_val:
            st.markdown(f"[Open Yahoo profile for {name}]({profile_val})")

    inc_ev = bool(S[ev_key])
    inc_pb = bool(S[pb_key])
    inc_pe = bool(S[pe_key])
    S["comps"][i]["inc_ev"] = inc_ev
    S["comps"][i]["inc_pb"] = inc_pb
    S["comps"][i]["inc_pe"] = inc_pe

    rows.append({
        "Company": name,
        "EV/EBITDA": ev, "P/B": pb, "P/E": pe,
        "Include_EV": inc_ev, "Include_PB": inc_pb, "Include_PE": inc_pe,
    })

df_comps = pd.DataFrame(
    rows,
    columns=["Company", "EV/EBITDA", "P/B", "P/E", "Include_EV", "Include_PB", "Include_PE"]
)

st.subheader("Entered Comparables")
st.dataframe(df_comps, use_container_width=True)

S["comps_num"] = int(num_comps)
S["comps_ev_list"] = df_comps["EV/EBITDA"].astype(float).tolist()
S["comps_pb_list"] = df_comps["P/B"].astype(float).tolist()
S["comps_pe_list"] = df_comps["P/E"].astype(float).tolist()
S["comps_inc_ev"] = df_comps["Include_EV"].astype(bool).tolist()
S["comps_inc_pb"] = df_comps["Include_PB"].astype(bool).tolist()
S["comps_inc_pe"] = df_comps["Include_PE"].astype(bool).tolist()


# =========================================================
# STEP 2 — AVERAGE & IMPLIED MULTIPLES
# =========================================================
st.header("Step 2 — Average & Implied Multiples")

ev_series = df_comps.loc[df_comps["Include_EV"] == True, "EV/EBITDA"]
pb_series = df_comps.loc[df_comps["Include_PB"] == True, "P/B"]
pe_series = df_comps.loc[df_comps["Include_PE"] == True, "P/E"]

avg_ev = filtered_average(ev_series)
avg_pb = filtered_average(pb_series)
avg_pe = filtered_average(pe_series)

if "discount_factor" not in st.session_state:
    st.session_state["discount_factor"] = 25.0

discount_pct = st.number_input(
    "Discount factor (%)",
    step=1.0,
    value=st.session_state["discount_factor"],
    key="discount_factor_widget",
)
st.session_state["discount_factor"] = discount_pct

discount = discount_pct / 100
implied_ev = avg_ev * (1 - discount)
implied_pb = avg_pb * (1 - discount)
implied_pe = avg_pe * (1 - discount)

st.dataframe(
    pd.DataFrame({
        "Multiple": ["EV/EBITDA", "P/B", "P/E"],
        "Average": [avg_ev, avg_pb, avg_pe],
        "Discount (%)": [discount_pct] * 3,
        "Implied": [implied_ev, implied_pb, implied_pe],
    }).style.format({"Average": "{:,.2f}", "Implied": "{:,.2f}"}),
    use_container_width=True,
)

st.session_state["implied_ev"] = float(implied_ev) if not pd.isna(implied_ev) else 0.0
st.session_state["implied_pb"] = float(implied_pb) if not pd.isna(implied_pb) else 0.0
st.session_state["implied_pe"] = float(implied_pe) if not pd.isna(implied_pe) else 0.0


# =========================================================
# TIMING SOURCE (from DCF)
# =========================================================
st.header("Timing Source (from DCF)")

dcf_timing_list = S.get("dcf_discount_periods_n", [])
default_base = float(S.get("comp_timing_base", 0.0))

if not dcf_timing_list:
    st.warning("⚠ No timing values detected from DCF. "
               "Either run the DCF model first or set a manual timing base.")
    base_timing = st.number_input(
        "Enter starting timing value for comparables (year 1):",
        value=default_base, step=0.01, format="%.4f",
        key="comp_timing_base_manual_no_dcf",
    )
else:
    timing_df = pd.DataFrame({
        "Forecast Year Index": list(range(len(dcf_timing_list))),
        "DCF Timing n": dcf_timing_list,
    })
    st.dataframe(timing_df, use_container_width=True)
    dcf_n0 = float(round(dcf_timing_list[0], 4))
    st.info(f"DCF First Timing Value (n₀) = **{dcf_n0} years**")

    timing_choice = st.radio(
        "Choose timing base for Comparables timing effect:",
        [f"Use DCF n₀ = {dcf_n0}", "Manually override starting timing value"],
        index=0 if default_base == 0.0 or np.isclose(default_base, dcf_n0) else 1,
        key="comp_timing_choice",
    )
    if timing_choice.startswith("Use DCF"):
        base_timing = dcf_n0
    else:
        base_timing = st.number_input(
            "Enter starting timing value for comparables (year 1):",
            value=default_base if default_base != 0.0 else dcf_n0,
            step=0.01, format="%.4f", key="comp_timing_base_manual",
        )

S["comp_timing_base"] = float(base_timing)
st.success(f"Timing base for comparables = **{base_timing:.4f}**")


# =========================================================
# STEP 3 — MAINTAINABLE EBITDA
# =========================================================
st.header("Step 3 — Maintainable EBITDA")

dcf_eb_all = S.get("dcf_ebitda_all", None)
if dcf_eb_all is None:
    dcf_eb_all = S.get("dcf_ebitda_forecast", {})

if not dcf_eb_all:
    st.warning("⚠ No EBITDA found from DCF — skipping EV/EBITDA method.")
    S["maintainable_ebitda"] = np.nan
else:
    eb_years_all = sorted(
        int(y) for y in dcf_eb_all.keys()
        if str(y).strip().isdigit() and len(str(y).strip()) == 4
    )
    if not eb_years_all:
        st.warning("⚠ DCF EBITDA found, but no valid 4-digit year keys — skipping EV/EBITDA method.")
        S["maintainable_ebitda"] = np.nan
    else:
        eb_min_year = min(eb_years_all)
        eb_max_year = max(eb_years_all)
        S.setdefault("comp_eb_start_year", eb_min_year)
        S.setdefault("comp_eb_end_year", eb_max_year)
        S.setdefault("comp_eb_weights", {})
        S.setdefault("comp_use_timing_eb", True)

        use_timing_eb = st.checkbox(
            "Apply timing effect from DCF to EBITDA?",
            value=bool(S.get("comp_use_timing_eb", True)),
            key="comp_use_timing_eb_checkbox",
        )
        S["comp_use_timing_eb"] = use_timing_eb

        prev_eb = S.get("_prev_comp_use_timing_eb", None)
        if prev_eb is None or bool(prev_eb) != bool(use_timing_eb):
            S["comp_use_timing_np"] = bool(use_timing_eb)
            S["comp_use_timing_np_checkbox"] = bool(use_timing_eb)
        S["_prev_comp_use_timing_eb"] = bool(use_timing_eb)

        if not use_timing_eb:
            S["comp_use_timing_np"] = False
            S["comp_use_timing_np_checkbox"] = False

        c_eb1, c_eb2 = st.columns(2)
        with c_eb1:
            eb_start_year = st.number_input("EBITDA Start Year", value=int(S["comp_eb_start_year"]),
                                            step=1, key="comp_eb_start_year_input")
        with c_eb2:
            eb_end_year = st.number_input("EBITDA End Year", value=int(S["comp_eb_end_year"]),
                                          step=1, key="comp_eb_end_year_input")

        eb_start_year = int(max(eb_start_year, eb_min_year))
        eb_end_year = int(min(eb_end_year, eb_max_year))
        if eb_end_year < eb_start_year:
            st.error("❌ EBITDA End Year must be ≥ Start Year.")
            st.stop()

        S["comp_eb_start_year"] = eb_start_year
        S["comp_eb_end_year"] = eb_end_year
        selected_eb_years = list(range(eb_start_year, eb_end_year + 1))

        st.subheader("EBITDA Weighting")
        rows_eb = []
        base_timing = float(S.get("comp_timing_base", 0.0))

        for idx, yr in enumerate(selected_eb_years):
            eb_val = float(dcf_eb_all.get(str(yr), 0.0))
            default_w = float(S["comp_eb_weights"].get(str(yr), 0.0))
            timing_val = base_timing + idx if use_timing_eb else 1.0

            c1, c2, c4 = st.columns([1, 2, 1])
            with c1:
                st.number_input(f"EB Year {yr}", value=int(yr), disabled=True, key=f"comp_eb_year_display_{yr}")
            with c2:
                st.number_input(f"EBITDA {yr}", value=eb_val, disabled=True,
                                format="%.2f", key=f"comp_eb_value_display_{yr}")
            with c4:
                weight_val = st.number_input(f"EB Weight {yr} (%)", value=float(default_w),
                                             step=0.1, format="%.2f", key=f"comp_eb_weight_{yr}")

            S["comp_eb_weights"][str(yr)] = float(weight_val)
            adj_eb = eb_val * timing_val
            weighted_eb = adj_eb * weight_val / 100.0
            rows_eb.append({
                "Year": int(yr), "EBITDA": eb_val,
                "Timing": timing_val if use_timing_eb else np.nan,
                "Weight (%)": weight_val, "Adjusted EBITDA": adj_eb, "Weighted EBITDA": weighted_eb,
            })

        df_eb = pd.DataFrame(rows_eb)
        if use_timing_eb:
            df_eb_display = df_eb[["Year", "EBITDA", "Timing", "Weight (%)", "Adjusted EBITDA", "Weighted EBITDA"]]
        else:
            df_eb_display = df_eb[["Year", "EBITDA", "Weight (%)", "Weighted EBITDA"]]
        df_eb_display = df_eb_display.copy()
        df_eb_display.index = df_eb_display.index + 1
        st.dataframe(format_numeric_columns(df_eb_display), use_container_width=True)

        maintainable_ebitda = float(df_eb["Weighted EBITDA"].sum())
        st.success(f"Maintainable EBITDA = {maintainable_ebitda:,.2f}")
        S["maintainable_ebitda"] = maintainable_ebitda


# =========================================================
# STEP 4 — MAINTAINABLE EARNINGS
# =========================================================
st.header("Step 4 — Maintainable Earnings")

dcf_np_all = S.get("dcf_profit_all", None)
if dcf_np_all is None:
    dcf_np_all = S.get("dcf_profit_forecast", {})

if not dcf_np_all:
    st.warning("⚠ No Earnings found from DCF — skipping P/E method.")
    S["maintainable_earnings"] = np.nan
else:
    np_years_all = sorted(
        int(y) for y in dcf_np_all.keys()
        if str(y).strip().isdigit() and len(str(y).strip()) == 4
    )
    if not np_years_all:
        st.warning("⚠ DCF Earnings found, but no valid 4-digit year keys — skipping P/E method.")
        S["maintainable_earnings"] = np.nan
    else:
        np_min_year = min(np_years_all)
        np_max_year = max(np_years_all)
        S.setdefault("comp_np_start_year", np_min_year)
        S.setdefault("comp_np_end_year", np_max_year)
        S.setdefault("comp_np_weights", {})
        S.setdefault("comp_use_timing_np", True)
        S.setdefault("comp_sync_np_to_eb", True)

        sync_np_to_eb = st.checkbox(
            "Auto-use the SAME years & weights as EBITDA (recommended)",
            value=bool(S.get("comp_sync_np_to_eb", True)),
            key="comp_sync_np_to_eb_checkbox",
        )
        S["comp_sync_np_to_eb"] = bool(sync_np_to_eb)

        if sync_np_to_eb:
            eb_sy = int(max(int(S.get("comp_eb_start_year", np_min_year)), np_min_year))
            eb_ey = int(min(int(S.get("comp_eb_end_year", np_max_year)), np_max_year))
            S["comp_np_start_year"] = eb_sy
            S["comp_np_end_year"] = eb_ey
            eb_w = S.get("comp_eb_weights", {}) or {}
            S["comp_np_weights"] = {str(y): float(eb_w.get(str(y), 0.0)) for y in range(eb_sy, eb_ey + 1)}
            for y in range(eb_sy, eb_ey + 1):
                S[f"comp_np_weight_{y}"] = float(S["comp_np_weights"].get(str(y), 0.0))
            st.info("✅ Earnings years & weights copied from EBITDA automatically.")

        use_timing_eb = bool(S.get("comp_use_timing_eb", True))
        if not use_timing_eb:
            S["comp_use_timing_np"] = False
            S["comp_use_timing_np_checkbox"] = False

        use_timing_np = st.checkbox(
            "Apply timing effect from DCF to Earnings?",
            value=bool(S.get("comp_use_timing_np", True)),
            key="comp_use_timing_np_checkbox",
            disabled=(not use_timing_eb),
        )
        S["comp_use_timing_np"] = bool(use_timing_np)

        if sync_np_to_eb:
            np_start_year = int(S["comp_np_start_year"])
            np_end_year = int(S["comp_np_end_year"])
            c_np1, c_np2 = st.columns(2)
            with c_np1:
                st.number_input("NP Start Year (auto from EBITDA)", value=np_start_year,
                                disabled=True, key="np_start_locked")
            with c_np2:
                st.number_input("NP End Year (auto from EBITDA)", value=np_end_year,
                                disabled=True, key="np_end_locked")
        else:
            c_np1, c_np2 = st.columns(2)
            with c_np1:
                np_start_year = st.number_input("NP Start Year",
                                                value=int(S.get("comp_np_start_year", np_min_year)),
                                                step=1, key="comp_np_start_year_input")
            with c_np2:
                np_end_year = st.number_input("NP End Year",
                                              value=int(S.get("comp_np_end_year", np_max_year)),
                                              step=1, key="comp_np_end_year_input")
            np_start_year = int(max(np_start_year, np_min_year))
            np_end_year = int(min(np_end_year, np_max_year))
            if np_end_year < np_start_year:
                st.error("❌ NP End Year cannot be before Start Year.")
                st.stop()
            S["comp_np_start_year"] = np_start_year
            S["comp_np_end_year"] = np_end_year

        selected_np_years = list(range(np_start_year, np_end_year + 1))
        st.subheader("Earnings Weighting")
        rows_np = []
        base_timing = float(S.get("comp_timing_base", 0.0))

        for idx, yr in enumerate(selected_np_years):
            np_val = float(dcf_np_all.get(str(yr), 0.0))
            default_w = float(S.get(f"comp_np_weight_{yr}", S["comp_np_weights"].get(str(yr), 0.0)))
            timing_val = base_timing + idx if use_timing_np else 1.0

            c1, c2, c4 = st.columns([1, 2, 1])
            with c1:
                st.number_input(f"Earnings Year {yr}", value=int(yr), disabled=True,
                                key=f"comp_np_year_display_{yr}")
            with c2:
                st.number_input(f"Earnings {yr}", value=np_val, disabled=True,
                                format="%.2f", key=f"comp_np_value_display_{yr}")
            with c4:
                weight_val = st.number_input(f"NP Weight {yr} (%)", value=float(default_w),
                                             step=0.1, format="%.2f", key=f"comp_np_weight_{yr}")

            S["comp_np_weights"][str(yr)] = float(weight_val)
            adj_np = np_val * timing_val
            weighted_np = adj_np * weight_val / 100.0
            rows_np.append({
                "Year": int(yr), "Earnings": np_val,
                "Timing": timing_val if use_timing_np else np.nan,
                "Weight (%)": weight_val, "Adjusted Earnings": adj_np, "Weighted Earnings": weighted_np,
            })

        df_np = pd.DataFrame(rows_np)
        if use_timing_np:
            df_np_display = df_np[["Year", "Earnings", "Timing", "Weight (%)", "Adjusted Earnings", "Weighted Earnings"]]
        else:
            df_np_display = df_np[["Year", "Earnings", "Weight (%)", "Weighted Earnings"]]
        df_np_display = df_np_display.copy()
        df_np_display.index = df_np_display.index + 1
        st.dataframe(format_numeric_columns(df_np_display), use_container_width=True)

        maintainable_earnings = float(df_np["Weighted Earnings"].sum())
        st.success(f"Maintainable Earnings = {maintainable_earnings:,.2f}")
        S["maintainable_earnings"] = maintainable_earnings


# =========================================================
# STEP 5 — BOOK VALUE & NET DEBT
# =========================================================
st.header("Step 5 — Book Value & Net Debt")

bank_outputs = (S.get("bank", {}) or {}).get("outputs", {}) or {}
bank_book_equity = bank_outputs.get("book_equity_0", None)
if bank_book_equity is not None:
    if "book_equity_input" not in S:
        S["book_equity"] = float(bank_book_equity)
        S["book_equity_input"] = float(bank_book_equity)

book_equity_default = float(S.get("book_equity", 0.0))
net_debt_default = float(S.get("net_debt", 0.0))

book_equity = st.number_input("Book Equity (USD)", value=book_equity_default,
                               step=1000.0, format="%.2f", key="book_equity_input")
S["book_equity"] = float(book_equity)
st.caption(f"💰 Book Equity: **{book_equity:,.2f} USD**")

net_debt = st.number_input("Net Debt (USD)", value=net_debt_default,
                            step=1000.0, format="%.2f", key="net_debt_input")
S["net_debt"] = float(net_debt)
st.caption(f"💳 Net Debt: **{net_debt:,.2f} USD**")


# =========================================================
# STEP 6 — FINAL EQUITY VALUES
# =========================================================
st.header("Step 6 — Computed Equity Values")

maintainable_ebitda = S.get("maintainable_ebitda", np.nan)
maintainable_earnings = S.get("maintainable_earnings", np.nan)

equity_ev = np.nan
equity_pb = np.nan
equity_pe = np.nan

if maintainable_ebitda is not None and np.isfinite(float(maintainable_ebitda)) and not pd.isna(implied_ev):
    equity_ev = implied_ev * float(maintainable_ebitda) - net_debt
if book_equity is not None and np.isfinite(float(book_equity)) and not pd.isna(implied_pb):
    equity_pb = implied_pb * float(book_equity)
if maintainable_earnings is not None and np.isfinite(float(maintainable_earnings)) and not pd.isna(implied_pe):
    equity_pe = implied_pe * float(maintainable_earnings)

S["value_ev_ebitda"] = float(equity_ev) if not pd.isna(equity_ev) else np.nan
S["value_pbv"] = float(equity_pb) if not pd.isna(equity_pb) else np.nan
S["value_pe"] = float(equity_pe) if not pd.isna(equity_pe) else np.nan

df_res = pd.DataFrame({
    "Method": ["EV/EBITDA", "P/B", "P/E"],
    "Equity Value (USD)": [equity_ev, equity_pb, equity_pe],
})
st.dataframe(format_numeric_columns(df_res), use_container_width=True)


# =========================================================
# DOWNLOAD EXCEL
# =========================================================
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _style_range(ws, cell_range, bold=False, fill=None, align_center=False, border=True, font_color=None):
    thin = Side(style="thin", color="000000")
    b = Border(left=thin, right=thin, top=thin, bottom=thin) if border else None
    for row in ws[cell_range]:
        for c in row:
            if bold:
                c.font = Font(bold=True, color=font_color or "000000")
            if fill is not None:
                c.fill = fill
            if align_center:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if b is not None:
                c.border = b


def build_comps_excel_with_formulas(S, df_comps) -> bytes:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0A1B33")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    money_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    mult_fmt = '0.00'

    # Sheet 1: Comps_Input
    ws1 = wb.active
    ws1.title = "Comps_Input"
    ws1["B1"] = "Comparable Company"
    ws1["B1"].font = title_font

    headers = ["Company", "Country", "EV/EBITDA", "P/B", "P/E", "Include_EV", "Include_PB", "Include_PE"]
    start_row = 3
    start_col = 2

    for j, h in enumerate(headers, start=start_col):
        c = ws1.cell(row=start_row, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    r = start_row + 1
    for _, row in df_comps.iterrows():
        ws1.cell(r, start_col + 0, row["Company"])
        ws1.cell(r, start_col + 1, "")
        ws1.cell(r, start_col + 2, float(row["EV/EBITDA"]) if pd.notna(row["EV/EBITDA"]) else None)
        ws1.cell(r, start_col + 3, float(row["P/B"]) if pd.notna(row["P/B"]) else None)
        ws1.cell(r, start_col + 4, float(row["P/E"]) if pd.notna(row["P/E"]) else None)
        ws1.cell(r, start_col + 5, bool(row["Include_EV"]))
        ws1.cell(r, start_col + 6, bool(row["Include_PB"]))
        ws1.cell(r, start_col + 7, bool(row["Include_PE"]))
        ws1.cell(r, start_col + 2).number_format = mult_fmt
        ws1.cell(r, start_col + 3).number_format = mult_fmt
        ws1.cell(r, start_col + 4).number_format = mult_fmt
        r += 1

    end_row = r - 1
    _style_range(ws1, f"B{start_row}:I{end_row}", border=True)
    for col, w in zip(["B", "C", "D", "E", "F", "G", "H", "I"], [30, 16, 12, 10, 10, 12, 12, 12]):
        ws1.column_dimensions[col].width = w

    # Sheet 2: Multiples
    ws2 = wb.create_sheet("Multiples")
    ws2["B1"] = "Multiples Summary"
    ws2["B1"].font = title_font
    ws2["B3"] = "Discount (%)"
    ws2["C3"] = float(S.get("discount_pct", 25.0)) / 100.0
    ws2["C3"].number_format = pct_fmt
    ws2["B3"].font = bold_font

    for j, h in enumerate(["Multiple", "Average", "Implied"], start=2):
        c = ws2.cell(row=5, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    drow1 = start_row + 1
    drow2 = end_row
    ws2["B6"] = "EV/EBITDA"
    ws2["C6"] = f'=IFERROR(AVERAGEIF(Comps_Input!$G${drow1}:$G${drow2},TRUE,Comps_Input!$D${drow1}:$D${drow2}),"")'
    ws2["D6"] = f'=IF(C6="","",C6*(1-$C$3))'
    ws2["B7"] = "P/B"
    ws2["C7"] = f'=IFERROR(AVERAGEIF(Comps_Input!$H${drow1}:$H${drow2},TRUE,Comps_Input!$E${drow1}:$E${drow2}),"")'
    ws2["D7"] = f'=IF(C7="","",C7*(1-$C$3))'
    ws2["B8"] = "P/E"
    ws2["C8"] = f'=IFERROR(AVERAGEIF(Comps_Input!$I${drow1}:$I${drow2},TRUE,Comps_Input!$F${drow1}:$F${drow2}),"")'
    ws2["D8"] = f'=IF(C8="","",C8*(1-$C$3))'

    for rr in [6, 7, 8]:
        ws2[f"C{rr}"].number_format = mult_fmt
        ws2[f"D{rr}"].number_format = mult_fmt
    _style_range(ws2, "B5:D8", border=True)
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14

    # Sheet 3: EBITDA_Maintainable
    ws3 = wb.create_sheet("EBITDA_Maintainable")
    ws3["B1"] = "Maintainable EBITDA (with timing + weights)"
    ws3["B1"].font = title_font
    ws3["B3"] = "Use Timing?"
    ws3["C3"] = bool(S.get("comp_use_timing_eb", True))
    ws3["B4"] = "Base Timing"
    ws3["C4"] = float(S.get("comp_timing_base", 1.0))
    ws3["B3"].font = bold_font
    ws3["B4"].font = bold_font

    dcf_eb_all = S.get("dcf_ebitda_all", None) or S.get("dcf_ebitda_forecast", {}) or {}
    eb_sy = int(S.get("comp_eb_start_year", 0) or 0)
    eb_ey = int(S.get("comp_eb_end_year", 0) or 0)
    eb_years = list(range(eb_sy, eb_ey + 1)) if eb_sy and eb_ey and eb_ey >= eb_sy else []

    for j, h in enumerate(["Year", "EBITDA", "Timing", "Weight (%)", "Adjusted EBITDA", "Weighted EBITDA"], start=2):
        c = ws3.cell(row=6, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    r0 = 7
    for idx, yr in enumerate(eb_years):
        ws3.cell(r0 + idx, 2, yr)
        ws3.cell(r0 + idx, 3, float(dcf_eb_all.get(str(yr), 0.0)))
        ws3.cell(r0 + idx, 4, f'=IF($C$3,$C$4+{idx},1)')
        w = float((S.get("comp_eb_weights", {}) or {}).get(str(yr), 0.0))
        ws3.cell(r0 + idx, 5, w / 100.0)
        ws3.cell(r0 + idx, 6, f"=C{r0 + idx}*D{r0 + idx}")
        ws3.cell(r0 + idx, 7, f"=F{r0 + idx}*E{r0 + idx}")
        ws3.cell(r0 + idx, 3).number_format = money_fmt
        ws3.cell(r0 + idx, 4).number_format = '0.0000'
        ws3.cell(r0 + idx, 5).number_format = pct_fmt
        ws3.cell(r0 + idx, 6).number_format = money_fmt
        ws3.cell(r0 + idx, 7).number_format = money_fmt

    last3 = r0 + len(eb_years) - 1 if eb_years else 7
    ws3["B" + str(last3 + 2)] = "Maintainable EBITDA"
    ws3["B" + str(last3 + 2)].font = bold_font
    ws3["G" + str(last3 + 2)] = f"=SUM(G{r0}:G{last3})"
    ws3["G" + str(last3 + 2)].font = bold_font
    ws3["G" + str(last3 + 2)].number_format = money_fmt
    _style_range(ws3, f"B6:G{last3}", border=True)
    for col, w in zip(["B", "C", "D", "E", "F", "G"], [10, 18, 12, 12, 18, 18]):
        ws3.column_dimensions[col].width = w

    # Sheet 4: Earnings_Maintainable
    ws4 = wb.create_sheet("Earnings_Maintainable")
    ws4["B1"] = "Maintainable Earnings (with timing + weights)"
    ws4["B1"].font = title_font
    ws4["B3"] = "Use Timing?"
    ws4["C3"] = bool(S.get("comp_use_timing_np", True))
    ws4["B4"] = "Base Timing"
    ws4["C4"] = float(S.get("comp_timing_base", 1.0))
    ws4["B3"].font = bold_font
    ws4["B4"].font = bold_font

    dcf_np_all = S.get("dcf_profit_all", None) or S.get("dcf_profit_forecast", {}) or {}
    np_sy = int(S.get("comp_np_start_year", 0) or 0)
    np_ey = int(S.get("comp_np_end_year", 0) or 0)
    np_years = list(range(np_sy, np_ey + 1)) if np_sy and np_ey and np_ey >= np_sy else []

    for j, h in enumerate(["Year", "Earnings", "Timing", "Weight (%)", "Adjusted Earnings", "Weighted Earnings"], start=2):
        c = ws4.cell(row=6, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    r0 = 7
    for idx, yr in enumerate(np_years):
        ws4.cell(r0 + idx, 2, yr)
        ws4.cell(r0 + idx, 3, float(dcf_np_all.get(str(yr), 0.0)))
        ws4.cell(r0 + idx, 4, f'=IF($C$3,$C$4+{idx},1)')
        w = float((S.get("comp_np_weights", {}) or {}).get(str(yr), 0.0))
        ws4.cell(r0 + idx, 5, w / 100.0)
        ws4.cell(r0 + idx, 6, f"=C{r0 + idx}*D{r0 + idx}")
        ws4.cell(r0 + idx, 7, f"=F{r0 + idx}*E{r0 + idx}")
        ws4.cell(r0 + idx, 3).number_format = money_fmt
        ws4.cell(r0 + idx, 4).number_format = '0.0000'
        ws4.cell(r0 + idx, 5).number_format = pct_fmt
        ws4.cell(r0 + idx, 6).number_format = money_fmt
        ws4.cell(r0 + idx, 7).number_format = money_fmt

    last4 = r0 + len(np_years) - 1 if np_years else 7
    ws4["B" + str(last4 + 2)] = "Maintainable Earnings"
    ws4["B" + str(last4 + 2)].font = bold_font
    ws4["G" + str(last4 + 2)] = f"=SUM(G{r0}:G{last4})"
    ws4["G" + str(last4 + 2)].font = bold_font
    ws4["G" + str(last4 + 2)].number_format = money_fmt
    _style_range(ws4, f"B6:G{last4}", border=True)
    for col, w in zip(["B", "C", "D", "E", "F", "G"], [10, 18, 12, 12, 18, 18]):
        ws4.column_dimensions[col].width = w

    # Sheet 5: Equity_Values
    ws5 = wb.create_sheet("Equity_Values")
    ws5["B1"] = "Computed Equity Values"
    ws5["B1"].font = title_font
    ws5["B3"] = "Book Equity"
    ws5["C3"] = float(S.get("book_equity", 0.0))
    ws5["B4"] = "Net Debt"
    ws5["C4"] = float(S.get("net_debt", 0.0))
    ws5["B3"].font = bold_font
    ws5["B4"].font = bold_font
    ws5["C3"].number_format = money_fmt
    ws5["C4"].number_format = money_fmt
    ws5["B6"] = "Maintainable EBITDA"
    ws5["C6"] = f"=EBITDA_Maintainable!G{last3 + 2}"
    ws5["B7"] = "Maintainable Earnings"
    ws5["C7"] = f"=Earnings_Maintainable!G{last4 + 2}"
    ws5["C6"].number_format = money_fmt
    ws5["C7"].number_format = money_fmt
    ws5["B6"].font = bold_font
    ws5["B7"].font = bold_font

    for j, h in enumerate(["Method", "Equity Value (USD)"], start=2):
        c = ws5.cell(row=9, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws5["B10"] = "EV/EBITDA"
    ws5["C10"] = '=IF(Multiples!D6="","",Multiples!D6*$C$6-$C$4)'
    ws5["B11"] = "P/B"
    ws5["C11"] = '=IF(Multiples!D7="","",Multiples!D7*$C$3)'
    ws5["B12"] = "P/E"
    ws5["C12"] = '=IF(Multiples!D8="","",Multiples!D8*$C$7)'

    for rr in [10, 11, 12]:
        ws5[f"C{rr}"].number_format = money_fmt
    _style_range(ws5, "B9:C12", border=True)
    ws5.column_dimensions["B"].width = 18
    ws5.column_dimensions["C"].width = 22

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


excel_bytes = build_comps_excel_with_formulas(S, df_comps)

st.download_button(
    label="⬇️ Download Comparables (Excel with formulas)",
    data=excel_bytes,
    file_name="comparables_with_formulas.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
