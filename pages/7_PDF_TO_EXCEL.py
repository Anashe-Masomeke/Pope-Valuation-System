import io, os, re, gc, traceback
import streamlit as st

st.set_page_config(page_title="PDF / Image → Excel | FBC", layout="wide")

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = True
if "user" not in st.session_state or not st.session_state.get("user"):
    st.session_state["user"] = {"username":"analyst","role":"analyst","full_name":"Analyst"}

try:
    import pytesseract
    from PIL import Image, ImageFilter, ImageEnhance
    LIBS_OK = True; _IMPORT_ERR = ""
except ImportError as _e:
    LIBS_OK = False; _IMPORT_ERR = str(_e)

PDF_LIBS_OK = False
if LIBS_OK:
    try:
        from pdf2image import convert_from_bytes
        PDF_LIBS_OK = True
    except ImportError:
        pass

try:
    import pdfplumber
    PDFPLUMBER_OK = True
except ImportError:
    PDFPLUMBER_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

_tess_env = os.environ.get("TESSERACT_CMD","")
if LIBS_OK:
    pytesseract.pytesseract.tesseract_cmd = _tess_env if _tess_env else "tesseract"
# ══════════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700;900&family=EB+Garamond:ital,wght@0,400;0,600;1,400&family=Material+Icons&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined');

/* ── 1. GLOBAL TYPOGRAPHY ─────────────────────────────────── */
html, body, .stApp, .block-container,
p, div, label,
h1, h2, h3, h4, h5, h6,
li, ul, ol, a, small,
.stDataFrame, .stTable {
  font-family: "EB Garamond", Georgia, "Times New Roman", serif !important;
  color: #1a1a2e;
}
h1, h2, h3, h4, .fbc-heading, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 {
  font-family: "Playfair Display", Georgia, serif !important;
  font-weight: 700 !important;
  letter-spacing: -0.01em !important;
}

/* ── 2. PAGE BACKGROUND ───────────────────────────────────── */
.stApp { background: #f5f7fb !important; }
.main .block-container { background: #f5f7fb !important; padding-top: 1.5rem !important; }

/* ── 3. SIDEBAR ───────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: linear-gradient(175deg, #001a5c 0%, #003399 45%, #0044cc 100%) !important;
    border-right: 2px solid rgba(245,180,0,0.25) !important;
    box-shadow: 4px 0 24px rgba(0,26,92,0.35) !important;
}
section[data-testid="stSidebar"]::before {
    content: "";
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #f5b400, #ffd040, #f5b400);
}
section[data-testid="stSidebar"] * {
    color: #e8f0ff !important;
    font-family: "EB Garamond", Georgia, serif !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
    font-family: "Playfair Display", serif !important;
    font-weight: 700 !important;
    text-shadow: 0 1px 4px rgba(0,0,0,0.3) !important;
}
section[data-testid="stSidebar"] .block-container {
    padding-top: 1.5rem !important;
}
section[data-testid="stSidebar"] a {
    color: #b8d0ff !important;
    transition: color 0.15s !important;
}
section[data-testid="stSidebar"] a:hover {
    color: #f5b400 !important;
}
section[data-testid="stSidebar"] hr {
    border: none !important;
    border-top: 1px solid rgba(245,180,0,0.25) !important;
    margin: 12px 0 !important;
}

/* ✅ Ensure Material Icons render correctly */
.material-icons,
.material-icons-outlined,
.material-symbols-outlined,
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stSidebarCollapseButton"] i {
    font-family: 'Material Icons', 'Material Symbols Outlined' !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    display: inline-block !important;
    white-space: nowrap !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}
[data-testid="stSidebarCollapseButton"] button {
    background: linear-gradient(135deg, #f5b400, #ffd040) !important;
    border: none !important;
    border-radius: 50% !important;
    width: 44px !important; height: 44px !important;
    box-shadow: 0 4px 14px rgba(245,180,0,0.50) !important;
    transition: all 0.2s ease !important;
}
[data-testid="stSidebarCollapseButton"] button:hover {
    transform: translateY(-1px) scale(1.06) !important;
    box-shadow: 0 8px 20px rgba(245,180,0,0.60) !important;
}
[data-testid="stSidebarCollapseButton"] svg {
    width: 22px !important; height: 22px !important;
    fill: #001a5c !important;
}
/* Move collapse arrow to right edge of sidebar */
[data-testid="stSidebarCollapseButton"] {
    position: absolute !important;
    top: 12px !important;
    right: 12px !important;
    left: auto !important;
    z-index: 999999 !important;
}
[data-testid="stSidebarCollapseButton"] button {
    background: linear-gradient(135deg, #003399, #0055ee) !important;
    border: none !important;
    border-radius: 50% !important;
    width: 46px !important;
    height: 46px !important;
    box-shadow: 0 6px 18px rgba(0,51,153,0.35) !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    transition: all 0.25s ease !important;
}

/* ── Hide the "keyboa..." app name text ── */
[data-testid="stSidebarHeader"] > *:not([data-testid="stSidebarCollapseButton"]) {
    display: none !important;
}

/* ── BUTTONS ── */
.stButton > button {
    background: linear-gradient(135deg, #003399 0%, #0055ee 100%) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-family: "EB Garamond", serif !important;
    font-size: 15px !important;
    padding: 10px 24px !important;
    box-shadow: 0 4px 14px rgba(0,51,153,0.30) !important;
}

.stButton > button * {
    color: #ffffff !important;
}
.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 22px rgba(0,51,153,0.40) !important;
}

/* ── INPUTS ── */
.stTextInput input, .stSelectbox select {
    border: 1.5px solid rgba(0,51,153,0.20) !important;
    border-radius: 8px !important;
    font-family: Georgia, serif !important;
    font-size: 15px !important;
    transition: border-color 0.15s !important;
}
.stTextInput input:focus {
    border-color: #003399 !important;
    box-shadow: 0 0 0 3px rgba(0,51,153,0.10) !important;
}

/* ── AUTH CARD ── */
.auth-card {
    background: #ffffff;
    border-radius: 20px;
    padding: 38px 44px;
    max-width: 460px;
    width: 100%;
    box-shadow: 0 20px 60px rgba(0,26,92,0.16);
    border-top: 4px solid #f5b400;
    margin: 0 auto;
}
.auth-title {
    font-family: "Playfair Display", serif !important;
    font-size: 26px !important;
    font-weight: 900 !important;
    color: #001a5c !important;
    text-align: center;
    margin-bottom: 4px !important;
}
.auth-subtitle {
    font-size: 14px !important;
    color: #003399 !important;
    text-align: center;
    font-style: italic;
    font-weight: 600 !important;
    margin-bottom: 20px !important;
}
.auth-divider {
    border: none;
    border-top: 1px solid rgba(0,51,153,0.12);
    margin: 18px 0;
}
.auth-link {
    font-size: 13px;
    color: #003399 !important;
    text-align: center;
    cursor: pointer;
    text-decoration: underline;
}

/* ── TOP NAV ── */
.top-nav {
    background: linear-gradient(90deg, #001a5c, #003399);
    border-radius: 16px;
    padding: 14px 28px;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    box-shadow: 0 6px 24px rgba(0,26,92,0.30);
    border-bottom: 3px solid #f5b400;
}
.top-title {
    font-size: 22px;
    font-weight: 900;
    color: #ffffff !important;
    font-family: "Playfair Display", serif !important;
    letter-spacing: -0.01em;
    text-shadow: 0 2px 8px rgba(0,0,0,0.25);
}
.top-logo-text {
    font-size: 13px;
    color: rgba(255,255,255,0.65) !important;
    font-style: italic;
    font-family: "EB Garamond", serif !important;
    margin-top: 2px;
}
.user-badge {
    background: rgba(245,180,0,0.22);
    border: 1.5px solid rgba(245,180,0,0.60);
    color: #ffd040 !important;
    font-size: 14px;
    font-weight: 700;
    padding: 6px 18px;
    border-radius: 999px;
    font-family: "EB Garamond", serif !important;
    white-space: nowrap;
    text-shadow: 0 1px 3px rgba(0,0,0,0.20);
}

/* ── PAGE HEADER ── */
.fbc-page-header {
    background: linear-gradient(135deg, #001a5c 0%, #003399 50%, #0044cc 100%);
    border-radius: 18px; padding: 26px 32px; margin-bottom: 24px;
    border-bottom: 3px solid #f5b400;
    box-shadow: 0 12px 40px rgba(0,26,92,0.28);
}
.fbc-page-header-title {
    font-family: "Playfair Display", serif !important;
    font-size: 26px !important; font-weight: 900 !important;
    color: #ffffff !important;
}
.fbc-page-header-sub {
    font-size: 14px; color: rgba(255,255,255,0.85) !important;
    margin-top: 6px; font-style: italic;
}

/* ── FEATURE CARDS ── */
.feature-box {
    background: #ffffff; padding: 20px 22px;
    border-radius: 14px; border-left: 6px solid #003399;
    box-shadow: 0 4px 12px rgba(0,0,0,0.07);
    transition: all 0.25s ease; margin-bottom: 14px;
}
.feature-box:hover {
    background: #f4f8ff;
    box-shadow: 0 8px 22px rgba(0,51,153,0.16);
    transform: translateY(-3px);
}
.feature-icon { font-size: 22px; margin-right: 8px; }

/* ── MISC ── */
[data-testid="metric-container"] {
    background: linear-gradient(135deg, #f0f5ff, #fff8e6) !important;
    border: 1px solid rgba(0,51,153,0.12) !important;
    border-radius: 14px !important;
    padding: 14px 16px !important;
}
.stDataFrame thead th { background: #003399 !important; color: white !important; }
.fbc-divider { border: none; border-top: 1px solid rgba(0,51,153,0.12); margin: 18px 0; }
.fbc-footer {
    text-align: center; padding: 22px; margin-top: 40px;
    color: #5a7099 !important; font-size: 13px;
    border-top: 1px solid rgba(0,51,153,0.10);
}
.fbc-footer b { color: #003399 !important; }
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-thumb { background: #003399; border-radius: 999px; }
.stTabs [data-baseweb="tab"] {
    border-radius: 10px 10px 0 0 !important; font-weight: 600 !important;
    font-family: Georgia, serif !important; color: #5a7099 !important;
}
.stTabs [aria-selected="true"] { background: #003399 !important; color: white !important; }
</style>
""", unsafe_allow_html=True)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700;900&family=EB+Garamond:ital,wght@0,400;0,600;1,400&display=swap');
html,body,.stApp,.block-container,p,div,label,h1,h2,h3,h4,h5,h6,li,ul,ol,a,small{
  font-family:"EB Garamond",Georgia,"Times New Roman",serif!important;color:#1a1a2e;}
h1,h2,h3,h4,.stMarkdown h1,.stMarkdown h2,.stMarkdown h3{
  font-family:"Playfair Display",Georgia,serif!important;font-weight:700!important;}
.stApp{background:#f5f7fb!important;}
.main .block-container{background:#f5f7fb!important;padding-top:1.5rem!important;}
section[data-testid="stSidebar"]{
  background:linear-gradient(175deg,#001a5c 0%,#003399 45%,#0044cc 100%)!important;
  border-right:2px solid rgba(245,180,0,.25)!important;}
section[data-testid="stSidebar"] *{color:#e8f0ff!important;font-family:"EB Garamond",Georgia,serif!important;}
section[data-testid="stSidebar"] h1,section[data-testid="stSidebar"] h2,section[data-testid="stSidebar"] h3{
  color:#fff!important;font-family:"Playfair Display",serif!important;font-weight:700!important;}
.stButton>button{background:linear-gradient(135deg,#003399,#0055ee)!important;color:#fff!important;
  border:none!important;border-radius:10px!important;font-weight:700!important;
  font-family:"EB Garamond",serif!important;font-size:15px!important;padding:10px 24px!important;
  box-shadow:0 4px 14px rgba(0,51,153,.30)!important;transition:all .2s ease!important;}
.stButton>button:hover{transform:translateY(-2px)!important;}
.stButton>button *,.stDownloadButton>button *{color:#fff!important;}
.stDownloadButton>button{background:linear-gradient(135deg,#003399,#0044cc)!important;
  color:#fff!important;border-radius:10px!important;font-weight:700!important;border:none!important;}
[data-testid="stFileUploader"]{background:rgba(0,51,153,.04);
  border:1px dashed rgba(0,51,153,.25);border-radius:14px;padding:18px 20px;}
::-webkit-scrollbar{width:6px;height:6px;}
::-webkit-scrollbar-thumb{background:#003399;border-radius:999px;}
.fbc-section{display:block;padding:16px 0;margin:28px 0 18px 0;border-bottom:2px solid rgba(0,51,153,.15);}
.fbc-section-title{font-family:"Playfair Display",serif!important;font-size:21px;font-weight:800;color:#001a5c!important;}
.fbc-divider{height:2px;background:linear-gradient(90deg,transparent,#f5b400,#003399,transparent);
  border:none;margin:20px 0;border-radius:999px;}
.fbc-footer{text-align:center;padding:22px;margin-top:40px;color:#5a7099!important;font-size:13px;
  border-top:1px solid rgba(0,51,153,.10);font-style:italic;}
.fbc-footer b{color:#003399!important;font-style:normal;}
.fbc-info-card{background:linear-gradient(135deg,#001a5c,#003399);border-radius:16px;
  padding:20px 26px;margin-bottom:20px;border-bottom:3px solid #f5b400;
  box-shadow:0 8px 24px rgba(0,26,92,.22);}
.fbc-info-card-title{font-family:"Playfair Display",serif!important;font-size:19px;
  font-weight:800;color:#fff!important;margin-bottom:8px;}
.fbc-info-card-body{color:rgba(255,255,255,.85)!important;font-size:14px;line-height:1.8;}
.fbc-warn-card{background:linear-gradient(135deg,#7f1d1d,#991b1b);border-radius:14px;
  padding:16px 22px;margin:14px 0;border-left:5px solid #f5b400;}
.fbc-warn-title{font-family:"Playfair Display",serif!important;font-weight:700;
  color:#fff!important;margin-bottom:4px;font-size:15px;}
.fbc-warn-body{color:rgba(255,255,255,.88)!important;font-size:13px;line-height:1.6;}
.result-card{background:#fff;border:1px solid rgba(0,51,153,.12);border-left:5px solid #003399;
  border-radius:14px;padding:16px 18px;margin-bottom:12px;box-shadow:0 4px 12px rgba(0,0,0,.06);}
.next-steps{background:linear-gradient(135deg,#f0f5ff,#fffdf0);
  border:1px solid rgba(0,51,153,.12);border-left:5px solid #f5b400;
  border-radius:14px;padding:18px 22px;margin-top:16px;}
.next-steps-title{font-family:"Playfair Display",serif!important;font-size:16px;
  font-weight:800;color:#001a5c;margin-bottom:8px;}
.next-steps-body{color:#374151;font-size:14px;line-height:1.9;}
.fbc-fail-card{background:#fff5f5;border:1px solid rgba(153,27,27,.25);border-left:5px solid #991b1b;
  border-radius:14px;padding:14px 18px;margin-bottom:10px;}
.fbc-fail-title{font-weight:700;color:#991b1b;font-family:"Playfair Display",serif!important;font-size:14px;}
.fbc-fail-body{color:#7f1d1d;font-size:13px;margin-top:2px;}
</style>
""", unsafe_allow_html=True)

def section(title):
    st.markdown(f'<div class="fbc-section"><div class="fbc-section-title">{title}</div></div>',
                unsafe_allow_html=True)

st.markdown("""
<div style="background:linear-gradient(90deg,#001233,#001a5c 40%,#003399);border-radius:16px;
  padding:18px 28px;margin-bottom:4px;display:flex;align-items:center;
  justify-content:space-between;box-shadow:0 6px 24px rgba(0,26,92,.32);
  border-bottom:3px solid #f5b400;">
  <div>
    <div style="font-family:'Playfair Display',serif;font-size:22px;font-weight:900;
      color:#fff;letter-spacing:-.01em;text-shadow:0 2px 8px rgba(0,0,0,.30);">
      📄 PDF &amp; Image → Excel Converter
    </div>
    <div style="font-family:'EB Garamond',serif;font-size:13px;font-style:italic;
      color:rgba(255,255,255,.65);margin-top:2px;">
      Smart extraction: column-aware text reading for digital PDFs · OCR for scanned documents
    </div>
  </div>
  <div style="background:rgba(245,180,0,.22);border:1.5px solid rgba(245,180,0,.60);
    color:#ffd040;font-size:13px;font-weight:700;padding:6px 18px;border-radius:999px;
    font-family:'EB Garamond',serif;white-space:nowrap;">FBC Securities</div>
</div>
<hr style="border:none;border-top:2px solid #dde6f5;margin:6px 0 20px 0;">
""", unsafe_allow_html=True)

if not OPENPYXL_OK:
    st.error("❌ openpyxl not installed.")
    st.stop()

# ═══════════════════════════════════════════════════════════════════════════════
# STRATEGY 1 — DIGITAL PDF EXTRACTION
# Handles:
#   • Single-column pages (normal statements)
#   • Multi-column / magazine layouts (e.g. IS and BS side-by-side on one page)
#   • 4-column value tables (Company 2025 / 2024, Group 2025 / 2024)
# ═══════════════════════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────────────────────
# Row / band helpers
# ─────────────────────────────────────────────────────────────────────────────

def _group_words_by_row(words, tol=3):
    if not words:
        return []
    ws = sorted(words, key=lambda w: w["top"])
    rows = [[ws[0]]]
    cur_top = ws[0]["top"]
    for w in ws[1:]:
        if abs(w["top"] - cur_top) <= tol:
            rows[-1].append(w)
            cur_top = sum(x["top"] for x in rows[-1]) / len(rows[-1])
        else:
            rows.append([w])
            cur_top = w["top"]
    return rows

def _pdf_find_column_splits(page, min_gap=40):
    """X-positions where a wide horizontal gap divides side-by-side statements (digital PDF)."""
    words = page.extract_words(x_tolerance=2, y_tolerance=3)
    if not words:
        return []
    xs = sorted(set(round(w["x0"]) for w in words))
    splits = []
    for i in range(len(xs) - 1):
        if xs[i + 1] - xs[i] >= min_gap:
            splits.append((xs[i] + xs[i + 1]) // 2)
    return splits
def _pdf_detect_value_columns(page):
    """
    Find the Note column x and the value-column x positions by locating the
    header row containing 2+ four-digit years (e.g. "2025 2024 2025 2024").
    Digital PDF version — takes a pdfplumber page/CroppedPage object.
    """
    words = page.extract_words(x_tolerance=2, y_tolerance=3)
    if not words:
        return None, []
    rows = _group_words_by_row(words)
    for row in rows[:25]:
        yr = [w for w in row if re.match(r"^20\d\d$", w["text"])]
        if len(yr) >= 2:
            yr.sort(key=lambda w: w["x0"])
            note_ws = [w for w in row if w["text"].lower() in ("note", "notes")]
            note_x = note_ws[0]["x0"] if note_ws else (yr[0]["x0"] - 60)
            return note_x, [w["x0"] for w in yr]
    return None, []
def _parse_val(token_text):
    """Parse a single already-merged numeric token like '(5,741,062)' or '5,611,571'."""
    raw = token_text.strip()
    if not raw or raw in ("-", "\u2013", "\u2014"):
        return None
    neg = raw.startswith("(") and raw.endswith(")")
    core = raw.strip("()").replace(",", "")
    try:
        v = float(core) if "." in core else int(core)
        return -v if neg else v
    except (ValueError, TypeError):
        return None


NOTE_RE = re.compile(r"^\d{1,2}(\.\d{1,2})?[A-Za-z]?$")   # e.g. "7", "14B", "7.1"
NUM_RE  = re.compile(r"^\(?-?\d[\d,\.]*\)?$")              # e.g. "(5,741,062)", "1,986", "-"
DASH_RE = re.compile(r"^[-\u2013\u2014]$")

def _pdf_extract_rows_4col(page, note_x, val_xs):
    """
    Extract rows once we know the Note column x and the (1, 2 or 4) value
    column x-positions for this band/page.

    Key fix vs. the old version: note tokens are identified FIRST by
    position (between label and the first value column) and EXCLUDED from
    numeric bucketing, so note numbers like "12", "16", "18" never get
    concatenated onto an adjacent value (e.g. "12450" -> note "12" + val "450").
    """
    words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
    if not words:
        return []

    n_cols = len(val_xs)
    rows = _group_words_by_row(words, tol=3)
    out = []

    # boundary between label text and the "note + values" zone on the right
    label_max_x = note_x - 5
    # boundary between the note token and the first value column
    note_max_x = val_xs[0] - 18

    for row in rows:
        row_sorted = sorted(row, key=lambda w: w["x0"])
        label_words = [w for w in row_sorted if w["x0"] < label_max_x]
        rhs_words = [w for w in row_sorted if w["x0"] >= label_max_x]

        note_tok = ""
        value_words = []
        for w in rhs_words:
            txt = w["text"]
            x = w["x0"]
            # Position decides note vs. value, not content: anything sitting
            # strictly left of the first value column (and right of the
            # label) is the note reference, never a value. A bare value can
            # never appear there because the value columns start at val_xs[0].
            if x < note_max_x and NOTE_RE.match(txt) and not note_tok:
                note_tok = txt
            else:
                value_words.append(w)

        # bucket remaining value words by nearest value-column x
        val_buckets = {i: [] for i in range(n_cols)}
        for w in value_words:
            dists = [abs(w["x0"] - vx) for vx in val_xs]
            nearest = dists.index(min(dists))
            val_buckets[nearest].append(w["text"])

        label = " ".join(w["text"] for w in label_words)
        vals = [_parse_val("".join(val_buckets[i])) for i in range(n_cols)]
        has_value = any(v is not None for v in vals)

        out.append({
            "label": label,
            "note": note_tok,
            "val1": vals[0] if len(vals) > 0 else "",
            "val2": vals[1] if len(vals) > 1 else "",
            "val3": vals[2] if len(vals) > 2 else "",
            "val4": vals[3] if len(vals) > 3 else "",
            "has_value": has_value,
        })
    return out


def _extract_rows_fallback(page, num_gap=60, label_gap=16, sub_gap=10):
    """2-column fallback for pages with no detectable year header row."""
    words = page.extract_words(x_tolerance=2, y_tolerance=3, keep_blank_chars=False)
    if not words:
        return []
    rows = _group_words_by_row(words, tol=3)
    out = []
    for r in rows:
        r2 = sorted(r, key=lambda w: w["x0"])
        n = len(r2)
        i = n - 1
        while i >= 0 and bool(NUM_RE.match(r2[i]["text"])):
            if i < n - 1 and r2[i + 1]["x0"] - r2[i]["x1"] > num_gap:
                break
            i -= 1
        num_start_idx = i + 1
        if num_start_idx >= n or num_start_idx == 0:
            line = " ".join(w["text"] for w in r2)
            out.append({"label": line, "note": "", "val1": "", "val2": "", "val3": "", "val4": "", "has_value": False})
            continue
        j = num_start_idx - 1
        while j > 0 and r2[j]["x0"] - r2[j - 1]["x1"] <= label_gap:
            j -= 1
        label = " ".join(w["text"] for w in r2[j:num_start_idx])
        numtoks = r2[num_start_idx:]
        groups = [[numtoks[0]]]
        for t in numtoks[1:]:
            prev = groups[-1][-1]
            if t["x0"] - prev["x1"] <= sub_gap:
                groups[-1].append(t)
            else:
                groups.append([t])
        note, val1, val2 = "", "", ""
        if len(groups) >= 3:
            note = "".join(w["text"] for w in groups[0])
            val1 = _parse_val("".join(w["text"] for w in groups[1]))
            val2 = _parse_val("".join(w["text"] for w in groups[2]))
        elif len(groups) == 2:
            val1 = _parse_val("".join(w["text"] for w in groups[0]))
            val2 = _parse_val("".join(w["text"] for w in groups[1]))
        elif len(groups) == 1:
            val1 = _parse_val("".join(w["text"] for w in groups[0]))
        out.append({
            "label": label, "note": note,
            "val1": val1 if val1 is not None else "",
            "val2": val2 if val2 is not None else "",
            "val3": "", "val4": "",
            "has_value": True,
        })
    return out


SECTION_PATTERNS = [
    ("Income Statement",  re.compile(r"(?i)(profit\s+or\s+loss|income\s+statement|statement\s+of\s+comprehensive\s+income)")),
    ("Balance Sheet",     re.compile(r"(?i)(financial\s+position|balance\s+sheet)")),
    ("Cash Flow",         re.compile(r"(?i)(cash\s*flows?)")),
    ("Changes in Equity", re.compile(r"(?i)(changes\s+in\s+equity)")),
]


def _classify_section(label, current):
    for name, pat in SECTION_PATTERNS:
        if pat.search(label):
            return name
    return current


def _is_short_heading(label, max_words=8):
    words = label.strip().split()
    return 0 < len(words) <= max_words


def _predetect_section_from_band(band_page):
    try:
        text = (band_page.extract_text() or "")[:300].replace("\n", " ")
    except Exception:
        return None
    for name, pat in SECTION_PATTERNS:
        if pat.search(text):
            return name
    return None


def extract_digital_pdf(pdf_bytes):
    """
    Reads every page of a digital PDF and extracts financial rows into
    section-named sheets. Handles side-by-side statements (e.g. Income
    Statement and Balance Sheet sharing one landscape page) and 2 or 4
    value columns (Company/Group x current/prior year).

    Each page (or band, for side-by-side layouts) is processed independently
    and wrapped in its own try/except so a problem on one page/band cannot
    silently push the whole document into the OCR fallback.

    Returns (sheets, raw_lines) or (None, None) if the PDF has no extractable
    text at all (i.e. it's a scanned image and needs OCR).
    """
    sections = {}
    order = []
    raw_lines = []
    current_section = "Other"
    found_any_text = False

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            try:
                if page.chars:
                    found_any_text = True
            except Exception:
                pass

            try:
                splits = _pdf_find_column_splits(page, min_gap=40)
            except Exception as _split_e:
                print(f"DIAG split EXCEPTION: {_split_e}", flush=True)
                splits = []
            print(f"DIAG page width={page.width!r}, splits={splits!r}", flush=True)
            boundaries = [0] + splits + [page.width]
            # Clamp to avoid float rounding pushing a boundary past the
            # actual page width, which raises on page.crop() in some
            # pdfplumber/pdfminer versions.
            boundaries = [max(0, min(b, page.width)) for b in boundaries]
            bands = [(boundaries[i], boundaries[i + 1]) for i in range(len(boundaries) - 1)]
            print(f"DIAG bands={bands!r}", flush=True)

            for band_x0, band_x1 in bands:
                try:
                    band_page = page.crop((band_x0, 0, band_x1, page.height))

                    band_section = _predetect_section_from_band(band_page)
                    if band_section and band_section != current_section:
                        current_section = band_section

                    try:
                        band_text = band_page.extract_text() or ""
                    except Exception:
                        band_text = ""
                    if band_text:
                        raw_lines.extend(band_text.split("\n"))
                    note_x, val_xs = _pdf_detect_value_columns(band_page)
                    if val_xs:
                        rows = _pdf_extract_rows_4col(band_page, note_x, val_xs)
                    else:
                        rows = _extract_rows_fallback(band_page)
                    kept_count = 0
                    for row in rows:
                        current_section = _classify_section(row["label"], current_section)
                        has_data = row.get("has_value") or _is_short_heading(row["label"])
                        if not has_data:
                            continue
                        kept_count += 1
                        if current_section not in sections:
                            sections[current_section] = []
                            order.append(current_section)
                        sections[current_section].append({
                            "label": row["label"],
                            "note": row["note"],
                            "val1": row["val1"],
                            "val2": row["val2"],
                            "val3": row.get("val3", ""),
                            "val4": row.get("val4", ""),
                        })
                    print(f"DIAG band ({band_x0:.0f},{band_x1:.0f}): val_xs={val_xs}, "
                              f"{len(rows)} rows extracted, {kept_count} kept", flush=True)
                except Exception as _diag_e:
                    print(f"DIAG band ({band_x0:.0f},{band_x1:.0f}) EXCEPTION: {_diag_e}", flush=True)
                    continue

    if not found_any_text:
        return None, None

    sheets = [(name, sections[name]) for name in order if sections[name]]
    return sheets, raw_lines

# ═══════════════════════════════════════════════════════════════════════════════
# STRATEGY 2: OCR PIPELINE — for scanned PDFs and images
# (unchanged from original)
# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# STRATEGY 2: OCR PIPELINE — for scanned PDFs and images
# ═══════════════════════════════════════════════════════════════════════════════
import pytesseract
from PIL import Image, ImageFilter, ImageEnhance

MAX_OCR_PIXELS = 16_000_000
NOISE_PAT = re.compile(r"^[^a-zA-Z0-9()\-.]+$")
NOISE_WORDS = {"it", "ot", "be", "a=", "bet", "Ml", "oe", "i", "—", "~~", "MIl", "Mi", "Ml"}
NOTE_RE = re.compile(r"^\d{1,2}(\.\d{1,2})?[A-Za-z]?$")


def _is_noise(t):
    return bool(NOISE_PAT.match(t)) or t in NOISE_WORDS


def _preprocess(pil_image, upscale=3, sharpen=True):
    img = pil_image.convert("RGB")
    w, h = img.width, img.height
    eff_upscale = upscale
    if w * h * (upscale ** 2) > MAX_OCR_PIXELS and w * h > 0:
        eff_upscale = max(1.0, (MAX_OCR_PIXELS / (w * h)) ** 0.5)
    new_w = max(1, int(w * eff_upscale))
    new_h = max(1, int(h * eff_upscale))
    img = img.resize((new_w, new_h), Image.LANCZOS)
    img = ImageEnhance.Contrast(img).enhance(1.4)
    if sharpen:
        img = img.filter(ImageFilter.SHARPEN)
    return img, eff_upscale


def _ocr_words(preprocessed_img, upscale):
    data = pytesseract.image_to_data(preprocessed_img, output_type=pytesseract.Output.DICT)
    words = []
    for i in range(len(data["text"])):
        txt = data["text"][i].strip()
        if not txt or _is_noise(txt):
            continue
        try:
            conf = float(data["conf"][i])
        except (ValueError, TypeError):
            conf = -1.0
        words.append({
            "text": txt,
            "left": data["left"][i] // upscale,
            "top": data["top"][i] // upscale,
            "width": data["width"][i] // upscale,
            "right": (data["left"][i] + data["width"][i]) // upscale,
            "conf": conf,
        })
    return words


def _group_rows(words, tol=6):
    if not words:
        return []
    ws = sorted(words, key=lambda w: w["top"])
    rows = [[ws[0]]]
    cur_top = ws[0]["top"]
    for w in ws[1:]:
        if abs(w["top"] - cur_top) <= tol:
            rows[-1].append(w)
            cur_top = sum(x["top"] for x in rows[-1]) / len(rows[-1])
        else:
            rows.append([w])
            cur_top = w["top"]
    for r in rows:
        r.sort(key=lambda w: w["left"])
    return rows


def _find_column_splits(words, min_gap=100):
    """Detect wide horizontal gaps that divide the image into side-by-side
    statements (mirrors the digital-PDF band-splitting logic)."""
    if not words:
        return []
    xs = sorted(set(w["left"] for w in words))
    splits = []
    for i in range(len(xs) - 1):
        if xs[i + 1] - xs[i] >= min_gap:
            splits.append((xs[i] + xs[i + 1]) // 2)
    return splits


def _detect_value_columns(rows):
    """
    Find the value-column x positions by locating a row with 2+ four-digit
    years (e.g. "2025 2024 2025 2024"). Returns (note_x, [val_x...]) or
    (None, []) if no such header row is found within the first ~25 rows.
    """
    for row in rows[:25]:
        yr = [w for w in row if re.match(r"^20\d\d$", w["text"])]
        if len(yr) >= 2:
            yr.sort(key=lambda w: w["left"])
            note_ws = [w for w in row if w["text"].lower() in ("note", "notes")]
            note_x = note_ws[0]["left"] if note_ws else (yr[0]["left"] - 80)
            return note_x, [w["left"] for w in yr]
    return None, []


def _merge_tokens(tokens):
    joined = " ".join(tokens)
    joined = re.sub(r"\$(?=\d)", "8", joined)
    joined = re.sub(r"(?<=\d)[Oo](?=\d)", "0", joined)
    joined = re.sub(r"[~:=`']+", "", joined)
    joined = joined.strip()
    # Strip stray OCR noise around a parenthesised negative, e.g.
    # "_(8,455,071)" -> "(8,455,071)", "(125,974) " -> "(125,974)"
    m = re.match(r"^[^\d(]*(\(?-?[\d,\.\s]+\)?)[^\d)]*$", joined)
    if m:
        joined = m.group(1).strip()
    return joined


def _parse_number(raw):
    s = raw.strip()
    if not s or s in ("-", "\u2013", "\u2014", ""):
        return None
    negative = s.startswith("(") and s.endswith(")")
    if negative:
        s = s[1:-1].strip()
    s_nospace = s.replace(" ", "")
    has_comma = "," in s_nospace
    has_dot = "." in s_nospace
    if not has_comma and not has_dot:
        cleaned = s_nospace
    elif has_comma and not has_dot:
        cleaned = s_nospace.replace(",", "")
    elif has_dot and not has_comma:
        parts = s_nospace.split(".")
        cleaned = s_nospace.replace(".", "") if (len(parts) > 2 or (len(parts) == 2 and len(parts[-1]) == 3)) else s_nospace
    else:
        cleaned = s_nospace.replace(",", "")
    if not re.match(r"^\d+(\.\d+)?$", cleaned):
        return raw
    try:
        fval = float(cleaned)
        result = int(fval) if fval == int(fval) else fval
        return -result if negative else result
    except Exception:
        return raw


def _extract_rows_4col(rows, note_x, val_xs, header_top, conf_threshold=60):
    """
    OCR equivalent of the digital-PDF _extract_rows_4col: bucket each row's
    words into label / note / up to 4 value columns using x-position only,
    so note numbers (12, 16, 18...) never get glued onto an adjacent value.

    Each value cell also gets a "risky" flag: True if the OCR confidence on
    any token in that cell was low, or if a non-empty token failed to parse
    into a clean number. Dashes/blanks are never flagged risky. This doesn't
    fix misreads but makes them visible so a person can spot-check just the
    flagged cells.
    """
    n_cols = len(val_xs)
    label_max_x = note_x - 10
    note_max_x = val_xs[0] - 25
    out = []

    for row in rows:
        is_pre_header = header_top > 0 and row[0]["top"] < header_top - 3
        if is_pre_header:
            line = " ".join(w["text"] for w in row)
            out.append({"label": line, "note": "", "val1": "", "val2": "", "val3": "", "val4": "",
                        "has_value": False, "risky1": False, "risky2": False, "risky3": False, "risky4": False})
            continue

        row_sorted = sorted(row, key=lambda w: w["left"])
        label_words = [w for w in row_sorted if w["left"] < label_max_x]
        rhs_words = [w for w in row_sorted if w["left"] >= label_max_x]

        note_tok = ""
        value_words = []
        for w in rhs_words:
            txt = w["text"]
            x = w["left"]
            if x < note_max_x and NOTE_RE.match(txt) and not note_tok:
                note_tok = txt
            else:
                value_words.append(w)

        val_buckets = {i: [] for i in range(n_cols)}
        for w in value_words:
            dists = [abs(w["left"] - vx) for vx in val_xs]
            nearest = dists.index(min(dists))
            val_buckets[nearest].append(w)

        label = " ".join(w["text"] for w in label_words)
        raw_vals = [_merge_tokens([w["text"] for w in val_buckets[i]]) for i in range(n_cols)]
        parsed = [_parse_number(r) if r else None for r in raw_vals]
        # _parse_number returns None for a literal dash/blank placeholder; in
        # that case the cell is correctly empty (not a parse failure).
        vals = [p if p is not None else "" for p in parsed]
        has_value = any(isinstance(v, (int, float)) for v in vals)

        riskies = []
        for i in range(n_cols):
            bucket = val_buckets[i]
            val = vals[i] if i < len(vals) else ""
            raw = raw_vals[i]
            is_dash = raw.strip() in ("-", "\u2013", "\u2014", "")
            if not bucket or is_dash:
                riskies.append(False)
                continue
            min_conf = min((w.get("conf", -1) for w in bucket), default=-1)
            low_conf = 0 <= min_conf < conf_threshold
            failed_parse = bool(raw) and not isinstance(val, (int, float))
            riskies.append(bool(low_conf or failed_parse))

        out.append({
            "label": label,
            "note": note_tok,
            "val1": vals[0] if len(vals) > 0 else "",
            "val2": vals[1] if len(vals) > 1 else "",
            "val3": vals[2] if len(vals) > 2 else "",
            "val4": vals[3] if len(vals) > 3 else "",
            "has_value": has_value,
            "risky1": riskies[0] if len(riskies) > 0 else False,
            "risky2": riskies[1] if len(riskies) > 1 else False,
            "risky3": riskies[2] if len(riskies) > 2 else False,
            "risky4": riskies[3] if len(riskies) > 3 else False,
        })
    return out


# ── 2-column fallback (legacy anchor-based detection, used when no year
#    header row is found in a band) ───────────────────────────────────────
MONTHS = {"january", "february", "march", "april", "may", "june", "july",
          "august", "september", "october", "november", "december"}
USD_PAT = re.compile(r"(?i)^(us[\$5s8sS]|u[\$5]s|usd|u\.s\.\$?)$")


def _x_clusters(words, gap=60):
    if not words:
        return []
    ws = sorted(words, key=lambda w: w["left"])
    clusters = [[ws[0]]]
    for w in ws[1:]:
        if w["left"] - clusters[-1][-1]["right"] < gap:
            clusters[-1].append(w)
        else:
            clusters.append([w])
    return clusters


def _detect_anchors_2col(rows, img_width):
    for ri, r in enumerate(rows):
        month_ws = [w for w in r if w["text"].lower() in MONTHS]
        if len(month_ws) < 2:
            continue
        month_ws = sorted(month_ws, key=lambda w: w["left"])
        usd_same = sorted([w for w in r if USD_PAT.match(w["text"])], key=lambda w: w["left"])
        if len(usd_same) >= 2:
            return int(month_ws[0]["left"] * 0.35), usd_same[0]["left"], usd_same[1]["left"], r[0]["top"]
        for look_r in rows[ri + 1:ri + 6]:
            usd_ws = sorted([w for w in look_r if USD_PAT.match(w["text"])], key=lambda w: w["left"])
            if len(usd_ws) >= 2:
                return int(month_ws[0]["left"] * 0.35), usd_ws[0]["left"], usd_ws[1]["left"], look_r[0]["top"]
            right_ws = [w for w in look_r if w["left"] > month_ws[0]["left"]]
            cls = _x_clusters(right_ws, gap=60)
            if len(cls) >= 2:
                return int(month_ws[0]["left"] * 0.35), cls[0][0]["left"], cls[-1][0]["left"], look_r[0]["top"]
    for r in rows:
        has_note = any(w["text"].lower() in ("note", "notes") for w in r)
        if not has_note:
            continue
        note_w = next(w for w in r if w["text"].lower() in ("note", "notes"))
        note_col = note_w["left"]
        right_words = sorted([w for w in r if w["left"] > note_col + 30], key=lambda w: w["left"])
        clusters = _x_clusters(right_words, gap=80)
        cl = [c[0]["left"] for c in clusters]
        if len(cl) >= 2:
            return note_col, cl[0], cl[-1], r[0]["top"]
        note_row_idx = rows.index(r)
        for look in rows[note_row_idx + 1:note_row_idx + 5]:
            rw = sorted([w for w in look if w["left"] > note_col + 30], key=lambda w: w["left"])
            cls2 = _x_clusters(rw, gap=80)
            if len(cls2) >= 2:
                return note_col, cls2[0][0]["left"], cls2[-1][0]["left"], r[0]["top"]
    for r in rows:
        usd = sorted([w for w in r if USD_PAT.match(w["text"])], key=lambda w: w["left"])
        if len(usd) >= 2:
            return int(usd[0]["left"] * 0.4), usd[0]["left"], usd[1]["left"], r[0]["top"]
    for r in rows:
        mw = sorted([w for w in r if w["text"].lower() in MONTHS], key=lambda w: w["left"])
        if len(mw) >= 2:
            return int(mw[0]["left"] * 0.4), mw[0]["left"], mw[1]["left"], r[0]["top"]
    num_re = re.compile(r"^\(?\d[\d\s,.']*\)?$")
    for r in rows:
        nw = sorted([w for w in r if num_re.match(w["text"]) and len(w["text"]) >= 3], key=lambda w: w["left"])
        if len(nw) >= 2:
            return int(nw[0]["left"] * 0.3), nw[0]["left"], nw[-1]["left"], r[0]["top"]
    return int(img_width * 0.38), int(img_width * 0.55), int(img_width * 0.75), 0


def _col_of_2col(word, note_col, val1_col, val2_col, midpoint):
    left = word["left"]
    text = word["text"]
    is_number = bool(re.match(r"^\(?\d[\d\s,.']*\)?$", text) and len(text) >= 2)
    if val2_col is not None and midpoint is not None and left >= midpoint:
        return "val2"
    if val1_col is not None and left >= val1_col - 15:
        return "val1"
    if note_col is not None and left >= note_col - 10 and not is_number:
        return "note"
    if is_number and val1_col is not None and left >= note_col:
        return "val1"
    return "label"


def _clean_note(tokens):
    return " ".join(t for t in tokens if re.match(r"^\d+\.?\d*$", t))


def _extract_rows_2col_fallback(rows, img_width, conf_threshold=60):
    note_col, val1_col, val2_col, header_top = _detect_anchors_2col(rows, img_width)
    midpoint = (val1_col + val2_col) / 2 - 5 if (val1_col is not None and val2_col is not None) else None
    out = []
    for r in rows:
        is_pre = header_top > 0 and r[0]["top"] < header_top
        if is_pre:
            line = " ".join(w["text"] for w in r)
            out.append({"label": line, "note": "", "val1": "", "val2": "", "val3": "", "val4": "",
                        "has_value": False, "risky1": False, "risky2": False, "risky3": False, "risky4": False})
            continue
        buckets = {"label": [], "note": [], "val1": [], "val2": []}
        for w in r:
            col = _col_of_2col(w, note_col, val1_col, val2_col, midpoint)
            buckets[col].append(w)
        label = " ".join(w["text"] for w in buckets["label"])
        note = _clean_note([w["text"] for w in buckets["note"]])
        raw1 = _merge_tokens([w["text"] for w in buckets["val1"]])
        raw2 = _merge_tokens([w["text"] for w in buckets["val2"]])
        p1 = _parse_number(raw1) if raw1 else None
        p2 = _parse_number(raw2) if raw2 else None
        v1 = p1 if p1 is not None else ""
        v2 = p2 if p2 is not None else ""
        has_value = isinstance(v1, (int, float)) or isinstance(v2, (int, float))

        def _cell_risky(bucket, raw, val):
            if not bucket or raw.strip() in ("-", "\u2013", "\u2014", ""):
                return False
            min_conf = min((w.get("conf", -1) for w in bucket), default=-1)
            low_conf = 0 <= min_conf < conf_threshold
            failed_parse = bool(raw) and not isinstance(val, (int, float))
            return bool(low_conf or failed_parse)

        risky1 = _cell_risky(buckets["val1"], raw1, v1)
        risky2 = _cell_risky(buckets["val2"], raw2, v2)
        out.append({"label": label, "note": note, "val1": v1, "val2": v2, "val3": "", "val4": "",
                    "has_value": has_value, "risky1": risky1, "risky2": risky2, "risky3": False, "risky4": False})
    return out


SECTION_PATTERNS = [
    ("Income Statement", re.compile(r"(?i)(profit\s+or\s+loss|income\s+statement|statement\s+of\s+comprehensive\s+income)")),
    ("Balance Sheet", re.compile(r"(?i)(financial\s+position|balance\s+sheet)")),
    ("Cash Flow", re.compile(r"(?i)(cash\s*flows?)")),
    ("Changes in Equity", re.compile(r"(?i)(changes\s+in\s+equity)")),
]


def _classify_section(label, current):
    for name, pat in SECTION_PATTERNS:
        if pat.search(label):
            return name
    return current


def _band_heading_text(rows, max_rows=6):
    lines = []
    for row in rows[:max_rows]:
        lines.append(" ".join(w["text"] for w in sorted(row, key=lambda w: w["left"])))
    return " ".join(lines)


def process_image_to_sections(pil_image, upscale=3):
    """
    Column-aware OCR pipeline. Detects side-by-side statements via a wide
    x-gap (like the digital-PDF band splitter) and, within each band,
    detects a 4-column (or 2-column) year header to bucket values by
    position instead of naive left/right halves. Each value cell carries a
    risky flag (risky1-risky4) based on OCR confidence and parse success.

    Returns (sections, raw_lines, debug) where sections is a list of
    (section_name, row_dicts) tuples — already split/labelled, ready to
    pass straight into rows_to_excel_bytes.
    """
    proc, up = _preprocess(pil_image, upscale)
    words = _ocr_words(proc, up)
    debug = {"upscale_used": up, "bands": []}

    if not words:
        return [], [], debug

    splits = _find_column_splits(words, min_gap=100)
    img_width = pil_image.width
    boundaries = [0] + splits + [img_width]
    bands = [(boundaries[i], boundaries[i + 1]) for i in range(len(boundaries) - 1)]

    sections = {}
    order = []
    raw_lines = []
    current_section = "Sheet1"

    for band_x0, band_x1 in bands:
        band_words = [w for w in words if band_x0 <= w["left"] < band_x1]
        if not band_words:
            continue
        rows = _group_rows(band_words, tol=6)

        heading = _band_heading_text(rows)
        band_section = _classify_section(heading, None)
        if band_section:
            current_section = band_section

        raw_lines.extend(" ".join(w["text"] for w in sorted(r, key=lambda w: w["left"])) for r in rows)

        note_x, val_xs = _detect_value_columns(rows)
        if val_xs:
            header_top = next((row[0]["top"] for row in rows
                                if any(re.match(r"^20\d\d$", w["text"]) for w in row)), 0)
            row_dicts = _extract_rows_4col(rows, note_x, val_xs, header_top)
            debug["bands"].append({"x0": band_x0, "x1": band_x1, "mode": "4col",
                                    "note_x": note_x, "val_xs": val_xs})
        else:
            row_dicts = _extract_rows_2col_fallback(rows, band_x1 - band_x0)
            debug["bands"].append({"x0": band_x0, "x1": band_x1, "mode": "2col-fallback"})

        for row in row_dicts:
            current_section = _classify_section(row["label"], current_section)
            label_words_count = len(row["label"].strip().split())
            has_data = row.get("has_value") or (0 < label_words_count <= 8)
            if not has_data:
                continue
            if current_section not in sections:
                sections[current_section] = []
                order.append(current_section)
            sections[current_section].append({
                "label": row["label"],
                "note": row["note"],
                "val1": row["val1"],
                "val2": row["val2"],
                "val3": row.get("val3", ""),
                "val4": row.get("val4", ""),
                "risky1": row.get("risky1", False),
                "risky2": row.get("risky2", False),
                "risky3": row.get("risky3", False),
                "risky4": row.get("risky4", False),
            })

    out_sections = [(name, sections[name]) for name in order if sections[name]]
    return out_sections, raw_lines, debug


def run_ocr_pipeline(pdf_bytes, dpi_choice, upscale_choice, show_debug):
    from pdf2image import convert_from_bytes as _cfb
    try:
        images = _cfb(pdf_bytes, dpi=dpi_choice)
    except Exception as e:
        if show_debug:
            st.warning(f"Render at {dpi_choice} DPI failed ({e}); retrying at 150 DPI…")
        images = _cfb(pdf_bytes, dpi=150)

    n_pages = len(images)
    page_bar = st.progress(0.0)
    sheets_data = []
    raw_lines = []
    for pg_i, image in enumerate(images, 1):
        page_bar.progress(pg_i / n_pages, text=f"OCR page {pg_i}/{n_pages}")
        try:
            sections, page_raw_lines, debug = process_image_to_sections(image, upscale=upscale_choice)
            if show_debug:
                st.info(f"Page {pg_i}: {len(debug['bands'])} band(s) · "
                        f"upscale used={debug['upscale_used']:.2f}x · "
                        + " · ".join(f"[{b['mode']}]" for b in debug["bands"]))
            if sections:
                for sec_name, rows in sections:
                    sheets_data.append((f"Pg{pg_i} {sec_name}"[:31], rows))
            else:
                sheets_data.append((f"Page {pg_i}", []))
            raw_lines.extend(page_raw_lines)
        except MemoryError:
            st.warning(f"⚠️ Page {pg_i} skipped — image was too large to process safely.")
        except Exception as e:
            st.warning(f"⚠️ Page {pg_i} OCR failed: {e}")
            if show_debug:
                st.code(traceback.format_exc())
        finally:
            del image
            gc.collect()
    page_bar.progress(1.0)
    return sheets_data, raw_lines
# ── Excel writer — updated to handle val3 / val4 columns ─────────────────────
def rows_to_excel_bytes(sheets, col_headers=None):
    if col_headers is None:
        col_headers = ["Item", "Note", "Period 1", "Period 2", "Period 3", "Period 4"]
    wb = Workbook(); wb.remove(wb.active)
    BLUE_FILL  = PatternFill("solid", fgColor="071426")
    LIGHT_FILL = PatternFill("solid", fgColor="F0F5FF")
    RISKY_FILL = PatternFill("solid", fgColor="FFD9A0")   # amber — low-confidence OCR cell
    RED_FONT   = Font(color="9C0006", bold=True)
    WHITE_FONT = Font(bold=True, color="FFFFFF")
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
    MONEY_FMT   = "#,##0;(#,##0)"
    used_names = set()
    any_risky_cells = False

    for title, row_dicts in sheets:
        safe_title = (title or "Sheet")[:31]
        base, n = safe_title, 1
        while safe_title in used_names:
            n += 1
            safe_title = f"{base[:28]}-{n}"
        used_names.add(safe_title)

        ws = wb.create_sheet(title=safe_title)

        # Detect whether this sheet has val3/val4 data
        has_4cols = any(
            row.get("val3") not in ("", None) or row.get("val4") not in ("", None)
            for row in row_dicts
        )
        n_val_cols = 4 if has_4cols else 2
        effective_headers = col_headers[:2 + n_val_cols]

        for ci, h in enumerate(effective_headers, 1):
            c = ws.cell(1, ci, h)
            c.font = WHITE_FONT; c.fill = BLUE_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        excel_ri = 2
        for row in row_dicts:
            label = row.get("label", "");
            note = row.get("note", "")
            v1 = row.get("val1", "");
            v2 = row.get("val2", "")
            v3 = row.get("val3", "");
            v4 = row.get("val4", "")
            riskies = [row.get("risky1", False), row.get("risky2", False),
                       row.get("risky3", False), row.get("risky4", False)]

            if not any([str(label).strip(), str(note).strip(),
                        str(v1).strip(), str(v2).strip(),
                        str(v3).strip(), str(v4).strip()]):
                continue
            try:
                ws.cell(excel_ri, 1, str(label)[:2000])
                ws.cell(excel_ri, 2, str(note)[:50])
                val_data = [v1, v2, v3, v4][:n_val_cols]
                risky_data = riskies[:n_val_cols]
                for idx, (val, is_risky) in enumerate(zip(val_data, risky_data)):
                    ci = idx + 3
                    cell = ws.cell(excel_ri, ci)
                    if isinstance(val, (int, float)):
                        cell.value = val;
                        cell.number_format = MONEY_FMT
                        cell.alignment = RIGHT_ALIGN
                    elif val not in ("", None):
                        rep = _parse_number(str(val))
                        if isinstance(rep, (int, float)):
                            cell.value = rep;
                            cell.number_format = MONEY_FMT
                            cell.alignment = RIGHT_ALIGN
                        else:
                            cell.value = str(val)[:2000];
                            cell.alignment = RIGHT_ALIGN
                    if is_risky:
                        cell.fill = RISKY_FILL
                        any_risky_cells = True
                if excel_ri % 2 == 0:
                    for ci in range(1, 2 + n_val_cols + 1):
                        if not (ci >= 3 and risky_data[ci - 3]):
                            ws.cell(excel_ri, ci).fill = LIGHT_FILL
                if str(label).isupper() or (not label and (v1 or v2)):
                    for ci in range(1, 2 + n_val_cols + 1):
                        ws.cell(excel_ri, ci).font = Font(bold=True)
                excel_ri += 1
            except Exception:
                continue

        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 17
        ws.column_dimensions["D"].width = 17
        if n_val_cols >= 3:
            ws.column_dimensions["E"].width = 17
        if n_val_cols >= 4:
            ws.column_dimensions["F"].width = 17
        ws.freeze_panes = "A2"

    buf = io.BytesIO();
    wb.save(buf);
    buf.seek(0)
    return buf.getvalue(), any_risky_cells

def raw_text_sheet(raw_lines, max_lines=2000):
    if not raw_lines:
        return None
    rows = [{"label": ln, "note": "", "val1": "", "val2": "", "val3": "", "val4": ""}
            for ln in raw_lines[:max_lines] if str(ln).strip()]
    return ("Raw Text", rows) if rows else None

# ═══════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI  (identical to original — no changes)
# ═══════════════════════════════════════════════════════════════════════════════

section("🔀 Choose Conversion Mode")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
mode = st.radio("mode",
    ["📄 PDF — digital or scanned document",
     "🖼️ Image / Screenshot (PNG, JPG, WEBP …)"],
    horizontal=True, label_visibility="collapsed")
is_image_mode = mode.startswith("🖼️")

st.markdown("""
<div class="fbc-info-card">
  <div class="fbc-info-card-title">🔬 How This Works</div>
  <div class="fbc-info-card-body">
    <b>Smart dual-engine extraction:</b><br>
    • <b>Digital PDFs</b> — column-aware text reading. Detects side-by-side statement
      layouts (e.g. Income Statement &amp; Balance Sheet on the same page) and
      extracts each independently. Supports 4-column tables (Company/Group × 2 years).<br>
    • <b>Scanned PDFs &amp; images</b> — OCR with 6-strategy column detection.<br>
    Every conversion also includes a <b>Raw Text</b> sheet with everything read
    from the document, so you always have a fallback.
  </div>
</div>
<div class="fbc-warn-card">
  <div class="fbc-warn-title">⚠️ Always verify totals before uploading to the DCF model</div>
  <div class="fbc-warn-body">Always cross-check column totals against the printed statement before feeding into valuation models.</div>
</div>
""", unsafe_allow_html=True)

st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)

section("⚙️ Options")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
col_o1, col_o2, col_o3, col_o4 = st.columns(4)
with col_o1:
    upscale_choice = st.selectbox("Upscale factor (OCR only)", [2,3], index=1)
with col_o2:
    dpi_choice = st.selectbox("Render DPI (scanned PDF OCR)", [200,300,400], index=1, disabled=is_image_mode)
with col_o3:
    col1_header = st.text_input("Col 3 header", value="Co. 2025")
    col2_header = st.text_input("Col 4 header", value="Co. 2024")
with col_o4:
    col3_header = st.text_input("Col 5 header", value="Gr. 2025")
    col4_header = st.text_input("Col 6 header", value="Gr. 2024")

col_headers = [
    "Item", "Note",
    col1_header or "Co. 2025",
    col2_header or "Co. 2024",
    col3_header or "Gr. 2025",
    col4_header or "Gr. 2024",
]
show_debug = st.checkbox("🐛 Show debug info", value=False)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# IMAGE MODE
# ─────────────────────────────────────────────────────────────────────────────
if is_image_mode:
    if not LIBS_OK:
        st.error(f"❌ Missing OCR libraries: {_IMPORT_ERR}")
        st.stop()
    section("🖼️ Upload Screenshots / Images")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    uploaded_images = st.file_uploader("Upload images",
        type=["png","jpg","jpeg","webp","bmp","tiff","tif"],
        accept_multiple_files=True, label_visibility="collapsed")
    if not uploaded_images:
        st.info("⬆️ Upload one or more screenshots to begin.")
        st.stop()
    section("🏷️ Name Each Sheet")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    sheet_names = []; default_names = ["Income Statement","Balance Sheet","Cash Flow"]
    prev_cols = st.columns(min(len(uploaded_images),3))
    for i, img_file in enumerate(uploaded_images):
        with prev_cols[i%3]:
            try:
                pil_prev = Image.open(img_file); img_file.seek(0)
                st.image(pil_prev, use_container_width=True, caption=img_file.name)
            except Exception:
                st.caption(img_file.name)
            sname = st.text_input(f"Sheet name {i+1}",
                value=default_names[i] if i<len(default_names) else f"Sheet {i+1}",
                key=f"img_sheet_name_{i}", label_visibility="collapsed")
            sheet_names.append(sname.strip() or f"Sheet {i+1}")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    col_btn, col_inf = st.columns([1,3])
    with col_btn:
        run_img = st.button("▶️ Run OCR & Convert", type="primary", use_container_width=True)
    with col_inf:
        st.markdown(f"<div style='padding-top:10px;color:#5a7099;font-style:italic;font-size:14px;'>"
                    f"{len(uploaded_images)} image(s) · {upscale_choice}× upscale</div>", unsafe_allow_html=True)
    if not run_img: st.stop()
    section("🔬 Converting…")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    sheets_data = []; raw_lines_all = []; failed_images = []
    bar = st.progress(0.0, text="Starting…")
    for i, img_file in enumerate(uploaded_images):
        sname = sheet_names[i]; status = st.empty()
        status.markdown(f"<span style='color:#003399;font-weight:700;'>OCR — {img_file.name} → '{sname}'…</span>", unsafe_allow_html=True)
        try:
            pil = Image.open(img_file)
            sections, page_raw_lines, debug = process_image_to_sections(pil, upscale=upscale_choice)
            if show_debug:
                st.info(f"'{sname}': {len(debug['bands'])} band(s) · upscale used={debug['upscale_used']:.2f}x · "
                        + " · ".join(f"[{b['mode']}]" for b in debug["bands"]))
            if sections:
                total_rows = 0
                for idx, (sec_name, rows) in enumerate(sections):
                    sheet_label = sname if len(sections) == 1 else f"{sname} - {sec_name}"
                    sheets_data.append((sheet_label[:31], rows))
                    total_rows += len(rows)
                status.success(f"✅ '{sname}' — {len(sections)} sheet(s), {total_rows} data rows extracted")
            else:
                status.warning(f"⚠️ '{sname}' — no data rows could be extracted")
            raw_lines_all.extend(page_raw_lines)
        except MemoryError:
            status.error(f"❌ {img_file.name}: image too large to process safely.")
            failed_images.append((img_file.name, "Image too large"))
        except Exception as exc:
            status.error(f"❌ Failed: {exc}")
            failed_images.append((img_file.name, str(exc)))
            if show_debug:
                st.code(traceback.format_exc())
        finally:
            gc.collect()
        bar.progress((i+1)/len(uploaded_images), text=f"Processed {i+1}/{len(uploaded_images)}")
    bar.progress(1.0, text="Done")
    raw_sheet = raw_text_sheet(raw_lines_all)
    if raw_sheet:
        sheets_data.append(raw_sheet)
    if not sheets_data:
        st.error("❌ No images converted.")
        st.stop()
    try:
        excel_bytes, has_risky = rows_to_excel_bytes(sheets_data, col_headers=col_headers)
    except Exception as e:
        st.error(f"❌ Failed to generate Excel: {e}")
        if show_debug:
            st.code(traceback.format_exc())
        st.stop()
    excel_name = "FBC_Screenshot_Extract.xlsx"
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    section("⬇️ Download Result")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    st.success(f"✅ {len(sheets_data)} sheet(s) extracted.")
    if has_risky:
        st.warning(
            "⚠️ Some cells are highlighted amber — OCR wasn't fully confident on those. Please double-check them against the source.")
    if failed_images:
        st.error(f"❌ {len(failed_images)} image(s) failed:")
        for fname, err in failed_images:
            st.markdown(f"""<div class="fbc-fail-card"><div class="fbc-fail-title">{fname}</div>
              <div class="fbc-fail-body">{err}</div></div>""", unsafe_allow_html=True)
    col_dl, col_nxt = st.columns([1,2])
    with col_dl:
        st.markdown(f"""<div class="result-card">
          <div style="font-family:'Playfair Display',serif;font-weight:700;color:#001a5c;font-size:15px;margin-bottom:4px;">📊 {excel_name}</div>
          <div style="color:#5a7099;font-size:13px;font-style:italic;">{len(sheets_data)} sheet(s) · {len(excel_bytes)//1024:,} KB</div></div>""", unsafe_allow_html=True)
        st.download_button("⬇️ Download Excel", data=excel_bytes, file_name=excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with col_nxt:
        st.markdown("""<div class="next-steps"><div class="next-steps-title">💡 Next Steps</div>
          <div class="next-steps-body"><b>1.</b> Download and open the Excel.<br>
          <b>2.</b> Verify column totals — check the <b>Raw Text</b> sheet if anything looks off.<br>
          <b>3.</b> Go to <b>📊 DCF Model</b> and upload the cleaned Excel.</div></div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# PDF MODE
# ─────────────────────────────────────────────────────────────────────────────
else:
    section("📤 Upload PDF(s)")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    uploaded_pdfs = st.file_uploader("Upload PDF(s)", type=["pdf"],
        accept_multiple_files=True, label_visibility="collapsed")
    if not uploaded_pdfs:
        st.info("⬆️ Upload one or more PDF files to begin.")
        st.stop()
    col_btn, col_inf = st.columns([1,3])
    with col_btn:
        run_pdf = st.button("▶️ Extract & Convert", type="primary", use_container_width=True)
    with col_inf:
        st.markdown(f"<div style='padding-top:10px;color:#5a7099;font-style:italic;font-size:14px;'>"
                    f"{len(uploaded_pdfs)} PDF(s) · auto-detects digital vs scanned · 4-column support</div>", unsafe_allow_html=True)
    if not run_pdf: st.stop()

    results = []; errors = []
    overall = st.progress(0.0, text="Starting…")

    for pdf_i, pdf_file in enumerate(uploaded_pdfs):
        pdf_name = pdf_file.name
        xl_name  = os.path.splitext(pdf_name)[0] + ".xlsx"
        st.markdown(f"<div style='font-family:Playfair Display,serif;font-size:17px;font-weight:700;color:#001a5c;margin:18px 0 6px 0;'>📄 {pdf_name}</div>", unsafe_allow_html=True)
        page_status = st.empty()
        try:
            pdf_bytes = pdf_file.getvalue()
            sheets_data = None; raw_lines = None; method_used = ""

            if PDFPLUMBER_OK:
                page_status.markdown("<span style='color:#003399;font-weight:700;'>🔍 Reading digital text…</span>",
                                     unsafe_allow_html=True)
                try:
                    sheets_data, raw_lines = extract_digital_pdf(pdf_bytes)
                    if show_debug:
                        n_sheets = len(sheets_data) if sheets_data else 0
                        n_raw = len(raw_lines) if raw_lines else 0
                        st.info(f"Digital extraction returned: sheets_data is not None = {sheets_data is not None} "
                                f"({n_sheets} sheets), raw_lines={n_raw} lines")
                    if sheets_data:
                        method_used = "digital text (column-aware, 4-col)"
                except Exception as e:
                    st.warning(f"Digital text extraction error: {e}")
                    st.code(traceback.format_exc())
                    sheets_data, raw_lines = None, None

            if not sheets_data:
                if not LIBS_OK:
                    st.error("❌ No extractable text found and OCR libraries not installed.")
                    errors.append((pdf_name, "No extractable text; OCR not available"))
                    overall.progress((pdf_i+1)/len(uploaded_pdfs))
                    continue
                if not PDF_LIBS_OK:
                    st.error("❌ pdf2image not installed — needed for scanned PDFs.")
                    errors.append((pdf_name, "pdf2image not installed"))
                    overall.progress((pdf_i+1)/len(uploaded_pdfs))
                    continue
                page_status.markdown("<span style='color:#003399;font-weight:700;'>🔬 Scanned PDF detected — running OCR…</span>", unsafe_allow_html=True)
                method_used = "OCR"
                sheets_data, raw_lines = run_ocr_pipeline(pdf_bytes, dpi_choice, upscale_choice, show_debug)

            if not sheets_data:
                errors.append((pdf_name, "No data rows could be extracted"))
                page_status.error(f"❌ Failed to generate Excel for {pdf_name}: no readable rows found.")
                overall.progress((pdf_i+1)/len(uploaded_pdfs))
                continue

            all_sheets = list(sheets_data)
            rsheet = raw_text_sheet(raw_lines)
            if rsheet:
                all_sheets.append(rsheet)

            try:
                excel_bytes_out, has_risky = rows_to_excel_bytes(all_sheets, col_headers=col_headers)
            except Exception as e:
                errors.append((pdf_name, f"Failed to build Excel: {e}"))
                page_status.error(f"❌ Failed to generate Excel for {pdf_name}.")
                if show_debug:
                    st.code(traceback.format_exc())
                overall.progress((pdf_i+1)/len(uploaded_pdfs))
                continue

            n_rows = sum(len(s[1]) for s in sheets_data)
            results.append((xl_name, excel_bytes_out, method_used, n_rows))
            page_status.success(f"✅ {pdf_name} → {xl_name} ({method_used}, {n_rows} data rows + raw text sheet)")
            if has_risky:
                st.warning(
                    f"⚠️ {pdf_name}: some cells are highlighted amber — OCR wasn't fully confident on those. Please double-check them against the source.")

        except MemoryError:
            errors.append((pdf_name, "Ran out of memory"))
            page_status.error(f"❌ {pdf_name}: file too large/complex to process safely.")
        except Exception as exc:
            errors.append((pdf_name, str(exc)))
            page_status.error(f"❌ {pdf_name}: {exc}")
            if show_debug:
                st.code(traceback.format_exc())
        finally:
            gc.collect()
        overall.progress((pdf_i+1)/len(uploaded_pdfs), text=f"Processed {pdf_i+1}/{len(uploaded_pdfs)}")

    overall.progress(1.0, text="All done")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
    section("⬇️ Download Results")
    st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)

    if results:
        st.success(f"✅ {len(results)} file(s) converted.")
        cols = st.columns(min(len(results),3))
        for i, (xl_name, xl_bytes, method, n_rows) in enumerate(results):
            with cols[i%3]:
                st.markdown(f"""<div class="result-card">
                  <div style="font-family:'Playfair Display',serif;font-weight:700;color:#001a5c;font-size:15px;margin-bottom:4px;">📊 {xl_name}</div>
                  <div style="color:#5a7099;font-size:13px;font-style:italic;">{method} · {n_rows} rows · {len(xl_bytes)//1024:,} KB</div></div>""", unsafe_allow_html=True)
                st.download_button("⬇️ Download", data=xl_bytes, file_name=xl_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_pdf_{i}", use_container_width=True)
    if errors:
        st.error(f"❌ {len(errors)} file(s) failed:")
        for fname, err in errors:
            st.markdown(f"""<div class="fbc-fail-card">
              <div class="fbc-fail-title">{fname}</div>
              <div class="fbc-fail-body">{err}</div></div>""", unsafe_allow_html=True)
    if results:
        st.markdown("""<div class="next-steps">
          <div class="next-steps-title">💡 Next Step — Upload into the DCF Model</div>
          <div class="next-steps-body">
            <b>1.</b> Download and open the Excel.<br>
            <b>2.</b> Verify totals — use the <b>Raw Text</b> sheet for anything that looks off.<br>
            <b>3.</b> Digital PDFs are split into named sheets: <b>Income Statement</b> /
               <b>Balance Sheet</b> / <b>Cash Flow</b> (scanned PDFs use Page 1, Page 2, …).<br>
            <b>4.</b> For 4-column statements (Company + Group), columns are:
               <b>Co. 2025 | Co. 2024 | Gr. 2025 | Gr. 2024</b> — customisable in Options above.<br>
            <b>5.</b> Head to <b>📊 DCF Model</b> and upload.
          </div></div>""", unsafe_allow_html=True)

st.markdown('<div class="fbc-footer">Powered by <b>FBC Securities</b> · Investment Research &amp; Valuation Dashboard</div>', unsafe_allow_html=True)
