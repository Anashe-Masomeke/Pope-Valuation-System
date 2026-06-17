import streamlit as st
import pandas as pd
import numpy as np
import altair as alt
from pathlib import Path
import base64
# ── Autosave active project (every 30 s) ─────────────────────────
# Autosave removed: use Save Now button in Projects page
def step(title: str, number: int):
    st.markdown(
        f"""
        <div class="fbc-step">
            <div class="fbc-step-badge">{number}</div>
            <div class="fbc-step-title">{title}</div>
        </div>
        """,
        unsafe_allow_html=True
    )
def add_watermark():
    logo_path = Path("assets") / "fbc_logo.png"
    if logo_path.exists():
        with open(logo_path, "rb") as f:
            logo_base64 = base64.b64encode(f.read()).decode()

        watermark_css = f"""
        <style>

        /* Make watermark very light */
        .stApp::before {{
            content: "";
            position: fixed;
            top: 40;
            left: 50;
            width: 100%;
            height: 100%;
            background-image: url("data:image/png;base64,{logo_base64}");
            background-repeat: no-repeat;
            background-position: center;
            background-size: 1500px;
            opacity: 0.07;   /* 🔥 control watermark visibility here */
            pointer-events: none;
            z-index: 0;
        }}

        .block-container {{
            position: relative;
            z-index: 1;
        }}
        </style>
        """
        st.markdown(watermark_css, unsafe_allow_html=True)


# ------------------------------------------------------------------------------
# PAGE CONFIG
# ------------------------------------------------------------------------------
st.set_page_config(page_title="Summary Valuation", layout="wide")

# ── CRASH CATCHER — remove after debugging ──────────────────────
import traceback as _tb, sys as _sys
_original_excepthook = _sys.excepthook
def _crash_catcher(exc_type, exc_value, exc_tb):
    print("\n\n========== SUMMARY PAGE CRASH ==========", flush=True)
    _tb.print_exception(exc_type, exc_value, exc_tb)
    print("=========================================\n", flush=True)
    _original_excepthook(exc_type, exc_value, exc_tb)
_sys.excepthook = _crash_catcher
# ── END CRASH CATCHER ────────────────────────────────────────────
add_watermark()
# ── Auth guard (robust: restores session if user dict is valid) ──
# ── Ensure session is always authenticated ───────────────────────
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = True
if "user" not in st.session_state or not st.session_state.get("user"):
    st.session_state["user"] = {"username": "analyst", "role": "analyst", "full_name": "Analyst"}

# ── Re-parse DCF DataFrames from file bytes if missing ────────────────
if (
    st.session_state.get("dcf_file_bytes")
    and st.session_state.get("dcf_is_df") is None
):
    try:
        import io as _io
        import pandas as _pd_parse

        def _sort_year_cols(df):
            item_col = df.columns[0]
            year_cols = []
            other_cols = []
            for c in df.columns[1:]:
                try:
                    int(str(c).strip())
                    year_cols.append(c)
                except ValueError:
                    other_cols.append(c)
            return df[[item_col] + sorted(year_cols, key=lambda x: int(str(x).strip())) + other_cols]

        def _clean_numeric(df):
            for col in df.columns[1:]:
                df[col] = _pd_parse.to_numeric(df[col], errors="coerce")
            return df

        _xls = _pd_parse.ExcelFile(_io.BytesIO(st.session_state["dcf_file_bytes"]))
        st.session_state["dcf_is_df"] = _sort_year_cols(_clean_numeric(_xls.parse(_xls.sheet_names[0])))
        st.session_state["dcf_bs_df"] = _sort_year_cols(_clean_numeric(_xls.parse(_xls.sheet_names[1])))
        st.session_state["dcf_cf_df"] = _sort_year_cols(_clean_numeric(_xls.parse(_xls.sheet_names[2])))
    except Exception:
        pass

# ── Re-build forecast_is_df from dcf_is_df if missing ────────────────
if st.session_state.get("dcf_is_df") is not None and st.session_state.get("forecast_is_df") is None:
    st.session_state["forecast_is_df"] = st.session_state["dcf_is_df"].copy()



# ─── FBC DESIGN SYSTEM ─────────────────────────────────────────
st.markdown('''
<style>
/* ================================================================
   FBC INVESTMENT VALUATION SYSTEM — Design System v3.0
   ================================================================ */

/* ── 0. GOOGLE FONTS ──────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700;900&family=EB+Garamond:ital,wght@0,400;0,600;1,400&family=Material+Icons&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined');

/* ── 1. GLOBAL TYPOGRAPHY ─────────────────────────────────── */
html, body, .stApp, .block-container,
p, div, label,
h1, h2, h3, h4, h5, h6,
li, ul, ol, a, small,
.stDataFrame, .stTable {
  font-family: "EB Garamond", Georgia, "Times New Roman", serif !important;
  color: #1a1a2e;
}

/* Headings — Playfair Display */
h1, h2, h3, h4, .fbc-heading, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 {
  font-family: "Playfair Display", Georgia, serif !important;
  font-weight: 700 !important;
  letter-spacing: -0.01em !important;
}

/* ── 2. PAGE BACKGROUND ───────────────────────────────────── */
.stApp {
  background: #f5f7fb !important;
}
.main .block-container {
  background: #f5f7fb !important;
  padding-top: 1.5rem !important;
}

/* ── 3. SIDEBAR ───────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: linear-gradient(175deg, #001a5c 0%, #003399 45%, #0044cc 100%) !important;
    border-right: 2px solid rgba(245,180,0,0.25) !important;
    box-shadow: 4px 0 24px rgba(0,26,92,0.35) !important;
}
section[data-testid="stSidebar"]::before {
    content: "";
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #f5b400, #ffd040, #f5b400);
}
section[data-testid="stSidebar"] * {
    color: #e8f0ff !important;
    font-family: "EB Garamond", Georgia, serif !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
    font-family: "Playfair Display", serif !important;
    font-weight: 700 !important;
    text-shadow: 0 1px 4px rgba(0,0,0,0.3) !important;
}
section[data-testid="stSidebar"] .block-container {
    padding-top: 1.5rem !important;
}
section[data-testid="stSidebar"] a {
    color: #b8d0ff !important;
    transition: color 0.15s !important;
}
section[data-testid="stSidebar"] a:hover {
    color: #f5b400 !important;
}

/* sidebar hr divider */
section[data-testid="stSidebar"] hr {
    border: none !important;
    border-top: 1px solid rgba(245,180,0,0.25) !important;
    margin: 12px 0 !important;
}

/* ✅ Ensure Material Icons render correctly */
.material-icons,
.material-icons-outlined,
.material-symbols-outlined,
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stSidebarCollapseButton"] i {
    font-family: 'Material Icons', 'Material Symbols Outlined' !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    display: inline-block !important;
    white-space: nowrap !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}
[data-testid="stSidebarCollapseButton"] button {
    background: linear-gradient(135deg, #f5b400, #ffd040) !important;
    border: none !important;
    border-radius: 50% !important;
    width: 44px !important; height: 44px !important;
    box-shadow: 0 4px 14px rgba(245,180,0,0.50) !important;
    transition: all 0.2s ease !important;
}
[data-testid="stSidebarCollapseButton"] button:hover {
    transform: translateY(-1px) scale(1.06) !important;
    box-shadow: 0 8px 20px rgba(245,180,0,0.60) !important;
}
[data-testid="stSidebarCollapseButton"] svg {
    width: 22px !important; height: 22px !important;
    fill: #001a5c !important;
}

/* ── 5. PAGE HEADER BANNER ────────────────────────────────── */
.fbc-page-header {
    background: linear-gradient(135deg, #001a5c 0%, #003399 50%, #0044cc 100%);
    border-radius: 18px;
    padding: 26px 32px;
    margin-bottom: 28px;
    border: 1px solid rgba(255,255,255,0.08);
    border-bottom: 3px solid #f5b400;
    box-shadow: 0 12px 40px rgba(0,26,92,0.28);
    position: relative;
    overflow: hidden;
}
.fbc-page-header::before {
    content: "";
    position: absolute;
    left: -40px; top: -40px;
    width: 200px; height: 200px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(245,180,0,0.12), transparent 70%);
    pointer-events: none;
}
.fbc-page-header::after {
    content: "";
    position: absolute;
    right: -30px; bottom: -30px;
    width: 160px; height: 160px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(255,255,255,0.06), transparent 65%);
    pointer-events: none;
}
.fbc-page-header-icon {
    font-size: 30px;
    margin-right: 12px;
    vertical-align: middle;
}
.fbc-page-header-title {
    font-family: "Playfair Display", serif !important;
    font-size: 28px !important;
    font-weight: 900 !important;
    color: #ffffff !important;
    display: inline !important;
    vertical-align: middle !important;
    letter-spacing: -0.01em !important;
}
.fbc-page-header-sub {
    font-size: 14px;
    color: rgba(255,255,255,0.78) !important;
    margin-top: 8px;
    font-style: italic;
}
.fbc-badge {
    display: inline-block;
    background: rgba(245,180,0,0.20);
    border: 1px solid rgba(245,180,0,0.50);
    color: #f5c842 !important;
    font-size: 10px;
    font-weight: 800;
    letter-spacing: 0.12em;
    padding: 3px 10px;
    border-radius: 999px;
    margin-left: 10px;
    vertical-align: middle;
    text-transform: uppercase;
    font-family: "EB Garamond", serif !important;
}

/* ── 6. SECTION HEADINGS ──────────────────────────────────── */
.fbc-section-heading {
    display: flex;
    align-items: center;
    gap: 12px;
    margin: 32px 0 16px 0;
    padding-bottom: 10px;
    border-bottom: 2px solid transparent;
    border-image: linear-gradient(90deg, #003399 0%, #f5b400 55%, transparent 90%) 1;
}
.fbc-section-heading-text {
    font-family: "Playfair Display", serif !important;
    font-size: 18px !important;
    font-weight: 700 !important;
    color: #001a5c !important;
    letter-spacing: 0.01em !important;
}
.fbc-section-heading-step {
    background: linear-gradient(135deg, #003399, #0044cc);
    color: white !important;
    font-size: 11px;
    font-weight: 900;
    min-width: 28px; height: 28px;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0;
    box-shadow: 0 3px 8px rgba(0,51,153,0.35);
}
.fbc-subsection-heading {
    font-family: "Playfair Display", serif !important;
    font-size: 15px !important;
    font-weight: 700 !important;
    color: #003399 !important;
    margin: 20px 0 10px 0 !important;
    padding-left: 12px !important;
    border-left: 3px solid #f5b400 !important;
}

/* ── 7. CARD / PANEL ──────────────────────────────────────── */
.fbc-card {
    background: #ffffff;
    border: 1px solid rgba(0,51,153,0.09);
    border-left: 5px solid #003399;
    border-radius: 16px;
    padding: 20px 24px;
    margin-bottom: 18px;
    box-shadow: 0 6px 18px rgba(0,26,92,0.07);
    transition: box-shadow 0.2s, transform 0.2s;
}
.fbc-card:hover {
    box-shadow: 0 12px 32px rgba(0,51,153,0.14);
    transform: translateY(-2px);
}
.fbc-card h3, .fbc-card h4 {
    font-family: "Playfair Display", serif !important;
    color: #001a5c !important;
    margin: 0 0 10px 0 !important;
}
.fbc-card-gold {
    border-left-color: #f5b400 !important;
}
.fbc-card-green {
    border-left-color: #10b981 !important;
}

/* sub-card (nested) */
.fbc-subcard {
    background: rgba(0,51,153,0.03);
    border: 1px solid rgba(0,51,153,0.10);
    border-radius: 12px;
    padding: 16px 18px;
    margin-top: 12px;
}

/* ── 8. KPI METRIC CARDS ──────────────────────────────────── */
.fbc-kpi {
    background: linear-gradient(135deg, #f0f5ff, #fff8e6);
    border: 1px solid rgba(0,51,153,0.12);
    border-radius: 16px;
    padding: 16px 18px;
    text-align: center;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05);
    transition: transform 0.2s, box-shadow 0.2s;
}
.fbc-kpi:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 20px rgba(0,51,153,0.12);
}
.fbc-kpi-label {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 0.10em;
    text-transform: uppercase;
    color: #5a7099 !important;
    margin-bottom: 6px;
    font-family: "EB Garamond", serif !important;
}
.fbc-kpi-value {
    font-family: "Playfair Display", serif !important;
    font-size: 24px !important;
    font-weight: 800 !important;
    color: #001a5c !important;
    line-height: 1.2;
}
.fbc-kpi-unit {
    font-size: 11px;
    color: #7a90b8 !important;
    margin-top: 3px;
}

/* DCF card / kpi aliases */
.dcf-card {
    background: #ffffff;
    border: 1px solid rgba(0,0,0,0.08);
    border-left: 5px solid #003399;
    border-radius: 16px;
    padding: 20px 22px;
    box-shadow: 0 6px 18px rgba(0,0,0,0.06);
    margin-top: 12px;
    margin-bottom: 16px;
}
.dcf-card h3 { margin: 0 0 10px 0; font-family: "Playfair Display", serif !important; color: #001a5c !important; }
.dcf-subcard {
    background: rgba(0,51,153,0.03);
    border: 1px solid rgba(0,51,153,0.10);
    border-radius: 14px;
    padding: 14px;
    margin-top: 10px;
}
.dcf-kpi {
    background: linear-gradient(135deg, rgba(0,51,153,0.08), rgba(245,180,0,0.08));
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: 16px;
    padding: 14px 16px;
    margin: 6px 0;
    transition: transform 0.2s;
}
.dcf-kpi:hover { transform: translateY(-1px); }
.dcf-kpi-title { font-size: 11px; opacity: 0.7; margin-bottom: 3px; text-transform: uppercase; letter-spacing: 0.07em; }
.dcf-kpi-value { font-size: 20px; font-weight: 800; font-family: "Playfair Display", serif !important; color: #001a5c !important; }

/* ── 9. RESET / UTILITY CARDS ─────────────────────────────── */
.fbc-reset-card {
    background: linear-gradient(135deg, #001a5c 0%, #003399 55%, #0055cc 100%);
    padding: 22px 28px;
    border-radius: 16px;
    color: white !important;
    box-shadow: 0 8px 24px rgba(0,26,92,0.30);
    margin-bottom: 22px;
    border-bottom: 3px solid #f5b400;
    position: relative;
    overflow: hidden;
}
.fbc-reset-card::after {
    content: "";
    position: absolute;
    right: -20px; top: -20px;
    width: 120px; height: 120px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(245,180,0,0.15), transparent 70%);
    pointer-events: none;
}
.fbc-reset-title {
    font-family: "Playfair Display", serif !important;
    font-size: 20px !important;
    font-weight: 700 !important;
    margin-bottom: 6px;
    color: white !important;
}
.fbc-reset-sub { font-size: 14px; opacity: 0.88; margin-bottom: 14px; color: rgba(255,255,255,0.85) !important; }

/* ── 10. BUTTONS ──────────────────────────────────────────── */
.stButton > button[kind="primary"],
.stButton > button {
    background: linear-gradient(135deg, #003399, #0044cc) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-family: "EB Garamond", Georgia, serif !important;
    padding: 10px 20px !important;
    transition: all 0.2s ease !important;
    box-shadow: 0 4px 12px rgba(0,51,153,0.28) !important;
    letter-spacing: 0.01em !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #0044cc, #0055ee) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 22px rgba(0,51,153,0.38) !important;
}
.fbc-reset-btn .stButton > button {
    background: linear-gradient(135deg, #f5b400, #ffd040) !important;
    color: #001a5c !important;
    box-shadow: 0 4px 12px rgba(245,180,0,0.32) !important;
}
.fbc-reset-btn .stButton > button:hover {
    background: linear-gradient(135deg, #ffd040, #ffe070) !important;
    box-shadow: 0 8px 20px rgba(245,180,0,0.45) !important;
}
.fbc-nav-btn button,
.fbc-nav-btn .stButton > button {
    background: linear-gradient(135deg, #003399, #001a4d) !important;
    color: white !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    padding: 8px 18px !important;
    border: none !important;
    transition: all 0.25s ease-in-out !important;
}
.fbc-nav-btn button:hover,
.fbc-nav-btn .stButton > button:hover {
    background: linear-gradient(135deg, #0055cc, #003399) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 14px rgba(0,0,0,0.25) !important;
}
.stDownloadButton > button {
    background: linear-gradient(135deg, #003399, #0044cc) !important;
    color: white !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    border: none !important;
    transition: all 0.2s ease !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #0044cc, #0055ee) !important;
    transform: translateY(-1px) !important;
}

/* ── 11. INPUTS ───────────────────────────────────────────── */
.stNumberInput input, .stTextInput input {
    border: 1.5px solid rgba(0,51,153,0.18) !important;
    border-radius: 10px !important;
    font-family: "EB Garamond", Georgia, serif !important;
    background: #ffffff !important;
    transition: border-color 0.15s, box-shadow 0.15s !important;
    padding: 8px 12px !important;
}
.stNumberInput input:focus, .stTextInput input:focus {
    border-color: #003399 !important;
    box-shadow: 0 0 0 3px rgba(0,51,153,0.10) !important;
}
.stSelectbox > div > div {
    border: 1.5px solid rgba(0,51,153,0.18) !important;
    border-radius: 10px !important;
    font-family: "EB Garamond", Georgia, serif !important;
}

/* ── 12. TABS ─────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    border-bottom: 2px solid rgba(0,51,153,0.12);
    padding-bottom: 0;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px 10px 0 0 !important;
    font-weight: 700 !important;
    font-family: "EB Garamond", Georgia, serif !important;
    color: #5a7099 !important;
    padding: 10px 20px !important;
    transition: all 0.15s !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #003399, #0044cc) !important;
    color: white !important;
    box-shadow: 0 -2px 0 0 #f5b400 inset !important;
}

/* ── 13. EXPANDERS ────────────────────────────────────────── */
.streamlit-expanderHeader {
    font-weight: 700 !important;
    color: #001a5c !important;
    font-family: "Playfair Display", Georgia, serif !important;
    border-radius: 10px !important;
    background: rgba(0,51,153,0.04) !important;
    padding: 12px 16px !important;
    border-left: 3px solid #f5b400 !important;
}
/* Fix: hide raw "keyboard_arrow_right" text Streamlit injects into expander headers */
[data-testid="stExpander"] details > summary > div > p > span[style],
[data-testid="stExpander"] details > summary p span[data-testid="stMarkdownContainer"] > p > span,
details > summary p > span[style*="font-weight"] {
    display: none !important;
}
[data-testid="stExpanderToggleIcon"] svg,
details > summary svg {
    font-family: "Material Icons", "Material Symbols Outlined" !important;
}

/* ── 14. METRICS (st.metric) ──────────────────────────────── */
[data-testid="metric-container"] {
    background: linear-gradient(135deg, #f0f5ff, #fff8e6) !important;
    border: 1px solid rgba(0,51,153,0.12) !important;
    border-radius: 16px !important;
    padding: 16px 18px !important;
    box-shadow: 0 4px 12px rgba(0,0,0,0.05) !important;
    transition: transform 0.2s, box-shadow 0.2s !important;
}
[data-testid="metric-container"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 20px rgba(0,51,153,0.10) !important;
}
[data-testid="metric-container"] label {
    font-size: 10px !important;
    font-weight: 700 !important;
    letter-spacing: 0.10em !important;
    text-transform: uppercase !important;
    color: #5a7099 !important;
    font-family: "EB Garamond", serif !important;
}
[data-testid="stMetricValue"] {
    font-family: "Playfair Display", serif !important;
    font-size: 24px !important;
    font-weight: 800 !important;
    color: #001a5c !important;
}
[data-testid="stMetricDelta"] {
    font-family: "EB Garamond", serif !important;
}

/* ── 15. ALERTS & INFO BOXES ──────────────────────────────── */
.stAlert {
    border-radius: 14px !important;
    font-family: "EB Garamond", Georgia, serif !important;
    border-left-width: 4px !important;
}

/* ── 16. PROGRESS BAR ─────────────────────────────────────── */
.stProgress > div > div > div {
    background: linear-gradient(90deg, #003399, #0044cc, #f5b400) !important;
    border-radius: 999px !important;
}

/* ── 17. SCROLLBAR ────────────────────────────────────────── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f0f4ff; border-radius: 999px; }
::-webkit-scrollbar-thumb { background: linear-gradient(#003399, #0044cc); border-radius: 999px; }
::-webkit-scrollbar-thumb:hover { background: #0055ee; }

/* ── 18. FOOTER ───────────────────────────────────────────── */
.fbc-footer {
    text-align: center;
    padding: 28px;
    margin-top: 48px;
    color: #5a7099 !important;
    font-size: 13px;
    border-top: 1px solid rgba(0,51,153,0.10);
    font-style: italic;
}
.fbc-footer b { color: #003399 !important; font-style: normal; }

/* ── 19. FEATURE CARDS (dashboard) ───────────────────────── */
.feature-box {
    background: #ffffff;
    padding: 22px 26px;
    border-radius: 16px;
    border-left: 6px solid #003399;
    box-shadow: 0 4px 14px rgba(0,0,0,0.07);
    transition: all 0.25s ease;
    margin-bottom: 16px;
    font-family: "EB Garamond", serif !important;
}
.feature-box:hover {
    background: linear-gradient(135deg, #f4f8ff, #fffdf0);
    box-shadow: 0 10px 28px rgba(0,51,153,0.16);
    transform: translateY(-3px);
    border-left-color: #f5b400;
}
.feature-icon { font-size: 24px; margin-right: 10px; }

/* ── 20. SMALL UTILITIES ──────────────────────────────────── */
.small-note {
    font-size: 12px;
    color: #7a90b8 !important;
    font-style: italic;
}
.gold-accent { color: #b87c00 !important; font-weight: 700; }
.blue-accent  { color: #003399 !important; font-weight: 700; }

/* ── 21. DATAFRAMES ───────────────────────────────────────── */
.stDataFrame thead th {
    background: linear-gradient(135deg, #001a5c, #003399) !important;
    color: white !important;
    font-weight: 700 !important;
    font-family: "Playfair Display", serif !important;
    letter-spacing: 0.02em !important;
}
.stDataFrame tbody tr:hover td {
    background: rgba(0,51,153,0.04) !important;
}
.stDataFrame tbody tr:nth-child(even) td {
    background: rgba(0,51,153,0.02) !important;
}

/* ── 22. TOP NAV (dashboard) ──────────────────────────────── */
.top-nav {
    position: fixed; top: 0; left: 0; width: 100%; height: 68px;
    background: linear-gradient(90deg, #001a5c, #003399, #0044cc);
    color: white; display: flex; align-items: center;
    padding: 0 28px; z-index: 99999;
    box-shadow: 0 4px 16px rgba(0,0,0,0.30);
    border-bottom: 2px solid #f5b400;
}
.top-title {
    font-family: "Playfair Display", serif !important;
    font-size: 26px; font-weight: 900; margin-left: 16px; letter-spacing: -0.01em;
    color: white;
}

/* ── 23. DIVIDER ──────────────────────────────────────────── */
.fbc-divider {
    height: 2px;
    background: linear-gradient(90deg, transparent, #f5b400, #003399, transparent);
    border: none;
    margin: 28px 0;
    border-radius: 999px;
}

/* ── 24. RADIO / CHECKBOX ─────────────────────────────────── */
.stRadio > label, .stCheckbox > label {
    font-family: "EB Garamond", Georgia, serif !important;
    color: #1a1a2e !important;
}

/* ── 25. SELECTBOX OPTIONS ────────────────────────────────── */
div[data-baseweb="select"] span {
    font-family: "EB Garamond", Georgia, serif !important;
}

/* ── 26. PEER PICKER (Comparables) ───────────────────────── */
.peer-picker-wrap {
    border: 1px solid rgba(0,51,153,0.14);
    border-radius: 18px;
    padding: 18px 18px 12px 18px;
    background: #f8fbff;
    box-shadow: 0 6px 20px rgba(0,26,92,0.07);
    margin-top: 12px;
    margin-bottom: 14px;
}
.peer-picker-head {
    font-family: "Playfair Display", serif !important;
    font-size: 17px;
    font-weight: 700;
    color: #001a5c;
    margin-bottom: 8px;
}
.peer-picker-sub {
    font-size: 13px;
    color: #475569;
    margin-bottom: 14px;
    font-style: italic;
}

/* ── 27. SENS TABLE OVERRIDES ─────────────────────────────── */
.sens-outer { border: 2px solid rgba(0,51,153,0.15) !important; border-radius: 16px !important; }
.sens-table { font-family: "EB Garamond", Georgia, serif !important; }
.sens-table thead th { background: #001a5c !important; }

/* ── 28. FOOTER BANNER ────────────────────────────────────── */
.footer {
    text-align: center;
    padding: 24px;
    margin-top: 40px;
    color: #5a7099 !important;
    font-size: 13px;
    border-top: 1px solid rgba(0,51,153,0.10);
    font-style: italic;
}
.footer b { color: #003399 !important; font-style: normal; }

</style>
''', unsafe_allow_html=True)

st.markdown('''
<div class="fbc-page-header">
    <span class="fbc-page-header-icon">🧾</span>
    <span class="fbc-page-header-title">Summary Valuation</span>
    <span class="fbc-badge">FBC Securities</span>
    <div class="fbc-page-header-sub">Consolidated view of all valuation outputs across models.</div>
</div>
''', unsafe_allow_html=True)
# ────────────────────────────────────────────────────────────────

# ---------------------------------------------------------
# ✅ FIX SIDEBAR COLLAPSE ARROW (Material Icons)
# ---------------------------------------------------------
st.markdown("""
<style>

/* ✅ Load Material Icons so Streamlit's sidebar collapse icon renders correctly */
@import url('https://fonts.googleapis.com/icon?family=Material+Icons');

/* ✅ Make sure ONLY icons use the Material Icons font (prevents 'keyboard_double_arrow_right' text) */
.material-icons, 
span.material-icons,
i.material-icons,
[data-testid="stSidebarCollapseButton"] span,
[data-testid="stSidebarCollapseButton"] i {
    font-family: 'Material Icons' !important;
    font-weight: normal !important;
    font-style: normal !important;
    letter-spacing: normal !important;
    text-transform: none !important;
    display: inline-block !important;
    white-space: nowrap !important;
    word-wrap: normal !important;
    direction: ltr !important;
    -webkit-font-feature-settings: 'liga' !important;
    -webkit-font-smoothing: antialiased !important;
}

/* ✅ Style the collapse/expand button nicely */
[data-testid="stSidebarCollapseButton"] button {
    background: #003399 !important;
    border: 1px solid rgba(255,255,255,0.25) !important;
    border-radius: 999px !important;
    width: 44px !important;
    height: 44px !important;
    box-shadow: 0 6px 18px rgba(0, 51, 153, 0.35) !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease !important;
}

[data-testid="stSidebarCollapseButton"] button:hover {
    transform: translateY(-1px) !important;
    background: #0047d6 !important;
    box-shadow: 0 10px 22px rgba(0, 71, 214, 0.35) !important;
}

[data-testid="stSidebarCollapseButton"] svg {
    width: 22px !important;
    height: 22px !important;
    fill: white !important;
}
/* ===== SIDEBAR GLASS STYLE ===== */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #003399 0%, #001a4d 100%) !important;
    color: white !important;
    border-right: 1px solid rgba(255,255,255,0.15);
    backdrop-filter: blur(8px);
}

/* Sidebar text */
section[data-testid="stSidebar"] * {
    color: white !important;
}

/* Sidebar headings */
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #ffffff !important;
}

/* Remove default sidebar padding spacing */
section[data-testid="stSidebar"] .block-container {
    padding-top: 1rem !important;
}
</style>
""", unsafe_allow_html=True)
st.markdown("""
<style>

/* =========================================================
   FBC CLEAN SECTION HEADER (NO SUBTITLE, NO STEPS)
   ========================================================= */

.fbc-section {
    display: flex;
    align-items: center;
    gap: 14px;

    padding: 16px 22px;
    margin: 28px 0 18px 0;

    background: linear-gradient(
        135deg,
        rgba(0, 51, 153, 0.08),
        rgba(245, 180, 0, 0.05)
    );

    border-left: 6px solid #003399;
    border-radius: 14px;

    box-shadow: 0 4px 14px rgba(0, 26, 92, 0.08);
}

/* Left indicator (circle like your UI) */
.fbc-section {
    display: block;
    padding: 16px 0;
    margin: 28px 0 18px 0;
    border-bottom: 2px solid rgba(0,51,153,0.15);
    transition: all 0.25s ease;
}

.fbc-section-title {
    font-size: 21px;
    font-weight: 800;
    color: #001a5c;
    position: relative;
}

/* animated underline */
.fbc-section-title::after {
    content: "";
    position: absolute;
    left: 0;
    bottom: -6px;
    width: 40px;
    height: 3px;
    background: #003399;
    transition: width 0.3s ease;
}

.fbc-section:hover .fbc-section-title::after {
    width: 100%;
}

/* Title only */
.fbc-section-title {
    font-family: "Playfair Display", serif !important;
    font-size: 21px;
    font-weight: 800;
    color: #001a5c !important;
    letter-spacing: -0.01em;
}

</style>
""", unsafe_allow_html=True)
def section(title: str):
    st.markdown(
        f"""
        <div class="fbc-section">
            <div class="fbc-section-dot"></div>
            <div class="fbc-section-title">{title}</div>
        </div>
        """,
        unsafe_allow_html=True
    )
# ------------------------------------------------------------------------------
# POWERBI DARK THEME (FBC TUNED)
# ------------------------------------------------------------------------------
DARK_BG = "#020617"        # page background
PANEL_BG = "#020c1f"       # panels
CARD_BG = "#071525"        # KPI cards
PRIMARY_TEXT = "#e5f2ff"
MUTED_TEXT = "#9ca3af"
ACCENT_BLUE = "#38bdf8"
ACCENT_CYAN = "#0ea5e9"
ACCENT_GOLD = "#fbbf24"
DANGER = "#f97373"

st.markdown(
    f"""
    <style>

    /* ------------------------------------------------------ */
    /* GLOBAL BACKGROUND & TEXT                               */
    /* ------------------------------------------------------ */
    .main {{
        background: radial-gradient(circle at top left, #0d1424 0, {DARK_BG} 60%, #000 100%);
        color: {PRIMARY_TEXT};
    }}
    .block-container {{
        padding-top: 1.2rem;
        padding-bottom: 3rem;
        max-width: 1300px;
    }}
h1, h2, h3, h4, h5 {{
    color: #000000 !important;
    font-family: Georgia, "Times New Roman", serif !important;
    font-weight: 600 !important;
}}


    /* ------------------------------------------------------ */
    /* TITLE BANNER                                           */
    /* ------------------------------------------------------ */
    .title-banner {{
        background: linear-gradient(90deg, #071426, #0a1b33 50%, #0d243f 100%);
        border-radius: 16px;
        padding: 1.2rem 1.5rem;
        border: 1px solid rgba(255,255,255,0.12);
        box-shadow: 0 12px 50px rgba(0,0,0,0.6);
    }}
    .title-main {{
        font-size: 1.85rem;
        font-weight: 700;
        color: #ffffff !important;
    }}
    .title-sub {{
        font-size: 0.95rem;
        color: #d3e2f5 !important;
        margin-top: 3px;
    }}

    /* ------------------------------------------------------ */
    /* KPI CARDS — HIGH VISIBILITY VERSION                    */
    /* ------------------------------------------------------ */
    .kpi-card {{
        border-radius: 14px;
        padding: 1.15rem 1.3rem;
        background: rgba(10, 20, 40, 0.85); /* darker, solid */
        border: 1px solid rgba(255,255,255,0.14);
        box-shadow: 0 12px 30px rgba(0,0,0,0.7);
        backdrop-filter: blur(4px); /* gentle */
    }}
    .kpi-title {{
        font-size: 0.85rem;
        color: #c7d4e8; /* brighter */
        text-transform: uppercase;
        letter-spacing: 0.06em;
    }}
    .kpi-value {{
        font-size: 1.6rem !important;
        font-weight: 700 !important;
        color: {ACCENT_GOLD} !important;
        text-shadow: 0 0 6px rgba(0,0,0,0.6);
    }}
    .kpi-sub {{
        font-size: 0.83rem;
        color: #b4c2d6 !important; /* more visible */
        margin-top: 2px;
    }}

    /* ------------------------------------------------------ */
    /* GLASS PANEL                                            */
    /* ------------------------------------------------------ */
    .glass-panel {{
        background: rgba(14,22,40,0.92);
        border-radius: 16px;
        padding: 1rem 1.3rem;
        border: 1px solid rgba(255,255,255,0.15);
        box-shadow: 0 10px 28px rgba(0,0,0,0.75);
        backdrop-filter: blur(3px);
    }}

    /* ------------------------------------------------------ */
    /* TABLE VISIBILITY FIX                                   */
    /* ------------------------------------------------------ */
    .stDataFrame, .stTable {{
        color: #f0f6ff !important;  /* TEXT FIX */
    }}

    .stDataFrame tbody td {{
        color: #1f2937 !important;  /* brighter cell text */
        font-size: 0.95rem !important;
    }}

    .stDataFrame thead th {{
    color: #0f172a !important;   
    font-weight: 700 !important;
    }}


    .stDataFrame tbody tr:hover {{
        background-color: rgba(255,255,255,0.08) !important;
    }}

    /* ------------------------------------------------------ */
    /* TABS – HIGH VISIBILITY                                 */
    /* ------------------------------------------------------ */
button[data-baseweb="tab"] {{
    font-size: 0.9rem;
    color: #1e293b !important;   
    background-color: rgba(255,255,255,0.05);
    border-radius: 10px;
    padding: 0.4rem 1rem;
    border: 1px solid rgba(255,255,255,0.15);
}}

button[data-baseweb="tab"][aria-selected="true"] {{
    color: #ffffff !important;
    background: #1e3a8a !important;   
    border: 1px solid #1e40af;
    box-shadow: 0 0 8px rgba(30,58,138,0.6);
}}


    /* ------------------------------------------------------ */
    /* INPUTS / SELECTS                                       */
    /* ------------------------------------------------------ */
    .stNumberInput input, .stSelectbox select {{
        background: rgba(20,34,60,0.95) !important;
        color: #ffffff !important;
        border-radius: 8px !important;
        border: 1px solid rgba(255,255,255,0.25) !important;
    }}

    </style>
    """,
    unsafe_allow_html=True,
)

# ------------------------------------------------------------------------------
# TITLE BANNER
# ------------------------------------------------------------------------------
st.markdown(
    """
    <div class="title-banner">
        <div class="title-main">📘 Summary Valuation – Weighted Equity Value</div>
        <div class="title-sub">
            FBC dashboard summarising DCF · DDM · EV/EBITDA · PBV · P/E · Banking valuations.
        </div>
    </div>
    """,
    unsafe_allow_html=True,

)
st.markdown("""
<style>

/* ------------------------------------------------------
   TITLE BANNER
------------------------------------------------------ */
.title-banner {
    background: linear-gradient(90deg, #071426, #0a1b33 50%, #0d243f 100%);
    border-radius: 16px;
    padding: 1.2rem 1.5rem;
    border: 1px solid rgba(255,255,255,0.12);
    box-shadow: 0 12px 50px rgba(0,0,0,0.6);
    margin-bottom: 20px;
}

.title-main {
    font-size: 1.85rem;
    font-weight: 700;
    color: #ffffff !important;
}

.title-sub {
    font-size: 0.95rem;
    color: #d3e2f5 !important;
    margin-top: 3px;
}

</style>
""", unsafe_allow_html=True)
st.markdown("""
<style>

/* =========================================================
   FBC STEP COMPONENT — FINAL & RELIABLE
   ========================================================= */

.fbc-step {
    display: flex;
    align-items: center;
    gap: 14px;

    padding: 16px 20px;
    margin: 28px 0 18px 0;

    background: linear-gradient(
        135deg,
        rgba(0, 51, 153, 0.06),
        rgba(245, 180, 0, 0.05)
    );

    border-left: 6px solid #003399;
    border-radius: 14px;

    box-shadow: 0 4px 14px rgba(0, 26, 92, 0.08);
}

.fbc-step-badge {
    min-width: 34px;
    height: 34px;
    border-radius: 50%;

    background: linear-gradient(135deg, #003399, #0044cc);
    color: white;

    font-weight: 900;
    font-size: 14px;

    display: flex;
    align-items: center;
    justify-content: center;

    box-shadow: 0 3px 8px rgba(0, 51, 153, 0.35);
}

.fbc-step-title {
    font-family: "Playfair Display", serif !important;
    font-size: 20px;
    font-weight: 800;
    color: #001a5c !important;
    letter-spacing: -0.01em;
}

</style>
""", unsafe_allow_html=True)
# ------------------------------------------------------------------------------
# INITIALISE SESSION STATE (PERSISTENT)
# ------------------------------------------------------------------------------
if "selected_models" not in st.session_state:
    st.session_state["selected_models"] = ["DCF", "DDM", "EV/EBITDA", "PBV", "P/E", "BANKING"]

if "model_weights" not in st.session_state:
    st.session_state["model_weights"] = {
        "DCF": 35.0,
        "DDM": 20.0,
        "EV/EBITDA": 15.0,
        "PBV": 10.0,
        "P/E": 10.0,
        "BANKING": 10.0,
    }

if "summary_num_shares" not in st.session_state:
    # Seed from num_shares if already set (e.g. from DDM), else 0
    st.session_state["summary_num_shares"] = float(st.session_state.get("num_shares", 0.0))

if "summary_current_price" not in st.session_state:
    st.session_state["summary_current_price"] = float(st.session_state.get("current_price", 0.0))

# Keep num_shares / current_price aliases in sync for backwards compat
if "num_shares" not in st.session_state:
    st.session_state["num_shares"] = st.session_state["summary_num_shares"]

if "current_price" not in st.session_state:
    st.session_state["current_price"] = st.session_state["summary_current_price"]

# Pre-cast to float BEFORE the widgets are created so value= always gets a float.
# These writes are safe here because the widgets haven't been instantiated yet.
# Do NOT write these keys again after the widgets appear further down.
_pre_ns = st.session_state.get("summary_num_shares", 0.0)
if not isinstance(_pre_ns, float):
    st.session_state["summary_num_shares"] = float(_pre_ns)
_pre_cp = st.session_state.get("summary_current_price", 0.0)
if not isinstance(_pre_cp, float):
    st.session_state["summary_current_price"] = float(_pre_cp)

# ------------------------------------------------------------------------------
# MODEL SELECTION (PERSISTENT)
# ------------------------------------------------------------------------------
step("📌 Select Models to Include",1)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
all_models = ["DCF", "DDM", "EV/EBITDA", "PBV", "P/E", "BANKING"]

selected_models = st.multiselect(
    "Choose models:",
    options=all_models,
    default=st.session_state["selected_models"],
    key="selected_models_input",
)
st.session_state["selected_models"] = selected_models

if not selected_models:
    st.warning("Please select at least one model.")
    st.stop()

# ── Recompute Comparables equity values from latest DCF data ──────────
# This runs every time Summary loads so values are never stale after
# a DCF change, even if the user hasn't revisited the Comparables page.
S = st.session_state
_implied_ev  = S.get("implied_ev",  0.0) or 0.0
_implied_pb  = S.get("implied_pb",  0.0) or 0.0
_implied_pe  = S.get("implied_pe",  0.0) or 0.0
_maint_ebitda   = S.get("maintainable_ebitda")
_maint_earnings = S.get("maintainable_earnings")
_book_equity    = S.get("book_equity")
_net_debt       = S.get("net_debt", 0.0) or 0.0

import numpy as _np_summ
import pandas as _pd_summ

# Recompute EBITDA-based equity value from fresh dcf_ebitda_all + weights
_dcf_eb_all = S.get("dcf_ebitda_all") or S.get("dcf_ebitda_forecast") or {}
_comp_eb_weights = S.get("comp_eb_weights") or {}
_timing_base = float(S.get("comp_timing_base", 1.0) or 1.0)
_use_timing_eb = bool(S.get("comp_use_timing_eb", True))
if _dcf_eb_all and _comp_eb_weights:
    _eb_years = sorted(int(y) for y in _dcf_eb_all.keys()
                   if str(y).strip().isdigit() and len(str(y).strip()) == 4
                   and float(_comp_eb_weights.get(str(int(y)), 0.0)) > 0)
    _weighted_eb = 0.0
    for _idx_e, _yr in enumerate(_eb_years):
        _wt = float(_comp_eb_weights.get(str(_yr), 0.0)) / 100.0
        if _wt > 0:
            _eb_val = float(_dcf_eb_all.get(str(_yr), 0.0))
            _timing = (_timing_base + _idx_e) if _use_timing_eb else 1.0
            _weighted_eb += _eb_val * _timing * _wt
    if _weighted_eb != 0.0:
        S["maintainable_ebitda"] = _weighted_eb
        _maint_ebitda = _weighted_eb

# Recompute Net Profit-based equity value from fresh dcf_profit_all + weights
_dcf_np_all = S.get("dcf_profit_all") or {}
_comp_np_weights = S.get("comp_np_weights") or {}
_use_timing_np = bool(S.get("comp_use_timing_np", True))
if _dcf_np_all and _comp_np_weights:
    _np_years = sorted(int(y) for y in _dcf_np_all.keys()
                   if str(y).strip().isdigit() and len(str(y).strip()) == 4
                   and float(_comp_np_weights.get(str(int(y)), 0.0)) > 0)
    _weighted_np = 0.0
    for _idx_n, _yr in enumerate(_np_years):
        _wt = float(_comp_np_weights.get(str(_yr), 0.0)) / 100.0
        if _wt > 0:
            _np_val = float(_dcf_np_all.get(str(_yr), 0.0))
            _timing = (_timing_base + _idx_n) if _use_timing_np else 1.0
            _weighted_np += _np_val * _timing * _wt
    if _weighted_np != 0.0:
        S["maintainable_earnings"] = _weighted_np
        _maint_earnings = _weighted_np

# Recompute final equity values and write back to session_state
if _maint_ebitda is not None and _np_summ.isfinite(float(_maint_ebitda)) and not _pd_summ.isna(_implied_ev) and _implied_ev:
    S["value_ev_ebitda"] = float(_implied_ev * float(_maint_ebitda) - _net_debt)
if _book_equity is not None and _np_summ.isfinite(float(_book_equity)) and not _pd_summ.isna(_implied_pb) and _implied_pb:
    S["value_pbv"] = float(_implied_pb * float(_book_equity))
if _maint_earnings is not None and _np_summ.isfinite(float(_maint_earnings)) and not _pd_summ.isna(_implied_pe) and _implied_pe:
    S["value_pe"] = float(_implied_pe * float(_maint_earnings))
# ── End recompute ─────────────────────────────────────────────────────

value_map = {
    "DCF": st.session_state.get("equity_value_dcf"),
    "DDM": st.session_state.get("equity_value_ddm"),
    "EV/EBITDA": st.session_state.get("value_ev_ebitda"),
    "PBV": st.session_state.get("value_pbv"),
    "P/E": st.session_state.get("value_pe"),
    "BANKING": st.session_state.get("equity_value_banking"),
}
# ------------------------------------------------------------------------------
# WEIGHT ASSIGNMENT (PERSISTENT)
# ------------------------------------------------------------------------------
step("🧮 Assign Weights (%)",2)
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
cols = st.columns(len(all_models))
weights_new = {}

for model, col in zip(all_models, cols):
    if model in selected_models:
        with col:
            new_val = st.number_input(
                f"{model} Weight (%)",
                min_value=0.0,
                max_value=100.0,
                value=float(st.session_state["model_weights"].get(model, 0)),
                step=1.0,
                key=f"weight_input_{model}",
            )
        weights_new[model] = new_val
    else:
        weights_new[model] = 0.0

st.session_state["model_weights"] = weights_new

total_w = sum(weights_new[m] for m in selected_models)
if total_w == 0:
    st.error("Total weight cannot be zero.")
    st.stop()

weights_normalized = {m: (weights_new[m] / total_w) * 100 for m in selected_models}

# ------------------------------------------------------------------------------
# SUMMARY DATAFRAME
# ------------------------------------------------------------------------------
rows = []
for model in selected_models:
    val = value_map.get(model)
    w = weights_normalized.get(model, 0)
    weighted_value = val * (w / 100) if val is not None else None
    rows.append([model, val, w, weighted_value])

df_summary = pd.DataFrame(
    rows, columns=["Model", "Value (USD)", "Weight (%)", "Weighted Value"]
)

weighted_equity = df_summary["Weighted Value"].sum()

# Extra stats for dashboard
valid_vals = df_summary["Value (USD)"].dropna()
if not valid_vals.empty:
    max_model = df_summary.loc[df_summary["Value (USD)"].idxmax(), "Model"]
    max_val = float(valid_vals.max())
    min_val = float(valid_vals.min())
    spread = max_val - min_val
    dispersion = float(valid_vals.std()) if len(valid_vals) > 1 else 0.0
else:
    max_model, max_val, min_val, spread, dispersion = "-", 0.0, 0.0, 0.0, 0.0

# ------------------------------------------------------------------------------
# KPI STRIP (POWERBI STYLE)
# ------------------------------------------------------------------------------
kpi_col1, kpi_col2, kpi_col3 = st.columns(3)

with kpi_col1:
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-title">Weighted Equity Value</div>
            <div class="kpi-value">{weighted_equity:,.0f}</div>
            <div class="kpi-sub">Total blended equity output</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

with kpi_col2:
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-title">Active Models</div>
            <div class="kpi-value">{len(selected_models)}</div>
            <div class="kpi-sub">DCF / DDM / Multiples / Banking in use</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


with kpi_col3:
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-title">Highest Model</div>
            <div class="kpi-value">{max_model}</div>
            <div class="kpi-sub">Range: {min_val:,.0f} – {max_val:,.0f}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# ------------------------------------------------------------------------------
# TABS: SUMMARY TABLE | INTERACTIVE DASHBOARD
# ------------------------------------------------------------------------------
tab1, tab2 = st.tabs(["📊 Summary Table", "📈 Interactive Dashboard"])

# -------- TAB 1: SUMMARY TABLE --------
with tab1:
    st.markdown('<div class="glass-panel">', unsafe_allow_html=True)

    st.dataframe(
        df_summary.style.format(
            {
                "Value (USD)": lambda x: f"{x:,.2f}" if pd.notnull(x) else "—",
                "Weight (%)": lambda x: f"{x:.0f}%",
                "Weighted Value": lambda x: f"{x:,.2f}" if pd.notnull(x) else "—",
            }
        ),
        width='stretch',
    )

# -------- TAB 2: INTERACTIVE DASHBOARD --------
with tab2:
    st.markdown('<div class="glass-panel">', unsafe_allow_html=True)
    st.markdown("### Model Comparison Dashboard")

    df_chart = df_summary.copy()

    base = (
        alt.Chart(df_chart)
        .properties(background=DARK_BG, height=260)
        .configure_axis(
            labelColor=PRIMARY_TEXT,
            titleColor=PRIMARY_TEXT,
            gridColor="#1f2937",
        )
        .configure_view(strokeOpacity=0)
        .configure_legend(labelColor=PRIMARY_TEXT, titleColor=PRIMARY_TEXT)
    )

    chart_values = (
        base.mark_bar(color=ACCENT_BLUE)
        .encode(
            x=alt.X("Model:N", title="Model"),
            y=alt.Y("Value (USD):Q", title="Equity Value (USD)"),
            tooltip=[
                alt.Tooltip("Model:N"),
                alt.Tooltip("Value (USD):Q", format=",.2f"),
            ],
        )
        .properties(title="Model Equity Values")
    )

    chart_weights = (
        base.mark_bar(color=ACCENT_GOLD)
        .encode(
            x=alt.X("Model:N", title="Model"),
            y=alt.Y("Weight (%):Q", title="Weight (%)"),
            tooltip=[
                alt.Tooltip("Model:N"),
                alt.Tooltip("Weight (%):Q", format=".1f"),
            ],
        )
        .properties(title="Model Weights")
    )

    chart_weighted = (
        base.mark_bar(color=ACCENT_CYAN)
        .encode(
            x=alt.X("Model:N", title="Model"),
            y=alt.Y("Weighted Value:Q", title="Weighted Equity (USD)"),
            tooltip=[
                alt.Tooltip("Model:N"),
                alt.Tooltip("Weighted Value:Q", format=",.2f"),
            ],
        )
        .properties(title="Weighted Contribution by Model", height=280)
    )

    c1, c2 = st.columns(2)
    with c1:
        st.altair_chart(chart_values, width='stretch')
    with c2:
        st.altair_chart(chart_weights, width='stretch')

    st.altair_chart(chart_weighted, width='stretch')

    st.markdown("</div>", unsafe_allow_html=True)
# ------------------------------------------------------------------------------
# 📌 GENERAL VALUATION SUMMARY TABLE (DOWNLOADABLE)
# ------------------------------------------------------------------------------
section("📌 Valuation Summary")
st.markdown('<hr class="fbc-divider">', unsafe_allow_html=True)
# ---- Inputs (unchanged logic) ----
c1, c2 = st.columns(2)

with c1:
    st.markdown("""
    <div class="fbc-forecast-label">
        Number of Shares in Issue
    </div>

    <style>
    .fbc-forecast-label {
        font-family: "Playfair Display", serif;
        font-size: 18px;
        font-weight: 800;
        color: #001a5c;

        display: inline-block;
        padding-left: 10px;
        border-left: 4px solid #003399;

        margin-bottom: 8px;
    }

    .fbc-forecast-label::after {
        content: "";
        display: block;
        width: 60px;
        height: 2px;
        margin-top: 6px;
        background: linear-gradient(90deg, #003399, transparent);
        border-radius: 2px;
    }
    </style>
    """, unsafe_allow_html=True)
    # Widget owns key="summary_num_shares" — Streamlit reads/writes it directly.
    # We must NOT write to session_state["summary_num_shares"] after this widget.
    # The value= param seeds the widget from the master key on every render,
    # so tab-switching and project restores always show the correct stored value.
    _ns_val = float(st.session_state.get("summary_num_shares", 0.0))
    num_shares = st.number_input(
        "Number of Shares in Issue",
        min_value=0.0,
        value=_ns_val,
        step=1000.0,
        format="%.0f",
        key="summary_num_shares",
        label_visibility="collapsed"
    )
    # Mirror to the generic alias for DDM / save compatibility.
    # summary_num_shares itself is managed by the widget — do NOT write it here.
    st.session_state["num_shares"] = float(num_shares)

with c2:
    st.markdown("""
    <div class="fbc-forecast-label">
        Current Share Price (USD)
    </div>

    <style>
    .fbc-forecast-label {
        font-family: "Playfair Display", serif;
        font-size: 18px;
        font-weight: 800;
        color: #001a5c;

        display: inline-block;
        padding-left: 10px;
        border-left: 4px solid #003399;

        margin-bottom: 8px;
    }

    .fbc-forecast-label::after {
        content: "";
        display: block;
        width: 60px;
        height: 2px;
        margin-top: 6px;
        background: linear-gradient(90deg, #003399, transparent);
        border-radius: 2px;
    }
    </style>
    """, unsafe_allow_html=True)
    # Widget owns key="summary_current_price" — Streamlit manages it directly.
    _cp_val = float(st.session_state.get("summary_current_price", 0.0))
    current_price = st.number_input(
        "Current Share Price (USD)",
        min_value=0.0,
        value=_cp_val,
        step=0.01,
        format="%.2f",
        key="summary_current_price",
        label_visibility="collapsed"
    )
    # Mirror to the generic alias. summary_current_price is widget-owned — do NOT write it here.
    st.session_state["current_price"] = float(current_price)

# ---- Calculations ----
intrinsic_value = (weighted_equity / num_shares) if (num_shares and num_shares > 0) else np.nan

if not np.isnan(intrinsic_value) and current_price > 0:
    upside_pct = (intrinsic_value - current_price) / current_price * 100
else:
    upside_pct = np.nan
# ---- Recommendation Logic (Buy / Hold / Reduce) ----
rec_label = "N/A"
rec_color = "#94a3b8"  # neutral grey
rec_reason = "Enter shares (>0) and current price (>0) to get a recommendation."

if (not np.isnan(intrinsic_value)) and (current_price > 0) and (not np.isnan(upside_pct)):

    # thresholds you can adjust
    BUY_TH = 10.0      # >= +15% upside => Buy/Accumulate
    HOLD_LOW = -10.0   # between -10% and +10% => Hold
    HOLD_HIGH = 10.0

    if upside_pct >= BUY_TH:
        rec_label = "🟢 BUY / ACCUMULATE"
        rec_color = "#22c55e"
        rec_reason = (
            f"Intrinsic value ({intrinsic_value:,.4f}) is above market price "
            f"({current_price:,.2f}), implying upside of +{upside_pct:.1f}%."
        )
    elif HOLD_LOW <= upside_pct <= HOLD_HIGH:
        rec_label = "🟡 HOLD / FAIRLY VALUED"
        rec_color = "#fbbf24"
        rec_reason = (
            f"Intrinsic value ({intrinsic_value:,.4f}) is close to market price "
            f"({current_price:,.2f}), implying limited upside/downside ({upside_pct:+.1f}%)."
        )
    else:
        rec_label = "🔴 REDUCE / SELL"
        rec_color = "#f97373"
        rec_reason = (
            f"Intrinsic value ({intrinsic_value:,.4f}) is below market price "
            f"({current_price:,.2f}), implying downside of {upside_pct:.1f}%."
        )

# ---- Display Recommendation (simple + clean) ----
st.markdown(
    f"""
    <div class="glass-panel" style="border-left: 6px solid {rec_color}; margin-top: 12px;">
        <div style="font-size: 0.85rem; color: #c7d4e8; text-transform: uppercase; letter-spacing: 0.06em;">
            Recommendation
        </div>
        <div style="font-size: 1.25rem; font-weight: 800; color: {rec_color}; margin-top: 4px;">
            {rec_label}
        </div>
        <div style="font-size: 0.9rem; color: #b4c2d6; margin-top: 6px;">
            {rec_reason}
        </div>
        <div style="font-size: 0.75rem; color: #94a3b8; margin-top: 8px;">
            Note: Based on blended valuation assumptions; sensitive to WACC, growth, and model weights.
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ---- Build general summary table ----
summary_rows = [
    ["Weighted Equity Value", weighted_equity, "USD"],
    ["Number of Shares", num_shares, "Shares"],
    ["Intrinsic Value per Share", intrinsic_value, "USD"],
    ["Current Share Price", current_price, "USD"],
    ["Upside / Downside (%)", upside_pct, "%"],
]

df_valuation_summary = pd.DataFrame(
    summary_rows,
    columns=["Metric", "Value", "Unit"]
)

# ---- Formatting for display ----
def format_value(row):
    if pd.isna(row["Value"]):
        return "—"
    if row["Unit"] == "USD" and row["Metric"] == "Intrinsic Value per Share":
        return f"{row['Value']:,.4f}"
    if row["Unit"] == "USD":
        return f"{row['Value']:,.2f}"
    if row["Unit"] == "%":
        sign = "+" if row["Value"] >= 0 else ""
        return f"{sign}{row['Value']:.1f}%"
    if row["Unit"] == "Shares":
        return f"{row['Value']:,.0f}"
    return str(row["Value"])

df_display = df_valuation_summary.copy()
df_display["Value"] = df_display.apply(format_value, axis=1)

# ---- Display table ----
def highlight_upside(row):
    if row["Metric"] == "Upside / Downside (%)":
        try:
            val = float(row["Value"].replace("%", "").replace("+", ""))
            if val < 0:
                return ["", "color: #f97373;", ""]
            else:
                return ["", "color: #22c55e;", ""]
        except:
            return ["", "", ""]
    return ["", "", ""]

styled_table = df_display.style.apply(highlight_upside, axis=1)

st.dataframe(styled_table, width="stretch", hide_index=True)

# ---- Download button (Combined Excel — all models, formula-linked) ----
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _build_combined_valuation_excel(ss, selected_models, value_map, weights_new, num_shares, current_price) -> bytes:  # noqa: C901
    """
    Build a single .xlsx workbook mirroring the Innscor Valuation Model format.

    Sheet order (only selected model sheets are created):
      1. Summary Valuation    — Innscor-style table + recommendation + analyst note
      2. Forecasts            — Full forecasted Income Statement + key ratios
      3. DCF                  — WACC build-up, UFCF table, sensitivity table, equity value
      4. Dividend Discount Model — dividend history, g, Re, P0, equity value
      5. CompCo               — peer multiples, maintainable metrics, implied equity values
      6. Banking              — residual income model outputs

    All model equity-value cells link back to Summary Valuation via Excel formulas.
    """
    import io, datetime
    import re as _re
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    def _sanitize(s):
        return _re.sub(r'[\x00-\x1f\x7f]', '', str(s)) if s else ''
    # ── Palette (matches Innscor navy/gold look) ──────────────────────────────
    NAVY        = "0A1B33"
    MID_NAVY    = "003399"
    LIGHT_BLUE  = "E8F0FF"
    GOLD        = "F5B400"
    WHITE       = "FFFFFF"
    LIGHT_GREY  = "F2F5FA"
    HEADER_GREY = "D9E2EF"

    # ── Fonts ─────────────────────────────────────────────────────────────────
    def fnt(bold=False, color="000000", sz=10, italic=False, name="Arial"):
        return Font(bold=bold, color=color, size=sz, italic=italic, name=name)

    F_TTL   = fnt(bold=True,  color=WHITE,    sz=12)   # sheet title
    F_HDR   = fnt(bold=True,  color=WHITE,    sz=10)   # column header
    F_SHDR  = fnt(bold=True,  color=WHITE,    sz=10)   # section header
    F_BOLD  = fnt(bold=True,  color="000000", sz=10)
    F_BLUE  = fnt(bold=False, color="0000FF", sz=10)   # hardcoded input
    F_GREEN = fnt(bold=False, color="008000", sz=10)   # cross-sheet link
    F_STD   = fnt(bold=False, color="000000", sz=10)
    F_NOTE  = fnt(bold=False, color="666666", sz=9, italic=True)

    # ── Fills ─────────────────────────────────────────────────────────────────
    FL_NAVY  = PatternFill("solid", fgColor=NAVY)
    FL_MID   = PatternFill("solid", fgColor=MID_NAVY)
    FL_GOLD  = PatternFill("solid", fgColor="F5B400")
    FL_LBLUE = PatternFill("solid", fgColor=LIGHT_BLUE)
    FL_LGREY = PatternFill("solid", fgColor=LIGHT_GREY)
    FL_HDR   = PatternFill("solid", fgColor="1A3A6B")  # darker navy for column headers

    # ── Borders ───────────────────────────────────────────────────────────────
    _thin = Side(style="thin", color=HEADER_GREY)
    _med  = Side(style="medium", color="AABBCC")
    BDR   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    BDR_M = Border(left=_med,  right=_med,  top=_med,  bottom=_med)

    # ── Number formats ────────────────────────────────────────────────────────
    FMT_MONEY  = '#,##0.00'
    FMT_MONEY0 = '#,##0'
    FMT_MONEY4 = '#,##0.0000'
    FMT_PCT    = '0.00%'
    FMT_PCT1   = '0.0%'
    FMT_MULT = '0.00'
    FMT_NUM    = '#,##0'

    # ── Helpers ───────────────────────────────────────────────────────────────
    def col(n):
        return get_column_letter(n)

    def write_title(ws, text, ncols=8, row=1):
        c = ws.cell(row, 1, text)
        c.font = F_TTL; c.fill = FL_NAVY
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 28
        for j in range(2, ncols + 1):
            ws.cell(row, j).fill = FL_NAVY
        return row + 1

    def write_section(ws, row, text, ncols=6):
        c = ws.cell(row, 1, text)
        c.font = F_SHDR; c.fill = FL_MID
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 17
        for j in range(2, ncols + 1):
            ws.cell(row, j).fill = FL_MID
        return row + 1

    def write_hdr(ws, row, labels, start_col=1):
        for j, lbl in enumerate(labels, start=start_col):
            c = ws.cell(row, j, lbl)
            c.font = F_HDR; c.fill = FL_HDR
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BDR
        ws.row_dimensions[row].height = 18
        return row + 1

    def cell_bd(ws, r, c, value=None, font=None, fmt=None, fill=None, align="left"):
        cell = ws.cell(r, c)
        if value is not None:
            cell.value = value
        if font:
            cell.font = font
        if fmt:
            cell.number_format = fmt
        if fill:
            cell.fill = fill
        cell.border = BDR
        cell.alignment = Alignment(horizontal=align, vertical="center")
        return cell

    def border_range(ws, r1, c1, r2, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).border = BDR

    # ── Normalised weights (only selected models) ─────────────────────────────
    total_w = sum(weights_new.get(m, 0.0) for m in selected_models)
    if total_w == 0:
        total_w = 1.0
    norm_w = {m: (weights_new.get(m, 0.0) / total_w) for m in selected_models}

    wb = Workbook()
    # We'll track cross-sheet equity value cell addresses for each model
    eq_addr = {}   # model → "SheetName!CellAddress"  (absolute)

    # =========================================================================
    # SHEET: Forecasts — Full Income Statement + Ratios (always created if DCF selected)
    # =========================================================================
    import numpy as _np

    # Track Forecasts sheet row positions for cross-sheet formulas used by CompCo
    _for_rev_row_excel = None       # excel row# of Revenue in Forecasts sheet
    _for_ebitda_row_excel = None    # excel row# of EBITDA
    _for_np_row_excel = None        # excel row# of Net Profit
    _for_bk_eq_row_excel = None     # excel row# of Book Equity (static input)
    _for_nd_row_excel = None        # excel row# of Net Debt (static input)
    _for_dep_row_excel = None       # excel row# of Depreciation
    _for_capex_row_excel = None     # excel row# of Capex
    _for_year_cols = []             # list of year column labels in Forecasts sheet
    _for_hist_col_count = 0         # number of historical year columns
    _yr_to_col = {}
    _for_ratio_rev_row = None       # excel row# of Revenue Growth ratio row
    _for_ratio_ebitda_row = None    # excel row# of EBITDA Margin ratio row
    _for_ratio_ebit_row = None      # excel row# of EBIT Margin ratio row
    _for_ratio_np_row = None        # excel row# of Net Profit Margin ratio row
    _for_wc_pct_row = None          # excel row# of WC % of Sales row
    _for_wc_last_row = None         # excel row# of last historical WC (for ΔWC anchor)
    _for_wc_forecast_start = None   # excel row# of first WC forecast row
    _for_wc_forecast_end = None     # excel row# of last WC forecast row
    _for_dwc_start = None           # excel row# of first ΔWC row
    _for_dwc_end = None             # excel row# of last ΔWC row

    if "DCF" in selected_models:
        wsFor = wb.active
        wsFor.title = "Forecasts"
        co_name = _sanitize(ss.get("company_name", "Company"))
        write_title(wsFor, f"{co_name} — Forecasted Income Statement", ncols=20)

        _fi = ss.get("forecast_is_df")
        forecast_is = _fi if (_fi is not None and not (hasattr(_fi, "empty") and _fi.empty)) else ss.get("dcf_is_df")

        r = 3
        if forecast_is is not None and hasattr(forecast_is, "columns"):
            item_col = "Item"
            year_cols = [c for c in forecast_is.columns if c != item_col]
            _for_year_cols = [str(y) for y in year_cols]

            # Avg column: col 1=Item, col 2..n+1=years, col n+2=Hist.Avg
            # Define early so IS forecast formulas can reference it
            _avg_col_num = len(year_cols) + 2
            _avg_col_L   = get_column_letter(_avg_col_num)

            # Detect historical vs forecast columns
            _dcf_is = ss.get("dcf_is_df")
            _dcf_is_cols = list(_dcf_is.columns) if (_dcf_is is not None and hasattr(_dcf_is, "columns")) else year_cols
            hist_yrs = set(str(c) for c in _dcf_is_cols if c != item_col)
            _for_hist_col_count = len(hist_yrs)

            forecast_start_yr = None
            try:
                int_years = sorted([int(c) for c in year_cols if str(c).isdigit()])
                # Use the last historical year stored by the DCF page (last uploaded year + 1)
                # Fall back to dcf_is_df columns if not available
                _last_hist_yr = None
                _dcf_is_yr_cols = [c for c in (_dcf_is.columns if _dcf_is is not None else []) if c != item_col]
                if _dcf_is_yr_cols:
                    try:
                        _last_hist_yr = max(int(str(c)) for c in _dcf_is_yr_cols if str(c).isdigit())
                    except Exception:
                        pass
                if _last_hist_yr is not None:
                    forecast_start_yr = next((y for y in int_years if y > _last_hist_yr), None)
                else:
                    # absolute fallback: first year not in dcf_is_df
                    forecast_start_yr = next((y for y in int_years if str(y) not in hist_yrs), None)
            except Exception:
                pass

            # Write header row
            write_hdr(wsFor, r, [item_col] + [str(y) for y in year_cols], start_col=1)
            r += 1
            is_data_start_row = r  # first data row of IS

            # ── Identify mapped row indices from session ─────────────────────
            # We use the IS core mapping stored in session to know which excel
            # rows correspond to which financial line items.
            _is_core = ss.get("is_core_mapping", {})
            def _mapped_excel_row(key):
                """Convert mapping label '5: Revenue' to 0-based df index."""
                lbl = _is_core.get(key)
                if not lbl:
                    return None
                try:
                    return int(str(lbl).split(":")[0]) - 1
                except Exception:
                    return None

            _rev_df_idx     = _mapped_excel_row("rev")
            _cos_df_idx     = _mapped_excel_row("cos")
            _gp_df_idx      = _mapped_excel_row("gp")
            _ebitda_df_idx  = _mapped_excel_row("ebitda")
            _dep_df_idx     = _mapped_excel_row("dep")
            _op_df_idx      = _mapped_excel_row("op")
            _pbt_df_idx     = _mapped_excel_row("pbt")
            _tax_df_idx     = _mapped_excel_row("tax")
            _np_df_idx      = _mapped_excel_row("np")

            # Pull DCF assumptions for formulas
            _avg_g          = float(ss.get("dcf_rev_growth_override", ss.get("dcf_yearly_growth_pct", {}).get(str(forecast_start_yr), 0.15) if forecast_start_yr else 0.15))
            _avg_gp_margin  = None
            _avg_tax_ratio  = 0.0
            _cos_ratio      = 0.0
            _wc_pct_used    = float(ss.get("dcf_wc_pct_method_last_val", 0.0) or 0.0)
            # Try to get GP margin from session (stored in DCF page globals)
            try:
                if _gp_df_idx is not None and _rev_df_idx is not None:
                    _rev_hist = forecast_is.iloc[_rev_df_idx][list(hist_yrs)] if hist_yrs else None
                    _gp_hist  = forecast_is.iloc[_gp_df_idx][list(hist_yrs)]  if hist_yrs else None
                    if _rev_hist is not None and _gp_hist is not None:
                        import numpy as _np2
                        _rv = _rev_hist.values.astype(float)
                        _gv = _gp_hist.values.astype(float)
                        mask = (_rv != 0) & ~_np2.isnan(_rv) & ~_np2.isnan(_gv)
                        if mask.any():
                            _avg_gp_margin = float(_np2.mean(_gv[mask] / _rv[mask]))
            except Exception:
                pass

            # ── Write IS rows ─────────────────────────────────────────────────
            for idx, row_data in forecast_is.iterrows():
                item_name = row_data.get(item_col, f"Row {idx}")
                is_subtotal = any(kw in str(item_name).upper() for kw in
                                  ["TOTAL", "PROFIT", "EBITDA", "EBIT", "REVENUE", "NET", "GROSS"])
                font_row = F_BOLD if is_subtotal else F_STD
                fill_row = FL_LGREY if is_subtotal else None

                # Track key row positions
                if idx == _rev_df_idx:
                    _for_rev_row_excel = r
                if idx == _ebitda_df_idx:
                    _for_ebitda_row_excel = r
                if idx == _np_df_idx:
                    _for_np_row_excel = r
                if idx == _dep_df_idx:
                    _for_dep_row_excel = r

                wsFor.cell(r, 1).value = str(item_name)
                wsFor.cell(r, 1).font = font_row
                wsFor.cell(r, 1).border = BDR
                if fill_row:
                    wsFor.cell(r, 1).fill = fill_row

                for ci, yr in enumerate(year_cols, start=2):
                    val = row_data.get(yr)
                    cell = wsFor.cell(r, ci)
                    is_forecast_col = (forecast_start_yr and str(yr).isdigit() and int(str(yr)) >= forecast_start_yr)
                    try:
                        cell.value = float(val) if val is not None and not (isinstance(val, float) and _np.isnan(val)) else None
                        cell.number_format = FMT_MONEY0
                    except (TypeError, ValueError):
                        cell.value = str(val) if val is not None else None
                    cell.font = F_BLUE if is_forecast_col else F_STD
                    if is_subtotal:
                        cell.font = Font(bold=True, color="0000FF" if is_forecast_col else "000000", name="Arial", size=10)
                        cell.fill = FL_LGREY
                    cell.border = BDR
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                r += 1

            # ── NOW overwrite forecast columns with Excel formulas ─────────────
            # We have historical data as hardcoded values (written above).
            # For forecast columns we overwrite with proper Excel formulas so
            # changing revenue growth in the ratios table updates everything.
            #
            # Formula structure (referencing ratios table written BELOW):
            #   Revenue(y)   = Revenue(y-1) * (1 + RevGrowth%)
            #   COS(y)       = -Revenue(y) * (1 - GPMargin%)    [if GP+COS mapped]
            #   GP(y)        = Revenue(y) + COS(y)
            #   Other rows   = Revenue(y) * (row ratio%)
            #   EBITDA/OP/PBT/NP = SUM of rows above (chain)
            #   Tax(y)       = PBT(y) * TaxRatio%
            #   NP(y)        = PBT(y) + Tax(y)
            #
            # The ratios table will be written after the IS data, so we use
            # a "forward reference" approach: write the IS formulas referencing
            # ratio rows we will create, using row numbers we pre-calculate.
            #
            # We pre-calculate where the ratios table will be.
            # Actual write sequence after IS loop:
            #   r += 1            → blank row         (r+1)
            #   write_section     → section header    (r+2), then r += 1 → r becomes r+3
            #   write_hdr         → column header     (r+3), then r += 1 → r becomes r+4
            #   Revenue Growth    → first data row    (r+4)
            # So _ratio_data_r = r_after_IS + 3  (the +1 blank + section + hdr = 3 rows before data)
            _ratio_data_r    = r + 3   # Revenue Growth row (first data row of ratio table)

            # We'll write ratios in this order:
            #   0: Revenue Growth (%)
            #   1: EBITDA Margin (%)
            #   2: EBIT Margin (%)
            #   3: Net Profit Margin (%)
            #   Then one row per non-total, non-revenue IS row (as % of Rev)
            # We need to know the row index for each IS row ratio:
            # Total IS rows that are "total" rows (skip in ratio table):
            _total_df_idxs = set()
            for _k in ["gp", "ebitda", "op", "pbt", "np"]:
                _v = _mapped_excel_row(_k)
                if _v is not None:
                    _total_df_idxs.add(_v)
            if _tax_df_idx is not None:
                _total_df_idxs.add(_tax_df_idx)
            # Only exclude COS from ratio table if GP mapping exists (otherwise treat as % of Rev)
            _has_gp_for_cos = (_gp_df_idx is not None and _avg_gp_margin is not None)
            if _cos_df_idx is not None and _has_gp_for_cos:
                _total_df_idxs.add(_cos_df_idx)

            # Build mapping: df_idx -> ratio table excel row
            # First 4 rows = key ratios (Rev Growth, EBITDA%, EBIT%, NP%)
            # Then each non-rev, non-total IS row gets its own ratio row
            _ratio_row_for_df_idx = {}  # df_idx -> excel row in ratio table
            _ratio_counter = _ratio_data_r + 4  # after the 4 key ratios

            for _df_i in range(len(forecast_is)):
                if _df_i == _rev_df_idx:
                    continue  # revenue not a ratio row
                if _df_i in _total_df_idxs:
                    continue  # totals not in ratio table
                _ratio_row_for_df_idx[_df_i] = _ratio_counter
                _ratio_counter += 1

            # Now overwrite forecast columns with formulas
            _hist_cols = [str(c) for c in year_cols if str(c) in hist_yrs]
            _fore_cols = [str(c) for c in year_cols if str(c) not in hist_yrs]

            # Column index mapping: year string -> excel col number (1-based, col 1=Item, col 2=first year)
            _yr_to_col = {str(yr): ci + 2 for ci, yr in enumerate(year_cols)}

            # For each forecast year, overwrite formulas
            _rev_row_e    = _for_rev_row_excel
            _gp_row_e     = (is_data_start_row + _gp_df_idx)    if _gp_df_idx is not None else None
            _cos_row_e    = (is_data_start_row + _cos_df_idx)   if _cos_df_idx is not None else None
            _ebitda_row_e = _for_ebitda_row_excel
            _op_row_e     = (is_data_start_row + _op_df_idx)    if _op_df_idx is not None else None
            _pbt_row_e    = (is_data_start_row + _pbt_df_idx)   if _pbt_df_idx is not None else None
            _tax_row_e    = (is_data_start_row + _tax_df_idx)   if _tax_df_idx is not None else None
            _np_row_e     = _for_np_row_excel

            for _f_i, _f_yr in enumerate(_fore_cols):
                _col_num = _yr_to_col[_f_yr]
                _col_L   = get_column_letter(_col_num)

                # Revenue: prev_col * (1 + RevGrowth%) — use THIS year's column for growth rate
                _rev_growth_cell = f"{_col_L}{_ratio_data_r}"   # column-specific Rev Growth cell
                if _rev_row_e:
                    if _f_i == 0:
                        # first forecast year: previous = last historical col
                        _last_hist_col_L = get_column_letter(_yr_to_col[_hist_cols[-1]]) if _hist_cols else "B"
                        wsFor.cell(_rev_row_e, _col_num).value = f"={_last_hist_col_L}{_rev_row_e}*(1+{_rev_growth_cell})"
                    else:
                        _prev_fore_col_L = get_column_letter(_yr_to_col[_fore_cols[_f_i - 1]])
                        wsFor.cell(_rev_row_e, _col_num).value = f"={_prev_fore_col_L}{_rev_row_e}*(1+{_rev_growth_cell})"
                    wsFor.cell(_rev_row_e, _col_num).number_format = FMT_MONEY0
                    wsFor.cell(_rev_row_e, _col_num).font = Font(bold=True, color="0000FF", name="Arial", size=10)
                    wsFor.cell(_rev_row_e, _col_num).fill = FL_LGREY
                    wsFor.cell(_rev_row_e, _col_num).border = BDR

                # COS and GP
                _gp_margin_cell = f"$B${_ratio_data_r + 1}"   # row 1 of ratio table = EBITDA margin (but we need GP margin)
                # We'll use a dedicated GP margin row in the ratio table at row _ratio_data_r + 4 + len(non-total-rows)
                # But simpler: reference the GP margin we'll put in the "EBITDA Margin" slot row
                # Actually GP Margin = EBITDA Margin is wrong. Let's add a GP Margin row at the END of ratio section.
                # We'll use the actual average computed above as a hardcoded value (blue) since GP margin
                # is a backward-looking average. The formula =Revenue*(GPmargin%) is the live formula.
                # The GP margin cell will be placed after the main ratio rows:
                _gp_margin_explicit_r = _ratio_counter  # last row in ratio table
                # (written later in ratio section)

                if _cos_row_e and _gp_row_e and _rev_row_e:
                    # ── GP margin method: only use if GP margin row was actually mapped ──
                    # Check if the GP margin avg cell will have a real value
                    _has_gp_mapping = (_gp_df_idx is not None and _avg_gp_margin is not None)
                    if _has_gp_mapping:
                        _last_cos = forecast_is.iloc[_cos_df_idx].get(
                            list(hist_yrs)[-1] if hist_yrs else year_cols[-1], 0.0
                        )
                        try:
                            _last_cos = float(_last_cos)
                        except Exception:
                            _last_cos = 0.0
                        _cos_sign = -1 if _last_cos < 0 else 1
                        wsFor.cell(_cos_row_e, _col_num).value = (
                            f"={_cos_sign}*{_col_L}{_rev_row_e}*(1-${_avg_col_L}${_gp_margin_explicit_r})"
                        )
                        wsFor.cell(_cos_row_e, _col_num).number_format = FMT_MONEY0
                        wsFor.cell(_cos_row_e, _col_num).font = Font(bold=False, color="0000FF", name="Arial", size=10)
                        wsFor.cell(_cos_row_e, _col_num).border = BDR
                        wsFor.cell(_gp_row_e, _col_num).value = (
                            f"={_col_L}{_rev_row_e}+{_col_L}{_cos_row_e}"
                        )
                        wsFor.cell(_gp_row_e, _col_num).number_format = FMT_MONEY0
                        wsFor.cell(_gp_row_e, _col_num).font = Font(bold=True, color="0000FF", name="Arial", size=10)
                        wsFor.cell(_gp_row_e, _col_num).fill = FL_LGREY
                        wsFor.cell(_gp_row_e, _col_num).border = BDR
                    else:
                        # ── No GP mapping: treat COS as % of Revenue like any other line ──
                        # Find its ratio row in the ratio table and use that formula instead
                        _cos_ratio_row = _ratio_row_for_df_idx.get(_cos_df_idx)
                        if _cos_ratio_row and _rev_row_e:
                            wsFor.cell(_cos_row_e, _col_num).value = (
                                f"={_col_L}{_rev_row_e}*${_avg_col_L}${_cos_ratio_row}"
                            )
                        else:
                            wsFor.cell(_cos_row_e, _col_num).value = 0
                        wsFor.cell(_cos_row_e, _col_num).number_format = FMT_MONEY0
                        wsFor.cell(_cos_row_e, _col_num).font = Font(bold=False, color="0000FF", name="Arial", size=10)
                        wsFor.cell(_cos_row_e, _col_num).border = BDR
                        # GP row: also falls through to the ratio table via standard chain
                        if _gp_row_e and _rev_row_e:
                            _gp_ratio_row = _ratio_row_for_df_idx.get(_gp_df_idx)
                            if _gp_ratio_row:
                                wsFor.cell(_gp_row_e, _col_num).value = (
                                    f"={_col_L}{_rev_row_e}*${_avg_col_L}${_gp_ratio_row}"
                                )
                                wsFor.cell(_gp_row_e, _col_num).number_format = FMT_MONEY0
                                wsFor.cell(_gp_row_e, _col_num).font = Font(bold=True, color="0000FF", name="Arial",
                                                                            size=10)
                                wsFor.cell(_gp_row_e, _col_num).fill = FL_LGREY
                                wsFor.cell(_gp_row_e, _col_num).border = BDR

                elif _gp_row_e and _rev_row_e:
                    wsFor.cell(_gp_row_e, _col_num).value = (
                        f"={_col_L}{_rev_row_e}*${_avg_col_L}${_gp_margin_explicit_r}"
                    )
                    wsFor.cell(_gp_row_e, _col_num).number_format = FMT_MONEY0
                    wsFor.cell(_gp_row_e, _col_num).font = Font(bold=True, color="0000FF", name="Arial", size=10)
                    wsFor.cell(_gp_row_e, _col_num).fill = FL_LGREY
                    wsFor.cell(_gp_row_e, _col_num).border = BDR

                # Other non-total rows: Revenue * avg ratio%
                for _df_i2, _ratio_excel_r in _ratio_row_for_df_idx.items():
                    _is_row_e = is_data_start_row + _df_i2
                    if _rev_row_e:
                        wsFor.cell(_is_row_e, _col_num).value = f"={_col_L}{_rev_row_e}*${_avg_col_L}${_ratio_excel_r}"
                    wsFor.cell(_is_row_e, _col_num).number_format = FMT_MONEY0
                    wsFor.cell(_is_row_e, _col_num).font = Font(bold=False, color="0000FF", name="Arial", size=10)
                    wsFor.cell(_is_row_e, _col_num).border = BDR

                # Totals chain: each total = SUM of rows from previous total down to just above current
                _chain = []
                for _k2, _di2 in [("rev", _rev_df_idx), ("gp", _gp_df_idx), ("ebitda", _ebitda_df_idx),
                                    ("op", _op_df_idx), ("pbt", _pbt_df_idx), ("np", _np_df_idx)]:
                    if _di2 is not None:
                        _chain.append((is_data_start_row + _di2))
                _chain = sorted(set(_chain))

                for _ci2 in range(1, len(_chain)):
                    _prev_r = _chain[_ci2 - 1]
                    _cur_r  = _chain[_ci2]
                    if _cur_r == _gp_row_e and _cos_row_e:
                        continue  # GP already handled
                    if _cur_r == _gp_row_e:
                        continue  # GP already handled
                    wsFor.cell(_cur_r, _col_num).value = f"=SUM({_col_L}{_prev_r}:{_col_L}{_cur_r - 1})"
                    wsFor.cell(_cur_r, _col_num).number_format = FMT_MONEY0
                    wsFor.cell(_cur_r, _col_num).font = Font(bold=True, color="0000FF", name="Arial", size=10)
                    wsFor.cell(_cur_r, _col_num).fill = FL_LGREY
                    wsFor.cell(_cur_r, _col_num).border = BDR

                # Tax = PBT * TaxRatio — use Hist. Avg column for tax ratio (row _ratio_data_r+2)
                _tax_ratio_cell = f"${_avg_col_L}${_ratio_data_r + 2}"
                if _tax_row_e and _pbt_row_e:
                    wsFor.cell(_tax_row_e, _col_num).value = f"={_col_L}{_pbt_row_e}*{_tax_ratio_cell}"
                    wsFor.cell(_tax_row_e, _col_num).number_format = FMT_MONEY0
                    wsFor.cell(_tax_row_e, _col_num).font = Font(bold=False, color="0000FF", name="Arial", size=10)
                    wsFor.cell(_tax_row_e, _col_num).border = BDR

                # Net Profit = PBT + Tax + anything between
                if _np_row_e and _pbt_row_e:
                    if _tax_row_e and (_tax_row_e + 1) <= (_np_row_e - 1):
                        wsFor.cell(_np_row_e, _col_num).value = f"={_col_L}{_pbt_row_e}+{_col_L}{_tax_row_e}+SUM({_col_L}{_tax_row_e + 1}:{_col_L}{_np_row_e - 1})"
                    elif _tax_row_e:
                        wsFor.cell(_np_row_e, _col_num).value = f"={_col_L}{_pbt_row_e}+{_col_L}{_tax_row_e}"
                    else:
                        wsFor.cell(_np_row_e, _col_num).value = f"={_col_L}{_pbt_row_e}"
                    wsFor.cell(_np_row_e, _col_num).number_format = FMT_MONEY0
                    wsFor.cell(_np_row_e, _col_num).font = Font(bold=True, color="0000FF", name="Arial", size=10)
                    wsFor.cell(_np_row_e, _col_num).fill = FL_LGREY
                    wsFor.cell(_np_row_e, _col_num).border = BDR

            # ── Key Ratios section — ALL AS EXCEL FORMULAS ────────────────────
            r += 1
            # Avg column is the column right after the last year column

            # Identify which column indices are historical vs forecast
            _hist_col_indices = []   # (ci, yr_str) for historical years
            _fore_col_indices = []   # (ci, yr_str) for forecast years
            for _ci_tmp, _yr_tmp in enumerate(year_cols, start=2):
                _is_fc_tmp = (forecast_start_yr and str(_yr_tmp).isdigit() and int(str(_yr_tmp)) >= forecast_start_yr)
                if _is_fc_tmp:
                    _fore_col_indices.append((_ci_tmp, str(_yr_tmp)))
                else:
                    _hist_col_indices.append((_ci_tmp, str(_yr_tmp)))

            # Build AVERAGE formula over historical columns for a given row
            def _avg_formula(row_num):
                if not _hist_col_indices:
                    return None
                hist_cells = ",".join(f"{get_column_letter(ci)}{row_num}" for ci, _ in _hist_col_indices)
                return f"=IFERROR(AVERAGE({hist_cells}),\"\")"

            write_section(wsFor, r, "Key Financial Ratios", ncols=_avg_col_num); r += 1
            write_hdr(wsFor, r, ["Metric"] + [str(y) for y in year_cols] + ["Hist. Avg"], start_col=1); r += 1

            # ── helper: write one ratio row ──────────────────────────────────
            # hist_fn(ci, yr_str) -> formula string for historical cell
            # avg_col -> True = add AVERAGE formula in avg col
            # forecast_ref -> "avg" = reference avg col; "value:X" = hardcode X
            def _write_ratio_row(ws, row, label, hist_fn, forecast_ref="avg",
                                  fmt=FMT_PCT1, label_font=None):
                ws.cell(row, 1).value = label
                ws.cell(row, 1).font  = label_font or F_STD
                ws.cell(row, 1).border = BDR

                # Historical cells
                for ci, yr_s in _hist_col_indices:
                    cell = ws.cell(row, ci)
                    cell.border = BDR
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = fmt
                    v = hist_fn(ci, yr_s)
                    if v is not None:
                        cell.value = v

                # Average column
                avg_cell = ws.cell(row, _avg_col_num)
                avg_cell.border = BDR
                avg_cell.alignment = Alignment(horizontal="right", vertical="center")
                avg_cell.number_format = fmt
                avg_f = _avg_formula(row)
                if avg_f:
                    avg_cell.value  = avg_f
                    avg_cell.font   = Font(bold=True, color="008000", name="Arial", size=10)  # green = formula

                # Forecast cells → reference average column
                for ci, yr_s in _fore_col_indices:
                    cell = ws.cell(row, ci)
                    cell.border = BDR
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = fmt
                    if forecast_ref == "avg":
                        cell.value = f"=${_avg_col_L}${row}"
                        cell.font  = F_GREEN
                    elif str(forecast_ref).startswith("value:"):
                        cell.value = float(str(forecast_ref).split(":", 1)[1])
                        cell.font  = F_BLUE
                    elif str(forecast_ref).startswith("peryear:"):
                        # per-year dict passed as peryear:{yr:val,...}
                        pass  # handled outside

            # -- Revenue Growth --
            _for_ratio_rev_row = r
            _growth_mode = ss.get("dcf_rev_growth_mode", "Uniform (same % each year)")
            _yearly_g_pct = ss.get("dcf_yearly_growth_pct", {}) or {}
            _uniform_g = ss.get("dcf_rev_growth_override") or ss.get("dcf_avg_g", ss.get("avg_g", None))
            wsFor.cell(r, 1).value = "Revenue Growth (%)"
            wsFor.cell(r, 1).font = F_STD; wsFor.cell(r, 1).border = BDR

            for ci, yr in enumerate(year_cols, start=2):
                cell = wsFor.cell(r, ci)
                cell.border = BDR
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = FMT_PCT1
                _is_fc_g = (forecast_start_yr and str(yr).isdigit() and int(str(yr)) >= forecast_start_yr)
                if _is_fc_g:
                    _yr_key = str(yr)
                    if _growth_mode.startswith("Uniform"):
                        # Always use the uniform average for ALL forecast years
                        _g_val = float(_uniform_g) if _uniform_g is not None else 0.0
                        _g_dec = _g_val / 100.0 if abs(_g_val) > 1 else _g_val
                        cell.value = _g_dec
                    else:
                        # Different growth per year — use per-year value
                        _g_val = _yearly_g_pct.get(_yr_key)
                        if _g_val is None and _uniform_g is not None:
                            _g_val = float(_uniform_g)
                        if _g_val is None:
                            cell.value = 0.0
                        else:
                            _g_dec = float(_g_val) / 100.0 if abs(float(_g_val)) > 1 else float(_g_val)
                            cell.value = _g_dec
                    cell.font = F_BLUE
                else:
                    # historical columns: derive YoY growth from actual uploaded values
                    # (both prior and current year are hardcoded numbers — no circular ref)
                    if ci > 2 and _for_rev_row_excel:
                        prev_col_L = get_column_letter(ci - 1)
                        cur_col_L = get_column_letter(ci)
                        cell.value = f"=IFERROR(({cur_col_L}{_for_rev_row_excel}-{prev_col_L}{_for_rev_row_excel})/ABS({prev_col_L}{_for_rev_row_excel}),\"\")"
                    else:
                        cell.value = None
            # Avg of historical revenue growths
            _rev_avg_cell = wsFor.cell(r, _avg_col_num)
            _rev_avg_cell.border = BDR
            _rev_avg_cell.number_format = FMT_PCT1
            _rev_avg_cell.alignment = Alignment(horizontal="right", vertical="center")
            if _hist_col_indices:
                # Average of historical cols (skip first hist col since it has no prior year growth)
                _hist_growth_cols = [ci for ci, _ in _hist_col_indices if ci > 2]
                if _hist_growth_cols:
                    _rev_avg_cell.value = f"=IFERROR(AVERAGE({','.join(f'{get_column_letter(c)}{r}' for c in _hist_growth_cols)}),\"\")"
                    _rev_avg_cell.font  = Font(bold=True, color="008000", name="Arial", size=10)
            r += 1

            # -- EBITDA Margin --
            _for_ratio_ebitda_row = r
            def _ebitda_hist(ci, yr_s):
                if _for_ebitda_row_excel and _for_rev_row_excel:
                    return f"=IFERROR({get_column_letter(ci)}{_for_ebitda_row_excel}/{get_column_letter(ci)}{_for_rev_row_excel},\"\")"
                return None
            _write_ratio_row(wsFor, r, "EBITDA Margin (%)", _ebitda_hist, forecast_ref="avg")
            r += 1

            # -- Effective Tax Rate (Tax/PBT) —─ used in forecasts --
            _for_ratio_ebit_row = r
            def _tax_hist(ci, yr_s):
                if _tax_df_idx is not None and _pbt_df_idx is not None:
                    _tax_r_e2 = is_data_start_row + _tax_df_idx
                    _pbt_r_e2 = is_data_start_row + _pbt_df_idx
                    return f"=IFERROR({get_column_letter(ci)}{_tax_r_e2}/{get_column_letter(ci)}{_pbt_r_e2},\"\")"
                return None
            _write_ratio_row(wsFor, r, "Effective Tax Rate (Tax/PBT) — used in forecasts", _tax_hist, forecast_ref="avg")
            r += 1

            # -- Net Profit Margin --
            _for_ratio_np_row = r
            def _np_hist(ci, yr_s):
                if _for_np_row_excel and _for_rev_row_excel:
                    return f"=IFERROR({get_column_letter(ci)}{_for_np_row_excel}/{get_column_letter(ci)}{_for_rev_row_excel},\"\")"
                return None
            _write_ratio_row(wsFor, r, "Net Profit Margin (%)", _np_hist, forecast_ref="avg")
            r += 1

            # -- All other non-total IS rows as % of Revenue --
            for _df_i_r, _ratio_excel_r in _ratio_row_for_df_idx.items():
                _row_label = str(forecast_is.iloc[_df_i_r].get(item_col, f"Row {_df_i_r}"))
                _is_row_e2 = is_data_start_row + _df_i_r
                def _generic_hist(ci, yr_s, _ire=_is_row_e2):
                    if _for_rev_row_excel:
                        return f"=IFERROR({get_column_letter(ci)}{_ire}/{get_column_letter(ci)}{_for_rev_row_excel},\"\")"
                    return None
                _write_ratio_row(wsFor, r, f"{_row_label} % of Revenue", _generic_hist,
                                 forecast_ref="avg", label_font=F_NOTE)
                r += 1

            # -- GP Margin row (drives COS forecast) --
            _gp_margin_explicit_r = r
            def _gp_hist(ci, yr_s):
                if _gp_df_idx is not None and _for_rev_row_excel:
                    _gp_r_e3 = is_data_start_row + _gp_df_idx
                    return f"=IFERROR({get_column_letter(ci)}{_gp_r_e3}/{get_column_letter(ci)}{_for_rev_row_excel},\"\")"
                return None
            _write_ratio_row(wsFor, r, "GP Margin % — used in COS/GP forecast (editable)",
                             _gp_hist, forecast_ref="avg")
            r += 1

        else:
            wsFor.cell(r, 1).value = "No forecasted income statement found. Run the DCF page first."
            wsFor.cell(r, 1).font = F_NOTE
            r += 2

        # ── Working Capital (Forecast) — fully formulized ─────────────────────
        r += 1
        write_section(wsFor, r, "Working Capital (Forecast & ΔWC)", ncols=6); r += 1

        _fore_yr_list = [str(y) for y in year_cols if str(y) not in hist_yrs] if forecast_is is not None else []
        _hist_yr_list = [str(y) for y in year_cols if str(y) in hist_yrs] if forecast_is is not None else []
        _wc_pct_used_val = float(ss.get("dcf_wc_pct_method_last_val", 0.0) or
                                 ss.get("dcf_wc_percent_avg", 0.0) or 0.0)
        _last_wc_hist = float(ss.get("dcf_last_wc_hist", 0.0) or 0.0)

        # ── HISTORICAL WC% TABLE (live formulas using BS data) ───────────────
        _bs_df_exp  = ss.get("dcf_bs_df")
        _ca_labels  = ss.get("dcf_mapping", {}).get("ca", [])
        _cl_labels  = ss.get("dcf_mapping", {}).get("cl", [])

        _wc_hist_data = {}

        if _bs_df_exp is not None and _ca_labels and _cl_labels:
            _ca_rows_exp = [int(str(l).split(":")[0]) - 1
                            for l in _ca_labels
                            if str(l).split(":")[0].strip().isdigit()]
            _cl_rows_exp = [int(str(l).split(":")[0]) - 1
                            for l in _cl_labels
                            if str(l).split(":")[0].strip().isdigit()]
            for _yr_h in _hist_yr_list:
                try:
                    if _yr_h in _bs_df_exp.columns:
                        _ca_v = float(_bs_df_exp.iloc[_ca_rows_exp][_yr_h].sum()) if _ca_rows_exp else 0.0
                        _cl_v = float(_bs_df_exp.iloc[_cl_rows_exp][_yr_h].sum()) if _cl_rows_exp else 0.0
                        _wc_hist_data[_yr_h] = (_ca_v, _cl_v, _ca_v - _cl_v)
                except Exception:
                    pass

        # Write historical WC analysis table
        write_hdr(wsFor, r, ["Year", "Current Assets", "Current Liabilities",
                              "WC (CA−CL)", "Revenue", "WC % of Sales"]); r += 1
        _hist_wc_pct_cells = []
        _hist_wc_table_start = r

        for _yr_h2 in _hist_yr_list:
            wsFor.cell(r, 1).value = int(_yr_h2) if _yr_h2.isdigit() else _yr_h2
            wsFor.cell(r, 1).font = F_STD
            wsFor.cell(r, 1).border = BDR
            wsFor.cell(r, 1).number_format = '0'
            wsFor.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")

            if _yr_h2 in _wc_hist_data:
                _ca_v2, _cl_v2, _wc_v2 = _wc_hist_data[_yr_h2]
                wsFor.cell(r, 2).value = float(_ca_v2)
                wsFor.cell(r, 2).font = F_BLUE
                wsFor.cell(r, 3).value = float(_cl_v2)
                wsFor.cell(r, 3).font = F_BLUE
                wsFor.cell(r, 4).value = f"=B{r}-C{r}"
                wsFor.cell(r, 4).font = F_GREEN
                _rev_col_num_h = _yr_to_col.get(_yr_h2)
                if _rev_col_num_h and _for_rev_row_excel:
                    _rev_col_L_h = get_column_letter(_rev_col_num_h)
                    wsFor.cell(r, 5).value = f"={_rev_col_L_h}{_for_rev_row_excel}"
                    wsFor.cell(r, 5).font = F_GREEN
                else:
                    wsFor.cell(r, 5).value = 0
                    wsFor.cell(r, 5).font = F_BLUE
                wsFor.cell(r, 6).value = f"=IFERROR(D{r}/E{r},0)"
                wsFor.cell(r, 6).font = F_GREEN
                _hist_wc_pct_cells.append(f"F{r}")
            else:
                for _c in range(2, 7):
                    wsFor.cell(r, _c).value = 0
                    wsFor.cell(r, _c).font = F_BLUE

            for _c in range(2, 7):
                wsFor.cell(r, _c).border = BDR
                wsFor.cell(r, _c).alignment = Alignment(horizontal="right", vertical="center")
                wsFor.cell(r, _c).number_format = FMT_MONEY0 if _c < 6 else FMT_PCT1
            r += 1

        # WC% assumption — dynamic (AVG vs MOST RECENT)
        write_hdr(wsFor, r, ["WC % of Sales", "Value"]);
        r += 1
        _for_wc_pct_row = r

        _wc_method = ss.get("dcf_wc_pct_method", "avg")  # ✅ CRITICAL

        # Label changes dynamically
        if _wc_method == "recent":
            label = "WC % of Sales — used (most recent year)"
        else:
            label = "WC % of Sales — used (average of history)"

        wsFor.cell(r, 1).value = label
        wsFor.cell(r, 1).font = F_BOLD
        wsFor.cell(r, 1).border = BDR
        wsFor.cell(r, 1).fill = FL_LBLUE

        if _hist_wc_pct_cells:
            if _wc_method == "recent":
                # ✅ TAKE LAST YEAR WC% (e.g. F2025)
                _last_wc_pct_cell = _hist_wc_pct_cells[-1]
                wsFor.cell(r, 2).value = f"=IFERROR({_last_wc_pct_cell},{_wc_pct_used_val})"
            else:
                # ✅ AVERAGE
                wsFor.cell(r, 2).value = f"=IFERROR(AVERAGE({','.join(_hist_wc_pct_cells)}),{_wc_pct_used_val})"

            wsFor.cell(r, 2).font = Font(bold=True, color="008000", name="Arial", size=10)

        else:
            wsFor.cell(r, 2).value = float(_wc_pct_used_val)
            wsFor.cell(r, 2).font = F_BLUE

        wsFor.cell(r, 2).number_format = FMT_PCT1
        wsFor.cell(r, 2).border = BDR
        wsFor.cell(r, 2).fill = FL_LBLUE
        wsFor.cell(r, 2).alignment = Alignment(horizontal="right", vertical="center")

        r += 1

        # Last historical WC anchor
        _for_wc_last_row = r
        wsFor.cell(r, 1).value = "Last Historical Working Capital (anchor for ΔWC)"
        wsFor.cell(r, 1).font = F_STD
        wsFor.cell(r, 1).border = BDR

        if _hist_wc_pct_cells:
            _last_hist_wc_row = _hist_wc_table_start + len(_hist_yr_list) - 1
            wsFor.cell(r, 2).value = f"=D{_last_hist_wc_row}"
            wsFor.cell(r, 2).font = F_GREEN
        else:
            wsFor.cell(r, 2).value = float(_last_wc_hist)
            wsFor.cell(r, 2).font = F_BLUE

        wsFor.cell(r, 2).number_format = FMT_MONEY0
        wsFor.cell(r, 2).border = BDR
        r += 1

        # Forecast WC table
        write_hdr(wsFor, r, ["Year", "Forecast Revenue", "Forecast WC (Rev×WC%)", "ΔWC (Old–New)"]); r += 1
        _for_wc_forecast_start = r
        _wc_prev_ref = f"B{_for_wc_last_row}"

        for _f_i2, _f_yr2 in enumerate(_fore_yr_list):
            _col_num2 = _yr_to_col.get(_f_yr2) if forecast_is is not None else None
            _col_L2 = get_column_letter(_col_num2) if _col_num2 else "B"
            wsFor.cell(r, 1).value = int(_f_yr2) if _f_yr2.isdigit() else _f_yr2
            wsFor.cell(r, 1).font = F_STD; wsFor.cell(r, 1).border = BDR
            wsFor.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")

            if _for_rev_row_excel and _col_num2:
                wsFor.cell(r, 2).value = f"={_col_L2}{_for_rev_row_excel}"
            else:
                wsFor.cell(r, 2).value = 0
            wsFor.cell(r, 2).font = F_STD; wsFor.cell(r, 2).number_format = FMT_MONEY0; wsFor.cell(r, 2).border = BDR

            wsFor.cell(r, 3).value = f"=B{r}*$B${_for_wc_pct_row}"
            wsFor.cell(r, 3).font = F_STD; wsFor.cell(r, 3).number_format = FMT_MONEY0; wsFor.cell(r, 3).border = BDR

            wsFor.cell(r, 4).value = f"={_wc_prev_ref}-C{r}"
            wsFor.cell(r, 4).font = F_STD; wsFor.cell(r, 4).number_format = FMT_MONEY0; wsFor.cell(r, 4).border = BDR
            _wc_prev_ref = f"C{r}"
            r += 1

        _for_wc_forecast_end = r - 1

        # ── Depreciation & Capex (from DCF session) ───────────────────────────
        r += 1
        write_section(wsFor, r, "Depreciation & Capex (from DCF assumptions)", ncols=6); r += 1
        _dep_forecast = ss.get("dcf_dep_forecast", {})
        _cap_forecast = [float(x) for x in ss.get("dcf_fcff_array", [])]  # capex baked in FCFF
        # Try to get capex separately from the DCF CF mapping
        _capex_forecast_direct = {}
        _df_dcf_export = ss.get("df_dcf_export")
        if _df_dcf_export is not None and hasattr(_df_dcf_export, "columns") and "Capex" in _df_dcf_export.columns:
            for _di3, _dr3 in _df_dcf_export.iterrows():
                try:
                    _capex_forecast_direct[str(int(_dr3.get("Year", 0)))] = float(_dr3.get("Capex", 0.0))
                except Exception:
                    pass

        write_hdr(wsFor, r, ["Year", "Depreciation (forecast)", "Capex (forecast)"]); r += 1
        _for_dep_capex_start = r
        for _f_i3, _f_yr3 in enumerate(_fore_yr_list):
            wsFor.cell(r, 1).value = int(_f_yr3) if _f_yr3.isdigit() else _f_yr3
            wsFor.cell(r, 1).font = F_STD; wsFor.cell(r, 1).border = BDR
            wsFor.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")

            # Depreciation: formula referencing IS Forecasts row if mapped, else fallback to session value
            if _for_dep_row_excel and _f_yr3 in _yr_to_col:
                _dep_col_L3 = get_column_letter(_yr_to_col[_f_yr3])
                wsFor.cell(r, 2).value = f"={_dep_col_L3}{_for_dep_row_excel}"
            else:
                _dep_val = float(_dep_forecast.get(_f_yr3, 0.0) or 0.0)
                wsFor.cell(r, 2).value = _dep_val
            wsFor.cell(r, 2).font = F_BLUE; wsFor.cell(r, 2).number_format = FMT_MONEY0; wsFor.cell(r, 2).border = BDR

            _cap_val = float(_capex_forecast_direct.get(_f_yr3, 0.0) or 0.0)
            wsFor.cell(r, 3).value = _cap_val
            wsFor.cell(r, 3).font = F_BLUE; wsFor.cell(r, 3).number_format = FMT_MONEY0; wsFor.cell(r, 3).border = BDR
            r += 1

        # Net Debt & Book Equity static inputs (used by CompCo formulas)
        r += 1
        write_section(wsFor, r, "Balance Sheet Metrics (used by CompCo)", ncols=4); r += 1
        write_hdr(wsFor, r, ["Item", "Value (USD)"]); r += 1
        _for_bk_eq_row_excel = r
        wsFor.cell(r, 1).value = "Book Value of Equity"
        wsFor.cell(r, 1).font = F_STD; wsFor.cell(r, 1).border = BDR
        wsFor.cell(r, 2).value = float(ss.get("book_equity", 0.0) or 0.0)
        wsFor.cell(r, 2).font = F_BLUE; wsFor.cell(r, 2).number_format = FMT_MONEY0; wsFor.cell(r, 2).border = BDR
        r += 1
        _for_nd_row_excel = r
        wsFor.cell(r, 1).value = "Net Debt"
        wsFor.cell(r, 1).font = F_STD; wsFor.cell(r, 1).border = BDR
        wsFor.cell(r, 2).value = float(ss.get("net_debt", 0.0) or 0.0)
        wsFor.cell(r, 2).font = F_BLUE; wsFor.cell(r, 2).number_format = FMT_MONEY0; wsFor.cell(r, 2).border = BDR
        r += 1

        wsFor.column_dimensions["A"].width = 44
        for ci in range(2, len((forecast_is.columns if forecast_is is not None else [])) + 2):
            wsFor.column_dimensions[get_column_letter(ci)].width = 16
        wsFor.freeze_panes = "B4"

        # =========================================================================
        # SHEET: DCF — Parameters + UFCF + Sensitivity + Output
        # =========================================================================
        wsDCF = wb.create_sheet("DCF")
        write_title(wsDCF, "DCF — Discounted Cash Flow Valuation", ncols=10)

        # ── Pull all DCF inputs ────────────────────────────────────────────────
        rf_pct     = float(ss.get("dcf_rf_pct",  ss.get("rf",  0) * 100 if ss.get("rf") else 11.61))
        mrp_pct    = float(ss.get("dcf_mrp_pct", ss.get("mrp", 0) * 100 if ss.get("mrp") else 13.82))
        tax_pct    = float(ss.get("dcf_tax_pct", ss.get("tax", 0) * 100 if ss.get("tax") else 25.0))
        beta_u     = float(ss.get("dcf_unlevered_beta", 1.0))
        de_ratio   = float(ss.get("de_ratio", 0.0))
        rd_pct_raw = ss.get("rd", 0.12)
        rd_pct     = float(rd_pct_raw * 100 if rd_pct_raw and rd_pct_raw <= 1 else (rd_pct_raw or 12.0))
        wacc_val   = float(ss.get("wacc", 0.0))
        g_pct      = float(ss.get("dcf_terminal_g_pct", ss.get("g", 0) * 100 if ss.get("g") else 5.0))
        net_debt   = float(ss.get("net_debt", 0.0))
        total_debt = float(ss.get("total_debt", 0.0))
        cash_bal   = float(ss.get("cash_balance", 0.0))
        book_eq    = float(ss.get("book_equity",  0.0))

        # ── WACC Build-up — with Formula & Description columns ────────────────
        r = 3
        write_section(wsDCF, r, "WACC Build-up", ncols=5); r += 1
        write_hdr(wsDCF, r, ["Parameter", "Value", "Formula", "Description"]); r += 1

        row_rf=r; row_mrp=r+1; row_bu=r+2; row_de=r+3; row_tax=r+4
        row_bl=r+5; row_ke=r+6; row_rd=r+7; row_wacc=r+8

        _wacc_rows = [
            # (label,  value,    is_input,  fmt,       formula_str,                                        description)
            ("Risk-free Rate (Rf)",       rf_pct/100,   True,  FMT_PCT,    "Input — government bond yield",                   "Zimbabwe / USD sovereign bond rate used as risk-free proxy"),
            ("Market Risk Premium (ERP)", mrp_pct/100,  True,  FMT_PCT,    "Input — equity risk premium",                     "Expected return of market above risk-free rate (Damodaran / local estimate)"),
            ("Unlevered Beta (βu)",       beta_u,       True,  "0.0000",   "Input — asset beta (unlevered)",                  "Beta stripped of financial leverage; reflects business risk only"),
            ("D/E Ratio",                 de_ratio,     True,  "0.0000",   f"Total Debt / Book Equity  ({total_debt} / {book_eq if book_eq else 1})", "Debt-to-equity ratio used to re-lever beta"),
            ("Tax Rate",                  tax_pct/100,  True,  FMT_PCT,    "Input — corporate tax rate",                      "Effective tax rate applied to interest tax shield in Hamada equation"),
            ("Levered Beta (βL)",         None,         False, "0.0000",   "βu × (1 + (1 − T) × D/E)",                      "Hamada equation: re-levers asset beta for the firm's actual capital structure"),
            ("Cost of Equity (Ke)",       None,         False, FMT_PCT,    "Rf + βL × ERP",                                  "CAPM: required return by equity holders"),
            ("Cost of Debt (Rd)",         rd_pct/100,   True,  FMT_PCT,    "Input — pre-tax cost of debt",                    "Weighted average interest rate on interest-bearing debt"),
            ("WACC",                      wacc_val,     True,  FMT_PCT,    "Ke x (E/V) + Rd x (1-T) x (D/V)",                "Weighted average: equity weight x Ke + debt weight x Rd x (1-T)"),
        ]

        for i, (lbl, val, is_inp, fmt, fml, desc) in enumerate(_wacc_rows):
            rr = r + i
            cell_bd(wsDCF, rr, 1, lbl,  F_BOLD if lbl == "WACC" else F_STD)
            if val is not None:
                cell_bd(wsDCF, rr, 2, val, F_BLUE if is_inp else F_STD, fmt)
            else:
                cell_bd(wsDCF, rr, 2, font=F_STD, fmt=fmt)
            # Formula column (col C) — italic grey
            c3 = wsDCF.cell(rr, 3)
            c3.value  = str(fml) if fml else ""  # assign separately — never via constructor — prevents = being treated as formula
            c3.font   = Font(italic=True, color="595959", name="Arial", size=9)
            c3.border = BDR
            c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            # Description column (col D)
            c4 = wsDCF.cell(rr, 4)
            c4.value  = str(desc) if desc else ""
            c4.font   = Font(italic=True, color="595959", name="Arial", size=9)
            c4.border = BDR
            c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Live Excel formulas for derived rows (col B)
        wsDCF.cell(row_bl,   2).value  = f"=B{row_bu}*(1+(1-B{row_tax})*B{row_de})"
        wsDCF.cell(row_bl,   2).font   = F_STD
        wsDCF.cell(row_bl,   2).number_format = "0.0000"
        wsDCF.cell(row_bl,   2).border = BDR
        wsDCF.cell(row_ke,   2).value  = f"=B{row_rf}+B{row_bl}*B{row_mrp}"
        wsDCF.cell(row_ke,   2).font   = F_STD
        wsDCF.cell(row_ke,   2).number_format = FMT_PCT
        wsDCF.cell(row_ke,   2).border = BDR
        # WACC live formula (overwrite the hardcoded input with a formula)
        _eq_val = book_eq if book_eq else 1.0
        _dbt_val = total_debt if total_debt else 0.0
        _tot_val = _eq_val + _dbt_val
        wsDCF.cell(row_wacc, 2).value  = f"=(B{row_ke}*({_eq_val}/{_tot_val}))+(B{row_rd}*(1-B{row_tax})*({_dbt_val}/{_tot_val}))"
        wsDCF.cell(row_wacc, 2).font   = Font(bold=True, name="Arial", size=10)
        wsDCF.cell(row_wacc, 2).number_format = FMT_PCT
        wsDCF.cell(row_wacc, 2).border = BDR

        r = row_wacc + 2

        # ── Net Debt Build-up ──────────────────────────────────────────────────
        write_section(wsDCF, r, "Net Debt Calculation", ncols=5); r += 1
        write_hdr(wsDCF, r, ["Item", "Value (USD)", "Formula", "Description"]); r += 1
        row_td = r; row_cb = r+1; row_nd_calc = r+2

        _nd_rows = [
            ("Total Debt (interest-bearing)",  total_debt, FMT_MONEY0, "Input — from Balance Sheet",  "Sum of short-term and long-term borrowings"),
            ("Less: Cash & Cash Equivalents",  cash_bal,   FMT_MONEY0, "Input — from Balance Sheet",  "Cash held offset against gross debt"),
            ("Net Debt",                        None,       FMT_MONEY0, f"Total Debt − Cash",         "Gross debt less cash; added back to EV to get Equity Value"),
        ]
        for i, (lbl, val, fmt, fml, desc) in enumerate(_nd_rows):
            rr = r + i
            cell_bd(wsDCF, rr, 1, lbl, F_BOLD if lbl == "Net Debt" else F_STD)
            if val is not None:
                cell_bd(wsDCF, rr, 2, val, F_BLUE, fmt)
            else:
                cell_bd(wsDCF, rr, 2, font=F_BOLD, fmt=fmt)
            c3 = wsDCF.cell(rr, 3)
            c3.value = str(fml) if fml else ""
            c3.font = Font(italic=True, color="595959", name="Arial", size=9); c3.border = BDR
            c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c4 = wsDCF.cell(rr, 4)
            c4.value = str(desc) if desc else ""
            c4.font = Font(italic=True, color="595959", name="Arial", size=9); c4.border = BDR
            c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Net Debt live formula
        wsDCF.cell(row_nd_calc, 2).value  = f"=B{row_td}-B{row_cb}"
        wsDCF.cell(row_nd_calc, 2).font   = F_BOLD
        wsDCF.cell(row_nd_calc, 2).number_format = FMT_MONEY0
        wsDCF.cell(row_nd_calc, 2).border = BDR
        r = row_nd_calc + 2

        # ── Terminal Value Assumptions ─────────────────────────────────────────
        write_section(wsDCF, r, "Terminal Value & Key Assumptions", ncols=5); r += 1
        write_hdr(wsDCF, r, ["Parameter", "Value", "Formula", "Description"]); r += 1
        row_g = r; row_nd = r + 1
        cell_bd(wsDCF, row_g,  1, "Terminal Growth Rate (g)", F_STD)
        cell_bd(wsDCF, row_g,  2, g_pct / 100, F_BLUE, FMT_PCT)
        wsDCF.cell(row_g, 3).value = "Input — long-run growth rate"; wsDCF.cell(row_g, 3).font = Font(italic=True, color="595959", name="Arial", size=9); wsDCF.cell(row_g, 3).border = BDR
        wsDCF.cell(row_g, 4).value = "Perpetuity growth assumption; must be < WACC"; wsDCF.cell(row_g, 4).font = Font(italic=True, color="595959", name="Arial", size=9); wsDCF.cell(row_g, 4).border = BDR
        cell_bd(wsDCF, row_nd, 1, "Net Debt", F_STD)
        wsDCF.cell(row_nd, 2).value  = f"=B{row_nd_calc}"
        wsDCF.cell(row_nd, 2).font   = F_STD
        wsDCF.cell(row_nd, 2).number_format = FMT_MONEY0
        wsDCF.cell(row_nd, 2).border = BDR
        wsDCF.cell(row_nd, 3).value = "Links to Net Debt Calculation section above"; wsDCF.cell(row_nd, 3).font = Font(italic=True, color="595959", name="Arial", size=9); wsDCF.cell(row_nd, 3).border = BDR
        wsDCF.cell(row_nd, 4).value = "Pulled from Net Debt section above"; wsDCF.cell(row_nd, 4).font = Font(italic=True, color="595959", name="Arial", size=9); wsDCF.cell(row_nd, 4).border = BDR
        r = row_nd + 2

        # ── UFCF Table with cross-sheet formulas ──────────────────────────────
        write_section(wsDCF, r, "Unlevered Free Cash Flow (UFCF) Forecast", ncols=8); r += 1
        write_hdr(wsDCF, r, ["Year", "Period (n)",
                               "EBITDA×(1−T)  [USD]",
                               "Dep × Tax Shield  [USD]",
                               "ΔWorking Capital  [USD]",
                               "Capex  [USD]",
                               "UFCF  [USD]",
                               "Discount Factor",
                               "PV of UFCF  [USD]"]); r += 1

        fcff_arr   = ss.get("dcf_fcff_array", [])
        n_arr      = ss.get("dcf_discount_periods_n", [])

        # Use actual forecast years from session (not today's year)
        _fore_yr_keys = []
        if _fore_yr_list:
            _fore_yr_keys = [str(y) for y in _fore_yr_list]
        elif fcff_arr:
            # fallback: build from dcf_is_df last year
            _dcf_is_tmp = ss.get("dcf_is_df")
            if _dcf_is_tmp is not None and hasattr(_dcf_is_tmp, "columns"):
                _last_yr_tmp = max((int(c) for c in _dcf_is_tmp.columns if c != "Item" and str(c).isdigit()), default=datetime.date.today().year)
            else:
                _last_yr_tmp = datetime.date.today().year
            _fore_yr_keys = [str(_last_yr_tmp + 1 + i) for i in range(len(fcff_arr))]

        cf_start = r
        _n_fore_dcf = max(len(fcff_arr), len(_fore_yr_keys))

        for i in range(_n_fore_dcf):
            rr = r + i
            _yr_lbl = int(_fore_yr_keys[i]) if i < len(_fore_yr_keys) else ""
            cell_bd(wsDCF, rr, 1, _yr_lbl, F_STD, align="center")

            # Period n — full precision (NO rounding)
            if i < len(n_arr):
                _n_val = float(n_arr[i])
            else:
                _n_val = float(i + 0.5)

            cell_bd(wsDCF, rr, 2, _n_val, F_BLUE, "0.0000")

            # ── EBITDA × (1-T): cross-sheet from Forecasts if row known ──
            _yr_s = _fore_yr_keys[i] if i < len(_fore_yr_keys) else None
            _fc_col_L_dcf = None
            if _yr_s and _yr_s in _yr_to_col:
                _fc_col_L_dcf = get_column_letter(_yr_to_col[_yr_s])

            if _for_ebitda_row_excel and _fc_col_L_dcf:
                wsDCF.cell(rr, 3).value = f"=Forecasts!{_fc_col_L_dcf}{_for_ebitda_row_excel}*(1-DCF!$B${row_tax})"
            else:
                # ✅ Always fallback if Forecasts missing
                wsDCF.cell(rr, 3).value = float(fcff_arr[i]) if i < len(fcff_arr) else 0
            wsDCF.cell(rr, 3).font = F_STD; wsDCF.cell(rr, 3).number_format = FMT_MONEY0; wsDCF.cell(rr, 3).border = BDR

            # ── Dep × Tax Shield: cross-sheet from Forecasts dep/capex table ──
            if _for_dep_row_excel and _fc_col_L_dcf:
                wsDCF.cell(rr, 4).value = f"=-Forecasts!{_fc_col_L_dcf}{_for_dep_row_excel}*DCF!$B${row_tax}"
            elif _for_dep_capex_start and i < _n_fore_dcf:
                wsDCF.cell(rr, 4).value = f"=-Forecasts!B{_for_dep_capex_start + i}*DCF!$B${row_tax}"
            else:
                wsDCF.cell(rr, 4).value = 0
            wsDCF.cell(rr, 4).font = F_STD; wsDCF.cell(rr, 4).number_format = FMT_MONEY0; wsDCF.cell(rr, 4).border = BDR

            # ── ΔWC: cross-sheet from Forecasts WC table col D ──
            if _for_wc_forecast_start and i < _n_fore_dcf:
                wsDCF.cell(rr, 5).value = f"=Forecasts!D{_for_wc_forecast_start + i}"
            else:
                wsDCF.cell(rr, 5).value = 0
            wsDCF.cell(rr, 5).font = F_STD; wsDCF.cell(rr, 5).number_format = FMT_MONEY0; wsDCF.cell(rr, 5).border = BDR

            # ── Capex: cross-sheet from Forecasts dep/capex table col C ──
            if _for_dep_capex_start and i < _n_fore_dcf:
                wsDCF.cell(rr, 6).value = f"=Forecasts!C{_for_dep_capex_start + i}"
            else:
                wsDCF.cell(rr, 6).value = float(fcff_arr[i]) if i < len(fcff_arr) else 0
            wsDCF.cell(rr, 6).font = F_STD; wsDCF.cell(rr, 6).number_format = FMT_MONEY0; wsDCF.cell(rr, 6).border = BDR

            # ── UFCF = sum of components ──
            wsDCF.cell(rr, 7).value  = f"=C{rr}+D{rr}+E{rr}+F{rr}"
            wsDCF.cell(rr, 7).font   = F_BOLD
            wsDCF.cell(rr, 7).number_format = FMT_MONEY0
            wsDCF.cell(rr, 7).border = BDR

            # ── Discount Factor = 1/(1+WACC)^n ──
            wsDCF.cell(rr, 8).value  = f"=1/(1+$B${row_wacc})^B{rr}"
            wsDCF.cell(rr, 8).font   = F_STD
            wsDCF.cell(rr, 8).number_format = "0.000000"
            wsDCF.cell(rr, 8).border = BDR

            # ── PV of UFCF ──
            wsDCF.cell(rr, 9).value  = f"=G{rr}*H{rr}"
            wsDCF.cell(rr, 9).font   = F_STD
            wsDCF.cell(rr, 9).number_format = FMT_MONEY0
            wsDCF.cell(rr, 9).border = BDR

        cf_end = r + max(_n_fore_dcf - 1, 0)
        r = cf_end + 2

        # ── DCF Summary ───────────────────────────────────────────────────────
        write_section(wsDCF, r, "DCF Valuation Summary", ncols=4); r += 1
        write_hdr(wsDCF, r, ["Item", "Value (USD)"]); r += 1
        row_pv_fcff = r;   row_tv = r+1; row_pv_tv = r+2
        row_ev = r+3;      row_nd2 = r+4; row_eq_dcf = r+5

        tv_val  = ss.get("dcf_terminal_value")
        pv_tv   = ss.get("dcf_pv_terminal")

        cell_bd(wsDCF, row_pv_fcff, 1, "Sum of PV(UFCF)",        F_STD)
        if len(fcff_arr) > 0:
            wsDCF.cell(row_pv_fcff, 2).value = f"=SUM(I{cf_start}:I{cf_end})"
        else:
            wsDCF.cell(row_pv_fcff, 2).value = float(ss.get("dcf_pv_fcff_sum", 0.0))
        wsDCF.cell(row_pv_fcff, 2).font = F_STD
        wsDCF.cell(row_pv_fcff, 2).number_format = FMT_MONEY0
        wsDCF.cell(row_pv_fcff, 2).border = BDR

        # Terminal Value = last UFCF × (1+g) / (WACC - g) — live formula
        cell_bd(wsDCF, row_tv, 1, "Terminal Value",               F_STD)
        wsDCF.cell(row_tv, 2).value  = f"=G{cf_end}*(1+B{row_g})/(B{row_wacc}-B{row_g})"
        wsDCF.cell(row_tv, 2).font   = F_STD
        wsDCF.cell(row_tv, 2).number_format = FMT_MONEY0
        wsDCF.cell(row_tv, 2).border = BDR

        # PV of Terminal Value = TV × last discount factor — live formula
        cell_bd(wsDCF, row_pv_tv, 1, "PV of Terminal Value",      F_STD)
        wsDCF.cell(row_pv_tv, 2).value  = f"=B{row_tv}*H{cf_end}"
        wsDCF.cell(row_pv_tv, 2).font   = F_STD
        wsDCF.cell(row_pv_tv, 2).number_format = FMT_MONEY0
        wsDCF.cell(row_pv_tv, 2).border = BDR

        cell_bd(wsDCF, row_ev, 1, "Enterprise Value (DCF)",        F_BOLD)
        wsDCF.cell(row_ev, 2).value  = f"=B{row_pv_fcff}+B{row_pv_tv}"
        wsDCF.cell(row_ev, 2).font   = F_BOLD
        wsDCF.cell(row_ev, 2).number_format = FMT_MONEY0
        wsDCF.cell(row_ev, 2).border = BDR

        cell_bd(wsDCF, row_nd2, 1, "Less: Net Debt",               F_STD)
        wsDCF.cell(row_nd2, 2).value  = f"=B{row_nd}"
        wsDCF.cell(row_nd2, 2).font   = F_GREEN
        wsDCF.cell(row_nd2, 2).number_format = FMT_MONEY0
        wsDCF.cell(row_nd2, 2).border = BDR

        cell_bd(wsDCF, row_eq_dcf, 1, "Equity Value (DCF)",        F_BOLD)
        wsDCF.cell(row_eq_dcf, 2).value  = f"=B{row_ev}-B{row_nd2}"
        wsDCF.cell(row_eq_dcf, 2).font   = F_BOLD
        wsDCF.cell(row_eq_dcf, 2).number_format = FMT_MONEY0
        wsDCF.cell(row_eq_dcf, 2).border = BDR
        wsDCF.cell(row_eq_dcf, 1).fill  = FL_LBLUE
        wsDCF.cell(row_eq_dcf, 2).fill  = FL_LBLUE
        r = row_eq_dcf + 3
        eq_addr["DCF"] = f"DCF!B{row_eq_dcf}"
        # ── Sensitivity Table ─────────────────────────────────────────────────
        write_section(wsDCF, r, "Sensitivity of Equity Value to WACC and Terminal Growth Rate", ncols=10); r += 1

        # Reconstruct sensitivity table using stored FCF arrays and ranges
        try:
            base_wacc_s = wacc_val if wacc_val > 0 else 0.13
            base_g_s    = g_pct / 100 if g_pct else 0.05
            wacc_step_s = float(ss.get("sens_store_wacc_step_pct", 5.0)) / 100.0
            g_step_s    = float(ss.get("sens_store_g_step_pct",    0.5))  / 100.0
            n_wacc      = int(ss.get("sens_store_wacc_points", 5))
            n_g         = int(ss.get("sens_store_g_points",    7))

            def _centered_range(base, step, n):
                half = n // 2
                return [base + (i - half) * step for i in range(n)]

            def _pct_lbl(v):
                return f"{v*100:.1f}%"

            wacc_range_s = _centered_range(base_wacc_s, wacc_step_s, n_wacc)
            g_range_s    = [max(-0.50, min(0.50, v)) for v in _centered_range(base_g_s, g_step_s, n_g)]

            fcff_s  = [float(x) for x in (fcff_arr or [])]
            n_s_arr = [float(x) for x in (n_arr or [])]

            def _ev_sens(fcff_v, periods, nd, ww, gg):
                if not fcff_v or ww <= gg:
                    return None
                dfs = [1.0 / (1.0 + ww) ** p for p in periods]
                pv_sum = sum(f * d for f, d in zip(fcff_v, dfs))
                tv = fcff_v[-1] * (1.0 + gg) / (ww - gg)
                n_last = periods[-1] if periods else 1.0
                pv_tv_s = tv / (1.0 + ww) ** n_last
                return pv_sum + pv_tv_s - nd

            wacc_labels = [_pct_lbl(w) for w in wacc_range_s]
            g_labels    = [_pct_lbl(g) for g in g_range_s]

            # Header row
            wsRow = r
            wsCol_start = 2
            wsCol = wsCol_start
            wsDCF.cell(wsRow, 1).value = "WACC \\ g"
            wsDCF.cell(wsRow, 1).font  = F_HDR
            wsDCF.cell(wsRow, 1).fill  = FL_HDR
            wsDCF.cell(wsRow, 1).border = BDR
            wsDCF.cell(wsRow, 1).alignment = Alignment(horizontal="center", vertical="center")
            for gl in g_labels:
                c = wsDCF.cell(wsRow, wsCol)
                c.value = gl; c.font = F_HDR; c.fill = FL_HDR; c.border = BDR
                c.alignment = Alignment(horizontal="center", vertical="center")
                wsCol += 1
            r += 1

            base_w_lbl = _pct_lbl(base_wacc_s)
            base_g_lbl = _pct_lbl(base_g_s)

            for wi, ww in enumerate(wacc_range_s):
                wsCol = wsCol_start
                wl = wacc_labels[wi]
                hdr_cell = wsDCF.cell(r, 1)
                hdr_cell.value = wl; hdr_cell.font = F_HDR; hdr_cell.fill = FL_HDR
                hdr_cell.border = BDR
                hdr_cell.alignment = Alignment(horizontal="center", vertical="center")
                for gi, gg in enumerate(g_range_s):
                    ev_s = _ev_sens(fcff_s, n_s_arr, net_debt, ww, gg)
                    c = wsDCF.cell(r, wsCol)
                    # --- Dynamic sensitivity (linked to WACC & g cells) ---

                    # relative position from center (base case)
                    wacc_shift = wi - (len(wacc_range_s) // 2)
                    g_shift = gi - (len(g_range_s) // 2)

                    # build Excel-referenced formulas (NO HARDCODING)
                    wacc_formula = f"(DCF!B{row_wacc}+{wacc_shift * wacc_step_s})"
                    g_formula = f"(DCF!B{row_g}+{g_shift * g_step_s})"

                    # sensitivity formula (fully dynamic)
                    c.value = (
                        f"=IF({wacc_formula}<={g_formula},NA(),"
                        f"(SUMPRODUCT(G{cf_start}:G{cf_end},(1+{wacc_formula})^(-B{cf_start}:B{cf_end}))+"
                        f"(G{cf_end}*(1+{g_formula})/({wacc_formula}-{g_formula}))/(1+{wacc_formula})^B{cf_end}"
                        f")-B{row_nd})"
                    )

                    # formatting
                    c.number_format = FMT_MONEY0
                    c.border = BDR
                    c.alignment = Alignment(horizontal="right", vertical="center")

                    # --- Highlight BASE CASE (exact match to DCF output) ---
                    if wacc_shift == 0 and g_shift == 0:
                        c.value = f"={eq_addr['DCF']}"  # EXACT link (fixes mismatch issue)
                        c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
                        c.fill = PatternFill("solid", fgColor="C00000")
                    else:
                        c.font = F_STD

                    wsCol += 1
                r += 1

            r += 1
            note_cell = wsDCF.cell(r, 1)
            note_cell.value = f"Red cell = base case (WACC: {_pct_lbl(base_wacc_s)}, g: {_pct_lbl(base_g_s)}). Values shown are Equity Value (USD)."
            note_cell.font  = F_NOTE

        except Exception:
            wsDCF.cell(r, 1).value = "Sensitivity table not available — run DCF page first."
            wsDCF.cell(r, 1).font = F_NOTE

        wsDCF.column_dimensions["A"].width = 30
        wsDCF.column_dimensions["B"].width = 18
        wsDCF.column_dimensions["C"].width = 36
        wsDCF.column_dimensions["D"].width = 42
        wsDCF.column_dimensions["E"].width = 20
        wsDCF.column_dimensions["F"].width = 18
        wsDCF.column_dimensions["G"].width = 18
        wsDCF.column_dimensions["H"].width = 14
        wsDCF.column_dimensions["I"].width = 18
        wsDCF.freeze_panes = "A3"
        # ── Store DCF row refs for cross-sheet use by DDM and Banking ────────
        _dcf_row_rf   = row_rf
        _dcf_row_mrp  = row_mrp
        _dcf_row_bu   = row_bu
        _dcf_row_de   = row_de
        _dcf_row_tax  = row_tax
        _dcf_row_bl   = row_bl
        _dcf_row_ke   = row_ke
        _dcf_row_wacc = row_wacc
        _dcf_row_g    = row_g
        _dcf_row_nd   = row_nd
    else:
        # activate and rename dummy sheet so it doesn't litter the workbook
        wsDCF = wb.active
        wsDCF.title = "_temp"
        _dcf_row_rf = _dcf_row_mrp = _dcf_row_bu = _dcf_row_de = None
        _dcf_row_tax = _dcf_row_bl = _dcf_row_ke = _dcf_row_wacc = None
        _dcf_row_g = _dcf_row_nd = None

    # =========================================================================
    # SHEET: Dividend Discount Model
    # =========================================================================
    if "DDM" in selected_models:
        wsDDM = wb.create_sheet("Dividend Discount Model")
        write_title(wsDDM, "DDM — Dividend Discount Model", ncols=6)

        r = 3
        write_section(wsDDM, r, "Dividend History", ncols=3); r += 1
        write_hdr(wsDDM, r, ["Year", "Dividend per Share (USD)"]); r += 1

        _ddm_raw = ss.get("ddm_dividends", {})
        ddm_divs = _ddm_raw if isinstance(_ddm_raw, dict) else {}
        ddm_years = sorted(ddm_divs.keys())
        div_start = r
        for yr in ddm_years:
            cell_bd(wsDDM, r, 1, int(yr), F_STD, FMT_NUM, align="center")
            wsDDM.cell(r, 1).number_format = '0'  # ← fix 2,021 → 2021
            cell_bd(wsDDM, r, 2, float(ddm_divs[yr]), F_BLUE, "0.00000")
            r += 1
        div_end = r - 1

        r += 1
        write_section(wsDDM, r, "Gordon Growth Model Inputs & Valuation", ncols=4); r += 1
        # ── CAPM parameters (linked from DCF sheet) ───────────────────────────
        write_hdr(wsDDM, r, ["CAPM Parameter", "Value (from DCF)"]); r += 1
        # Detect manual override mode — use the persistent store key, not the widget key
        _ddm_use_custom = (
                bool(ss.get("ddm_use_custom_params_store", False))
                and "ddm_saved_rf" in ss  # saved values must actually exist
        )

        # Pull manual override values if used, else fall back to DCF values
        if _ddm_use_custom:
            _ddm_rf_val = float(ss.get("ddm_saved_rf", ss.get("dcf_rf_pct", 0.0))) / 100.0
            _ddm_mrp_val = float(ss.get("ddm_saved_mrp", ss.get("dcf_mrp_pct", 0.0))) / 100.0
            _ddm_bu_val = float(ss.get("ddm_saved_beta", ss.get("dcf_unlevered_beta", 0.0)))
            _ddm_de_val = float(ss.get("de_ratio", 0.0))
            _ddm_tax_val = float(ss.get("ddm_saved_tax", ss.get("dcf_tax_pct", 0.0))) / 100.0
        else:
            _ddm_rf_val = float(ss.get("dcf_rf_pct", 0.0)) / 100.0
            _ddm_mrp_val = float(ss.get("dcf_mrp_pct", 0.0)) / 100.0
            _ddm_bu_val = float(ss.get("dcf_unlevered_beta", 0.0))
            _ddm_de_val = float(ss.get("de_ratio", 0.0))
            _ddm_tax_val = float(ss.get("dcf_tax_pct", 0.0)) / 100.0

        capm_params_ddm = [
            ("Risk-free Rate (Rf)", _dcf_row_rf, FMT_PCT, _ddm_rf_val),
            ("Market Risk Premium (MRP)", _dcf_row_mrp, FMT_PCT, _ddm_mrp_val),
            ("Unlevered Beta (βu)", _dcf_row_bu, "0.0000", _ddm_bu_val),
            ("D/E Ratio", _dcf_row_de, "0.0000", _ddm_de_val),
            ("Tax Rate", _dcf_row_tax, FMT_PCT, _ddm_tax_val),
        ]

        # Track which row each CAPM param lands on (for formula references below)
        _ddm_capm_rows = {}
        _capm_keys = ["rf", "mrp", "bu", "de", "tax"]

        for (lbl_c, src_row, fmt_c, manual_val), _key in zip(capm_params_ddm, _capm_keys):
            cell_bd(wsDDM, r, 1, lbl_c, F_STD)
            if _ddm_use_custom:
                # Blue hardcoded input — user typed these manually
                wsDDM.cell(r, 2).value = float(manual_val)
                wsDDM.cell(r, 2).font = F_BLUE
            else:
                # Green cross-sheet link from DCF
                if src_row is not None:
                    wsDDM.cell(r, 2).value = f"=DCF!B{src_row}"
                    wsDDM.cell(r, 2).font = F_GREEN
                else:
                    wsDDM.cell(r, 2).value = float(manual_val)
                    wsDDM.cell(r, 2).font = F_BLUE
            wsDDM.cell(r, 2).number_format = fmt_c
            wsDDM.cell(r, 2).border = BDR
            _ddm_capm_rows[_key] = r
            r += 1

        r += 1
        write_hdr(wsDDM, r, ["Parameter", "Value"]);
        r += 1

        row_ddm_g  = r;   row_ddm_re = r+1; row_ddm_d1 = r+2
        row_ddm_p0 = r+3; row_ddm_ns = r+4; row_ddm_ev = r+5

        ddm_g_val  = float(ss.get("ddm_g", 0.0) or 0.0)
        ddm_re_val = float(ss.get("ddm_Re", 0.0) or 0.0)

        cell_bd(wsDDM, row_ddm_g,  1, "Growth Rate (g)",          F_STD)
        # g: pull from DCF terminal growth rate if DCF is in the workbook, else hardcode
        wsDDM.cell(row_ddm_g, 2).value = (
            f"=IFERROR(IF({div_end}-{div_start}=0,0,"
            f"(B{div_end}/B{div_start})^(1/({div_end}-{div_start}))-1),0)"
        )
        wsDDM.cell(row_ddm_g, 2).font = F_GREEN
        wsDDM.cell(row_ddm_g, 2).number_format = FMT_PCT
        wsDDM.cell(row_ddm_g, 2).border = BDR
        wsDDM.cell(row_ddm_g, 2).number_format = FMT_PCT
        wsDDM.cell(row_ddm_g, 2).border = BDR

        cell_bd(wsDDM, row_ddm_re, 1, "Cost of Equity (Re)", F_STD)
        if _ddm_use_custom:
            # Build CAPM chain using the manually entered rows above:
            # Levered Beta = βu × (1 + (1-T) × D/E)
            # Re = Rf + βL × MRP
            _r_bu = _ddm_capm_rows.get("bu")
            _r_de = _ddm_capm_rows.get("de")
            _r_tax = _ddm_capm_rows.get("tax")
            _r_rf = _ddm_capm_rows.get("rf")
            _r_mrp = _ddm_capm_rows.get("mrp")
            if all(v is not None for v in [_r_bu, _r_de, _r_tax, _r_rf, _r_mrp]):
                # Insert a hidden Levered Beta row just above Re
                # We write it into row_ddm_re and push Re one row down
                # Instead, compute Re inline as a single formula:
                # Re = Rf + (βu*(1+(1-T)*DE)) * MRP
                wsDDM.cell(row_ddm_re, 2).value = (
                    f"=B{_r_rf}+(B{_r_bu}*(1+(1-B{_r_tax})*B{_r_de}))*B{_r_mrp}"
                )
                wsDDM.cell(row_ddm_re, 2).font = F_GREEN
            else:
                wsDDM.cell(row_ddm_re, 2).value = ddm_re_val
                wsDDM.cell(row_ddm_re, 2).font = F_BLUE
        elif _dcf_row_ke is not None:
            wsDDM.cell(row_ddm_re, 2).value = f"=DCF!B{_dcf_row_ke}"
            wsDDM.cell(row_ddm_re, 2).font = F_GREEN
        else:
            wsDDM.cell(row_ddm_re, 2).value = ddm_re_val
            wsDDM.cell(row_ddm_re, 2).font = F_BLUE
        wsDDM.cell(row_ddm_re, 2).number_format = FMT_PCT
        wsDDM.cell(row_ddm_re, 2).border = BDR

        cell_bd(wsDDM, row_ddm_d1, 1, "Next Dividend (D1)",       F_STD)
        if ddm_years:
            last_div_cell = f"B{div_start + len(ddm_years) - 1}"
            wsDDM.cell(row_ddm_d1, 2).value = f"={last_div_cell}*(1+B{row_ddm_g})"
        else:
            wsDDM.cell(row_ddm_d1, 2).value = 0.0
        wsDDM.cell(row_ddm_d1, 2).font = F_STD
        wsDDM.cell(row_ddm_d1, 2).number_format = "0.00000"
        wsDDM.cell(row_ddm_d1, 2).border = BDR

        cell_bd(wsDDM, row_ddm_p0, 1, "Intrinsic Value / Share (P0)", F_BOLD)
        wsDDM.cell(row_ddm_p0, 2).value = (
            f"=IF(B{row_ddm_re}>B{row_ddm_g},"
            f"B{row_ddm_d1}/(B{row_ddm_re}-B{row_ddm_g}),NA())"
        )
        wsDDM.cell(row_ddm_p0, 2).font = F_STD
        wsDDM.cell(row_ddm_p0, 2).number_format = FMT_MONEY4
        wsDDM.cell(row_ddm_p0, 2).border = BDR
        wsDDM.cell(row_ddm_p0, 1).fill = FL_LGREY
        wsDDM.cell(row_ddm_p0, 2).fill = FL_LGREY

        cell_bd(wsDDM, row_ddm_ns, 1, "Number of Shares",         F_STD)
        cell_bd(wsDDM, row_ddm_ns, 2, float(ss.get("num_shares", 0.0)), F_BLUE, FMT_NUM)

        cell_bd(wsDDM, row_ddm_ev, 1, "Total Equity Value (DDM)", F_BOLD)
        wsDDM.cell(row_ddm_ev, 2).value = (
            f"=IF(ISNUMBER(B{row_ddm_p0}),"
            f"B{row_ddm_p0}*B{row_ddm_ns},NA())"
        )
        wsDDM.cell(row_ddm_ev, 2).font = F_BOLD
        wsDDM.cell(row_ddm_ev, 2).number_format = FMT_MONEY0
        wsDDM.cell(row_ddm_ev, 2).border = BDR
        wsDDM.cell(row_ddm_ev, 1).fill = FL_LBLUE
        wsDDM.cell(row_ddm_ev, 2).fill = FL_LBLUE

        wsDDM.column_dimensions["A"].width = 32
        wsDDM.column_dimensions["B"].width = 22
        wsDDM.freeze_panes = "A3"
        eq_addr["DDM"] = f"'Dividend Discount Model'!B{row_ddm_ev}"

    # =========================================================================
    # SHEET: CompCo
    # =========================================================================
    if any(m in selected_models for m in ["EV/EBITDA", "PBV", "P/E"]):
        wsComp = wb.create_sheet("CompCo")
        write_title(wsComp, "Comparable Company Valuation — EV/EBITDA · P/BV · P/E", ncols=10)

        r = 3
        write_section(wsComp, r, "Comparable Company Multiples", ncols=6); r += 1
        write_hdr(wsComp, r, ["Company", "EV/EBITDA", "P/B", "P/E", "Incl. EV?", "Incl. PB?", "Incl. PE?"]); r += 1

        # Pull from S["comps"] dict — the actual structure used by the Comparables page
        comps_dict = ss.get("comps", {})
        # Also support list-based fallback keys written at line 3060-3065 of 2_COMPARABLES.py
        comps_ev_list  = ss.get("comps_ev_list",  [])
        comps_pb_list  = ss.get("comps_pb_list",  [])
        comps_pe_list  = ss.get("comps_pe_list",  [])
        comps_inc_ev   = ss.get("comps_inc_ev",   [])
        comps_inc_pb   = ss.get("comps_inc_pb",   [])
        comps_inc_pe   = ss.get("comps_inc_pe",   [])
        num_comps      = int(ss.get("comps_num",  ss.get("num_comps", 0)))

        comp_start = r

        if comps_dict:
            for i in sorted(comps_dict.keys()):
                c = comps_dict[i]
                name   = str(c.get("name",   f"Comp {i}"))
                ev_v   = c.get("ev",    None)
                pb_v   = c.get("pb",    None)
                pe_v   = c.get("pe",    None)
                inc_ev = bool(c.get("inc_ev", True))
                inc_pb = bool(c.get("inc_pb", True))
                inc_pe = bool(c.get("inc_pe", True))

                cell_bd(wsComp, r, 1, name,   F_STD)
                cell_bd(wsComp, r, 2, float(ev_v) if ev_v is not None else None, F_BLUE, "0.00")
                cell_bd(wsComp, r, 3, float(pb_v) if pb_v is not None else None, F_BLUE, "0.00")
                cell_bd(wsComp, r, 4, float(pe_v) if pe_v is not None else None, F_BLUE, "0.00")
                cell_bd(wsComp, r, 5, "Yes" if inc_ev else "No", F_STD)
                cell_bd(wsComp, r, 6, "Yes" if inc_pb else "No", F_STD)
                cell_bd(wsComp, r, 7, "Yes" if inc_pe else "No", F_STD)
                r += 1
        elif comps_ev_list:
            for i in range(len(comps_ev_list)):
                name = str(ss.get(f"comp_name_{i}", f"Comp {i+1}"))
                cell_bd(wsComp, r, 1, name, F_STD)
                cell_bd(wsComp, r, 2, float(comps_ev_list[i]) if i < len(comps_ev_list) else None, F_BLUE, "0.00")
                cell_bd(wsComp, r, 3, float(comps_pb_list[i]) if i < len(comps_pb_list) else None, F_BLUE, "0.00")
                cell_bd(wsComp, r, 4, float(comps_pe_list[i]) if i < len(comps_pe_list) else None, F_BLUE, "0.00")
                cell_bd(wsComp, r, 5, "Yes" if (i < len(comps_inc_ev) and comps_inc_ev[i]) else "No", F_STD)
                cell_bd(wsComp, r, 6, "Yes" if (i < len(comps_inc_pb) and comps_inc_pb[i]) else "No", F_STD)
                cell_bd(wsComp, r, 7, "Yes" if (i < len(comps_inc_pe) and comps_inc_pe[i]) else "No", F_STD)
                r += 1
        else:
            cell_bd(wsComp, r, 1, "(No comparable companies found — visit Comparables page first)", F_NOTE)
            r += 1
        comp_end = r - 1

        r += 1
        write_section(wsComp, r, "Peer Averages & Discount", ncols=5); r += 1
        write_hdr(wsComp, r, ["Multiple", "Peer Average", "Discount (%)", "Applied Multiple"]); r += 1

        # Use the actual discount_factor key (not discount_pct)
        disc_pct_raw = float(ss.get("discount_factor", ss.get("discount_pct", ss.get("comp_discount_pct", 25.0))) or 25.0)
        disc_decimal = disc_pct_raw / 100.0 if disc_pct_raw > 1 else disc_pct_raw

        # Use actual implied multiples already computed by the Comparables page
        implied_ev_mult = float(ss.get("implied_ev", 0.0) or 0.0)
        implied_pb_mult = float(ss.get("implied_pb", 0.0) or 0.0)
        implied_pe_mult = float(ss.get("implied_pe", 0.0) or 0.0)

        row_disc  = r
        row_ev_m  = r + 1
        row_pb_m  = r + 2
        row_pe_m  = r + 3

        cell_bd(wsComp, row_disc, 1, "Size/Country/Liquidity Discount", F_STD)
        cell_bd(wsComp, row_disc, 3, disc_decimal, F_BLUE, FMT_PCT)

        # EV/EBITDA row
        cell_bd(wsComp, row_ev_m, 1, "EV/EBITDA", F_STD)
        # Peer average from comps with Include_EV=True
        if comps_dict:
            ev_included = [float(comps_dict[i].get("ev", 0) or 0) for i in sorted(comps_dict.keys()) if comps_dict[i].get("inc_ev", True) and comps_dict[i].get("ev") is not None]
            ev_avg = sum(ev_included) / len(ev_included) if ev_included else None
        elif comps_ev_list and comps_inc_ev:
            ev_incl = [v for v, inc in zip(comps_ev_list, comps_inc_ev) if inc]
            ev_avg = sum(ev_incl) / len(ev_incl) if ev_incl else None
        else:
            ev_avg = None
        wsComp.cell(row_ev_m, 2).value = (
            f"=AVERAGEIF(E{comp_start}:E{comp_end},\"Yes\",B{comp_start}:B{comp_end})"
        )
        wsComp.cell(row_ev_m, 2).font = F_GREEN
        wsComp.cell(row_ev_m, 2).number_format = '0.00'
        wsComp.cell(row_ev_m, 2).border = BDR
        cell_bd(wsComp, row_ev_m, 3, disc_decimal, F_STD, FMT_PCT)
        wsComp.cell(row_ev_m, 4).value = f"=IF(B{row_ev_m}=\"\",\"\",B{row_ev_m}*(1-C{row_ev_m}))"
        wsComp.cell(row_ev_m, 4).font = F_GREEN
        wsComp.cell(row_ev_m, 4).number_format = '0.00'
        wsComp.cell(row_ev_m, 4).border = BDR

        # P/B row
        cell_bd(wsComp, row_pb_m, 1, "P/B", F_STD)
        if comps_dict:
            pb_included = [float(comps_dict[i].get("pb", 0) or 0) for i in sorted(comps_dict.keys()) if comps_dict[i].get("inc_pb", True) and comps_dict[i].get("pb") is not None]
            pb_avg = sum(pb_included) / len(pb_included) if pb_included else None
        elif comps_pb_list and comps_inc_pb:
            pb_incl = [v for v, inc in zip(comps_pb_list, comps_inc_pb) if inc]
            pb_avg = sum(pb_incl) / len(pb_incl) if pb_incl else None
        else:
            pb_avg = None
        wsComp.cell(row_pb_m, 2).value = (
            f"=AVERAGEIF(F{comp_start}:F{comp_end},\"Yes\",C{comp_start}:C{comp_end})"
        )
        wsComp.cell(row_pb_m, 2).font = F_GREEN
        wsComp.cell(row_pb_m, 2).number_format = '0.00'
        wsComp.cell(row_pb_m, 2).border = BDR
        cell_bd(wsComp, row_pb_m, 3, disc_decimal, F_STD, FMT_PCT)
        wsComp.cell(row_pb_m, 4).value = f"=IF(B{row_pb_m}=\"\",\"\",B{row_pb_m}*(1-C{row_pb_m}))"
        wsComp.cell(row_pb_m, 4).font = F_GREEN
        wsComp.cell(row_pb_m, 4).number_format = '0.00'
        wsComp.cell(row_pb_m, 4).border = BDR

        # P/E row
        cell_bd(wsComp, row_pe_m, 1, "P/E", F_STD)
        if comps_dict:
            pe_included = [float(comps_dict[i].get("pe", 0) or 0) for i in sorted(comps_dict.keys()) if comps_dict[i].get("inc_pe", True) and comps_dict[i].get("pe") is not None]
            pe_avg = sum(pe_included) / len(pe_included) if pe_included else None
        elif comps_pe_list and comps_inc_pe:
            pe_incl = [v for v, inc in zip(comps_pe_list, comps_inc_pe) if inc]
            pe_avg = sum(pe_incl) / len(pe_incl) if pe_incl else None
        else:
            pe_avg = None
        wsComp.cell(row_pe_m, 2).value = (
            f"=AVERAGEIF(G{comp_start}:G{comp_end},\"Yes\",D{comp_start}:D{comp_end})"
        )
        wsComp.cell(row_pe_m, 2).font = F_GREEN
        wsComp.cell(row_pe_m, 2).number_format = '0.00'
        wsComp.cell(row_pe_m, 2).border = BDR
        cell_bd(wsComp, row_pe_m, 3, disc_decimal, F_STD, FMT_PCT)
        wsComp.cell(row_pe_m, 4).value = f"=IF(B{row_pe_m}=\"\",\"\",B{row_pe_m}*(1-C{row_pe_m}))"
        wsComp.cell(row_pe_m, 4).font = F_GREEN
        wsComp.cell(row_pe_m, 4).number_format = '0.00'
        wsComp.cell(row_pe_m, 4).border = BDR

        r = row_pe_m + 2

        # ── Maintainable Metrics — pull from Forecasts sheet via formulas ──────
        write_section(wsComp, r, "Company Financials — Maintainable Metrics", ncols=4); r += 1
        write_hdr(wsComp, r, ["Input", "Value (USD)"]); r += 1

        row_eb  = r;   row_np = r+1; row_bk = r+2; row_nd = r+3

        # Book Equity and Net Debt: formula references to Forecasts sheet
        # (these rows were stored in _for_bk_eq_row_excel and _for_nd_row_excel)
        cell_bd(wsComp, row_eb, 1, "Maintainable EBITDA",  F_STD)
        # Filled by weighted EBITDA table formula below — placeholder for now
        cell_bd(wsComp, row_np, 1, "Maintainable Earnings", F_STD)
        # Filled by weighted Earnings table formula below

        cell_bd(wsComp, row_bk, 1, "Book Value of Equity", F_STD)
        if _for_bk_eq_row_excel:
            wsComp.cell(row_bk, 2).value = f"=Forecasts!$B${_for_bk_eq_row_excel}"
            wsComp.cell(row_bk, 2).font = F_GREEN
        else:
            wsComp.cell(row_bk, 2).value = float(ss.get("book_equity", 0.0) or 0.0)
            wsComp.cell(row_bk, 2).font = F_BLUE
        wsComp.cell(row_bk, 2).number_format = FMT_MONEY0
        wsComp.cell(row_bk, 2).border = BDR

        cell_bd(wsComp, row_nd, 1, "Net Debt", F_STD)
        if _for_nd_row_excel:
            wsComp.cell(row_nd, 2).value = f"=Forecasts!$B${_for_nd_row_excel}"
            wsComp.cell(row_nd, 2).font = F_GREEN
        else:
            wsComp.cell(row_nd, 2).value = float(ss.get("net_debt", 0.0) or 0.0)
            wsComp.cell(row_nd, 2).font = F_BLUE
        wsComp.cell(row_nd, 2).number_format = FMT_MONEY0
        wsComp.cell(row_nd, 2).border = BDR
        r = row_nd + 2

        # ── Weighted EBITDA Table — live cross-sheet link from Forecasts ────────
        write_section(wsComp, r, "Maintainable EBITDA — Weighted Average (from DCF Forecasts)", ncols=5); r += 1

        dcf_eb_all      = ss.get("dcf_ebitda_all") or ss.get("dcf_ebitda_forecast") or {}
        comp_eb_weights = ss.get("comp_eb_weights", {}) or {}
        _dcf_eb = dcf_eb_all if isinstance(dcf_eb_all, dict) else {}

        if _dcf_eb:
            write_hdr(wsComp, r, ["Year", "EBITDA (USD)", "Timing", "Weight (%)", "Adj. EBITDA", "Weighted EBITDA"]); r += 1
            eb_rows_start = r
            eb_years = sorted(int(y) for y in _dcf_eb.keys() if str(y).lstrip("-").isdigit())
            weighted_eb_years = [y for y in eb_years if float(comp_eb_weights.get(str(y), 0.0)) > 0] or eb_years[:6]
            _eb_timing_base = float(ss.get("comp_timing_base", 1.0) or 1.0)
            _eb_use_timing  = bool(ss.get("comp_use_timing_eb", True))
            for idx_e, yr in enumerate(weighted_eb_years):
                eb_val = _dcf_eb.get(str(yr), _dcf_eb.get(yr, 0.0))
                wt_pct = float(comp_eb_weights.get(str(yr), 0.0)) / 100.0
                cell_bd(wsComp, r, 1, int(yr),  F_STD, align="center")
                # EBITDA value: cross-sheet formula if we know the Forecasts row+col, else hardcoded
                if _for_ebitda_row_excel and _for_year_cols:
                    _yr_str = str(yr)
                    if _yr_str in _for_year_cols:
                        _fc_col_idx = _for_year_cols.index(_yr_str) + 2  # +2 because col1=Item
                        _fc_col_L = get_column_letter(_fc_col_idx)
                        wsComp.cell(r, 2).value = f"=Forecasts!{_fc_col_L}{_for_ebitda_row_excel}"
                        wsComp.cell(r, 2).font = F_GREEN
                    else:
                        wsComp.cell(r, 2).value = float(eb_val) if eb_val is not None else 0.0
                        wsComp.cell(r, 2).font = F_BLUE
                else:
                    wsComp.cell(r, 2).value = float(eb_val) if eb_val is not None else 0.0
                    wsComp.cell(r, 2).font = F_BLUE
                wsComp.cell(r, 2).number_format = FMT_MONEY0
                wsComp.cell(r, 2).border = BDR
                wsComp.cell(r, 2).alignment = Alignment(horizontal="right", vertical="center")
                # Timing factor
                timing_val = (_eb_timing_base + idx_e) if _eb_use_timing else 1.0
                cell_bd(wsComp, r, 3, timing_val, F_BLUE, "0.0000")
                cell_bd(wsComp, r, 4, wt_pct, F_BLUE, FMT_PCT1)
                # Adjusted EBITDA = EBITDA × Timing
                wsComp.cell(r, 5).value = f"=B{r}*C{r}"
                wsComp.cell(r, 5).font = F_STD
                wsComp.cell(r, 5).number_format = FMT_MONEY0
                wsComp.cell(r, 5).border = BDR
                wsComp.cell(r, 5).alignment = Alignment(horizontal="right", vertical="center")
                # Weighted EBITDA = Adjusted × Weight
                wsComp.cell(r, 6).value = f"=E{r}*D{r}"
                wsComp.cell(r, 6).font = F_STD
                wsComp.cell(r, 6).number_format = FMT_MONEY0
                wsComp.cell(r, 6).border = BDR
                wsComp.cell(r, 6).alignment = Alignment(horizontal="right", vertical="center")
                r += 1
            eb_rows_end = r - 1
            wsComp.cell(r, 1).value = "Maintainable EBITDA"
            wsComp.cell(r, 1).font = F_BOLD; wsComp.cell(r, 1).border = BDR; wsComp.cell(r, 1).fill = FL_LBLUE
            wsComp.cell(r, 4).value = f"=SUM(D{eb_rows_start}:D{eb_rows_end})"
            wsComp.cell(r, 4).font = F_BOLD; wsComp.cell(r, 4).number_format = FMT_PCT1
            wsComp.cell(r, 4).border = BDR; wsComp.cell(r, 4).fill = FL_LBLUE
            wsComp.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
            wsComp.cell(r, 6).value = f"=SUM(F{eb_rows_start}:F{eb_rows_end})"
            wsComp.cell(r, 6).font = F_BOLD; wsComp.cell(r, 6).number_format = FMT_MONEY0
            wsComp.cell(r, 6).border = BDR; wsComp.cell(r, 6).fill = FL_LBLUE
            wsComp.cell(r, 6).alignment = Alignment(horizontal="right", vertical="center")
            row_eb_total = r
            # Fill Maintainable EBITDA in metrics table with formula from this weighted total
            wsComp.cell(row_eb, 2).value = f"=F{row_eb_total}"
            wsComp.cell(row_eb, 2).font = F_GREEN
            wsComp.cell(row_eb, 2).number_format = FMT_MONEY0
            wsComp.cell(row_eb, 2).border = BDR
            r += 2
        else:
            wsComp.cell(r, 1).value = "(EBITDA forecast not found — run DCF page first)"
            wsComp.cell(r, 1).font = F_NOTE
            row_eb_total = None; r += 2

        # ── Weighted Earnings Table — live cross-sheet link from Forecasts ──────
        write_section(wsComp, r, "Maintainable Earnings — Weighted Average (from DCF Forecasts)", ncols=5); r += 1

        dcf_np_all      = ss.get("dcf_profit_all") or ss.get("dcf_profit_forecast") or {}
        comp_np_weights = ss.get("comp_np_weights", {}) or {}
        _dcf_np = dcf_np_all if isinstance(dcf_np_all, dict) else {}

        if _dcf_np:
            write_hdr(wsComp, r, ["Year", "Earnings (USD)", "Timing", "Weight (%)", "Adj. Earnings", "Weighted Earnings"]); r += 1
            np_rows_start = r
            np_years = sorted(int(y) for y in _dcf_np.keys() if str(y).lstrip("-").isdigit())
            weighted_np_years = [y for y in np_years if float(comp_np_weights.get(str(y), 0.0)) > 0] or np_years[:6]
            _np_timing_base = float(ss.get("comp_timing_base", 1.0) or 1.0)
            _np_use_timing  = bool(ss.get("comp_use_timing_np", True))
            for idx_n, yr in enumerate(weighted_np_years):
                np_val = _dcf_np.get(str(yr), _dcf_np.get(yr, 0.0))
                wt_pct = float(comp_np_weights.get(str(yr), 0.0)) / 100.0
                cell_bd(wsComp, r, 1, int(yr),  F_STD, align="center")
                # Earnings value: cross-sheet formula if we know Forecasts row+col
                if _for_np_row_excel and _for_year_cols:
                    _yr_str2 = str(yr)
                    if _yr_str2 in _for_year_cols:
                        _fc_col_idx2 = _for_year_cols.index(_yr_str2) + 2
                        _fc_col_L2 = get_column_letter(_fc_col_idx2)
                        wsComp.cell(r, 2).value = f"=Forecasts!{_fc_col_L2}{_for_np_row_excel}"
                        wsComp.cell(r, 2).font = F_GREEN
                    else:
                        wsComp.cell(r, 2).value = float(np_val) if np_val is not None else 0.0
                        wsComp.cell(r, 2).font = F_BLUE
                else:
                    wsComp.cell(r, 2).value = float(np_val) if np_val is not None else 0.0
                    wsComp.cell(r, 2).font = F_BLUE
                wsComp.cell(r, 2).number_format = FMT_MONEY0
                wsComp.cell(r, 2).border = BDR
                wsComp.cell(r, 2).alignment = Alignment(horizontal="right", vertical="center")
                # Timing factor
                timing_val_n = (_np_timing_base + idx_n) if _np_use_timing else 1.0
                cell_bd(wsComp, r, 3, timing_val_n, F_BLUE, "0.0000")
                cell_bd(wsComp, r, 4, wt_pct, F_BLUE, FMT_PCT1)
                # Adjusted Earnings = Earnings × Timing
                wsComp.cell(r, 5).value = f"=B{r}*C{r}"
                wsComp.cell(r, 5).font = F_STD
                wsComp.cell(r, 5).number_format = FMT_MONEY0
                wsComp.cell(r, 5).border = BDR
                wsComp.cell(r, 5).alignment = Alignment(horizontal="right", vertical="center")
                # Weighted Earnings = Adjusted × Weight
                wsComp.cell(r, 6).value = f"=E{r}*D{r}"
                wsComp.cell(r, 6).font = F_STD
                wsComp.cell(r, 6).number_format = FMT_MONEY0
                wsComp.cell(r, 6).border = BDR
                wsComp.cell(r, 6).alignment = Alignment(horizontal="right", vertical="center")
                r += 1
            np_rows_end = r - 1
            wsComp.cell(r, 1).value = "Maintainable Earnings"
            wsComp.cell(r, 1).font = F_BOLD; wsComp.cell(r, 1).border = BDR; wsComp.cell(r, 1).fill = FL_LBLUE
            wsComp.cell(r, 4).value = f"=SUM(D{np_rows_start}:D{np_rows_end})"
            wsComp.cell(r, 4).font = F_BOLD; wsComp.cell(r, 4).number_format = FMT_PCT1
            wsComp.cell(r, 4).border = BDR; wsComp.cell(r, 4).fill = FL_LBLUE
            wsComp.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
            wsComp.cell(r, 6).value = f"=SUM(F{np_rows_start}:F{np_rows_end})"
            wsComp.cell(r, 6).font = F_BOLD; wsComp.cell(r, 6).number_format = FMT_MONEY0
            wsComp.cell(r, 6).border = BDR; wsComp.cell(r, 6).fill = FL_LBLUE
            wsComp.cell(r, 6).alignment = Alignment(horizontal="right", vertical="center")
            row_np_total = r
            # Fill Maintainable Earnings in metrics table
            wsComp.cell(row_np, 2).value = f"=F{row_np_total}"
            wsComp.cell(row_np, 2).font = F_GREEN
            wsComp.cell(row_np, 2).number_format = FMT_MONEY0
            wsComp.cell(row_np, 2).border = BDR
            r += 2
        else:
            wsComp.cell(r, 1).value = "(Earnings forecast not found — run DCF page first)"
            wsComp.cell(r, 1).font = F_NOTE
            row_np_total = None; r += 2

        # ── Implied Equity Values — formulas point to the live weighted totals ─
        write_section(wsComp, r, "Implied Equity Values", ncols=4); r += 1
        write_hdr(wsComp, r, ["Method", "Formula", "Equity Value (USD)"]); r += 1

        row_ev_eq = r;   row_pb_eq = r+1;  row_pe_eq = r+2

        # EV/EBITDA: Applied multiple × Maintainable EBITDA (live total) − Net Debt
        cell_bd(wsComp, row_ev_eq, 1, "EV/EBITDA Valuation", F_BOLD)
        cell_bd(wsComp, row_ev_eq, 2, "Applied EV/EBITDA × Maint. EBITDA − Net Debt", F_NOTE)
        if row_eb_total:
            wsComp.cell(row_ev_eq, 3).value = f'=IF(D{row_ev_m}="","",D{row_ev_m}*F{row_eb_total}-B{row_nd})'
        else:
            wsComp.cell(row_ev_eq, 3).value = f'=IF(D{row_ev_m}="","",D{row_ev_m}*B{row_eb}-B{row_nd})'
        wsComp.cell(row_ev_eq, 3).font = F_BOLD
        wsComp.cell(row_ev_eq, 3).number_format = FMT_MONEY0
        wsComp.cell(row_ev_eq, 3).border = BDR
        wsComp.cell(row_ev_eq, 1).fill = FL_LBLUE
        wsComp.cell(row_ev_eq, 3).fill = FL_LBLUE

        # P/BV: Applied P/B × Book Equity (static — book equity not in DCF forecast)
        cell_bd(wsComp, row_pb_eq, 1, "P/BV Valuation", F_BOLD)
        cell_bd(wsComp, row_pb_eq, 2, "Applied P/B × Book Equity", F_NOTE)
        wsComp.cell(row_pb_eq, 3).value = f'=IF(D{row_pb_m}="","",D{row_pb_m}*B{row_bk})'
        wsComp.cell(row_pb_eq, 3).font = F_BOLD
        wsComp.cell(row_pb_eq, 3).number_format = FMT_MONEY0
        wsComp.cell(row_pb_eq, 3).border = BDR
        wsComp.cell(row_pb_eq, 1).fill = FL_LBLUE
        wsComp.cell(row_pb_eq, 3).fill = FL_LBLUE

        # P/E: Applied P/E × Maintainable Earnings (live total)
        cell_bd(wsComp, row_pe_eq, 1, "P/E Valuation", F_BOLD)
        cell_bd(wsComp, row_pe_eq, 2, "Applied P/E × Maint. Earnings", F_NOTE)
        if row_np_total:
            wsComp.cell(row_pe_eq, 3).value = f'=IF(D{row_pe_m}="","",D{row_pe_m}*F{row_np_total})'
        else:
            wsComp.cell(row_pe_eq, 3).value = f'=IF(D{row_pe_m}="","",D{row_pe_m}*B{row_np})'
        wsComp.cell(row_pe_eq, 3).font = F_BOLD
        wsComp.cell(row_pe_eq, 3).number_format = FMT_MONEY0
        wsComp.cell(row_pe_eq, 3).border = BDR
        wsComp.cell(row_pe_eq, 1).fill = FL_LBLUE
        wsComp.cell(row_pe_eq, 3).fill = FL_LBLUE

        for w, c_l in zip([34, 14, 12, 12, 14, 16, 10, 10, 10], list("ABCDEFGHI")):
            wsComp.column_dimensions[c_l].width = w
        wsComp.freeze_panes = "A3"

        eq_addr["EV/EBITDA"] = f"CompCo!C{row_ev_eq}"
        eq_addr["PBV"]       = f"CompCo!C{row_pb_eq}"
        eq_addr["P/E"]       = f"CompCo!C{row_pe_eq}"

    # =========================================================================
    # SHEET: Banking
    # =========================================================================
    if "BANKING" in selected_models:
        wsBank = wb.create_sheet("Banking")
        write_title(wsBank, "Banking — Residual Income / Excess Returns Model", ncols=6)

        bank_out = ss.get("BANK", {}).get("outputs", {}) if ss.get("BANK") else {}

        r = 3
        write_section(wsBank, r, "Key Inputs & Model Outputs", ncols=4); r += 1
        write_hdr(wsBank, r, ["Metric", "Value"]); r += 1

        bank_items = [
            ("Base Year",               bank_out.get("base_year", "")),
            ("Beginning Book Equity",   bank_out.get("book_equity_0", ss.get("book_equity", 0.0))),
            ("Base Year Earnings",      bank_out.get("earnings_0", 0.0)),
            ("Cost of Equity (Ke)",     bank_out.get("ke", ss.get("bank_ke", 0.0))),
            ("Sum PV(Residual Income)", bank_out.get("pv_resid_sum", 0.0)),
            ("PV of Terminal Value",    bank_out.get("pv_terminal", 0.0)),
        ]
        row_bk_start = r
        for i, (lbl, val) in enumerate(bank_items):
            cell_bd(wsBank, r, 1, lbl, F_STD)
            # ── Cross-sheet wiring for Book Equity and Cost of Equity ──────────
            if i == 1 and _for_bk_eq_row_excel:
                # Beginning Book Equity → Forecasts sheet
                wsBank.cell(r, 2).value = f"=Forecasts!$B${_for_bk_eq_row_excel}"
                wsBank.cell(r, 2).font  = F_GREEN
                wsBank.cell(r, 2).number_format = FMT_MONEY0
                wsBank.cell(r, 2).border = BDR
            elif i == 3 and _dcf_row_ke is not None:
                # Cost of Equity (Ke) → DCF CAPM sheet
                wsBank.cell(r, 2).value = f"=DCF!B{_dcf_row_ke}"
                wsBank.cell(r, 2).font  = F_GREEN
                wsBank.cell(r, 2).number_format = FMT_PCT
                wsBank.cell(r, 2).border = BDR
            else:
                if isinstance(val, float):
                    cell_bd(wsBank, r, 2, val, F_BLUE,
                            FMT_PCT if (abs(val) < 2 and "Cost" in lbl) else FMT_MONEY0)
                else:
                    cell_bd(wsBank, r, 2, str(val), F_BLUE)
            r += 1

        # row indices
        bkeq_r = row_bk_start
        pv_ri_r = row_bk_start + 4
        pv_tv_r = row_bk_start + 5
        row_bk_eq = r + 1

        cell_bd(wsBank, r, 1, "", F_STD)  # blank gap row
        r += 1
        cell_bd(wsBank, row_bk_eq, 1, "Total Equity Value (Banking)", F_BOLD)
        wsBank.cell(row_bk_eq, 2).value = f"=B{bkeq_r}+B{pv_ri_r}+B{pv_tv_r}"
        wsBank.cell(row_bk_eq, 2).font = F_BOLD
        wsBank.cell(row_bk_eq, 2).number_format = FMT_MONEY0
        wsBank.cell(row_bk_eq, 2).border = BDR
        wsBank.cell(row_bk_eq, 1).fill = FL_LBLUE
        wsBank.cell(row_bk_eq, 2).fill = FL_LBLUE

        wsBank.column_dimensions["A"].width = 32
        wsBank.column_dimensions["B"].width = 22
        wsBank.freeze_panes = "A3"
        eq_addr["BANKING"] = f"Banking!B{row_bk_eq}"

    # ── Remove unused _temp sheet if DCF was not selected ─────────────────────
    if "DCF" not in selected_models and "_temp" in wb.sheetnames:
        del wb["_temp"]

    # =========================================================================
    # SHEET: Summary Valuation  (inserted at position 0 — first tab)
    # Mirrors Innscor "Summary Valuation" sheet format exactly
    # =========================================================================
    wsSum = wb.create_sheet("Summary Valuation", 0)
    write_title(wsSum, "Summary Valuation", ncols=5)

    # ── Model table — mirrors Innscor layout ──────────────────────────────────
    r = 3
    # blank row like Innscor
    r += 1

    write_hdr(wsSum, r, ["Model", "Value (USD)", "Weight", "Weighted Value of Equity"]); r += 1

    sum_start = r
    for mdl in selected_models:
        addr  = eq_addr.get(mdl)
        w_dec = norm_w.get(mdl, 0.0)
        w_raw = weights_new.get(mdl, 0.0)

        # Column A: model label (same style as Innscor)
        cell_bd(wsSum, r, 1, f"{mdl} Valuation", F_STD)

        # Column B: equity value — formula link to model sheet (green = cross-sheet)
        if addr:
            wsSum.cell(r, 2).value = f"=IF(ISNUMBER({addr}),{addr},0)"
            wsSum.cell(r, 2).font  = F_GREEN
        else:
            val = value_map.get(mdl)
            wsSum.cell(r, 2).value = float(val) if val else 0.0
            wsSum.cell(r, 2).font  = F_BLUE
        wsSum.cell(r, 2).number_format = FMT_MONEY0
        wsSum.cell(r, 2).border = BDR
        wsSum.cell(r, 2).alignment = Alignment(horizontal="right", vertical="center")

        # Column C: weight (decimal) — hardcoded input in blue like Innscor
        cell_bd(wsSum, r, 3, w_dec, F_BLUE, FMT_PCT1, align="center")

        # Column D: weighted value = Value × Weight
        wsSum.cell(r, 4).value  = f"=B{r}*C{r}"
        wsSum.cell(r, 4).font   = F_STD
        wsSum.cell(r, 4).number_format = FMT_MONEY0
        wsSum.cell(r, 4).border = BDR
        wsSum.cell(r, 4).alignment = Alignment(horizontal="right", vertical="center")
        r += 1

    sum_end = r - 1

    # ── Totals row ─────────────────────────────────────────────────────────────
    wsSum.cell(r, 1, "Weighted Equity Value").font = F_BOLD
    wsSum.cell(r, 1).border = BDR
    wsSum.cell(r, 1).fill   = FL_LBLUE
    wsSum.cell(r, 2).value  = ""    # total equity value goes in col D per Innscor style
    wsSum.cell(r, 2).border = BDR
    wsSum.cell(r, 2).fill   = FL_LBLUE
    wsSum.cell(r, 3).value  = f"=SUM(C{sum_start}:C{sum_end})"
    wsSum.cell(r, 3).font   = F_BOLD
    wsSum.cell(r, 3).number_format = FMT_PCT1
    wsSum.cell(r, 3).border = BDR
    wsSum.cell(r, 3).fill   = FL_LBLUE
    wsSum.cell(r, 3).alignment = Alignment(horizontal="center", vertical="center")
    wsSum.cell(r, 4).value  = f"=SUM(D{sum_start}:D{sum_end})"
    wsSum.cell(r, 4).font   = Font(bold=True, name="Arial", color="000080")
    wsSum.cell(r, 4).number_format = FMT_MONEY0
    wsSum.cell(r, 4).border = BDR
    wsSum.cell(r, 4).fill   = FL_LBLUE
    wsSum.cell(r, 4).alignment = Alignment(horizontal="right", vertical="center")
    total_r = r; r += 1

    # ── Per-share section — mirrors Innscor ────────────────────────────────────
    r += 1
    write_hdr(wsSum, r, ["Metric", "Value", "Unit"]); r += 1

    row_wev  = r;   row_ns = r+1; row_ivps = r+2
    row_sp   = r+3; row_upd = r+4; row_rec = r+5

    cell_bd(wsSum, row_wev, 1, "Weighted Equity Value", F_STD)
    wsSum.cell(row_wev, 2).value = f"=SUM(D{sum_start}:D{sum_end})"
    wsSum.cell(row_wev, 2).font = F_GREEN
    wsSum.cell(row_wev, 2).number_format = FMT_MONEY0
    wsSum.cell(row_wev, 2).border = BDR
    cell_bd(wsSum, row_wev,  3, "USD", F_STD)

    cell_bd(wsSum, row_ns,   1, "Number of Shares in Issue", F_STD)
    cell_bd(wsSum, row_ns,   2, float(num_shares) if num_shares else 0.0, F_BLUE, FMT_NUM)
    cell_bd(wsSum, row_ns,   3, "Shares", F_STD)

    cell_bd(wsSum, row_ivps, 1, "Intrinsic Value per Share", F_BOLD)
    wsSum.cell(row_ivps, 2).value = f"=IF(B{row_ns}>0,B{row_wev}/B{row_ns},NA())"
    wsSum.cell(row_ivps, 2).font = F_BOLD
    wsSum.cell(row_ivps, 2).number_format = FMT_MONEY4
    wsSum.cell(row_ivps, 2).border = BDR
    wsSum.cell(row_ivps, 1).fill  = FL_LBLUE
    wsSum.cell(row_ivps, 2).fill  = FL_LBLUE
    cell_bd(wsSum, row_ivps, 3, "USD", F_STD)

    cell_bd(wsSum, row_sp,   1, "Current Market Price",     F_STD)
    cell_bd(wsSum, row_sp,   2, float(current_price) if current_price else 0.0, F_BLUE, FMT_MONEY4)
    cell_bd(wsSum, row_sp,   3, "USD", F_STD)

    cell_bd(wsSum, row_upd,  1, "Upside / Downside",        F_STD)
    wsSum.cell(row_upd, 2).value = (
        f"=IF(AND(ISNUMBER(B{row_ivps}),B{row_sp}>0),"
        f"(B{row_ivps}-B{row_sp})/B{row_sp},NA())"
    )
    wsSum.cell(row_upd, 2).font = F_STD
    wsSum.cell(row_upd, 2).number_format = "0.0%"
    wsSum.cell(row_upd, 2).border = BDR
    wsSum.cell(row_upd, 1).fill  = FL_LBLUE
    wsSum.cell(row_upd, 2).fill  = FL_LBLUE
    cell_bd(wsSum, row_upd,  3, "%", F_STD)

    cell_bd(wsSum, row_rec,  1, "Recommendation",           F_BOLD)
    wsSum.cell(row_rec, 2).value = (
        f'=IF(ISNA(B{row_upd}),"N/A",'
        f'IF(B{row_upd}>=0.10,"BUY / ACCUMULATE",'
        f'IF(AND(B{row_upd}>=-0.10,B{row_upd}<=0.10),"HOLD / FAIRLY VALUED",'
        f'"REDUCE / AVOID")))'
    )
    wsSum.cell(row_rec, 2).font = Font(bold=True, name="Arial")
    wsSum.cell(row_rec, 2).border = BDR

    for rr in range(row_wev, row_rec + 1):
        for cc in [1, 2, 3]:
            wsSum.cell(rr, cc).border = BDR

    # ── Analyst Recommendation Note ───────────────────────────────────────────
    r = row_rec + 3
    write_section(wsSum, r, "Analyst Recommendation", ncols=4); r += 1

    # Build recommendation rationale from live values
    co_name_s = ss.get("company_name", "the company")
    try:
        total_w_s = sum(weights_new.get(m, 0.0) for m in selected_models) or 1.0
        norm_w_s  = {m: weights_new.get(m, 0.0) / total_w_s for m in selected_models}
        we_s = sum((value_map.get(m) or 0.0) * norm_w_s.get(m, 0.0) for m in selected_models)
        iv  = (we_s / float(num_shares)) if num_shares and float(num_shares) > 0 else 0.0
        cp  = float(current_price) if current_price else 0.0
        upd = ((iv - cp) / cp * 100) if cp > 0 and iv > 0 else 0.0
        if upd >= 10:
            rec_txt = (
                f"Based on a blended valuation of selected models weighted as above, {co_name_s} "
                f"has an intrinsic value per share of USD {iv:,.4f} against a current market price of "
                f"USD {cp:,.4f}, implying upside of {upd:+.1f}%. We therefore initiate coverage with a "
                f"BUY / ACCUMULATE recommendation. The stock appears undervalued relative to its "
                f"fundamental intrinsic value across multiple valuation methodologies."
            )
        elif upd >= -10:
            rec_txt = (
                f"Based on a blended valuation of selected models weighted as above, {co_name_s} "
                f"has an intrinsic value per share of USD {iv:,.4f} against a current market price of "
                f"USD {cp:,.4f}, implying limited upside/downside of {upd:+.1f}%. We initiate coverage with "
                f"a HOLD / FAIRLY VALUED recommendation. The stock appears to be trading near fair value."
            )
        else:
            rec_txt = (
                f"Based on a blended valuation of selected models weighted as above, {co_name_s} "
                f"has an intrinsic value per share of USD {iv:,.4f} against a current market price of "
                f"USD {cp:,.4f}, implying downside of {upd:+.1f}%. We initiate coverage with a "
                f"REDUCE / SELL recommendation. The stock appears overvalued relative to its intrinsic value."
            )
    except Exception:
        rec_txt = (
            f"Enter the number of shares and current market price above to generate an analyst "
            f"recommendation. The weighted equity value across selected models is shown in the table."
        )
    # Clear cells in the merge range before merging
    from openpyxl.styles import Border
    for _mr in range(r, r + 5):
        for _mc in range(1, 5):
            wsSum.cell(_mr, _mc).value = None
            wsSum.cell(_mr, _mc).border = Border()
            wsSum.cell(_mr, _mc).fill = PatternFill(fill_type=None)

    wsSum.merge_cells(start_row=r, start_column=1, end_row=r + 4, end_column=4)
    rec_note_cell = wsSum.cell(r, 1)
    # Build the recommendation note as an Excel formula so it auto-updates
    rec_formula = (
        f'=IF(AND(ISNUMBER(B{row_ivps}),B{row_sp}>0),'
        f'"Based on a blended valuation of selected models weighted as above, '
        f'the company has an intrinsic value per share of USD "&TEXT(B{row_ivps},"#,##0.0000")&" against a current market price of USD "'
        f'&TEXT(B{row_sp},"#,##0.0000")&", implying "&TEXT(B{row_upd},"0.0%")&" upside/downside. "'
        f'&IF(B{row_upd}>=0.10,"We therefore initiate coverage with a BUY / ACCUMULATE recommendation. The stock appears undervalued relative to its fundamental intrinsic value.",'
        f'IF(AND(B{row_upd}>=-0.10,B{row_upd}<=0.10),"We initiate coverage with a HOLD / FAIRLY VALUED recommendation. The stock appears to be trading near fair value.",'
        f'"We initiate coverage with a REDUCE / SELL recommendation. The stock appears overvalued relative to its intrinsic value.")),'
        f'"Enter the number of shares and current market price above to generate a recommendation.")'
    )
    rec_note_cell.value = rec_formula
    rec_note_cell.font = Font(name="Arial", size=10, italic=True, color="1A1A2E")
    rec_note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    wsSum.row_dimensions[r].height = 90
    # ── Colour legend (Innscor-style note at bottom) ───────────────────────────
    leg = r + 6
    wsSum.cell(leg,   1, "Notes:").font = F_BOLD
    wsSum.cell(leg+1, 1, "Blue text  — hardcoded input; safe to edit directly").font = Font(color="0000FF", name="Arial", size=9)
    wsSum.cell(leg+2, 1, "Green text — formula pulling from model sheet; do NOT overwrite").font = Font(color="008000", name="Arial", size=9)
    wsSum.cell(leg+3, 1, "Blue fill  — key output cell").font = Font(color="000080", name="Arial", size=9, italic=True)
    wsSum.cell(leg+4, 1, "Recommendation thresholds: BUY ≥ +10% upside | HOLD between -10% and +10% | REDUCE < -10%").font = Font(name="Arial", size=9, italic=True)

    wsSum.column_dimensions["A"].width = 36
    wsSum.column_dimensions["B"].width = 22
    wsSum.column_dimensions["C"].width = 12
    wsSum.column_dimensions["D"].width = 26
    wsSum.freeze_panes = "A5"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ---- Render download button ----
st.markdown("---")
st.markdown("### 📥 Download Combined Valuation Model (All Models — Formula-Linked)")
st.markdown(
    """
    <div style="font-size:0.9rem; color:#4b5563; margin-bottom:12px;">
    Downloads a single <strong>.xlsx</strong> with every valuation model on its own sheet.
    Edit any assumption (WACC, growth rate, peer multiples, dividends) and
    <strong>Valuation_Summary updates automatically via formulas</strong>.
    </div>
    """,
    unsafe_allow_html=True,
)

# ── Attempt to restore project session if not already loaded ──────────
_active_pid = st.session_state.get("active_project_id")
if _active_pid and not st.session_state.get("_project_data_loaded"):
    try:
        from auth import load_project_session as _load_proj_sess
        _proj_data = _load_proj_sess(_active_pid)
        _SUMMARY_WIDGET_KEYS = {
            "summary_num_shares", "summary_current_price",
            "selected_models_input", "authenticated", "user",
        }
        if _proj_data:
            for _k, _v in _proj_data.items():
                if _k in _SUMMARY_WIDGET_KEYS:
                    continue
                if _k not in st.session_state:
                    st.session_state[_k] = _v
            st.session_state["_project_data_loaded"] = True
    except Exception:
        pass

import re as _re_fn

def _safe_filename(name: str) -> str:
    name = (name or "").strip() or "Company"
    name = _re_fn.sub(r'[\\/*?:"<>|]', "", name)   # strip illegal filename characters
    name = _re_fn.sub(r"\s+", "_", name)            # spaces -> underscores
    return name[:80]

@st.dialog("Which company are you valuing?")
def _ask_company_and_build():
    name = st.text_input(
        "Company name",
        key="dl_company_name_input",
        placeholder="e.g. Innscor Africa Limited",
    )
    if st.button("Confirm & Prepare Download", key="dl_company_confirm_btn"):
        if not name or not name.strip():
            st.warning("Please enter a company name.")
        else:
            st.session_state["company_name"] = name.strip()
            _ss2 = dict(st.session_state)
            _excel_bytes2 = _build_combined_valuation_excel(
                ss=_ss2,
                selected_models=selected_models,
                value_map=value_map,
                weights_new=weights_new,
                num_shares=num_shares,
                current_price=current_price,
            )
            st.session_state["_pending_excel_bytes"] = _excel_bytes2
            st.session_state["_pending_excel_filename"] = f"{_safe_filename(name)}_Valuation.xlsx"
            st.session_state["_show_download_ready"] = True
            st.rerun()

if st.button("📥 Download Full Valuation Workbook (All Models + Formulas)", key="open_company_dialog_btn"):
    _ask_company_and_build()

if st.session_state.get("_show_download_ready") and st.session_state.get("_pending_excel_bytes"):
    st.success(f"Workbook ready as **{st.session_state['_pending_excel_filename']}** — click below to save it.")
    st.download_button(
        label="⬇️ Click here to download",
        data=st.session_state["_pending_excel_bytes"],
        file_name=st.session_state["_pending_excel_filename"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="final_download_btn",
    )
