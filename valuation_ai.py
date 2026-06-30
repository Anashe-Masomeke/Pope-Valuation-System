"""
FBC Suite — Combined Desktop App
──────────────────────────────────
  📊  Sharestock Upload Converter   (Tab 1)
  ✉   Deal Note Email Automator    (Tab 2)

Requirements:
    pip install pandas openpyxl fpdf2 pywin32 pymupdf

    updating new .exe file:
    pyinstaller --onefile --noconsole fbc_suite.py --name fbc-suite
"""

# ════════════════════════════════════════════════════════════════════════════
#  AUTO-UPDATE
# ════════════════════════════════════════════════════════════════════════════
import sys, os, subprocess, urllib.request

VERSION       = 29
GITHUB_USER   = "Anashe-Masomeke"
GITHUB_REPO   = "fbc-suite"
GITHUB_BRANCH = "main"
EXE_NAME      = "fbc-suite.exe"

_EXE = f"https://github.com/{GITHUB_USER}/{GITHUB_REPO}/releases/latest/download/{EXE_NAME}"
_VER = (f"https://raw.githubusercontent.com/"
        f"{GITHUB_USER}/{GITHUB_REPO}/{GITHUB_BRANCH}/version.txt")

def _remote_ver():
    try:
        with urllib.request.urlopen(_VER, timeout=4) as r:
            return int(r.read().decode().strip())
    except Exception:
        return -1

def check_and_apply_update():
    rv = _remote_ver()
    if rv <= VERSION:
        return

    current_exe = os.path.abspath(sys.argv[0])
    exe_dir     = os.path.dirname(current_exe)

    new_exe_path = os.path.join(exe_dir, f"fbc-suite-v{rv}.exe")
    bat_path     = os.path.join(exe_dir, "_fbc_updater.bat")

    import tkinter as tk
    from tkinter import messagebox

    splash = tk.Tk()
    splash.title("FBC Suite")
    splash.resizable(False, False)
    splash.configure(bg=SIDEBAR_BG)
    w, h = 320, 110
    x = (splash.winfo_screenwidth()  - w) // 2
    y = (splash.winfo_screenheight() - h) // 2
    splash.geometry(f"{w}x{h}+{x}+{y}")
    tk.Label(splash, text="Updating FBC Suite…", bg=SIDEBAR_BG, fg=WHITE,
             font=("Segoe UI", 11, "bold")).pack(pady=(22, 6))
    tk.Label(splash, text=f"v{VERSION}  →  v{rv}", bg=SIDEBAR_BG, fg="#90CAF9",
             font=("Segoe UI", 9)).pack()
    splash.update()

    try:
        MIN_SIZE = 20 * 1024 * 1024
        with urllib.request.urlopen(_EXE, timeout=180) as resp:
            with open(new_exe_path, "wb") as f:
                while True:
                    chunk = resp.read(65536)
                    if not chunk:
                        break
                    f.write(chunk)
                    splash.update()

        size = os.path.getsize(new_exe_path)
        if size < MIN_SIZE:
            os.remove(new_exe_path)
            raise Exception(f"Download incomplete ({size // 1024} KB).")

        bat_lines = [
            "@echo off",
            "ping 127.0.0.1 -n 4 > nul",
            f'start "" "{new_exe_path}"',
            "ping 127.0.0.1 -n 2 > nul",
            'del "%~f0"',
        ]
        with open(bat_path, "w") as f:
            f.write("\n".join(bat_lines) + "\n")

        subprocess.Popen(
            ["cmd.exe", "/c", bat_path],
            creationflags=subprocess.CREATE_NO_WINDOW,
            close_fds=True
        )
        splash.destroy()
        sys.exit(0)

    except Exception as e:
        splash.destroy()
        for fp in [new_exe_path, bat_path]:
            try: os.remove(fp)
            except Exception: pass
        print(f"[Auto-update] Failed silently, continuing on v{VERSION}: {e}")

# ════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ════════════════════════════════════════════════════════════════════════════
import re, json, csv, threading
import shutil as _shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

# ── Voice imports (graceful layered fallback) ────────────────────────────────
_VOICE_READY   = False
_sr            = None
_tts           = None
_whisper_model = None
_vosk_model    = None
_RECOGNISER    = "none"
_fuzz          = None

def _init_voice():
    global _VOICE_READY, _sr, _tts, _fuzz, _RECOGNISER
    try:
        import speech_recognition as sr
        _sr = sr
    except ImportError:
        return
    try:
        import whisper as _w
        _RECOGNISER = "whisper"
    except ImportError:
        pass
    if _RECOGNISER == "none":
        try:
            import vosk as _v  # noqa: F401
            _RECOGNISER = "vosk"
        except ImportError:
            pass
    if _RECOGNISER == "none":
        _RECOGNISER = "google"
    try:
        from rapidfuzz import process as rfp, fuzz as rff
        _fuzz = (rfp, rff)
    except ImportError:
        pass
    _VOICE_READY = True

threading.Thread(target=_init_voice, daemon=True).start()

_whisper_lock = threading.Lock()

def _get_whisper():
    global _whisper_model
    with _whisper_lock:
        if _whisper_model is None:
            import whisper
            _whisper_model = whisper.load_model("base.en")
    return _whisper_model

_vosk_lock = threading.Lock()
_VOSK_MODEL_PATH = os.path.join(os.path.expanduser("~"), "vosk-model-small-en-us-0.15")

def _get_vosk():
    global _vosk_model
    with _vosk_lock:
        if _vosk_model is None:
            import vosk
            if not os.path.isdir(_VOSK_MODEL_PATH):
                raise FileNotFoundError(
                    f"Vosk model not found at:\n{_VOSK_MODEL_PATH}\n\n"
                    "Download from: alphacephei.com/vosk/models\n"
                    "Extract to your home folder.")
            _vosk_model = vosk.Model(_VOSK_MODEL_PATH)
    return _vosk_model

def _require(pkg, install_name=None):
    import importlib
    try:
        return importlib.import_module(pkg)
    except ImportError:
        name = install_name or pkg
        raise ImportError(
            f"Missing package '{name}'. Open terminal and run:\n  pip install {name}")

# ════════════════════════════════════════════════════════════════════════════
#  SHARED COLOURS
# ════════════════════════════════════════════════════════════════════════════
FBC_DARK   = "#003B6F"
FBC_MID    = "#0066B3"
FBC_ACCENT = "#00A3E0"
GREEN_DARK = "#1A6B3A"
RED_DARK   = "#B71C1C"
WHITE      = "#FFFFFF"
BG         = "#F0F4F8"
CARD_BG    = "#FFFFFF"
SEP_CLR    = "#D0DAE8"
TAG_BLUE   = "#E8F1FB"
COL1_HDR   = "#003B6F"
COL2_HDR   = "#1A3A6B"
BOTTOM     = "#0D2B4E"
SIDEBAR_BG      = "#001F3F"
SIDEBAR_ACTIVE  = "#0066B3"
SIDEBAR_HOVER   = "#003B6F"
SIDEBAR_TEXT    = "#B0C8E8"
SIDEBAR_TEXT_ON = "#FFFFFF"

# ════════════════════════════════════════════════════════════════════════════
#  LOGIN DIALOG
# ════════════════════════════════════════════════════════════════════════════
APP_PASSWORD = "enock"
MAX_ATTEMPTS = 6

class LoginDialog(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("FBC Suite — Login")
        self.resizable(False, False)
        self.configure(bg=SIDEBAR_BG)
        self._attempts = 0
        self.authenticated = False
        self._build()
        self.update_idletasks()
        w, h = 380, 340
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build(self):
        hdr = tk.Frame(self, bg=FBC_ACCENT, pady=18)
        hdr.pack(fill="x")
        tk.Label(hdr, text="FBC", bg=FBC_DARK, fg=WHITE,
                 font=("Segoe UI", 20, "bold"), padx=12, pady=6).pack()
        tk.Label(hdr, text="Suite", bg=FBC_ACCENT, fg=WHITE,
                 font=("Segoe UI", 11)).pack(pady=(2, 0))
        body = tk.Frame(self, bg=SIDEBAR_BG, padx=36, pady=24)
        body.pack(fill="both", expand=True)
        tk.Label(body, text="Enter Password", bg=SIDEBAR_BG, fg=SIDEBAR_TEXT,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")
        pw_row = tk.Frame(body, bg=SIDEBAR_BG)
        pw_row.pack(fill="x", pady=(6, 0))
        self._pw_var = tk.StringVar()
        self._show_pw = False
        self.entry_pw = tk.Entry(pw_row, textvariable=self._pw_var, show="●",
                                 font=("Segoe UI", 12), bg="#0D2B4E", fg=WHITE,
                                 insertbackground=WHITE, relief="flat",
                                 highlightbackground=FBC_MID, highlightthickness=1)
        self.entry_pw.pack(side="left", fill="x", expand=True, ipady=8, padx=(0, 4))
        self.entry_pw.focus()
        self.btn_eye = tk.Button(pw_row, text="👁", command=self._toggle_show,
                                 bg="#0D2B4E", fg=SIDEBAR_TEXT, relief="flat",
                                 font=("Segoe UI", 12), cursor="hand2",
                                 activebackground=FBC_MID, activeforeground=WHITE,
                                 padx=6)
        self.btn_eye.pack(side="left")
        self.lbl_err = tk.Label(body, text="", bg=SIDEBAR_BG, fg="#FF6B6B",
                                font=("Segoe UI", 9))
        self.lbl_err.pack(anchor="w", pady=(6, 0))
        self.lbl_attempts = tk.Label(body, text="", bg=SIDEBAR_BG, fg="#607080",
                                     font=("Segoe UI", 8))
        self.lbl_attempts.pack(anchor="w")
        self.btn_login = tk.Button(body, text="  🔓  Login  ",
                                   command=self._attempt_login,
                                   bg=FBC_MID, fg=WHITE, relief="flat",
                                   font=("Segoe UI", 11, "bold"),
                                   cursor="hand2", pady=10, activebackground=FBC_ACCENT)
        self.btn_login.pack(fill="x", pady=(16, 0))
        self.entry_pw.bind("<Return>", lambda _: self._attempt_login())
        tk.Label(self, text=f"v{VERSION}", bg=SIDEBAR_BG, fg="#2A4A6A",
                 font=("Segoe UI", 8)).pack(side="bottom", pady=6)

    def _toggle_show(self):
        self._show_pw = not self._show_pw
        self.entry_pw.config(show="" if self._show_pw else "●")
        self.btn_eye.config(text="🙈" if self._show_pw else "👁")

    def _attempt_login(self):
        entered = self._pw_var.get().strip().lower()
        if entered == APP_PASSWORD.lower():
            self.authenticated = True
            self.destroy()
            return
        self._attempts += 1
        remaining = MAX_ATTEMPTS - self._attempts
        if remaining <= 0:
            messagebox.showerror("Access Denied",
                "Too many incorrect attempts.\nThe application will now close.")
            self.destroy()
            return
        self.lbl_err.config(text="❌  Incorrect password. Please try again.")
        self.lbl_attempts.config(
            text=f"  {remaining} attempt{'s' if remaining > 1 else ''} remaining")
        self._pw_var.set("")
        self.entry_pw.focus()
        self._shake()

    def _shake(self, times=6, distance=8):
        x0 = self.winfo_x()
        y0 = self.winfo_y()
        def step(n):
            if n == 0:
                self.geometry(f"+{x0}+{y0}")
                return
            offset = distance if n % 2 == 0 else -distance
            self.geometry(f"+{x0 + offset}+{y0}")
            self.after(40, lambda: step(n - 1))
        step(times)

    def _on_close(self):
        self.authenticated = False
        self.destroy()


# ════════════════════════════════════════════════════════════════════════════
#  RECIPIENTS CONFIG  (persistent per-user JSON files)
# ════════════════════════════════════════════════════════════════════════════

# --- Sarestock / Deals Confirmation recipients --------------------------------
SARESTOCK_RECIP_FILE = os.path.join(os.path.expanduser("~"), ".fbc_sarestock_recipients.json")

_SARESTOCK_DEFAULT_TO = ["Anesu.Zingundu@fbc.co.zw"]
_SARESTOCK_DEFAULT_CC = [
    "Enock.Rukarwa@fbc.co.zw", "Manatsa.Tagwireyi@fbc.co.zw",
    "Norman.Chirima@fbc.co.zw", "Richard.Mashava@fbc.co.zw",
    "Anashe.Masomeke@fbc.co.zw",
]

def load_sarestock_recipients():
    try:
        with open(SARESTOCK_RECIP_FILE) as f:
            d = json.load(f)
        return d.get("to", _SARESTOCK_DEFAULT_TO), d.get("cc", _SARESTOCK_DEFAULT_CC)
    except Exception:
        return list(_SARESTOCK_DEFAULT_TO), list(_SARESTOCK_DEFAULT_CC)

def save_sarestock_recipients(to_list, cc_list):
    with open(SARESTOCK_RECIP_FILE, "w") as f:
        json.dump({"to": to_list, "cc": cc_list}, f, indent=2)

# --- Custodian recipients overrides ------------------------------------------
CUSTODIAN_RECIP_FILE = os.path.join(os.path.expanduser("~"), ".fbc_custodian_recipients.json")

def load_custodian_overrides():
    try:
        with open(CUSTODIAN_RECIP_FILE) as f:
            return json.load(f)
    except Exception:
        return {}

def save_custodian_overrides(overrides):
    with open(CUSTODIAN_RECIP_FILE, "w") as f:
        json.dump(overrides, f, indent=2)


# ════════════════════════════════════════════════════════════════════════════
#  SHARED RECIPIENTS EDITOR DIALOG
# ════════════════════════════════════════════════════════════════════════════
class RecipientsDialog(tk.Toplevel):
    def __init__(self, parent, title, to_list, cc_list, on_save,
                 to_label="To (primary recipients)",
                 cc_label="CC (copied recipients)"):
        super().__init__(parent)
        self.title(title)
        self.configure(bg=BG)
        self.resizable(False, False)
        self.grab_set()
        self._on_save = on_save
        self._has_to  = to_list is not None
        self._to_list = list(to_list) if to_list else []
        self._cc_list = list(cc_list)
        self._to_label = to_label
        self._cc_label = cc_label
        self._build()
        self.update_idletasks()
        w = 560
        h = self.winfo_reqheight()
        x = parent.winfo_rootx() + (parent.winfo_width()  - w) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build(self):
        tk.Label(self, text=f"  {self.title()}", bg=FBC_DARK, fg=WHITE,
                 font=("Segoe UI", 11, "bold"), pady=10).pack(fill="x")

        body = tk.Frame(self, bg=BG, padx=18, pady=12)
        body.pack(fill="both", expand=True)

        if self._has_to:
            self._to_frame = self._section(body, self._to_label, self._to_list)
        self._cc_frame = self._section(body, self._cc_label, self._cc_list)

        hint = tk.Label(body,
            text="One address per line.  Accepted formats:\n"
                 "  plain@email.com   or   Display Name <plain@email.com>",
            bg=BG, fg="#607080", font=("Segoe UI", 8), justify="left")
        hint.pack(anchor="w", pady=(6, 0))

        bot = tk.Frame(self, bg=BG, padx=18, pady=10)
        bot.pack(fill="x")
        tk.Button(bot, text="💾  Save", command=self._save,
                  bg=GREEN_DARK, fg=WHITE, relief="flat",
                  font=("Segoe UI", 10, "bold"), cursor="hand2",
                  padx=16, pady=7).pack(side="right")
        tk.Button(bot, text="Cancel", command=self.destroy,
                  bg="#607080", fg=WHITE, relief="flat",
                  font=("Segoe UI", 10), cursor="hand2",
                  padx=12, pady=7).pack(side="right", padx=(0, 8))

    def _section(self, parent, label, initial_list):
        tk.Label(parent, text=label, bg=BG, fg=FBC_DARK,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(8, 2))
        txt = tk.Text(parent, height=4, font=("Segoe UI", 9),
                      relief="flat", bg=WHITE,
                      highlightbackground=FBC_MID, highlightthickness=1,
                      wrap="none")
        txt.insert("1.0", "\n".join(initial_list))
        txt.pack(fill="x", pady=(0, 4))
        return txt

    def _parse(self, widget):
        raw = widget.get("1.0", "end").strip()
        return [ln.strip() for ln in raw.splitlines() if ln.strip()]

    def _save(self):
        to_result = self._parse(self._to_frame) if self._has_to else None
        cc_result = self._parse(self._cc_frame)
        self._on_save(to_result, cc_result)
        self.destroy()


# ════════════════════════════════════════════════════════════════════════════
#  SARESTOCK LOGIC
# ════════════════════════════════════════════════════════════════════════════

EO_HEADERS = [
    "Exchange","Market","Symbol","Buy/Sell","Participant","Custodian","Client",
    "Trader","Short Sell","Price","Volume","Yield %","Accrued Interest","Order No.",
    "Ticket No.","Date/Time","Execution Date/Time","Type","Filled Volume",
    "Remaining Volume","Disc. Volume","Trigger Price","Order Initiator","Pricing Mechanism"
]
PREVIEW_COLS = ["Exchange","Market","Participant","Custodian","Client",
                "Symbol","Buy/Sell","Price","Volume","Ticket No."]

SARESTOCK_EMAIL_SUBJECT = "DEALS CONFIRMATION"

def get_sarestock_email_body(sender_name=""):
    name = sender_name.strip() or "FBC Securities"
    return f"Good day,\r\n\r\nKindly find attached for deals confirmation.\r\n\r\nRegards,\r\n{name}."

FIELD_MAP = [
    ("Security","Symbol"),("SCA Code","Custodian"),("Buy/Sell","Buy/Sell"),
    ("Quantity","Volume + Filled Vol."),("Price","Yield"),
    ("Ticket No.","Match Reference"),("Trader","Trader + Order Init."),
    ("VFX → VFEX","Exchange (+E)"),("VFEX = FBCSZWVX","Participant (fixed)"),
    ("ZSE = FBCSZWHX","Participant (fixed)"),("…-02 → …-0002","Client (zero-pad)"),
    ("DD/MM/YYYY …","Date/Time (auto)"),
]

def get_exchange(market):
    u=(market or "").strip().upper(); return "VFEX" if u=="VFX" else (u or "ZSE")

def get_participant(exch): return "FBCSZWVX" if exch=="VFEX" else "FBCSZWHX"

def get_market(sym,exch):
    s=(sym or "").upper().strip()
    if exch=="VFEX" or s.endswith(".VX"): return "REG"
    if s.endswith(".ZW"):
        if any(r in s for r in ["FHML","ZMRE","STFL","IPFL","HAFP","REVH"]): return "REIT"
        if any(o in s for o in ["SEED","CFI","CAFCA"]): return "ODD"
        return "REG"
    return "REG"

def pad_client(c):
    s=str(c or "").strip(); d=s.rfind("-")
    return s if d==-1 else f"{s[:d]}-{s[d+1:].zfill(4)}"

def get_now():
    d=datetime.now()
    return f"{d.day}/{d.month}/{d.year} {d.hour}:{d.minute:02d}"

def stamp():
    d=datetime.now(); return f"{d.day}_{d.month}_{d.year}"

def transform_rows(raw_rows):
    now = get_now()
    out = []
    for r in raw_rows:
        exch = get_exchange(r.get("Market", ""))
        sym  = r.get("Security", "")
        out.append({
            "Exchange": exch, "Market": get_market(sym, exch), "Symbol": sym,
            "Buy/Sell": r.get("Buy/Sell", ""), "Participant": get_participant(exch),
            "Custodian": r.get("SCA Code", ""), "Client": str(r.get("CSD Account", "") or "").strip(),
            "Trader": r.get("Trader", ""), "Short Sell": "NO",
            "Price": r.get("Yield", ""), "Volume": r.get("Quantity", ""),
            "Yield %": "0", "Accrued Interest": "0", "Order No.": r.get("Trade Leg", "").lstrip("0") or "0",
            "Ticket No.": r.get("Match Reference", "").lstrip("0") or "0",
            "Date/Time": now, "Execution Date/Time": now,
            "Type": "Limit", "Filled Volume": r.get("Quantity", ""),
            "Remaining Volume": "0", "Disc. Volume": "0", "Trigger Price": "0",
            "Order Initiator": r.get("Trader", ""), "Pricing Mechanism": ""
        })
    return out, now

def generate_csv(rows, out_dir, label):
    label = label.upper()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"ExportExecutedOrders_{label}_{ts}.csv")
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=EO_HEADERS)
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in EO_HEADERS})
    return path

def generate_matched_excel(source_path, raw_rows, out_dir):
    exch = get_exchange(raw_rows[0].get("Market", "")) if raw_rows else ""
    label = "VFEX" if exch == "VFEX" else "ZSE"
    ext = os.path.splitext(source_path)[1] if source_path else ".xlsx"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(out_dir, f"MATCHED TRADES, {label}_{ts}{ext}")
    src_norm = os.path.normcase(os.path.abspath(source_path))
    dst_norm = os.path.normcase(os.path.abspath(dest))
    if src_norm == dst_norm:
        return dest
    try:
        _shutil.copy2(source_path, dest)
    except PermissionError:
        raise PermissionError(
            f"Could not copy '{os.path.basename(source_path)}' — "
            "please close it in Excel and try again.")
    return dest

ANESU_COLUMNS = ["Market", "CSD Account", "SCA Code", "Name",
                  "Security", "Buy/Sell", "Quantity", "Yield"]

def generate_anesu_excel(raw_rows, out_dir):
    _require("pandas")
    import pandas as pd
    from openpyxl.styles import Font
    exch_raw = (raw_rows[0].get("Market", "") or "").strip().upper() if raw_rows else ""
    label = "VFEX" if exch_raw in ("VFX", "VFEX") else "ZSE"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"CONFIRMATION TEMPLATE, {label}_{ts}.xlsx")
    trimmed = [{col: r.get(col, "") for col in ANESU_COLUMNS} for r in raw_rows]
    df = pd.DataFrame(trimmed, columns=ANESU_COLUMNS)
    df.to_excel(path, index=False)

    wb = _require("openpyxl").load_workbook(path)
    ws = wb.active
    font = Font(name="Aptos Narrow", size=11)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    wb.save(path)
    return path

def generate_pdf(raw_rows, raw_headers, out_dir):
    _require("fpdf", "fpdf2")
    from fpdf import FPDF

    FONT_SIZE  = 6.5
    LINE_H     = 5.0
    HEADER_H   = 6.5
    CHAR_W     = FONT_SIZE * 0.50
    MAX_CHARS  = 28
    MIN_COL_MM = 8
    MAX_COL_MM = 55
    MARGIN     = 8
    PAGE_H_MM  = 210

    def _safe(text):
        return str(text).encode("latin-1", errors="replace").decode("latin-1")

    exch_raw = (raw_rows[0].get("Market", "") or "").strip().upper() if raw_rows else ""
    exch = "VFEX" if exch_raw in ("VFX", "VFEX") else "ZSE"
    out_path = os.path.join(out_dir, f"MATCHED TRADES, {exch}.pdf")

    def _col_w(hdr, rows, key):
        mx = len(str(hdr))
        for r in rows:
            mx = max(mx, len(str(r.get(key, "") or "")))
        mx = min(mx, MAX_CHARS)
        return max(MIN_COL_MM, min(MAX_COL_MM, mx * CHAR_W))

    col_widths = [_col_w(h, raw_rows, h) for h in raw_headers]
    total_content_w = sum(col_widths)
    page_w = total_content_w + 2 * MARGIN
    ts_label   = datetime.now().strftime("%d %b %Y  %H:%M")
    total_rows = len(raw_rows)

    pdf = FPDF(orientation="L", unit="mm", format=(PAGE_H_MM, page_w))
    pdf.set_margins(MARGIN, MARGIN, MARGIN)
    pdf.set_auto_page_break(auto=True, margin=14)

    def _page_header():
        pdf.set_font("Courier", style="B", size=6.5)
        pdf.set_text_color(90, 90, 90)
        label = _safe(
            f"MATCHED TRADES - {exch}  |  "
            f"{total_rows} row(s)  |  {ts_label}  |  "
            f"All {len(raw_headers)} columns — scroll right to see full table")
        pdf.cell(0, 4.5, label, border=0, align="L")
        pdf.ln(6)
        pdf.set_text_color(0, 0, 0)

    def _draw_col_headers():
        pdf.set_font("Courier", style="B", size=FONT_SIZE)
        pdf.set_text_color(0, 0, 0)
        for i, h in enumerate(raw_headers):
            pdf.cell(col_widths[i], HEADER_H, _safe(str(h)[:MAX_CHARS]), border=0, align="L")
        pdf.ln()
        y = pdf.get_y()
        pdf.set_draw_color(150, 150, 150)
        pdf.line(MARGIN, y, page_w - MARGIN, y)
        pdf.set_draw_color(0, 0, 0)
        pdf.ln(0.8)

    def _draw_data_rows():
        pdf.set_font("Courier", size=FONT_SIZE)
        pdf.set_text_color(0, 0, 0)
        for row in raw_rows:
            if pdf.get_y() > pdf.h - 16:
                pdf.add_page()
                _page_header()
                _draw_col_headers()
                pdf.set_font("Courier", size=FONT_SIZE)
                pdf.set_text_color(0, 0, 0)
            for i, h in enumerate(raw_headers):
                val = _safe(str(row.get(h, "") or "")[:MAX_CHARS])
                pdf.cell(col_widths[i], LINE_H, val, border=0, align="L")
            pdf.ln()

    pdf.add_page()
    _page_header()
    _draw_col_headers()
    _draw_data_rows()
    pdf.output(out_path)
    return out_path


def open_sarestock_outlook(file_paths, sender_name=""):
    _require("win32com.client","pywin32")
    import win32com.client as win32
    to_list, cc_list = load_sarestock_recipients()
    outlook = win32.Dispatch("outlook.application")
    mail    = outlook.CreateItem(0)
    mail.Subject = SARESTOCK_EMAIL_SUBJECT
    mail.Body    = get_sarestock_email_body(sender_name)
    mail.To  = "; ".join(to_list)
    mail.CC  = "; ".join(cc_list)
    for fp in file_paths:
        if fp and os.path.exists(fp): mail.Attachments.Add(fp)
    mail.Display(True)


# ════════════════════════════════════════════════════════════════════════════
#  EMAILER LOGIC
# ════════════════════════════════════════════════════════════════════════════
SENDER_NAME_FILE = os.path.join(os.path.expanduser("~"), ".fbc_sender_name.txt")

def load_sender_name():
    try:
        with open(SENDER_NAME_FILE) as f:
            return f.read().strip()
    except Exception:
        return ""

def save_sender_name(name):
    with open(SENDER_NAME_FILE, "w") as f:
        f.write(name.strip())

KNOWN_CUSTODIANS = ["FBCZSEZW","CBZCZWHX","STINZWVX","CBCZSEZW","FBCSZWVX","NSSAZWHX"]

CUSTODIAN_PREFIX_MAP = [
    ("FBC","FBCZSEZW"),("CBC","CBCZSEZW"),("CBZ","CBZCZWHX"),
    ("STIN","STINZWVX"),("STIZ","STINZWVX"),("NSSA","NSSAZWHX"),
]

_FBC_CC = [
    "Manatsa Tagwireyi <Manatsa.Tagwireyi@fbc.co.zw>",
    "Norman Chirima <Norman.Chirima@fbc.co.zw>",
    "Enock Rukarwa <Enock.Rukarwa@fbc.co.zw>",
    "Richard Mashava <Richard.Mashava@fbc.co.zw>",
    "Anesu Zingundu <Anesu.Zingundu@fbc.co.zw>",
]

CUSTODIAN_ROUTING = {
    "FBCZSEZW":{"label":"FBC Securities (ZSE)",
        "to":["Faith Chikati <Faith.Chikati@fbc.co.zw>"],
        "cc":["Custodial Services <CustodialServices@fbc.co.zw>"]+_FBC_CC},
    "CBZCZWHX":{"label":"CBZ (ZSE)",
        "to":["Sharleen Kapininga <skapininga@cbz.co.zw>","Phillipa Gurure <pgurure@cbz.co.zw>"],
        "cc":["Custodial Services <custodialservices@cbz.co.zw>"]+_FBC_CC},
    "STINZWVX":{"label":"Stanbic",
        "to":["Maigurira, Debra D <maigurirad@stanbic.com>","Chibvongodze, Kudakwashe K <chibvongodzek@stanbic.com>"],
        "cc":["custodyzim <custodyzim@standardbank.co.za>"]+_FBC_CC},
    "CBCZSEZW":{"label":"CABS / Old Mutual",
        "to":["Darlington Tatenda Maenda <darlingtonm@oldmutual.co.zw>"],
        "cc":["Custodial Services Division <custodialservicesdivision@cabs.co.zw>"]+_FBC_CC},
    "FBCSZWVX":{"label":"FBC Securities (VFEX)",
        "to":["Faith Chikati <Faith.Chikati@fbc.co.zw>"],
        "cc":["Custodial Services <CustodialServices@fbc.co.zw>"]+_FBC_CC},
    "NSSAZWHX":{"label":"NSSA",
        "to": ["Charles Mugabe <mugabec@nssa.org.zw>", "Malvern Murombedzi <murombedzim@nssa.org.zw>"],
        "cc": ["Kudzaishe P. Manenji <manenjiK@nssa.org.zw>"]+_FBC_CC},
}

def get_effective_custodian_routing(code):
    base     = CUSTODIAN_ROUTING.get(code, {})
    override = load_custodian_overrides().get(code, {})
    if not base:
        return None
    return {
        "label": base["label"],
        "to":    override.get("to", base["to"]),
        "cc":    override.get("cc", base["cc"]),
    }

def get_custodian_body(multi=False, sender_name=""):
    name = sender_name.strip() or "FBC Securities"
    if multi:
        return f"Good day,\r\n\r\nKindly find attached today's deal notes.\r\n\r\nRegards,\r\n{name}."
    return f"Good day,\r\n\r\nKindly find attached today's deal note.\r\n\r\nRegards,\r\n{name}."

def parse_custodian_from_pdf(pdf_path):
    try:
        import fitz
        doc=fitz.open(pdf_path); text="".join(p.get_text() for p in doc); doc.close()
        for code in KNOWN_CUSTODIANS:
            if code in text: return code
        candidates=re.findall(r'\b([A-Z]{4,10})\b',text)
        for c in candidates:
            for prefix,canonical in CUSTODIAN_PREFIX_MAP:
                if c.startswith(prefix): return canonical
    except Exception: pass
    return None

def parse_deal_info_from_pdf(pdf_path):
    info={"deal_number":"","counter":"","deal_date":""}
    try:
        import fitz
        doc=fitz.open(pdf_path); text="".join(p.get_text() for p in doc); doc.close()
        m=re.search(r'Deal Number\s+(\d+)',text)
        if m: info["deal_number"]=m.group(1)
        m=re.search(r'Deal Date\s+([\d/]+)',text)
        if m: info["deal_date"]=m.group(1)
        m=re.search(r'\b([A-Z]{2,6}\.ZW|[A-Z]{2,6}\.VX)\b',text)
        if m: info["counter"]=m.group(1)
    except Exception: pass
    return info

def open_outlook(to_list,cc_list,subject,body,attachments):
    try:
        import win32com.client as win32
        outlook=win32.Dispatch("outlook.application"); mail=outlook.CreateItem(0)
        mail.To="; ".join(to_list); mail.CC="; ".join(cc_list)
        mail.Subject=subject; mail.Body=body
        for path in attachments:
            if os.path.exists(path): mail.Attachments.Add(path)
        mail.Display(True)
    except ImportError: raise ImportError("pywin32 not installed.\n\nRun:  pip install pywin32")
    except Exception as e: raise RuntimeError(f"Outlook error: {e}")


# ════════════════════════════════════════════════════════════════════════════
#  SARESTOCK PAGE
# ════════════════════════════════════════════════════════════════════════════
class SarestockPage(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=BG)
        self.raw_rows = [];  self.raw_headers = [];  self.conv_rows = []
        self.source_path = None; self.gen_csv = self.gen_pdf = self.gen_mt_xlsx = self.gen_anesu_xlsx = None
        self.raw_rows2 = []; self.raw_headers2 = []; self.conv_rows2 = []
        self.source_path2 = None; self.gen_csv2 = self.gen_pdf2 = self.gen_mt_xlsx2 = self.gen_anesu_xlsx2 = None
        self.out_dir=os.path.join(os.path.expanduser("~"),"Downloads")
        self._build()

    def _build(self):
        info=tk.Frame(self,bg=FBC_MID,padx=16,pady=8); info.pack(fill="x")
        tk.Label(info,text="📊  Sarestock Upload Converter",bg=FBC_MID,fg=WHITE,
                 font=("Segoe UI",11,"bold")).pack(side="left")
        self.paned=tk.PanedWindow(self,orient="horizontal",bg=SEP_CLR,sashwidth=4,sashrelief="flat")
        self.paned.pack(fill="both",expand=True)
        self.left_frame,self.left_canvas,self.left_body=self._scroll_pane(self.paned)
        self.paned.add(self.left_frame,stretch="always")
        self.right_frame,self.right_canvas,self.right_body=self._scroll_pane(self.paned)
        self.paned.add(self.right_frame,stretch="always")
        self._build_bottom_bar()
        self._build_left_column()
        self._build_right_column()

    def _scroll_pane(self,parent):
        frame=tk.Frame(parent,bg=BG)
        canvas=tk.Canvas(frame,bg=BG,highlightthickness=0)
        vsb=ttk.Scrollbar(frame,orient="vertical",command=canvas.yview)
        inner=tk.Frame(canvas,bg=BG)
        inner.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0),window=inner,anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left",fill="both",expand=True); vsb.pack(side="right",fill="y")
        canvas.bind("<Enter>",lambda e,c=canvas:c.bind_all("<MouseWheel>",
            lambda ev:c.yview_scroll(-1*(ev.delta//120),"units")))
        canvas.bind("<Leave>",lambda e,c=canvas:c.unbind_all("<MouseWheel>"))
        return frame,canvas,inner

    def _build_bottom_bar(self):
        bar=tk.Frame(self,bg=BOTTOM,pady=10); bar.pack(fill="x",side="bottom")
        path_row=tk.Frame(bar,bg=BOTTOM); path_row.pack(fill="x",padx=16,pady=(0,6))
        tk.Label(path_row,text="Files saved to:",bg=BOTTOM,fg="#8BAAC8",
                 font=("Segoe UI",8)).pack(side="left")
        self.lbl_outdir=tk.Label(path_row,text=self.out_dir,bg=BOTTOM,fg="#90CAF9",
                                  font=("Segoe UI",8)); self.lbl_outdir.pack(side="left",padx=6)
        tk.Button(path_row,text="Change...",command=self._pick_outdir,bg="#1A3A6B",fg="#90CAF9",
                  relief="flat",font=("Segoe UI",8),cursor="hand2",padx=6,pady=2).pack(side="left")
        tk.Button(path_row, text="⚙ Configure Recipients",
                  command=self._configure_recipients,
                  bg=FBC_ACCENT, fg=WHITE, relief="flat",
                  font=("Segoe UI", 8, "bold"), cursor="hand2",
                  padx=8, pady=2).pack(side="left", padx=(12, 0))
        tk.Button(path_row, text="Clear Uploads", command=self._clear_uploads,
                  bg=RED_DARK, fg=WHITE, relief="flat",
                  font=("Segoe UI", 8, "bold"), cursor="hand2",
                  padx=8, pady=2).pack(side="right")
        btn_row=tk.Frame(bar,bg=BOTTOM); btn_row.pack(fill="x",padx=16)
        btn_row.columnconfigure(0,weight=1); btn_row.columnconfigure(1,weight=1); btn_row.columnconfigure(2,weight=2)
        self.btn_email=tk.Button(btn_row,text="Send — ZSE Only",command=self._send_email,
            bg=GREEN_DARK,fg=WHITE,font=("Segoe UI",10,"bold"),relief="flat",pady=9,
            cursor="hand2",state="disabled")
        self.btn_email.grid(row=0,column=0,sticky="ew",padx=(0,6))
        self.btn_email2=tk.Button(btn_row,text="Send — VFEX Only",command=self._send_email2,
            bg="#1A3A6B",fg=WHITE,font=("Segoe UI",10,"bold"),relief="flat",pady=9,
            cursor="hand2",state="disabled")
        self.btn_email2.grid(row=0,column=1,sticky="ew",padx=(0,6))
        self.btn_email_both=tk.Button(btn_row,text="Send BOTH ZSE + VFEX in One Email",
            command=self._send_email_both,bg=FBC_MID,fg=WHITE,font=("Segoe UI",11,"bold"),
            relief="flat",pady=9,cursor="hand2",state="disabled")
        self.btn_email_both.grid(row=0,column=2,sticky="ew")
        self.lbl_recip_summary = tk.Label(bar, text="", bg=BOTTOM, fg="#5D7A99",
                                          font=("Segoe UI", 8))
        self.lbl_recip_summary.pack(pady=(4, 0))
        self._refresh_recip_summary()

    def _refresh_recip_summary(self):
        to_list, cc_list = load_sarestock_recipients()
        to_str = "; ".join(to_list) if to_list else "(none)"
        cc_str = f"{len(cc_list)} CC address{'es' if len(cc_list)!=1 else ''}"
        self.lbl_recip_summary.config(
            text=f"Deals Confirmation  →  To: {to_str}  |  {cc_str}  "
                 f"  (click ⚙ Configure Recipients to change)")

    def _configure_recipients(self):
        to_list, cc_list = load_sarestock_recipients()
        def on_save(new_to, new_cc):
            save_sarestock_recipients(new_to or [], new_cc)
            self._refresh_recip_summary()
            messagebox.showinfo("Saved",
                "Deals Confirmation recipients updated.\n\n"
                "All future sends will use the new addresses.", parent=self)
        RecipientsDialog(
            self,
            title="⚙  Deals Confirmation — Configure Recipients",
            to_list=to_list,
            cc_list=cc_list,
            on_save=on_save,
            to_label="To  (primary recipients — required)",
            cc_label="CC  (copied recipients)",
        )

    def _clear_uploads(self):
        has_data = bool(self.source_path or self.source_path2)
        if not has_data:
            messagebox.showinfo("Nothing to Clear", "No files are currently loaded.")
            return
        if not messagebox.askyesno("Clear Uploads",
                "Clear all uploaded matched trades files and start fresh?\n\n"
                "This does NOT delete any files from disk."):
            return
        self.raw_rows=[]; self.raw_headers=[]; self.conv_rows=[]
        self.source_path=None
        self.gen_csv = self.gen_pdf = self.gen_mt_xlsx = self.gen_anesu_xlsx = None
        self.raw_rows2=[]; self.raw_headers2=[]; self.conv_rows2=[]
        self.source_path2=None
        self.gen_csv2 = self.gen_pdf2 = self.gen_mt_xlsx2 = self.gen_anesu_xlsx2 = None
        self.prev_outer1.pack_forget()
        self.prev_outer2.pack_forget()
        for w in self.info_bar1.winfo_children(): w.pack_forget()
        self.info_bar1.pack_forget()
        for w in self.info_bar2.winfo_children(): w.pack_forget()
        self.info_bar2.pack_forget()
        self.lbl_csv_done.config(text=""); self.lbl_pdf_done.config(text="")
        self.lbl_csv2_done.config(text=""); self.lbl_pdf2_done.config(text="")
        self.btn_csv.config(text="Download CSV", bg=FBC_MID, state="disabled")
        self.btn_pdf.config(text="Download PDF", bg=RED_DARK, state="disabled")
        self.btn_csv2.config(text="Download CSV", bg=FBC_MID, state="disabled")
        self.btn_pdf2.config(text="Download PDF", bg=RED_DARK, state="disabled")
        for b in (self.btn_email, self.btn_email2, self.btn_email_both):
            b.config(state="disabled")
        self.btn_email.config(text="Send — ZSE Only")
        self.btn_email2.config(text="Send — VFEX Only")
        messagebox.showinfo("Cleared", "Both upload slots cleared. Ready for a new upload.")

    def _build_left_column(self):
        p=self.left_body
        hdr=tk.Frame(p,bg=COL1_HDR); hdr.pack(fill="x",padx=12,pady=(12,0))
        tk.Label(hdr,text=" 1 ",bg=WHITE,fg=COL1_HDR,font=("Segoe UI",9,"bold"),
                 padx=4,pady=4).pack(side="left",padx=(8,0),pady=6)
        tk.Label(hdr,text="  FIRST EXCHANGE  (ZSE or VFEX)",bg=COL1_HDR,fg=WHITE,
                 font=("Segoe UI",10,"bold")).pack(side="left",pady=6)
        ucard=self._card(p,COL1_HDR)
        dz=tk.Frame(ucard,bg="#F4F8FE",relief="groove",bd=2); dz.pack(fill="x",pady=(0,10))
        inner=tk.Frame(dz,bg="#F4F8FE"); inner.pack(pady=16)
        tk.Label(inner,text="Upload Matched Trades File",bg="#F4F8FE",fg=FBC_MID,
                 font=("Segoe UI",10,"bold")).pack(pady=(4,0))
        tk.Label(inner,text=".csv or .xlsx",bg="#F4F8FE",fg="#8096B0",font=("Segoe UI",8)).pack()
        tk.Button(inner,text="  Browse...  ",command=self._pick_file,bg=FBC_MID,fg=WHITE,
                  font=("Segoe UI",10,"bold"),relief="flat",padx=14,pady=6,cursor="hand2").pack(pady=(8,0))
        self.info_bar1=tk.Frame(ucard,bg=TAG_BLUE,highlightbackground=FBC_ACCENT,highlightthickness=1)
        self.lbl_file1=tk.Label(self.info_bar1,text="",bg=TAG_BLUE,fg=FBC_DARK,font=("Segoe UI",9,"bold"))
        self.lbl_rows1=tk.Label(self.info_bar1,text="",bg=TAG_BLUE,fg=FBC_MID,font=("Consolas",8))
        self.btn_reupload1=tk.Button(self.info_bar1,text="Change",command=self._pick_file,
                                     bg=TAG_BLUE,fg=FBC_MID,relief="flat",font=("Segoe UI",8),cursor="hand2")
        dcard=self._card(p,COL1_HDR)
        tk.Label(dcard,text="DOWNLOAD",bg=CARD_BG,fg="#8096B0",font=("Segoe UI",8,"bold")).pack(anchor="w")
        btn_row=tk.Frame(dcard,bg=CARD_BG); btn_row.pack(fill="x",pady=(6,2))
        self.btn_csv=self._col_btn(btn_row,"Download CSV",self._dl_csv,FBC_MID)
        self.btn_pdf=self._col_btn(btn_row,"Download PDF",self._dl_pdf,RED_DARK)
        for b in (self.btn_csv,self.btn_pdf): b.config(state="disabled")
        self.lbl_csv_done=tk.Label(dcard,text="",bg=CARD_BG,fg=GREEN_DARK,font=("Segoe UI",8)); self.lbl_csv_done.pack(anchor="w")
        self.lbl_pdf_done=tk.Label(dcard,text="",bg=CARD_BG,fg=GREEN_DARK,font=("Segoe UI",8)); self.lbl_pdf_done.pack(anchor="w")
        self.prev_outer1=tk.Frame(p,bg=BG)
        self._build_preview_shell(self.prev_outer1,COL1_HDR,"PREVIEW — FIRST EXCHANGE","prev_body1","lbl_showing1")
        self.prev_outer1.pack_forget()

    def _build_right_column(self):
        p=self.right_body
        hdr=tk.Frame(p,bg=COL2_HDR); hdr.pack(fill="x",padx=12,pady=(12,0))
        tk.Label(hdr,text=" 2 ",bg=WHITE,fg=COL2_HDR,font=("Segoe UI",9,"bold"),
                 padx=4,pady=4).pack(side="left",padx=(8,0),pady=6)
        tk.Label(hdr,text="  SECOND EXCHANGE  (ZSE or VFEX)",bg=COL2_HDR,fg=WHITE,
                 font=("Segoe UI",10,"bold")).pack(side="left",pady=6)
        ucard=self._card(p,COL2_HDR)
        dz=tk.Frame(ucard,bg="#F4F8FE",relief="groove",bd=2); dz.pack(fill="x",pady=(0,10))
        inner=tk.Frame(dz,bg="#F4F8FE"); inner.pack(pady=16)
        tk.Label(inner,text="Upload Matched Trades File",bg="#F4F8FE",fg=FBC_MID,
                 font=("Segoe UI",10,"bold")).pack(pady=(4,0))
        tk.Label(inner,text=".csv or .xlsx",bg="#F4F8FE",fg="#8096B0",font=("Segoe UI",8)).pack()
        tk.Button(inner,text="  Browse...  ",command=self._pick_file2,bg=COL2_HDR,fg=WHITE,
                  font=("Segoe UI",10,"bold"),relief="flat",padx=14,pady=6,cursor="hand2").pack(pady=(8,0))
        self.info_bar2=tk.Frame(ucard,bg=TAG_BLUE,highlightbackground=FBC_ACCENT,highlightthickness=1)
        self.lbl_file2=tk.Label(self.info_bar2,text="",bg=TAG_BLUE,fg=FBC_DARK,font=("Segoe UI",9,"bold"))
        self.lbl_rows2=tk.Label(self.info_bar2,text="",bg=TAG_BLUE,fg=FBC_MID,font=("Consolas",8))
        self.btn_reupload2=tk.Button(self.info_bar2,text="Change",command=self._pick_file2,
                                     bg=TAG_BLUE,fg=FBC_MID,relief="flat",font=("Segoe UI",8),cursor="hand2")
        dcard=self._card(p,COL2_HDR)
        tk.Label(dcard,text="DOWNLOAD",bg=CARD_BG,fg="#8096B0",font=("Segoe UI",8,"bold")).pack(anchor="w")
        btn_row=tk.Frame(dcard,bg=CARD_BG); btn_row.pack(fill="x",pady=(6,2))
        self.btn_csv2=self._col_btn(btn_row,"Download CSV",self._dl_csv2,FBC_MID)
        self.btn_pdf2=self._col_btn(btn_row,"Download PDF",self._dl_pdf2,RED_DARK)
        for b in (self.btn_csv2,self.btn_pdf2): b.config(state="disabled")
        self.lbl_csv2_done=tk.Label(dcard,text="",bg=CARD_BG,fg=GREEN_DARK,font=("Segoe UI",8)); self.lbl_csv2_done.pack(anchor="w")
        self.lbl_pdf2_done=tk.Label(dcard,text="",bg=CARD_BG,fg=GREEN_DARK,font=("Segoe UI",8)); self.lbl_pdf2_done.pack(anchor="w")
        self.prev_outer2=tk.Frame(p,bg=BG)
        self._build_preview_shell(self.prev_outer2,COL2_HDR,"PREVIEW — SECOND EXCHANGE","prev_body2","lbl_showing2")
        self.prev_outer2.pack_forget()

    def _card(self,parent,accent):
        wrapper=tk.Frame(parent,bg=BG); wrapper.pack(fill="x",padx=12,pady=(4,0))
        tk.Frame(wrapper,bg=accent,height=2).pack(fill="x")
        body=tk.Frame(wrapper,bg=CARD_BG,padx=14,pady=12,highlightbackground=SEP_CLR,highlightthickness=1)
        body.pack(fill="x"); return body

    def _col_btn(self,parent,text,cmd,bg):
        b=tk.Button(parent,text=text,command=cmd,bg=bg,fg=WHITE,
                    font=("Segoe UI",9,"bold"),relief="flat",padx=10,pady=7,cursor="hand2")
        b.pack(side="left",padx=(0,8),pady=2); return b

    def _build_preview_shell(self,outer,color,title,body_attr,label_attr):
        outer.pack(fill="x",padx=12,pady=(4,0))
        hdr=tk.Frame(outer,bg=color); hdr.pack(fill="x")
        tk.Label(hdr,text=f"  {title}",bg=color,fg=WHITE,
                 font=("Segoe UI",9,"bold")).pack(side="left",pady=5,padx=6)
        lbl=tk.Label(hdr,text="",bg=color,fg="#90CAF9",font=("Segoe UI",8))
        lbl.pack(side="right",padx=8,pady=5); setattr(self,label_attr,lbl)
        body=tk.Frame(outer,bg=CARD_BG,padx=14,pady=10,highlightbackground=SEP_CLR,highlightthickness=1)
        body.pack(fill="x"); setattr(self,body_attr,body)

    def _build_preview(self,body_attr,label_attr,rows,tickets,now):
        body=getattr(self,body_attr); lbl=getattr(self,label_attr)
        for w in body.winfo_children(): w.destroy()
        summ=tk.Frame(body,bg=CARD_BG); summ.pack(fill="x",pady=(0,8))
        def ibox(parent,ltext,val,col):
            f=tk.Frame(parent,bg="#F0F7FF",padx=10,pady=6,highlightbackground=SEP_CLR,highlightthickness=1)
            f.grid(row=0,column=col,sticky="nsew",padx=(0,6)); parent.columnconfigure(col,weight=1)
            tk.Label(f,text=ltext,bg="#F0F7FF",fg="#8096B0",font=("Segoe UI",7,"bold")).pack(anchor="w")
            tk.Label(f,text=val,bg="#F0F7FF",fg=FBC_DARK,font=("Segoe UI",11,"bold")).pack(anchor="w")
        ibox(summ,"ROWS",str(len(rows)),0); ibox(summ,"DATE/TIME",now,1)
        ibox(summ,"TICKET RANGE",f"{tickets[0]} to {tickets[-1]}",2)
        style=ttk.Style()
        try: style.theme_use("default")
        except Exception: pass
        style.configure("Treeview.Heading",background=FBC_DARK,foreground=WHITE,relief="flat",font=("Segoe UI",8,"bold"))
        style.map("Treeview.Heading",background=[("active",FBC_MID)],foreground=[("active",WHITE)])
        style.configure("Treeview",font=("Segoe UI",8),rowheight=22)
        frm=tk.Frame(body,bg=CARD_BG); frm.pack(fill="x")
        xsb=ttk.Scrollbar(frm,orient="horizontal"); ysb=ttk.Scrollbar(frm,orient="vertical")
        tv=ttk.Treeview(frm,columns=PREVIEW_COLS,show="headings",height=min(len(rows),7),
                        xscrollcommand=xsb.set,yscrollcommand=ysb.set)
        xsb.config(command=tv.xview); ysb.config(command=tv.yview)
        for col in PREVIEW_COLS: tv.heading(col,text=col); tv.column(col,width=95,minwidth=70,anchor="w")
        for i,row in enumerate(rows):
            vals=[row.get(c,"") for c in PREVIEW_COLS]
            bs=row.get("Buy/Sell","").strip().lower()
            tag="buy" if bs=="buy" else("sell" if bs=="sell" else("even" if i%2==0 else "odd"))
            tv.insert("","end",values=vals,tags=(tag,))
        tv.tag_configure("even",background="#F7FAFF"); tv.tag_configure("odd",background=CARD_BG)
        tv.tag_configure("buy",foreground=GREEN_DARK,background="#F2FBF5")
        tv.tag_configure("sell",foreground=RED_DARK,background="#FFF5F5")
        tv.grid(row=0,column=0,sticky="nsew"); ysb.grid(row=0,column=1,sticky="ns")
        xsb.grid(row=1,column=0,sticky="ew"); frm.columnconfigure(0,weight=1)
        lbl.config(text=f"showing {len(rows)} rows")

    def _pick_file(self):
        path=filedialog.askopenfilename(title="Select First Exchange Matched Trades File",
            filetypes=[("CSV / Excel","*.csv *.xlsx *.xls"),("All files","*.*")])
        if path: self._load_file(path)

    def _load_file(self,path):
        try:
            pd=_require("pandas")
            df=pd.read_csv(path,dtype=str).fillna("") if path.lower().endswith(".csv") else pd.read_excel(path,dtype=str).fillna("")
            self.raw_headers=list(df.columns); self.raw_rows=df.to_dict("records")
            if not self.raw_rows: raise ValueError("File is empty.")
            self.source_path=path
            self.conv_rows, now = transform_rows(self.raw_rows)
            tickets = [r.get("Ticket No.", "").lstrip("0") or "0" for r in self.conv_rows]
            self.gen_csv = self.gen_pdf = self.gen_mt_xlsx = self.gen_anesu_xlsx = None
            self.lbl_csv_done.config(text=""); self.lbl_pdf_done.config(text="")
            self.btn_csv.config(text="Download CSV",bg=FBC_MID)
            self.btn_pdf.config(text="Download PDF",bg=RED_DARK)
            fname=os.path.basename(path); exch=get_exchange(self.raw_rows[0].get("Market",""))
            for w in self.info_bar1.winfo_children(): w.pack_forget()
            self.info_bar1.pack(fill="x",pady=(0,6))
            tk.Label(self.info_bar1,text="OK",bg=TAG_BLUE,font=("Segoe UI",10)).pack(side="left",padx=(6,2),pady=4)
            self.lbl_file1.config(text=fname); self.lbl_file1.pack(side="left",pady=4)
            self.lbl_rows1.config(text=f"  {len(self.conv_rows)} rows  |  {tickets[0]} to {tickets[-1]}")
            self.lbl_rows1.pack(side="left",pady=4); self.btn_reupload1.pack(side="right",padx=6,pady=4)
            for b in (self.btn_csv,self.btn_pdf): b.config(state="normal")
            self.btn_email.config(text=f"Send — {exch} Only",state="normal")
            self.prev_outer1.pack(fill="x",padx=12,pady=(4,0))
            self._build_preview("prev_body1","lbl_showing1",self.conv_rows,tickets,now)
            if self.source_path2: self.btn_email_both.config(state="normal")
        except Exception as e: messagebox.showerror("Error loading file",str(e))

    def _dl_csv(self):
        try:
            exch=get_exchange(self.raw_rows[0].get("Market",""))
            self.gen_csv=generate_csv(self.conv_rows,self.out_dir,exch)
            self.lbl_csv_done.config(text=f"OK  {os.path.basename(self.gen_csv)} saved")
            self.btn_csv.config(text="CSV Downloaded",bg="#1B5E20")
        except Exception as e: messagebox.showerror("CSV Error",str(e))

    def _dl_pdf(self):
        try:
            self.gen_pdf=generate_pdf(self.raw_rows,self.raw_headers,self.out_dir)
            self.lbl_pdf_done.config(text=f"OK  {os.path.basename(self.gen_pdf)} saved")
            self.btn_pdf.config(text="PDF Downloaded",bg="#7B1010")
        except Exception as e: messagebox.showerror("PDF Error",str(e))

    def _ensure_email_files(self):
        if not self.gen_pdf:
            self.gen_pdf = generate_pdf(self.raw_rows, self.raw_headers, self.out_dir)
            self.lbl_pdf_done.config(text=f"OK  {os.path.basename(self.gen_pdf)} saved")
        if not self.gen_mt_xlsx:
            self.gen_mt_xlsx = generate_matched_excel(self.source_path, self.raw_rows, self.out_dir)
        if not self.gen_anesu_xlsx:
            self.gen_anesu_xlsx = generate_anesu_excel(self.raw_rows, self.out_dir)

    def _send_email(self):
        try:
            self._ensure_email_files()
            open_sarestock_outlook([self.gen_pdf, self.gen_mt_xlsx, self.gen_anesu_xlsx],
                                   sender_name=load_sender_name())
        except ImportError: messagebox.showerror("pywin32 not installed","Run:  pip install pywin32")
        except Exception as e: messagebox.showerror("Outlook Error",str(e))

    def _pick_file2(self):
        path=filedialog.askopenfilename(title="Select Second Exchange Matched Trades File",
            filetypes=[("CSV / Excel","*.csv *.xlsx *.xls"),("All files","*.*")])
        if path: self._load_file2(path)

    def _load_file2(self,path):
        try:
            pd=_require("pandas")
            df=pd.read_csv(path,dtype=str).fillna("") if path.lower().endswith(".csv") else pd.read_excel(path,dtype=str).fillna("")
            self.raw_headers2=list(df.columns); self.raw_rows2=df.to_dict("records")
            if not self.raw_rows2: raise ValueError("File is empty.")
            self.source_path2=path
            self.conv_rows2, now = transform_rows(self.raw_rows2)
            tickets = [r.get("Ticket No.", "").lstrip("0") or "0" for r in self.conv_rows2]
            self.gen_csv2 = self.gen_pdf2 = self.gen_mt_xlsx2 = self.gen_anesu_xlsx2 = None
            self.lbl_csv2_done.config(text=""); self.lbl_pdf2_done.config(text="")
            self.btn_csv2.config(text="Download CSV",bg=FBC_MID)
            self.btn_pdf2.config(text="Download PDF",bg=RED_DARK)
            fname=os.path.basename(path); exch2=get_exchange(self.raw_rows2[0].get("Market",""))
            for w in self.info_bar2.winfo_children(): w.pack_forget()
            self.info_bar2.pack(fill="x",pady=(0,6))
            tk.Label(self.info_bar2,text="OK",bg=TAG_BLUE,font=("Segoe UI",10)).pack(side="left",padx=(6,2),pady=4)
            self.lbl_file2.config(text=fname); self.lbl_file2.pack(side="left",pady=4)
            self.lbl_rows2.config(text=f"  {len(self.conv_rows2)} rows  |  {tickets[0]} to {tickets[-1]}")
            self.lbl_rows2.pack(side="left",pady=4); self.btn_reupload2.pack(side="right",padx=6,pady=4)
            for b in (self.btn_csv2,self.btn_pdf2): b.config(state="normal")
            self.btn_email2.config(text=f"Send — {exch2} Only",state="normal")
            self.prev_outer2.pack(fill="x",padx=12,pady=(4,0))
            self._build_preview("prev_body2","lbl_showing2",self.conv_rows2,tickets,now)
            if self.source_path: self.btn_email_both.config(state="normal")
        except Exception as e: messagebox.showerror("Error loading 2nd file",str(e))

    def _dl_csv2(self):
        try:
            exch2=get_exchange(self.raw_rows2[0].get("Market",""))
            self.gen_csv2=generate_csv(self.conv_rows2,self.out_dir,exch2)
            self.lbl_csv2_done.config(text=f"OK  {os.path.basename(self.gen_csv2)} saved")
            self.btn_csv2.config(text="CSV Downloaded",bg="#1B5E20")
        except Exception as e: messagebox.showerror("CSV Error (2nd)",str(e))

    def _dl_pdf2(self):
        try:
            self.gen_pdf2=generate_pdf(self.raw_rows2,self.raw_headers2,self.out_dir)
            self.lbl_pdf2_done.config(text=f"OK  {os.path.basename(self.gen_pdf2)} saved")
            self.btn_pdf2.config(text="PDF Downloaded",bg="#7B1010")
        except Exception as e: messagebox.showerror("PDF Error (2nd)",str(e))

    def _ensure_email_files2(self):
        if not self.gen_pdf2:
            self.gen_pdf2 = generate_pdf(self.raw_rows2, self.raw_headers2, self.out_dir)
            self.lbl_pdf2_done.config(text=f"OK  {os.path.basename(self.gen_pdf2)} saved")
        if not self.gen_mt_xlsx2:
            self.gen_mt_xlsx2 = generate_matched_excel(self.source_path2, self.raw_rows2, self.out_dir)
        if not self.gen_anesu_xlsx2:
            self.gen_anesu_xlsx2 = generate_anesu_excel(self.raw_rows2, self.out_dir)

    def _send_email2(self):
        try:
            self._ensure_email_files2()
            open_sarestock_outlook([self.gen_pdf2, self.gen_mt_xlsx2, self.gen_anesu_xlsx2],
                                   sender_name=load_sender_name())
        except ImportError: messagebox.showerror("pywin32 not installed","Run:  pip install pywin32")
        except Exception as e: messagebox.showerror("Outlook Error",str(e))

    def _send_email_both(self):
        try:
            self._ensure_email_files(); self._ensure_email_files2()
            open_sarestock_outlook([self.gen_pdf, self.gen_mt_xlsx, self.gen_anesu_xlsx,
                                    self.gen_pdf2, self.gen_mt_xlsx2, self.gen_anesu_xlsx2],
                                   sender_name=load_sender_name())
        except ImportError: messagebox.showerror("pywin32 not installed","Run:  pip install pywin32")
        except Exception as e: messagebox.showerror("Outlook Error",str(e))

    def _pick_outdir(self):
        d=filedialog.askdirectory(title="Choose output folder",initialdir=self.out_dir)
        if d: self.out_dir=d; self.lbl_outdir.config(text=d)


# ════════════════════════════════════════════════════════════════════════════
#  EMAILER PAGE  (custodian-only — client email removed)
# ════════════════════════════════════════════════════════════════════════════
class EmailerPage(tk.Frame):
    def __init__(self,parent):
        super().__init__(parent,bg=BG)
        self.deal_items=[]; self.pdf_folder=""
        self.sender_name = load_sender_name()
        self.sent_custodians = set()
        self._build()

    def _build(self):
        bar=tk.Frame(self,bg=FBC_MID,padx=16,pady=8); bar.pack(fill="x")
        tk.Label(bar,text="Deal Note Email Automator",bg=FBC_MID,fg=WHITE,
                 font=("Segoe UI",11,"bold")).pack(side="left")

        name_bar = tk.Frame(self, bg=FBC_DARK, padx=16, pady=6)
        name_bar.pack(fill="x")
        tk.Label(name_bar, text="Your Name (used in email sign-off):",
                 bg=FBC_DARK, fg=SIDEBAR_TEXT, font=("Segoe UI", 9, "bold")).pack(side="left")
        self._sender_var = tk.StringVar(value=self.sender_name)
        name_entry = tk.Entry(name_bar, textvariable=self._sender_var,
                              font=("Segoe UI", 10), bg="#0D2B4E", fg=WHITE,
                              insertbackground=WHITE, relief="flat",
                              highlightbackground=FBC_MID, highlightthickness=1,
                              width=26)
        name_entry.pack(side="left", padx=(8, 6), ipady=4)
        tk.Button(name_bar, text="Save Name", command=self._save_sender_name,
                  bg=GREEN_DARK, fg=WHITE, relief="flat",
                  font=("Segoe UI", 9, "bold"), cursor="hand2",
                  padx=10, pady=4).pack(side="left")
        self.lbl_name_saved = tk.Label(name_bar, text="", bg=FBC_DARK, fg="#90EE90",
                                       font=("Segoe UI", 8))
        self.lbl_name_saved.pack(side="left", padx=8)
        self.lbl_name_hint = tk.Label(name_bar,
            text=f"Saved: {self.sender_name}" if self.sender_name else "No name saved yet",
            bg=FBC_DARK,
            fg="#90CAF9" if self.sender_name else "#FF9966",
            font=("Segoe UI", 8))
        self.lbl_name_hint.pack(side="right")

        fp=tk.Frame(self,bg=WHITE,padx=16,pady=12); fp.pack(fill="x",padx=16,pady=(12,0))
        btn_row=tk.Frame(fp,bg=WHITE); btn_row.pack(fill="x")
        tk.Button(btn_row,text="Choose Deal Notes Folder",command=self._pick_folder,
                  bg=FBC_MID,fg=WHITE,relief="flat",font=("Segoe UI",10,"bold"),
                  cursor="hand2",padx=14,pady=8).pack(side="left")
        tk.Label(btn_row,text="  or  ",bg=WHITE,fg="#8096B0",font=("Segoe UI",9)).pack(side="left")
        tk.Button(btn_row,text="Select Individual Deal Note(s)",command=self._pick_individual_files,
                  bg="#4051B5",fg=WHITE,relief="flat",font=("Segoe UI",10,"bold"),
                  cursor="hand2",padx=14,pady=8).pack(side="left")
        self.btn_clear=tk.Button(btn_row,text="Clear All Uploads",command=self._clear_uploads,
                  bg=RED_DARK,fg=WHITE,relief="flat",font=("Segoe UI",9,"bold"),
                  cursor="hand2",padx=10,pady=8,state="disabled")
        self.btn_clear.pack(side="right")
        info_row=tk.Frame(fp,bg=WHITE); info_row.pack(fill="x",pady=(6,0))
        self.lbl_folder=tk.Label(info_row,text="No files loaded",bg=WHITE,fg="#8096B0",font=("Segoe UI",9))
        self.lbl_folder.pack(side="left")
        self.lbl_found=tk.Label(info_row,text="",bg=WHITE,fg=FBC_MID,font=("Consolas",9))
        self.lbl_found.pack(side="left",padx=10)

        self._build_custodian_tab()

    def _build_custodian_tab(self):
        p=self
        tk.Label(p,text="Groups all PDFs by custodian — one email per custodian with all their deal notes attached.",
                 bg=BG,fg="#607080",font=("Segoe UI",9)).pack(anchor="w",padx=16,pady=(10,0))
        self.btn_send_all_cust=tk.Button(p,text="Send ALL Custodian Emails",
            command=self._cust_send_all,bg=GREEN_DARK,fg=WHITE,font=("Segoe UI",11,"bold"),
            relief="flat",padx=16,pady=10,cursor="hand2",state="disabled")
        self.btn_send_all_cust.pack(fill="x",padx=16,pady=(8,4))
        self.lbl_cust_hint=tk.Label(p,text="Load a folder above to begin.",bg=BG,fg="#607080",font=("Segoe UI",9))
        self.lbl_cust_hint.pack(anchor="w",padx=16,pady=(0,8))
        canvas=tk.Canvas(p,bg=BG,highlightthickness=0)
        sb=ttk.Scrollbar(p,orient="vertical",command=canvas.yview)
        self.cust_body=tk.Frame(canvas,bg=BG)
        self.cust_body.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0),window=self.cust_body,anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left",fill="both",expand=True,padx=(16,0))
        sb.pack(side="right",fill="y",pady=4)

    def _render_custodian_tab(self):
        for w in self.cust_body.winfo_children(): w.destroy()
        self.cust_btn_map={}
        groups={}
        for item in self.deal_items: groups.setdefault(item["custodian"],[]).append(item)
        for code,items in sorted(groups.items()):
            routing = get_effective_custodian_routing(code)
            sent = code in self.sent_custodians
            head_color = "#1A6B3A" if sent else (FBC_MID if routing else RED_DARK)
            card_bg    = "#F0FBF4" if sent else WHITE
            card=tk.Frame(self.cust_body,bg=card_bg,pady=0,padx=0)
            card.pack(fill="x",padx=4,pady=(0,10))
            head=tk.Frame(card,bg=head_color,pady=7,padx=12); head.pack(fill="x")
            label=routing["label"] if routing else "UNKNOWN CUSTODIAN"
            count=len(items)
            status_badge = "  ✓ SENT" if sent else ""
            tk.Label(head,text=f"{code}  —  {label}{status_badge}",bg=head_color,fg=WHITE,
                     font=("Segoe UI",10,"bold")).pack(side="left")
            tk.Label(head,text=f"{count} deal note{'s' if count>1 else ''}",
                     bg=head_color,fg=WHITE,font=("Segoe UI",9)).pack(side="right")
            inner=tk.Frame(card,bg=card_bg,padx=12,pady=8); inner.pack(fill="x")
            for it in items:
                tk.Label(inner,text=f"  {it['fname']}",bg=card_bg,fg="#2D3748",font=("Segoe UI",9)).pack(anchor="w")
            if routing:
                subj=f"DEAL NOTE{'S' if count>1 else ''} - {datetime.now().strftime('%d %b %Y')}"
                tk.Label(inner,text=f"Subject: {subj}",bg=card_bg,fg="#607080",
                         font=("Segoe UI",8,"italic")).pack(anchor="w",pady=(6,0))
                tk.Label(inner,text=f"To: {'; '.join(routing['to'])}",bg=card_bg,
                         fg="#607080",font=("Segoe UI",8)).pack(anchor="w")
                cc_summary = f"CC: {len(routing['cc'])} address{'es' if len(routing['cc'])!=1 else ''}"
                tk.Label(inner,text=cc_summary,bg=card_bg,fg="#8096B0",font=("Segoe UI",8)).pack(anchor="w")
                action_row = tk.Frame(inner, bg=card_bg)
                action_row.pack(anchor="w", pady=(8,0))
                btn_text = f"✓ Sent  ({count} file{'s' if count>1 else ''})" if sent \
                           else f"Open in Outlook  ({count} file{'s' if count>1 else ''})"
                btn_bg   = "#2E7D32" if sent else FBC_MID
                btn=tk.Button(action_row, text=btn_text,
                    command=lambda c=code: self._cust_send_one(c),
                    bg=btn_bg, fg=WHITE, relief="flat",
                    font=("Segoe UI",9,"bold"),
                    cursor="arrow" if sent else "hand2",
                    state="disabled" if sent else "normal",
                    disabledforeground=WHITE,
                    padx=10, pady=6)
                btn.pack(side="left", padx=(0, 8))
                self.cust_btn_map[code] = btn
                tk.Button(action_row, text="⚙ Edit Recipients",
                          command=lambda c=code: self._edit_custodian_recipients(c),
                          bg=FBC_ACCENT, fg=WHITE, relief="flat",
                          font=("Segoe UI",8,"bold"), cursor="hand2",
                          padx=8, pady=6).pack(side="left")
            else:
                tk.Label(inner,text="No routing configured for this custodian code.",
                         bg=card_bg,fg=RED_DARK,font=("Segoe UI",9)).pack(anchor="w")

        known  = sum(1 for c in groups if c in CUSTODIAN_ROUTING)
        unsent = sum(1 for c in groups if c in CUSTODIAN_ROUTING and c not in self.sent_custodians)
        self.btn_send_all_cust.config(
            state="normal" if unsent else "disabled",
            text=f"Send ALL {known} Custodian Email{'s' if known!=1 else ''} in Outlook"
                 + (f"  ({known - unsent} already sent)" if known > unsent else ""))
        self.lbl_cust_hint.config(
            text=f"  {len(self.deal_items)} deal note(s) across {len(groups)} custodian(s).",
            fg=GREEN_DARK)

    def _edit_custodian_recipients(self, code):
        routing = get_effective_custodian_routing(code)
        if not routing:
            messagebox.showwarning("Unknown", f"No default routing for {code}.", parent=self)
            return
        def on_save(new_to, new_cc):
            overrides = load_custodian_overrides()
            overrides[code] = {"to": new_to or routing["to"], "cc": new_cc}
            save_custodian_overrides(overrides)
            self._render_custodian_tab()
            messagebox.showinfo("Saved",
                f"Recipients for {code} updated.\nNew addresses will be used on the next send.",
                parent=self)
        RecipientsDialog(
            self,
            title=f"⚙  {routing['label']} — Edit Recipients",
            to_list=routing["to"],
            cc_list=routing["cc"],
            on_save=on_save,
            to_label=f"To  (primary recipients for {code})",
            cc_label=f"CC  (copied recipients for {code})",
        )

    def _cust_send_one(self,code):
        routing = get_effective_custodian_routing(code)
        if not routing: messagebox.showwarning("Unknown",f"No routing for {code}."); return
        items=[it for it in self.deal_items if it["custodian"]==code]
        count=len(items)
        subj=f"DEAL NOTE{'S' if count>1 else ''} - {datetime.now().strftime('%d %b %Y')}"
        body=get_custodian_body(multi=(count>1), sender_name=self._get_sender())
        try:
            open_outlook(routing["to"],routing["cc"],subj,body,[it["path"] for it in items])
            self.sent_custodians.add(code)
            if code in self.cust_btn_map:
                self.cust_btn_map[code].config(text="✓ Sent",bg="#2E7D32")
                self.after(300, self._render_custodian_tab)
        except ImportError as e: messagebox.showerror("pywin32 not installed",str(e))
        except Exception as e: messagebox.showerror("Outlook Error",str(e))

    def _cust_send_all(self):
        codes=sorted(set(it["custodian"] for it in self.deal_items
                         if it["custodian"] in CUSTODIAN_ROUTING
                         and it["custodian"] not in self.sent_custodians))
        if not codes: messagebox.showinfo("Nothing","No unsent custodian emails found."); return
        if not messagebox.askyesno("Send All",f"Open {len(codes)} Outlook window(s), one per custodian?\n\nContinue?"): return
        for code in codes: self._cust_send_one(code)
        messagebox.showinfo("Done",f"  {len(codes)} Outlook window(s) opened.")

    def _save_sender_name(self):
        name = self._sender_var.get().strip()
        if not name:
            messagebox.showwarning("Empty Name","Please enter your name before saving.", parent=self)
            return
        self.sender_name = name
        save_sender_name(name)
        self.lbl_name_saved.config(text="Saved!")
        self.lbl_name_hint.config(text=f"Saved: {name}", fg="#90CAF9")
        self.after(2500, lambda: self.lbl_name_saved.config(text=""))

    def _get_sender(self):
        return self._sender_var.get().strip() or self.sender_name or "FBC Securities"

    def _pick_folder(self):
        folder=filedialog.askdirectory(title="Select folder containing deal note PDFs")
        if not folder: return
        pdfs=sorted(f for f in os.listdir(folder) if f.lower().endswith(".pdf"))
        if not pdfs: messagebox.showwarning("No PDFs","No PDF files found in that folder."); return
        self.pdf_folder=folder
        self.lbl_folder.config(text=f"  {os.path.basename(folder)}",fg=FBC_DARK)
        self.lbl_found.config(text=f"Scanning {len(pdfs)} PDF(s)...")
        self._disable_send_buttons()
        self.btn_clear.config(state="normal")
        threading.Thread(target=self._scan,args=(folder,pdfs),daemon=True).start()

    def _pick_individual_files(self):
        paths=filedialog.askopenfilenames(
            title="Select Deal Note PDF(s)",
            filetypes=[("PDF files","*.pdf"),("All files","*.*")])
        if not paths: return
        pdf_paths=list(paths)
        already={it["path"] for it in self.deal_items}
        new_paths=[p for p in pdf_paths if p not in already]
        if not new_paths:
            messagebox.showinfo("No New Files","All selected files are already loaded."); return
        self.lbl_found.config(text=f"Scanning {len(new_paths)} new PDF(s)...")
        self.btn_clear.config(state="normal")
        names=", ".join(os.path.basename(p) for p in new_paths[:3])
        if len(new_paths)>3: names+=f" +{len(new_paths)-3} more"
        self.lbl_folder.config(text=f"  {names}",fg=FBC_DARK)
        threading.Thread(target=self._scan_files,args=(new_paths,),daemon=True).start()

    def _clear_uploads(self):
        if not self.deal_items: return
        if not messagebox.askyesno("Clear All Uploads",
            f"Remove all {len(self.deal_items)} loaded deal note(s) and start fresh?\n\n"
            "This does NOT delete the files from disk."): return
        self.deal_items=[]; self.pdf_folder=""
        self.sent_custodians.clear()
        self.lbl_folder.config(text="No files loaded",fg="#8096B0")
        self.lbl_found.config(text="")
        self.btn_clear.config(state="disabled")
        self._disable_send_buttons()
        for w in self.cust_body.winfo_children(): w.destroy()
        self.lbl_cust_hint.config(text="Load files above to begin.",fg="#607080")
        self.btn_send_all_cust.config(state="disabled",text="Send ALL Custodian Emails")

    def _disable_send_buttons(self):
        self.btn_send_all_cust.config(state="disabled")

    def _scan(self,folder,pdfs):
        items=[]
        for fname in pdfs:
            path=os.path.join(folder,fname)
            items.append({
                "fname": fname, "path": path,
                "custodian": parse_custodian_from_pdf(path) or "UNKNOWN",
                "deal_info": parse_deal_info_from_pdf(path), "sent": False,
            })
        self.deal_items=items
        self._after_scan(len(items))

    def _scan_files(self,paths):
        new_items=[]
        for path in paths:
            fname=os.path.basename(path)
            new_items.append({
                "fname": fname, "path": path,
                "custodian": parse_custodian_from_pdf(path) or "UNKNOWN",
                "deal_info": parse_deal_info_from_pdf(path), "sent": False,
            })
        self.deal_items.extend(new_items)
        self._after_scan(len(self.deal_items))

    def _after_scan(self,total):
        self.after(0,self._render_custodian_tab)
        self.after(0,lambda:self.lbl_found.config(text=f"  {total} PDF(s) loaded"))


# ════════════════════════════════════════════════════════════════════════════
#  VOICE ENGINE
# ════════════════════════════════════════════════════════════════════════════
_tts_lock = threading.Lock()

def speak(text: str):
    global _tts
    if not _VOICE_READY:
        return
    try:
        if _tts is None:
            import pyttsx3
            _tts = pyttsx3.init()
            _tts.setProperty("rate", 165)
        def _run():
            with _tts_lock:
                try:
                    _tts.say(text)
                    _tts.runAndWait()
                except Exception:
                    pass
        threading.Thread(target=_run, daemon=True).start()
    except Exception:
        pass

class VoiceBar(tk.Frame):
    MIC_IDLE   = ("🎤  Hold to speak", FBC_MID,   WHITE)
    MIC_LISTEN = ("🔴  Listening…",   "#B71C1C",  WHITE)
    MIC_THINK  = ("⏳  Processing…",  "#555555",  WHITE)
    MIC_NODEPS = ("🎤  Voice (install deps)", "#2A4A6A", SIDEBAR_TEXT)

    def __init__(self, parent, dispatch_cb, hotkey_widget=None):
        super().__init__(parent, bg=SIDEBAR_BG)
        self._cb     = dispatch_cb
        self._active = False
        self._build()
        if hotkey_widget:
            hotkey_widget.bind_all("<Control-space>", lambda e: self._toggle())

    def _build(self):
        tk.Frame(self, bg=FBC_MID, height=1).pack(fill="x", padx=10, pady=(8,4))
        tk.Label(self, text="VOICE ASSISTANT", bg=SIDEBAR_BG,
                 fg="#2A4A6A", font=("Segoe UI",7,"bold")).pack()
        self.btn = tk.Button(
            self,
            text=self.MIC_IDLE[0]   if _VOICE_READY else self.MIC_NODEPS[0],
            bg=self.MIC_IDLE[1]     if _VOICE_READY else self.MIC_NODEPS[1],
            fg=self.MIC_IDLE[2]     if _VOICE_READY else self.MIC_NODEPS[2],
            relief="flat", font=("Segoe UI",9,"bold"),
            cursor="hand2" if _VOICE_READY else "arrow",
            padx=6, pady=10, wraplength=170,
            command=self._toggle if _VOICE_READY else self._show_install,
        )
        self.btn.pack(fill="x", padx=10, pady=4)
        self.lbl_heard = tk.Label(self, text="", bg=SIDEBAR_BG, fg="#90CAF9",
                                  font=("Segoe UI",8), wraplength=170, justify="left")
        self.lbl_heard.pack(fill="x", padx=10)
        tk.Label(self, text="Ctrl+Space to activate",
                 bg=SIDEBAR_BG, fg="#2A4A6A", font=("Segoe UI",7)).pack(pady=(2,6))

    def _set_state(self, s):
        self.btn.config(text=s[0], bg=s[1], fg=s[2])

    def _show_install(self):
        messagebox.showinfo("Voice — Install Required",
            "To enable voice control, open a terminal and run:\n\n"
            "  pip install SpeechRecognition pyttsx3 pyaudio\n\n"
            "If pyaudio fails:\n"
            "  pip install pipwin\n"
            "  pipwin install pyaudio\n\n"
            "Then restart FBC Suite.")

    def _toggle(self):
        if self._active: return
        self._active = True
        self._set_state(self.MIC_LISTEN)
        threading.Thread(target=self._listen, daemon=True).start()

    def _listen(self):
        r = _sr.Recognizer(); mic = _sr.Microphone(); text = ""
        try:
            with mic as source:
                r.adjust_for_ambient_noise(source, duration=0.3)
                audio = r.listen(source, timeout=6, phrase_time_limit=8)
            self.after(0, lambda: self._set_state(self.MIC_THINK))
            text = r.recognize_google(audio).lower().strip()
        except _sr.WaitTimeoutError: text = ""
        except _sr.UnknownValueError: text = ""
        except Exception as exc:
            text = ""
            self.after(0, lambda: self.lbl_heard.config(text=f"Error: {exc}"))
        finally:
            self._active = False
            self.after(0, lambda: self._set_state(self.MIC_IDLE))
        if text:
            self.after(0, lambda t=text: self._on_heard(t))

    def _on_heard(self, text: str):
        self.lbl_heard.config(text=f'"{text}"')
        self._cb(text)
        self.after(6000, lambda: self.lbl_heard.config(text=""))


# ════════════════════════════════════════════════════════════════════════════
#  MAIN APP SHELL
# ════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"FBC Suite  v{VERSION}")
        self.state("zoomed")
        self.configure(bg=SIDEBAR_BG)
        self._active_page = None
        self._build()

    def _build(self):
        sidebar = tk.Frame(self, bg=SIDEBAR_BG, width=210)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        logo = tk.Frame(sidebar, bg=SIDEBAR_BG, pady=20)
        logo.pack(fill="x")
        tk.Label(logo, text="FBC", bg=FBC_ACCENT, fg=WHITE,
                 font=("Segoe UI",16,"bold"), padx=10, pady=6).pack()
        tk.Label(logo, text="Suite", bg=SIDEBAR_BG, fg=SIDEBAR_TEXT,
                 font=("Segoe UI",10)).pack(pady=(4,0))
        tk.Frame(sidebar, bg=FBC_MID, height=1).pack(fill="x", padx=16, pady=(0,10))
        self.nav_buttons = {}
        for label, key in [("Converter","converter"),("Deal Note\nEmailer","emailer")]:
            btn = tk.Button(sidebar, text=label,
                command=lambda k=key: self._switch(k),
                bg=SIDEBAR_BG, fg=SIDEBAR_TEXT,
                activebackground=SIDEBAR_ACTIVE, activeforeground=WHITE,
                font=("Segoe UI",10,"bold"), relief="flat",
                cursor="hand2", pady=16, width=16, justify="center")
            btn.pack(fill="x", padx=8, pady=2)
            btn.bind("<Enter>", lambda e,b=btn,k=key: b.config(
                bg=SIDEBAR_ACTIVE if self._active_page==k else SIDEBAR_HOVER))
            btn.bind("<Leave>", lambda e,b=btn,k=key: b.config(
                bg=SIDEBAR_ACTIVE if self._active_page==k else SIDEBAR_BG))
            self.nav_buttons[key] = btn
        self.voice_bar = VoiceBar(sidebar, self._voice_dispatch, hotkey_widget=self)
        self.voice_bar.pack(side="bottom", fill="x")
        tk.Label(sidebar, text=f"v{VERSION}", bg=SIDEBAR_BG, fg="#2A4A6A",
                 font=("Segoe UI",8)).pack(side="bottom", pady=4)
        self.content = tk.Frame(self, bg=BG)
        self.content.pack(side="left", fill="both", expand=True)
        self.pages = {
            "converter": SarestockPage(self.content),
            "emailer":   EmailerPage(self.content),
        }
        self._switch("converter")

    def _voice_dispatch(self, text: str):
        t = text.lower()
        if any(w in t for w in ("converter","convert","sarestock","first tab")):
            self._switch("converter"); speak("Switched to Converter."); return
        if any(w in t for w in ("email","emailer","deal note","second tab")):
            self._switch("emailer"); speak("Switched to Deal Note Emailer."); return
        conv = self.pages["converter"]
        if any(w in t for w in ("browse","upload file","load file","open file","first exchange","pick file")):
            self._switch("converter"); speak("Opening file browser."); conv._pick_file(); return
        if any(w in t for w in ("second exchange","second file","upload second","load second","vfex file","pick second")):
            self._switch("converter"); speak("Opening second file browser."); conv._pick_file2(); return
        if any(w in t for w in ("download csv","save csv","get csv")):
            self._switch("converter")
            if conv.conv_rows: speak("Downloading CSV."); conv._dl_csv()
            elif conv.conv_rows2: speak("Downloading second CSV."); conv._dl_csv2()
            else: speak("No file loaded yet.")
            return
        if any(w in t for w in ("download pdf","save pdf","get pdf")):
            self._switch("converter")
            if conv.conv_rows: speak("Downloading PDF."); conv._dl_pdf()
            elif conv.conv_rows2: speak("Downloading second PDF."); conv._dl_pdf2()
            else: speak("No file loaded yet.")
            return
        if any(w in t for w in ("send zse","send z s e","zse email")):
            self._switch("converter"); speak("Opening ZSE email."); conv._send_email(); return
        if any(w in t for w in ("send vfex","vfex email")):
            self._switch("converter"); speak("Opening VFEX email."); conv._send_email2(); return
        if any(w in t for w in ("send both","both emails","zse and vfex","send everything converter")):
            self._switch("converter"); speak("Opening combined email."); conv._send_email_both(); return
        if any(w in t for w in ("clear converter","clear uploads converter")):
            self._switch("converter"); speak("Clearing converter uploads."); conv._clear_uploads(); return
        em = self.pages["emailer"]
        if any(w in t for w in ("load folder","pick folder","open folder","select folder","load pdfs","browse folder")):
            self._switch("emailer"); speak("Opening folder browser."); em._pick_folder(); return
        if any(w in t for w in ("load files","pick files","individual files","select files","browse files")):
            self._switch("emailer"); speak("Opening file picker."); em._pick_individual_files(); return
        if any(w in t for w in ("send all custodian","custodian emails","send custodians","all custodians")):
            self._switch("emailer"); speak("Sending all custodian emails."); em._cust_send_all(); return
        if any(w in t for w in ("clear emailer","clear deal","clear files emailer")):
            self._switch("emailer"); speak("Clearing loaded files."); em._clear_uploads(); return
        if any(w in t for w in ("help","commands","what can you do","what can i say")):
            speak("You can say: switch to converter, switch to emailer, browse file, download CSV, "
                  "download PDF, send ZSE, send VFEX, send both, load folder, send all custodians, "
                  "or clear files.")
            _VoiceHelpDialog(self); return
        if any(w in t for w in ("status","how many","how many files","what's loaded")):
            parts = []
            if conv.conv_rows: parts.append(f"{len(conv.conv_rows)} rows in first exchange")
            if conv.conv_rows2: parts.append(f"{len(conv.conv_rows2)} rows in second exchange")
            if em.deal_items: parts.append(f"{len(em.deal_items)} deal notes loaded")
            speak((", ".join(parts)+".") if parts else "Nothing loaded yet."); return
        speak("Sorry, I didn't catch that. Say 'help' for a list of commands.")

    def _switch(self, key):
        for page in self.pages.values(): page.pack_forget()
        self.pages[key].pack(fill="both", expand=True)
        self._active_page = key
        for k, btn in self.nav_buttons.items():
            btn.config(bg=SIDEBAR_ACTIVE if k==key else SIDEBAR_BG,
                       fg=WHITE         if k==key else SIDEBAR_TEXT)


class _VoiceHelpDialog(tk.Toplevel):
    COMMANDS = [
        ("Navigation",        ["switch to converter", "switch to emailer"]),
        ("Converter",         ["browse file / second file","download CSV / PDF",
                               "send ZSE / send VFEX / send both",
                               "clear converter"]),
        ("Deal Note Emailer", ["load folder / load files","send all custodians",
                               "clear emailer"]),
        ("General",           ["status — how many files loaded",
                               "help — show this dialog",
                               "Ctrl+Space — activate mic anywhere"]),
    ]
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Voice Commands — FBC Suite")
        self.configure(bg=SIDEBAR_BG)
        self.resizable(False, False)
        self.grab_set()
        tk.Label(self, text="🎤  Voice Commands", bg=FBC_ACCENT, fg=WHITE,
                 font=("Segoe UI",12,"bold"), pady=10).pack(fill="x")
        body = tk.Frame(self, bg=SIDEBAR_BG, padx=20, pady=14)
        body.pack(fill="both", expand=True)
        for section, cmds in self.COMMANDS:
            tk.Label(body, text=section.upper(), bg=SIDEBAR_BG, fg=FBC_ACCENT,
                     font=("Segoe UI",8,"bold")).pack(anchor="w", pady=(8,2))
            for c in cmds:
                tk.Label(body, text=f"  • {c}", bg=SIDEBAR_BG, fg=SIDEBAR_TEXT,
                         font=("Segoe UI",9)).pack(anchor="w")
        tk.Button(self, text="Close", command=self.destroy,
                  bg=FBC_MID, fg=WHITE, relief="flat",
                  font=("Segoe UI",10,"bold"), pady=8, cursor="hand2").pack(
                      fill="x", padx=20, pady=14)
        self.update_idletasks()
        w, h = 360, self.winfo_reqheight()
        x = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        y = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")


# ════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    check_and_apply_update()
    login = LoginDialog()
    login.mainloop()
    if not login.authenticated:
        sys.exit(0)
    app = App()
    app.mainloop()
